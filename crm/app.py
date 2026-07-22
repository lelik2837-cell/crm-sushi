import ast
import calendar
import csv
import io
import operator as _op
import os
import re
import sys
import logging
import smtplib
import secrets
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from functools import wraps

# Гарантируем что папка crm/ в пути — нужно для sber_api.py
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from datetime import date, datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, g
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
import sqlite3
import threading
import json as _json_lib

logging.basicConfig(level=logging.INFO)

app = Flask(__name__)
app.secret_key = 'sushi-crm-secret-2024-change-in-prod'
app.config['TEMPLATES_AUTO_RELOAD'] = True
# Railway и другие прокси-хосты: корректно определяем HTTPS и хост
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

DATABASE = os.environ.get('DATABASE_PATH', os.path.join(os.path.dirname(os.path.abspath(__file__)), 'crm.db'))

SMTP_HOST     = 'smtp.gmail.com'
SMTP_PORT     = 587
SMTP_USER     = 'papasushi42@gmail.com'
SMTP_PASSWORD = 'ciznzwgctjqpvbli'
SMTP_FROM     = 'CRMPAPA <papasushi42@gmail.com>'


@app.template_filter('datefmt')
def datefmt(val):
    """Конвертирует YYYY-MM-DD → DD-MM-YYYY для отображения."""
    if not val:
        return ''
    try:
        return datetime.strptime(str(val)[:10], '%Y-%m-%d').strftime('%d-%m-%Y')
    except Exception:
        return str(val)


ROLE_LABELS = {
    'admin': 'Администратор',
    'sushi': 'Сушист',
    'packer': 'Упаковщик',
    'courier': 'Курьер',
    'cleaner': 'Уборщица',
    'cook': 'Повар',
}


def _reload_role_labels(conn):
    global ROLE_LABELS
    rows = conn.execute(
        'SELECT code, name FROM positions WHERE is_active=1 ORDER BY sort_order, name'
    ).fetchall()
    if rows:
        ROLE_LABELS = {r['code']: r['name'] for r in rows}

# Hardcoded defaults — seeded into DB on first run
_DEFAULT_EXPENSE_CATEGORIES = [
    ('repair_plumbing', 'Ремонт сантех.', 1),
    ('repair_grease', 'Чистка жироуловителя', 2),
    ('repair_electric', 'Ремонт электрика', 3),
    ('repair_fridge', 'Ремонт холод.оборуд.', 4),
    ('repair_other', 'Ремонт другой', 5),
    ('shop', 'Магазин / Аптека', 6),
    ('staff', 'Стафф', 6.5),
    ('taxi', 'Такси', 7),
    ('cash_plus', 'Плюсы в кассу', 8),
    ('oil', 'За масло отработанное', 9),
    ('fish', 'Рыба (головы, хребты)', 10),
    ('change', 'Размен внёс Алексей', 11),
    ('other', 'Другое', 12),
]

ACTION_LABELS = {
    'revenue_update': ('Выручка',      'info'),
    'expense_add':    ('Расход +',     'success'),
    'expense_update': ('Расход ✎',     'warning'),
    'expense_delete': ('Расход −',     'danger'),
    'staff_add':      ('Сотрудник +',  'success'),
    'staff_update':   ('Сотрудник ✎',  'warning'),
    'staff_delete':   ('Сотрудник −',  'danger'),
    'shift_open':     ('Открытие',     'success'),
    'shift_close':    ('Закрытие',     'secondary'),
    'shift_reopen':   ('Переоткрытие', 'warning'),
    'salary_paid':    ('Выплата ЗП',   'primary'),
    'morning_cash_update': ('Утром в кассе', 'info'),
}

ROUTE_REASON_LABELS = {
    'delivery_run': 'Развоз',
    'relocation':   'Перемещение',
    'store':        'Заезд в магазин',
    'order':        'По заказу',
}

# Реестр пунктов меню для настраиваемых прав доступа по ролям (см. role_menu_permissions).
# (code, label, group ('dash'|'reports'|'settings'), branch_scoped)
MENU_ITEMS = [
    ('dashboard',             'Дашборд',                 'dash',     True),
    ('fot_dashboard',         'ФОТ',                      'dash',     True),
    ('lfl_dashboard',         'LFL',                      'dash',     True),
    ('ratings_dashboard',     'Рейтинги',                 'dash',     True),
    ('wait_dashboard',        'Ожидание',                 'dash',     True),
    ('promo_dashboard',       'Промокоды',                'dash',     True),
    ('shifts_archive',        'Смены',                    'reports',  True),
    ('reports_shifts',        'Выручка',                  'reports',  True),
    ('reports_salary',        'Зарплаты',                 'reports',  True),
    ('expenses_report',       'Другие расходы',           'reports',  True),
    ('cash_flow_report',      'Движение наличных',        'reports',  True),
    ('wait_time_report',      'Время ожидания',           'reports',  True),
    ('report_reconciliation', 'Сверка итогов',            'reports',  True),
    ('reconciliation_cashless', 'Сверка безнала',         'reports',  True),
    ('pnl_report',            'P&L отчёт',                'reports',  True),
    ('change_settings',       'Размен',                   'reports',  True),
    ('purchases',             'Накладные',                'reports',  True),
    ('bank',                  'Банк',                     'reports',  True),
    ('call_center',           'Колл-центр',               'reports',  False),
    ('contact_center_report', 'Контакт-центр',            'reports',  False),
    ('employees',             'Сотрудники',               'settings', True),
    ('history',               'История изменений',        'settings', True),
    ('import_shifts',         'Импорт смен',              'settings', False),
    ('orders_report',         'Отчёт по заказам',         'settings', False),
    ('branches',              'Филиалы',                  'settings', False),
    ('users',                 'Пользователи',             'settings', False),
    ('settings_rates',        'Ставки сотрудников',       'settings', False),
    ('settings_api',          'Интеграция 1С',            'settings', False),
    ('settings_categories',   'Категории расходов',       'settings', False),
    ('settings_bonuses',      'Премии',                   'settings', False),
    ('flyer_promo_settings',  'Листовка (промокод)',      'settings', False),
    ('revenue_manual',        'Старая выручка',           'settings', False),
    ('gsheet_settings',       'Экспорт в Google Sheets',  'settings', False),
]
MENU_ITEMS_BY_CODE = {m[0]: {'label': m[1], 'group': m[2], 'branch_scoped': m[3]} for m in MENU_ITEMS}
MENU_GROUP_LABELS = {'dash': 'Дашборд', 'reports': 'Отчёты', 'settings': 'Настройки'}

# Подпункты (вкладки) внутри некоторых пунктов меню — на «Роли и доступ» можно
# либо отметить весь раздел целиком (галочка у самого пункта в MENU_ITEMS —
# тогда доступны все вкладки), либо точечно выдать доступ на отдельные вкладки.
# (parent_code -> [(subitem_code, label, branch_scoped), ...])
MENU_SUBITEMS = {
    'call_center': [
        ('call_center_schedule',  'График',                 False),
        ('call_center_shifts',    'Смены',                  False),
        ('call_center_employees', 'Сотрудники колл-центра', False),
    ],
}
MENU_SUBITEMS_FLAT = [code for _subs in MENU_SUBITEMS.values() for code, _label, _bs in _subs]
ROLE_CONFIGURABLE = ('admin', 'director', 'callcenter')
LOGIN_ROLE_LABELS = {'admin': 'Администратор', 'director': 'Управляющий', 'callcenter': 'Оператор колл-центра'}

# Дефолтная матрица прав — воспроизводит СЕГОДНЯШНЕЕ поведение 1-в-1,
# чтобы включение фичи ничего не сломало на существующих базах.
# dashboard/employees/purchases/history/shifts_archive и раньше не были защищены
# @owner_required — их роут пускал любую залогиненную роль (просто без ссылки в меню),
# поэтому admin/director по умолчанию должны видеть именно их, а не только director->employees.
_ALWAYS_ACCESSIBLE_BEFORE = {'dashboard', 'employees', 'purchases', 'history', 'shifts_archive'}
_DEFAULT_ROLE_VISIBLE = {
    'director': set(_ALWAYS_ACCESSIBLE_BEFORE),
    'admin': set(_ALWAYS_ACCESSIBLE_BEFORE),
    # callcenter — новая роль без унаследованного поведения, доступ настраивается
    # владельцем с нуля через Настройки → Роли и доступ.
    'callcenter': set(),
}

FORMULA_VARS = {
    'revenue':        'Общая выручка',
    'cash':           'Наличные из выручки',
    'card':           'Безналичные',
    'online':         'Онлайн',
    'delivery':       'Выручка доставки',
    'pickup':         'Выручка самовывоза',
    'orders':         'Кол-во заказов',
    'shifts':         'Кол-во смен',
    'expenses_cash':  'Расходы наличными',
    'expenses_card':  'Расходы картой',
    'expenses_total': 'Расходы всего',
    'pay_admin':      'ФОТ администраторов',
    'pay_cook':       'ФОТ поваров',
    'pay_sushi':      'ФОТ сушистов',
    'pay_courier':    'ФОТ курьеров',
    'pay_cleaner':    'ФОТ уборщиц',
    'pay_packer':     'ФОТ упаковщиков',
    'pay_total':      'Весь ФОТ',
    'profit':         'Прибыль (выручка − расходы − ФОТ)',
}

_SAFE_OPS = {
    ast.Add: _op.add, ast.Sub: _op.sub,
    ast.Mult: _op.mul, ast.Div: _op.truediv,
    ast.Pow: _op.pow, ast.Mod: _op.mod,
    ast.USub: _op.neg, ast.UAdd: lambda x: x,
}


def safe_eval(formula, variables):
    def _eval(node):
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return float(node.value)
        if isinstance(node, ast.Name):
            if node.id not in variables:
                raise ValueError(f"Unknown: {node.id}")
            return float(variables[node.id] or 0)
        if isinstance(node, ast.BinOp) and type(node.op) in _SAFE_OPS:
            l, r = _eval(node.left), _eval(node.right)
            if isinstance(node.op, ast.Div) and r == 0:
                return 0.0
            return _SAFE_OPS[type(node.op)](l, r)
        if isinstance(node, ast.UnaryOp) and type(node.op) in _SAFE_OPS:
            return _SAFE_OPS[type(node.op)](_eval(node.operand))
        raise ValueError(f"Unsupported node: {type(node)}")
    try:
        return _eval(ast.parse(formula.strip(), mode='eval').body)
    except Exception:
        return None


def get_db():
    conn = sqlite3.connect(DATABASE, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')
    conn.execute('PRAGMA busy_timeout = 10000')
    conn.execute('PRAGMA journal_mode = WAL')
    return conn


def get_branch_groups(conn):
    groups = conn.execute(
        'SELECT * FROM branch_groups ORDER BY sort_order, name'
    ).fetchall()
    result = []
    for g in groups:
        members = conn.execute(
            '''SELECT b.id, b.name FROM branch_group_members bgm
               JOIN branches b ON b.id = bgm.branch_id
               WHERE bgm.group_id = ? AND b.is_active = 1
               ORDER BY b.name''',
            (g['id'],)
        ).fetchall()
        result.append({
            'id': g['id'],
            'name': g['name'],
            'abbr': g['abbr'] or '',
            'sort_order': g['sort_order'],
            'branches': [dict(m) for m in members],
            'branch_ids': [m['id'] for m in members],
        })
    return result


def get_branch_raw_map(conn):
    """Сопоставление branch_raw (текст подразделения из CSV-выгрузки «Отчёт по заказам»)
    с branches.id — задаётся вручную на странице «Филиалы» (см. save_branch_raw_mapping)."""
    return {
        r['branch_raw']: r['branch_id']
        for r in conn.execute('SELECT branch_raw, branch_id FROM branch_raw_map').fetchall()
    }


def get_expense_categories(conn, branch_id=None):
    all_cats = conn.execute(
        'SELECT id, code, label, type, parent_id, show_contractors, show_shift '
        'FROM expense_categories WHERE is_active=1 ORDER BY sort_order, label'
    ).fetchall()
    if not branch_id:
        return all_cats
    # Load branch restrictions
    restricted = {}
    for row in conn.execute('SELECT category_id, branch_id FROM expense_category_branches').fetchall():
        restricted.setdefault(row['category_id'], []).append(row['branch_id'])
    # Keep categories with no restriction or with matching branch
    return [c for c in all_cats if not restricted.get(c['id']) or branch_id in restricted[c['id']]]


# ─── BANK HELPERS ─────────────────────────────────────────────────────────────

def _detect_encoding(raw_bytes):
    for enc in ('utf-8-sig', 'utf-8', 'cp1251'):
        try:
            raw_bytes.decode(enc)
            return enc
        except UnicodeDecodeError:
            continue
    return 'utf-8'


def _parse_date_str(s):
    s = str(s).strip().split(' ')[0].split('T')[0]
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y.%m.%d'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def _clean_num(s):
    if not s:
        return None
    s = str(s).strip().replace('\xa0', '').replace(' ', '').replace(' ', '').replace(' ', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def _map_csv_columns(fieldnames):
    cols = {(f or '').lower().strip(): f for f in fieldnames}
    DATE_P  = ['дата операции', 'дата проведения', 'дата платежа', 'дата', 'date_oper', 'o_date', 'date']
    DEBIT_P = ['сумма списания', 'расход', 'дебет', 'сумма дебет', 'списание']
    CRED_P  = ['сумма зачисления', 'приход', 'кредит', 'сумма кредит', 'зачисление']
    AMT_P   = ['сумма операции', 'сумма платежа', 'sum_rur', 'sum_val', 'сумма', 'amount']
    TYPE_P  = ['вид операции', 'приход/расход', 'тип операции', 'д/к', 'd_c', 'dc', 'тип']
    DESC_P  = ['назначение платежа', 'text70', 'назначение', 'описание', 'description', 'наименование']
    CTR_P   = ['контрагент', 'pol_name', 'получатель', 'plat_name', 'плательщик', 'наименование контрагента']
    INN_P   = ['инн контрагента', 'инн плательщика', 'инн получателя', 'инн', 'inn', 'inn_pol', 'inn_plat']

    def find(patterns):
        for p in patterns:
            if p in cols:
                return cols[p]
            for k, v in cols.items():
                if p in k:
                    return v
        return None

    return {
        'date': find(DATE_P), 'debit': find(DEBIT_P), 'credit': find(CRED_P),
        'amount': find(AMT_P), 'type': find(TYPE_P),
        'description': find(DESC_P), 'counterparty': find(CTR_P),
        'inn': find(INN_P),
    }


def _parse_bank_csv(raw_bytes):
    enc = _detect_encoding(raw_bytes)
    text = raw_bytes.decode(enc)
    lines = [l for l in text.splitlines() if l.strip()]
    if not lines:
        raise ValueError('Файл пустой')

    DATE_WORDS = ('дата', 'date')
    AMT_WORDS  = ('сумм', 'amount', 'sum', 'приход', 'расход', 'поступлен', 'списан', 'дебет', 'кредит')

    # Find the header row (first row containing a date keyword + amount keyword and multiple columns)
    header_idx = 0
    for i, line in enumerate(lines[:30]):
        low = line.lower()
        has_date = any(w in low for w in DATE_WORDS)
        has_amt  = any(w in low for w in AMT_WORDS)
        n_cols   = max(line.count(';'), line.count(','), line.count('\t'))
        if has_date and has_amt and n_cols >= 2:
            header_idx = i
            break

    data_text = '\n'.join(lines[header_idx:])

    # Try semicolon, then comma, then tab as delimiter
    rows = []
    reader = None
    for delim in (';', ',', '\t'):
        r = csv.DictReader(io.StringIO(data_text), delimiter=delim)
        rs = list(r)
        fnames = [f for f in (r.fieldnames or []) if f and f.strip()]
        if rs and len(fnames) >= 2:
            rows, reader = rs, r
            break

    if not rows:
        preview = '\n'.join(lines[:5])
        raise ValueError(f'Не удалось распознать формат CSV. Начало файла:\n{preview}')

    fieldnames = reader.fieldnames or []
    col = _map_csv_columns(fieldnames)

    # Auto-detect date column if mapping failed
    if not col['date']:
        for f in fieldnames:
            if not f:
                continue
            vals = [row.get(f, '') for row in rows[:5] if row.get(f)]
            if vals and any(_parse_date_str(v) for v in vals):
                col['date'] = f
                break

    # Auto-detect amount columns if mapping failed
    if not col['amount'] and not col.get('debit') and not col.get('credit'):
        for f in fieldnames:
            if not f or f == col.get('date'):
                continue
            nums = [_clean_num(row.get(f)) for row in rows[:5]]
            if any(v is not None and v != 0 for v in nums):
                col['amount'] = f
                break

    result = []
    for row in rows:
        date_val = _parse_date_str(row.get(col.get('date') or '\x00', '') or '')
        if not date_val:
            continue
        if col.get('debit') or col.get('credit'):
            d = _clean_num(row.get(col.get('debit') or '\x00', '')) or 0
            c = _clean_num(row.get(col.get('credit') or '\x00', '')) or 0
            if c > 0:
                amount = c
            elif d > 0:
                amount = -d
            else:
                continue
        elif col.get('amount'):
            amount = _clean_num(row.get(col['amount'], ''))
            if amount is None:
                continue
            if col.get('type'):
                tv = (row.get(col['type']) or '').strip().lower()
                if tv in ('d', 'д') or any(w in tv for w in ('списан', 'расход', 'дебет', ' д ')):
                    amount = -abs(amount)
                elif tv in ('c', 'к') or any(w in tv for w in ('зачисл', 'приход', 'кредит', ' к ')):
                    amount = abs(amount)
        else:
            continue
        desc = (row.get(col.get('description') or '\x00', '') or '').strip()
        ctr  = (row.get(col.get('counterparty') or '\x00', '') or '').strip()
        inn  = re.sub(r'\D', '', (row.get(col.get('inn') or '\x00', '') or '').strip())
        result.append({'date': date_val, 'amount': amount, 'description': desc, 'counterparty': ctr, 'inn': inn})

    if not result and rows:
        col_info = '; '.join(f'{k}={v}' for k, v in col.items() if v) or 'ни одна не определена'
        first_cols = ', '.join(str(f) for f in fieldnames[:8] if f)
        raise ValueError(
            f'Строки найдены ({len(rows)} шт.), но транзакции не распознаны. '
            f'Колонки файла: {first_cols}. '
            f'Маппинг: {col_info}.'
        )

    return result


def _detect_branch_card(conn, txn):
    """Если транзакция — расход по карте филиала, возвращает (card4, branch_name), иначе None."""
    desc = (txn.get('description') or '') + ' ' + (txn.get('counterparty') or '')
    if 'PURCHASE' not in desc.upper():
        return None
    card_sequences = re.findall(r'[A-Za-z0-9]{10,24}', desc)
    cards = conn.execute(
        'SELECT bc.card_number, b.name as branch_name '
        'FROM branch_cards bc JOIN branches b ON b.id=bc.branch_id WHERE bc.is_active=1'
    ).fetchall()
    for card in cards:
        num = card['card_number'].replace(' ', '')
        if any(seq.endswith(num) or (len(num) >= 8 and num in seq) for seq in card_sequences):
            card4 = num[-4:] if len(num) >= 4 else num
            return (card4, card['branch_name'])
    return None


# Слова которые нельзя использовать как ключевые — слишком общие, дают ложные совпадения
_KW_BLACKLIST = {
    'ооо', 'оао', 'пао', 'зао', 'ао', 'ип', 'ичп', 'гуп', 'муп', 'фгуп', 'фгбу',
    'нко', 'ано', 'нао', 'кфх', 'тсж', 'снт', 'пк', 'пот', 'оп',
    'llc', 'ltd', 'inc', 'corp', 'gmbh', 'ag',
}

def _filter_keywords(kw_list):
    """Убирает из списка ключевых слов общие юридические формы."""
    return [k for k in kw_list if k.lower() not in _KW_BLACKLIST and len(k) > 2]


def _match_contractors(conn, txns):
    contractors = conn.execute('SELECT id, name, category, keywords, inn FROM contractors WHERE is_active=1').fetchall()

    def _refresh():
        nonlocal contractors
        contractors = conn.execute('SELECT id, name, category, keywords, inn FROM contractors WHERE is_active=1').fetchall()

    def _set_inn(cid, inn):
        """Записывает ИНН контрагенту если у него его ещё нет."""
        if inn:
            conn.execute(
                'UPDATE contractors SET inn=? WHERE id=? AND (inn IS NULL OR inn="")',
                (inn, cid)
            )

    for txn in txns:
        inn         = (txn.get('inn') or '').strip()
        # Матчим только по counterparty, description — назначение платежа, не имя контрагента
        cp_text     = (txn.get('counterparty') or '').lower().strip()
        branch_card = txn.pop('_branch_card', None)  # сохраняем и убираем из dict
        is_card     = txn.pop('is_card', False)       # флаг карточной покупки из API

        # 1. Матч по ИНН (наивысший приоритет)
        if inn:
            row = conn.execute(
                'SELECT id, category FROM contractors WHERE inn=? AND is_active=1', (inn,)
            ).fetchone()
            if row:
                txn['contractor_id'] = row['id']
                txn['category'] = txn.get('category') or row['category']
                continue

        # 2. Матч по ключевым словам (только против поля counterparty)
        matched_ctr = None
        for c in contractors:
            raw_kw = (c['keywords'] or '').strip()
            name_lc = (c['name'] or '').lower()
            if raw_kw:
                # Ключевые слова заданы вручную — разбиваем по запятой, фильтруем общие слова
                kws = _filter_keywords([k.strip().lower() for k in raw_kw.split(',') if k.strip()])
            else:
                # Ключевых слов нет — используем имя целиком как одну фразу
                kws = [name_lc] if name_lc else []
            if any(kw and kw in cp_text for kw in kws):
                matched_ctr = c
                break

        if matched_ctr:
            txn['contractor_id'] = matched_ctr['id']
            txn['category'] = txn.get('category') or matched_ctr['category']
            # Если у контрагента нет ИНН — заполняем из выписки
            if inn and not (matched_ctr['inn'] or '').strip():
                _set_inn(matched_ctr['id'], inn)
                _refresh()
            continue

        # 3. Карточная покупка или расход по карте филиала — контрагента не создаём
        if branch_card or is_card:
            continue

        # 4. Новый контрагент: ищем по ИНН или имени, иначе создаём
        cp = (txn.get('counterparty') or '').strip()
        if not cp:
            continue

        existing = None
        if inn:
            existing = conn.execute('SELECT id FROM contractors WHERE inn=?', (inn,)).fetchone()
        if not existing:
            existing = conn.execute('SELECT id FROM contractors WHERE LOWER(name)=LOWER(?)', (cp,)).fetchone()

        if existing:
            _set_inn(existing['id'], inn)
            cid = existing['id']
        else:
            conn.execute(
                'INSERT INTO contractors (name, keywords, inn) VALUES (?,?,?)',
                (cp, cp, inn or None)
            )
            cid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            _refresh()

        txn['contractor_id'] = cid

    return txns


def _match_terminal(conn, txn):
    text = (txn.get('description') or '') + ' ' + (txn.get('counterparty') or '')

    # 1. Совпадение по номеру карты из настроек филиала
    # Извлекаем все длинные буквенно-цифровые последовательности (маскированные номера карт Сбербанка)
    # Пример: "по карте MIR 2202209264BD8602" → '2202209264BD8602'
    card_sequences = re.findall(r'[A-Za-z0-9]{10,24}', text)

    cards = conn.execute(
        'SELECT bc.id as card_id, bc.card_number, bc.card_name, bc.branch_id, b.name as branch_name '
        'FROM branch_cards bc JOIN branches b ON b.id=bc.branch_id WHERE bc.is_active=1'
    ).fetchall()
    for card in cards:
        num = card['card_number'].replace(' ', '')
        # Проверяем все длинные последовательности: номер карты должен быть суффиксом
        matched = any(seq.endswith(num) or (len(num) >= 8 and num in seq)
                      for seq in card_sequences)
        if matched:
            label = f'{card["branch_name"]} – {card["card_name"] or card["card_number"]}'
            t = conn.execute(
                'SELECT id FROM bank_terminals WHERE terminal_number=? AND branch_id=?',
                (num, card['branch_id'])
            ).fetchone()
            if t:
                return t['id']
            conn.execute(
                'INSERT INTO bank_terminals (terminal_number, name, branch_id) VALUES (?,?,?)',
                (num, label, card['branch_id'])
            )
            return conn.execute('SELECT last_insert_rowid()').fetchone()[0]

    # 2. Совпадение по номерам мерчанта/терминала из настроек филиала
    branches = conn.execute(
        'SELECT id, name, merchant_numbers FROM branches '
        'WHERE merchant_numbers IS NOT NULL AND merchant_numbers != ""'
    ).fetchall()
    for branch in branches:
        numbers = [n.strip() for n in (branch['merchant_numbers'] or '').split(',') if n.strip()]
        for num in numbers:
            if num and num in text:
                t = conn.execute(
                    'SELECT id FROM bank_terminals WHERE terminal_number=? AND branch_id=?',
                    (num, branch['id'])
                ).fetchone()
                if t:
                    return t['id']
                conn.execute(
                    'INSERT INTO bank_terminals (terminal_number, name, branch_id) VALUES (?,?,?)',
                    (num, f'{branch["name"]} – {num}', branch['id'])
                )
                return conn.execute('SELECT last_insert_rowid()').fetchone()[0]

    # 3. Фолбэк: стандартный TID паттерн
    m = re.search(r'TID\s*[:\-]?\s*(\d{6,12})', text, re.IGNORECASE)
    if m:
        tid = m.group(1)
        t = conn.execute('SELECT id FROM bank_terminals WHERE terminal_number=?', (tid,)).fetchone()
        if t:
            return t['id']
    return None


# ──────────────────────────────────────────────────────────────────────────────

def build_cats_groups(cats):
    """Build grouped structure for dropdown: [(parent_dict, [child_dict, ...]), ...]."""
    children_map = {}
    for c in cats:
        if c['parent_id']:
            children_map.setdefault(c['parent_id'], []).append(dict(c))
    groups = []
    for c in cats:
        if not c['parent_id']:
            groups.append((dict(c), children_map.get(c['id'], [])))
    return groups


def filter_cats_by_flag(cats, flag):
    """Keep categories visible for a given place of use (e.g. 'show_shift').
    A top-level category that has children is a pure group label — it is never
    directly selectable itself, and is only kept (as an <optgroup>) if it has at
    least one visible child; its own flag is irrelevant. A top-level category
    with no children at all is a plain, directly selectable item, kept only if
    its own flag is set. Child categories are kept only if their own flag is set."""
    parent_ids_with_any_child = {c['parent_id'] for c in cats if c['parent_id']}
    parent_ids_with_visible_child = {c['parent_id'] for c in cats if c['parent_id'] and c[flag]}
    result = []
    for c in cats:
        if c['parent_id']:
            if c[flag]:
                result.append(c)
        elif c['id'] in parent_ids_with_any_child:
            if c['id'] in parent_ids_with_visible_child:
                result.append(c)
        elif c[flag]:
            result.append(c)
    return result


def _manual_rev_total(conn, date_from, date_to, bids=None):
    """Сумма revenue_manual за период там, где нет смены с уже внесённой выручкой (fallback)."""
    bf = f"AND m.branch_id IN ({','.join(str(b) for b in bids)})" if bids else ""
    row = conn.execute(f'''
        SELECT COALESCE(SUM(m.amount), 0) AS total
        FROM revenue_manual m
        WHERE m.date BETWEEN ? AND ? {bf}
        AND NOT EXISTS (
            SELECT 1 FROM shifts s JOIN shift_revenue r ON r.shift_id=s.id
            WHERE s.date=m.date AND s.branch_id=m.branch_id AND (r.total_revenue > 0 OR s.status='closed')
        )
    ''', [date_from, date_to]).fetchone()
    return float(row['total'] or 0)


def _manual_rev_by_month(conn, date_from, date_to, bids=None):
    """Словарь (year, month) -> amount из revenue_manual (fallback)."""
    bf = f"AND m.branch_id IN ({','.join(str(b) for b in bids)})" if bids else ""
    rows = conn.execute(f'''
        SELECT CAST(strftime('%Y', m.date) AS INTEGER) AS year,
               CAST(strftime('%m', m.date) AS INTEGER) AS month,
               COALESCE(SUM(m.amount), 0) AS total
        FROM revenue_manual m
        WHERE m.date BETWEEN ? AND ? {bf}
        AND NOT EXISTS (
            SELECT 1 FROM shifts s JOIN shift_revenue r ON r.shift_id=s.id
            WHERE s.date=m.date AND s.branch_id=m.branch_id AND (r.total_revenue > 0 OR s.status='closed')
        )
        GROUP BY year, month
    ''', [date_from, date_to]).fetchall()
    return {(r['year'], r['month']): float(r['total']) for r in rows}


def _manual_rev_by_day(conn, date_from, date_to, bids=None):
    """Словарь date -> amount из revenue_manual (fallback)."""
    bf = f"AND m.branch_id IN ({','.join(str(b) for b in bids)})" if bids else ""
    rows = conn.execute(f'''
        SELECT m.date, COALESCE(SUM(m.amount), 0) AS total
        FROM revenue_manual m
        WHERE m.date BETWEEN ? AND ? {bf}
        AND NOT EXISTS (
            SELECT 1 FROM shifts s JOIN shift_revenue r ON r.shift_id=s.id
            WHERE s.date=m.date AND s.branch_id=m.branch_id AND (r.total_revenue > 0 OR s.status='closed')
        )
        GROUP BY m.date
    ''', [date_from, date_to]).fetchall()
    return {r['date']: float(r['total']) for r in rows}


def get_kpi_values(conn, branch_id, date_from, date_to):
    bf = f"AND s.branch_id={int(branch_id)}" if branch_id else ""
    rev = conn.execute(f'''
        SELECT COALESCE(SUM(r.total_revenue),0) as revenue,
               COALESCE(SUM(r.cash_amount),0) as cash,
               COALESCE(SUM(r.card_amount),0) as card,
               COALESCE(SUM(r.online_amount),0) as online,
               COALESCE(SUM(r.delivery_revenue),0) as delivery,
               COALESCE(SUM(r.pickup_revenue),0) as pickup,
               COALESCE(SUM(r.delivery_orders+r.pickup_orders),0) as orders,
               COUNT(DISTINCT s.id) as shifts
        FROM shifts s LEFT JOIN shift_revenue r ON r.shift_id=s.id
        WHERE s.date BETWEEN ? AND ? {bf}
    ''', (date_from, date_to)).fetchone()

    exp = conn.execute(f'''
        SELECT COALESCE(SUM(e.amount_cash),0) as expenses_cash,
               COALESCE(SUM(e.amount_card),0) as expenses_card
        FROM expenses e JOIN shifts s ON s.id=e.shift_id
        WHERE s.date BETWEEN ? AND ? {bf}
    ''', (date_from, date_to)).fetchone()

    pay_rows = conn.execute(f'''
        SELECT es.role_snapshot, COALESCE(SUM(es.total_amount),0) as total
        FROM employee_shifts es JOIN shifts s ON s.id=es.shift_id
        WHERE s.date BETWEEN ? AND ? {bf}
        GROUP BY es.role_snapshot
    ''', (date_from, date_to)).fetchall()

    pay_by_role = {r['role_snapshot']: r['total'] for r in pay_rows}
    pay_total = sum(pay_by_role.values())

    bids_kpi = [int(branch_id)] if branch_id else None
    manual_rev = _manual_rev_total(conn, date_from, date_to, bids_kpi)

    vals = dict(rev)
    vals['revenue'] = float(vals['revenue']) + manual_rev
    vals.update({
        'expenses_cash':  exp['expenses_cash'],
        'expenses_card':  exp['expenses_card'],
        'expenses_total': exp['expenses_cash'] + exp['expenses_card'],
        'pay_admin':      pay_by_role.get('admin', 0),
        'pay_cook':       pay_by_role.get('cook', 0),
        'pay_sushi':      pay_by_role.get('sushi', 0),
        'pay_courier':    pay_by_role.get('courier', 0),
        'pay_cleaner':    pay_by_role.get('cleaner', 0),
        'pay_packer':     pay_by_role.get('packer', 0),
        'pay_total':      pay_total,
        'profit':         vals['revenue'] - exp['expenses_cash'] - exp['expenses_card'] - pay_total,
    })
    return vals


def init_db():
    with open(os.path.join(os.path.dirname(__file__), 'schema.sql')) as f:
        with get_db() as conn:
            conn.executescript(f.read())
    with get_db() as conn:
        # Seed owner user
        owner = conn.execute("SELECT id FROM users WHERE role='owner'").fetchone()
        if not owner:
            conn.execute(
                "INSERT INTO users (username, password_hash, role, full_name) VALUES (?,?,?,?)",
                ('owner', generate_password_hash('admin123', method='pbkdf2:sha256'), 'owner', 'Владелец')
            )
            conn.execute("INSERT INTO branches (name) VALUES (?)", ('КВАДРАТ',))

        # Add auto_bonus column if missing (migration for existing DBs)
        existing_cols = [r[1] for r in conn.execute("PRAGMA table_info(employee_shifts)").fetchall()]
        if 'auto_bonus' not in existing_cols:
            conn.execute("ALTER TABLE employee_shifts ADD COLUMN auto_bonus REAL DEFAULT 0")

        # Доп. маршруты курьера (перемещение/заезд в магазин/по заказу) —
        # разбивка км/заказов сверх «Данных из гуляша». employee_shifts.km/orders
        # остаются агрегатом (гуляш + сумма/количество этих маршрутов).
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS courier_extra_routes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_shift_id INTEGER NOT NULL REFERENCES employee_shifts(id) ON DELETE CASCADE,
                reason TEXT NOT NULL DEFAULT 'relocation',
                km REAL DEFAULT 0,
                comment TEXT DEFAULT '',
                sort_order INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        ''')
        _cer_cols = [r[1] for r in conn.execute("PRAGMA table_info(courier_extra_routes)").fetchall()]
        if 'comment' not in _cer_cols:
            conn.execute("ALTER TABLE courier_extra_routes ADD COLUMN comment TEXT DEFAULT ''")

        # Add branch_id to bonus_rules if missing
        br_cols = [r[1] for r in conn.execute("PRAGMA table_info(bonus_rules)").fetchall()]
        if 'branch_id' not in br_cols:
            conn.execute("ALTER TABLE bonus_rules ADD COLUMN branch_id INTEGER REFERENCES branches(id)")

        # Add first_name / last_name to employees if missing
        emp_cols = [r[1] for r in conn.execute("PRAGMA table_info(employees)").fetchall()]
        if 'last_name' not in emp_cols:
            conn.execute("ALTER TABLE employees ADD COLUMN last_name TEXT DEFAULT ''")
        if 'first_name' not in emp_cols:
            conn.execute("ALTER TABLE employees ADD COLUMN first_name TEXT DEFAULT ''")
        # Migrate existing full_name → last_name + first_name
        conn.execute("""
            UPDATE employees SET
                last_name = CASE WHEN INSTR(full_name,' ')>0
                    THEN TRIM(SUBSTR(full_name,1,INSTR(full_name,' ')-1))
                    ELSE full_name END,
                first_name = CASE WHEN INSTR(full_name,' ')>0
                    THEN TRIM(SUBSTR(full_name,INSTR(full_name,' ')+1))
                    ELSE '' END
            WHERE last_name = '' OR last_name IS NULL
        """)

        # Create rate_templates and related tables if missing
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS rate_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                role TEXT NOT NULL,
                name TEXT NOT NULL,
                rate REAL DEFAULT 0,
                rate_per_km REAL DEFAULT 10,
                rate_per_order REAL DEFAULT 100,
                is_active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS rate_template_branches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                template_id INTEGER NOT NULL REFERENCES rate_templates(id) ON DELETE CASCADE,
                branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
                UNIQUE(template_id, branch_id)
            );
            CREATE TABLE IF NOT EXISTS rate_template_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                template_id INTEGER NOT NULL REFERENCES rate_templates(id) ON DELETE CASCADE,
                rate REAL DEFAULT 0,
                rate_per_km REAL DEFAULT 10,
                rate_per_order REAL DEFAULT 100,
                valid_from DATE NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        ''')

        # Positions (должности) — replaces hardcoded ROLE_LABELS
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS positions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                abbr TEXT DEFAULT '',
                sort_order INTEGER DEFAULT 0,
                is_active INTEGER DEFAULT 1
            );
        ''')
        try:
            conn.execute("ALTER TABLE positions ADD COLUMN abbr TEXT DEFAULT ''")
        except Exception:
            pass
        # Seed from ROLE_LABELS if table is empty
        if conn.execute('SELECT COUNT(*) FROM positions').fetchone()[0] == 0:
            for i, (code, name) in enumerate(list(ROLE_LABELS.items())):
                conn.execute(
                    'INSERT OR IGNORE INTO positions (code, name, sort_order) VALUES (?,?,?)',
                    (code, name, i * 10)
                )
        conn.commit()
        # Reload ROLE_LABELS from DB so additions survive restarts
        _reload_role_labels(conn)

        # Create taxi and address tables (safe no-ops if already exist — handled by schema IF NOT EXISTS)
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS employee_address_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES employees(id),
                address TEXT NOT NULL,
                valid_from DATE NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS taxi_trips (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                shift_id INTEGER NOT NULL REFERENCES shifts(id),
                amount REAL DEFAULT 0,
                payment_type TEXT DEFAULT 'cash',
                in_gulyash INTEGER DEFAULT 0,
                note TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS taxi_trip_employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                trip_id INTEGER NOT NULL REFERENCES taxi_trips(id) ON DELETE CASCADE,
                employee_id INTEGER REFERENCES employees(id),
                name_snapshot TEXT NOT NULL,
                address_snapshot TEXT
            );
        ''')

        # Seed default bonus rules for cooks
        if conn.execute("SELECT COUNT(*) FROM bonus_rules").fetchone()[0] == 0:
            conn.execute("INSERT INTO bonus_rules (role,threshold_pct,bonus_pct) VALUES ('cook',8.0,1.5)")
            conn.execute("INSERT INTO bonus_rules (role,threshold_pct,bonus_pct) VALUES ('cook',10.0,1.0)")

        # Seed expense categories
        existing = conn.execute("SELECT COUNT(*) FROM expense_categories").fetchone()[0]
        if existing == 0:
            for code, label, sort in _DEFAULT_EXPENSE_CATEGORIES:
                conn.execute(
                    "INSERT OR IGNORE INTO expense_categories (code, label, sort_order) VALUES (?,?,?)",
                    (code, label, sort)
                )

            conn.execute('''CREATE TABLE IF NOT EXISTS contractor_categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                sort_order INTEGER DEFAULT 0
            )''')
            _DEFAULT_CTR_CATS = [
                ('Продукты', 1), ('Упаковка', 2), ('Налоги', 3),
                ('Зарплата', 4), ('Аренда', 5), ('Коммунальные', 6),
                ('Транспорт', 7), ('Реклама', 8), ('Оборудование', 9), ('Прочее', 10),
            ]
            for name, sort in _DEFAULT_CTR_CATS:
                conn.execute(
                    "INSERT OR IGNORE INTO contractor_categories (name, sort_order) VALUES (?,?)",
                    (name, sort)
                )

        # Create user_branches table (many-to-many users ↔ branches)
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS user_branches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
                UNIQUE(user_id, branch_id)
            );
            CREATE TABLE IF NOT EXISTS employee_branches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES employees(id) ON DELETE CASCADE,
                branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
                UNIQUE(employee_id, branch_id)
            );
        ''')
        conn.execute('''
            INSERT OR IGNORE INTO user_branches (user_id, branch_id)
            SELECT id, branch_id FROM users WHERE branch_id IS NOT NULL
        ''')
        conn.execute('''
            INSERT OR IGNORE INTO employee_branches (employee_id, branch_id)
            SELECT id, branch_id FROM employees WHERE branch_id IS NOT NULL
        ''')

        # Add allowed_ip to branches if missing
        branch_cols = [r[1] for r in conn.execute("PRAGMA table_info(branches)").fetchall()]
        if 'allowed_ip' not in branch_cols:
            conn.execute("ALTER TABLE branches ADD COLUMN allowed_ip TEXT DEFAULT NULL")
        if 'merchant_numbers' not in branch_cols:
            conn.execute("ALTER TABLE branches ADD COLUMN merchant_numbers TEXT DEFAULT ''")

        # Add type and parent_id to expense_categories if missing
        ec_cols = [r[1] for r in conn.execute("PRAGMA table_info(expense_categories)").fetchall()]
        if 'type' not in ec_cols:
            conn.execute("ALTER TABLE expense_categories ADD COLUMN type TEXT DEFAULT 'expense'")
        if 'parent_id' not in ec_cols:
            conn.execute("ALTER TABLE expense_categories ADD COLUMN parent_id INTEGER REFERENCES expense_categories(id)")
        if 'show_contractors' not in ec_cols:
            conn.execute("ALTER TABLE expense_categories ADD COLUMN show_contractors INTEGER DEFAULT 1")
        if 'show_shift' not in ec_cols:
            conn.execute("ALTER TABLE expense_categories ADD COLUMN show_shift INTEGER DEFAULT 1")

        # Create branch groups tables
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS branch_groups (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                sort_order INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS branch_group_members (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                group_id INTEGER NOT NULL REFERENCES branch_groups(id) ON DELETE CASCADE,
                branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
                UNIQUE(group_id, branch_id)
            );
            CREATE TABLE IF NOT EXISTS branch_raw_map (
                branch_raw TEXT PRIMARY KEY,
                branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE CASCADE
            );
            -- Разовая уборка: короткоживущая параллельная система групп по branch_raw
            -- (branch_raw_groups/branch_raw_group_members), заменена на branch_raw_map выше.
            DROP TABLE IF EXISTS branch_raw_group_members;
            DROP TABLE IF EXISTS branch_raw_groups;
            CREATE TABLE IF NOT EXISTS branch_cards (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
                card_number TEXT NOT NULL,
                card_name TEXT DEFAULT '',
                is_active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        ''')

        try:
            conn.execute("ALTER TABLE shift_revenue ADD COLUMN actual_cash_comment TEXT DEFAULT ''")
        except Exception:
            pass

        try:
            conn.execute("ALTER TABLE shift_revenue ADD COLUMN plus_amount REAL DEFAULT 0")
        except Exception:
            pass

        conn.executescript('''
            CREATE TABLE IF NOT EXISTS cash_plus_entries (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                shift_id    INTEGER NOT NULL REFERENCES shifts(id) ON DELETE CASCADE,
                amount      REAL    NOT NULL DEFAULT 0,
                description TEXT    DEFAULT '',
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        ''')

        try:
            conn.execute("ALTER TABLE shift_revenue ADD COLUMN morning_cash REAL DEFAULT 0")
        except Exception:
            pass

        try:
            conn.execute("ALTER TABLE shift_revenue ADD COLUMN kassa_nal REAL DEFAULT NULL")
        except Exception:
            pass

        # Add category and cash/card split to cash_plus_entries
        plus_cols = [r[1] for r in conn.execute("PRAGMA table_info(cash_plus_entries)").fetchall()]
        if 'category' not in plus_cols:
            conn.execute("ALTER TABLE cash_plus_entries ADD COLUMN category TEXT DEFAULT ''")
        if 'amount_cash' not in plus_cols:
            conn.execute("ALTER TABLE cash_plus_entries ADD COLUMN amount_cash REAL DEFAULT 0")
            conn.execute("ALTER TABLE cash_plus_entries ADD COLUMN amount_card REAL DEFAULT 0")
            conn.execute("UPDATE cash_plus_entries SET amount_cash = amount WHERE amount > 0")

        try:
            conn.execute("ALTER TABLE purchases ADD COLUMN payer TEXT DEFAULT ''")
        except Exception:
            pass

        try:
            conn.execute("ALTER TABLE purchases ADD COLUMN import_hash TEXT")
            conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_purchases_import_hash ON purchases(import_hash) WHERE import_hash IS NOT NULL")
        except Exception:
            pass

        try:
            conn.execute("ALTER TABLE contractors ADD COLUMN is_card_merchant INTEGER DEFAULT 0")
        except Exception:
            pass

        try:
            conn.execute("ALTER TABLE contractors ADD COLUMN inn TEXT DEFAULT ''")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_contractors_inn ON contractors(inn) WHERE inn IS NOT NULL AND inn != ''")
        except Exception:
            pass

        try:
            conn.execute("ALTER TABLE branches ADD COLUMN abbr TEXT DEFAULT ''")
        except Exception:
            pass

        try:
            conn.execute("ALTER TABLE branch_groups ADD COLUMN abbr TEXT DEFAULT ''")
        except Exception:
            pass

        try:
            conn.execute("ALTER TABLE users ADD COLUMN email TEXT DEFAULT ''")
        except Exception:
            pass

        # Expand role CHECK constraint to include 'director' if needed
        schema_row = conn.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='users'"
        ).fetchone()
        if schema_row and 'director' not in schema_row['sql']:
            conn.executescript('''
                PRAGMA foreign_keys = OFF;
                CREATE TABLE IF NOT EXISTS users_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    role TEXT NOT NULL CHECK(role IN ('owner','admin','employee','director')),
                    full_name TEXT NOT NULL,
                    branch_id INTEGER REFERENCES branches(id),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    email TEXT DEFAULT ''
                );
                INSERT OR IGNORE INTO users_new
                    SELECT id, username, password_hash, role, full_name, branch_id, created_at, COALESCE(email,'')
                    FROM users;
                DROP TABLE users;
                ALTER TABLE users_new RENAME TO users;
                PRAGMA foreign_keys = ON;
            ''')
            logging.info('Migrated users table: added director role')

        # Expand role CHECK constraint to include 'callcenter' (замена роли 'employee')
        schema_row = conn.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='users'"
        ).fetchone()
        if schema_row and 'callcenter' not in schema_row['sql']:
            conn.executescript('''
                PRAGMA foreign_keys = OFF;
                CREATE TABLE IF NOT EXISTS users_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    role TEXT NOT NULL CHECK(role IN ('owner','admin','callcenter','director')),
                    full_name TEXT NOT NULL,
                    branch_id INTEGER REFERENCES branches(id),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    email TEXT DEFAULT ''
                );
                INSERT OR IGNORE INTO users_new
                    SELECT id, username, password_hash,
                           CASE WHEN role='employee' THEN 'callcenter' ELSE role END,
                           full_name, branch_id, created_at, COALESCE(email,'')
                    FROM users;
                DROP TABLE users;
                ALTER TABLE users_new RENAME TO users;
                PRAGMA foreign_keys = ON;
            ''')
            logging.info('Migrated users table: added callcenter role, migrated employee -> callcenter')

        conn.executescript('''
            CREATE TABLE IF NOT EXISTS password_reset_tokens (
                id      INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                token   TEXT NOT NULL UNIQUE,
                expires_at TIMESTAMP NOT NULL,
                used    INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS invite_tokens (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                token      TEXT NOT NULL UNIQUE,
                role       TEXT NOT NULL DEFAULT 'admin',
                branch_ids TEXT NOT NULL DEFAULT '[]',
                created_by INTEGER NOT NULL REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                expires_at TIMESTAMP NOT NULL,
                used       INTEGER DEFAULT 0,
                used_by    INTEGER REFERENCES users(id)
            );
        ''')

        # Create bank module tables
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS contractors (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                category TEXT DEFAULT '',
                keywords TEXT DEFAULT '',
                is_active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS contractor_categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                sort_order INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS bank_terminals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                terminal_number TEXT NOT NULL UNIQUE,
                name TEXT DEFAULT '',
                branch_id INTEGER REFERENCES branches(id),
                is_active INTEGER DEFAULT 1
            );
            CREATE TABLE IF NOT EXISTS bank_accounts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                bank_name TEXT DEFAULT '',
                account_number TEXT DEFAULT '',
                is_active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS bank_account_branches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bank_account_id INTEGER NOT NULL REFERENCES bank_accounts(id) ON DELETE CASCADE,
                branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
                UNIQUE(bank_account_id, branch_id)
            );
            CREATE TABLE IF NOT EXISTS bank_statements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bank_account_id INTEGER NOT NULL REFERENCES bank_accounts(id) ON DELETE CASCADE,
                filename TEXT NOT NULL,
                uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                uploaded_by INTEGER REFERENCES users(id),
                date_from DATE,
                date_to DATE,
                row_count INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS bank_transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                statement_id INTEGER NOT NULL REFERENCES bank_statements(id) ON DELETE CASCADE,
                bank_account_id INTEGER NOT NULL REFERENCES bank_accounts(id),
                txn_date DATE NOT NULL,
                amount REAL NOT NULL,
                description TEXT DEFAULT '',
                counterparty TEXT DEFAULT '',
                contractor_id INTEGER REFERENCES contractors(id),
                category TEXT DEFAULT '',
                terminal_id INTEGER REFERENCES bank_terminals(id),
                is_ignored INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS api_settings (
                key   TEXT PRIMARY KEY,
                value TEXT
            );
            CREATE TABLE IF NOT EXISTS api_1c_tokens (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                branch_id   INTEGER NOT NULL REFERENCES branches(id),
                token       TEXT    NOT NULL UNIQUE,
                description TEXT,
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS api_1c_log (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                token       TEXT,
                branch_id   INTEGER,
                method      TEXT,
                path        TEXT,
                body        TEXT,
                status      TEXT DEFAULT 'received',
                parsed_ok   INTEGER DEFAULT 0,
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS api_revenue_tokens (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                branch_id   INTEGER NOT NULL REFERENCES branches(id),
                token       TEXT    NOT NULL UNIQUE,
                description TEXT,
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS api_revenue_log (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                token       TEXT,
                branch_id   INTEGER,
                method      TEXT,
                path        TEXT,
                body        TEXT,
                status      TEXT DEFAULT 'received',
                parsed_ok   INTEGER DEFAULT 0,
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS api_waittime_tokens (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                branch_id   INTEGER NOT NULL REFERENCES branches(id),
                token       TEXT    NOT NULL UNIQUE,
                description TEXT,
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS api_waittime_log (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                token       TEXT,
                branch_id   INTEGER,
                method      TEXT,
                path        TEXT,
                body        TEXT,
                status      TEXT DEFAULT 'received',
                parsed_ok   INTEGER DEFAULT 0,
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS wait_time_log (
                id                INTEGER PRIMARY KEY AUTOINCREMENT,
                branch_id         INTEGER NOT NULL REFERENCES branches(id),
                promised_minutes  REAL,
                estimated_minutes REAL,
                recorded_at       TIMESTAMP NOT NULL,
                created_at        TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS purchase_suppliers (
                id   INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE
            );
            CREATE TABLE IF NOT EXISTS purchases (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                branch_id      INTEGER NOT NULL REFERENCES branches(id),
                supplier       TEXT    NOT NULL,
                amount         REAL    NOT NULL DEFAULT 0,
                date           DATE    NOT NULL,
                invoice_number TEXT,
                note           TEXT,
                created_at     TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                created_by     INTEGER REFERENCES users(id)
            );
        ''')

        # Feature: multiple roles per employee
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS employee_roles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES employees(id) ON DELETE CASCADE,
                role TEXT NOT NULL,
                rate REAL DEFAULT 0,
                rate_per_km REAL DEFAULT 10,
                rate_per_order REAL DEFAULT 100,
                is_active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(employee_id, role)
            );
        ''')
        _er_cols = [r[1] for r in conn.execute("PRAGMA table_info(employee_roles)").fetchall()]
        if 'rate_template_id' not in _er_cols:
            conn.execute("ALTER TABLE employee_roles ADD COLUMN rate_template_id INTEGER REFERENCES rate_templates(id)")
        if 'pay_monthly' not in _er_cols:
            conn.execute("ALTER TABLE employee_roles ADD COLUMN pay_monthly INTEGER DEFAULT 0")
        conn.execute('''
            CREATE TABLE IF NOT EXISTS employee_pay_monthly_branches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES employees(id) ON DELETE CASCADE,
                role TEXT NOT NULL,
                branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
                UNIQUE(employee_id, role, branch_id)
            )
        ''')
        # История переключений «ежедневно/ежемесячно» по датам — чтобы смена режима
        # сегодня не задним числом влияла на уже прошедшие смены (см. pay_staff/_effective_pay_monthly).
        conn.execute('''
            CREATE TABLE IF NOT EXISTS employee_pay_monthly_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES employees(id) ON DELETE CASCADE,
                role TEXT NOT NULL,
                pay_monthly INTEGER NOT NULL DEFAULT 0,
                effective_from DATE NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        _emp_tcols = [r[1] for r in conn.execute("PRAGMA table_info(employees)").fetchall()]
        if 'rate_template_id' not in _emp_tcols:
            conn.execute("ALTER TABLE employees ADD COLUMN rate_template_id INTEGER REFERENCES rate_templates(id)")

        # Feature: import batches for Excel imports
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS import_batches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                branch_id INTEGER NOT NULL REFERENCES branches(id),
                filename TEXT NOT NULL,
                imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                imported_by INTEGER REFERENCES users(id),
                shifts_created INTEGER DEFAULT 0,
                expenses_created INTEGER DEFAULT 0,
                employees_created INTEGER DEFAULT 0,
                employee_shifts_created INTEGER DEFAULT 0
            );
        ''')
        _shifts_cols = [r[1] for r in conn.execute("PRAGMA table_info(shifts)").fetchall()]
        if 'import_batch_id' not in _shifts_cols:
            conn.execute("ALTER TABLE shifts ADD COLUMN import_batch_id INTEGER REFERENCES import_batches(id)")

        conn.execute('''
            CREATE TABLE IF NOT EXISTS import_staging (
                token TEXT PRIMARY KEY,
                branch_id INTEGER,
                data TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Feature: fire/restore employees
        _emp_cols2 = [r[1] for r in conn.execute("PRAGMA table_info(employees)").fetchall()]
        if 'is_fired' not in _emp_cols2:
            conn.execute("ALTER TABLE employees ADD COLUMN is_fired INTEGER DEFAULT 0")
        if 'fired_at' not in _emp_cols2:
            conn.execute("ALTER TABLE employees ADD COLUMN fired_at TIMESTAMP")
        if 'fired_comment' not in _emp_cols2:
            conn.execute("ALTER TABLE employees ADD COLUMN fired_comment TEXT DEFAULT ''")

        # Feature: monthly salary flag
        if 'pay_monthly' not in _emp_cols2:
            conn.execute("ALTER TABLE employees ADD COLUMN pay_monthly INTEGER DEFAULT 0")

        # Feature: employee phone number
        if 'phone' not in _emp_cols2:
            conn.execute("ALTER TABLE employees ADD COLUMN phone TEXT DEFAULT ''")

        # One shift per branch per day: unique index
        conn.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_shifts_branch_date ON shifts(branch_id, date)"
        )

        # Колл-центр — полностью отдельный набор таблиц (операторы не привязаны
        # к филиалу и не должны фигурировать в обычном отчёте по сотрудникам).
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS call_center_employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                full_name TEXT NOT NULL,
                rate REAL DEFAULT 0,
                is_active INTEGER DEFAULT 1,
                is_fired INTEGER DEFAULT 0,
                fired_at TIMESTAMP,
                fired_comment TEXT DEFAULT '',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS call_center_rate_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES call_center_employees(id) ON DELETE CASCADE,
                rate REAL NOT NULL,
                effective_from DATE NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS call_center_schedule (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES call_center_employees(id) ON DELETE CASCADE,
                date DATE NOT NULL,
                planned_start TEXT DEFAULT '10:00',
                planned_end TEXT DEFAULT '22:00',
                UNIQUE(employee_id, date)
            );
        ''')
        # Смены колл-центра теперь считаются автоматически из графика (часы = конец−начало,
        # ставка — актуальная на дату), отдельная таблица с ручными часами больше не нужна.
        conn.execute("DROP TABLE IF EXISTS call_center_shifts")
        _ccs_cols = [r[1] for r in conn.execute("PRAGMA table_info(call_center_schedule)").fetchall()]
        if 'planned_start' not in _ccs_cols:
            conn.execute("ALTER TABLE call_center_schedule ADD COLUMN planned_start TEXT DEFAULT '10:00'")
        if 'planned_end' not in _ccs_cols:
            conn.execute("ALTER TABLE call_center_schedule ADD COLUMN planned_end TEXT DEFAULT '22:00'")

        # Per-account Sber auto-sync columns
        _ba_cols = [r[1] for r in conn.execute("PRAGMA table_info(bank_accounts)").fetchall()]
        if 'sber_auto_sync' not in _ba_cols:
            conn.execute("ALTER TABLE bank_accounts ADD COLUMN sber_auto_sync INTEGER DEFAULT 0")
        if 'sber_last_sync' not in _ba_cols:
            conn.execute("ALTER TABLE bank_accounts ADD COLUMN sber_last_sync TEXT DEFAULT ''")
        if 'sber_last_result' not in _ba_cols:
            conn.execute("ALTER TABLE bank_accounts ADD COLUMN sber_last_result TEXT DEFAULT ''")

        # Per-shift terminal breakdown for beznal
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS shift_terminals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                shift_id INTEGER NOT NULL REFERENCES shifts(id) ON DELETE CASCADE,
                terminal_number TEXT DEFAULT '',
                amount REAL DEFAULT 0,
                sort_order INTEGER DEFAULT 0
            );
        ''')

        # Feature: change schedules
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS change_schedule (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                branch_id INTEGER REFERENCES branches(id),
                weekday INTEGER,
                amount REAL NOT NULL DEFAULT 0,
                valid_from TEXT NOT NULL,
                valid_to TEXT,
                label TEXT DEFAULT '',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS change_date_overrides (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                branch_id INTEGER NOT NULL REFERENCES branches(id),
                date TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(branch_id, date)
            );
        ''')

        # Feature: промокод листовки (листовка в заказ) по дням/филиалам
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS flyer_promocodes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                branch_id INTEGER NOT NULL REFERENCES branches(id),
                date TEXT NOT NULL,
                code TEXT NOT NULL DEFAULT '',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(branch_id, date)
            );
        ''')

        # Правила разбора банковских операций
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS bank_parse_rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bank_account_id INTEGER REFERENCES bank_accounts(id),
                name TEXT NOT NULL,
                direction TEXT NOT NULL DEFAULT 'any',
                keyword TEXT NOT NULL,
                commission_included INTEGER DEFAULT 1,
                commission_pattern TEXT DEFAULT '',
                category TEXT DEFAULT '',
                sort_order INTEGER DEFAULT 0,
                is_active INTEGER DEFAULT 1
            );
        ''')
        # Паттерны извлечения сумм из описания операции (замена одиночной «комиссии» —
        # теперь можно вытащить несколько сумм с разными категориями/филиалами из одного текста).
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS bank_parse_rule_patterns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                rule_id INTEGER NOT NULL REFERENCES bank_parse_rules(id) ON DELETE CASCADE,
                example_text TEXT NOT NULL,
                example_value TEXT NOT NULL,
                regex_pattern TEXT NOT NULL,
                direction TEXT NOT NULL DEFAULT 'expense',
                category TEXT DEFAULT '',
                sort_order INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS bank_parse_rule_pattern_branches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pattern_id INTEGER NOT NULL REFERENCES bank_parse_rule_patterns(id) ON DELETE CASCADE,
                branch_id INTEGER REFERENCES branches(id),
                branch_group_id INTEGER REFERENCES branch_groups(id)
            );
            CREATE TABLE IF NOT EXISTS bank_parse_rule_branches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                rule_id INTEGER NOT NULL REFERENCES bank_parse_rules(id) ON DELETE CASCADE,
                branch_id INTEGER REFERENCES branches(id),
                branch_group_id INTEGER REFERENCES branch_groups(id)
            );
            CREATE TABLE IF NOT EXISTS bank_parse_extracted_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bank_transaction_id INTEGER NOT NULL REFERENCES bank_transactions(id) ON DELETE CASCADE,
                pattern_id INTEGER NOT NULL REFERENCES bank_parse_rule_patterns(id) ON DELETE CASCADE,
                branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
                amount REAL NOT NULL,
                direction TEXT NOT NULL,
                category TEXT DEFAULT '',
                txn_date DATE NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(bank_transaction_id, pattern_id, branch_id)
            );
        ''')

        # Миграция: раньше филиал выбирался у каждого паттерна отдельно
        # (bank_parse_rule_pattern_branches), теперь — только у правила целиком
        # (паттерн всегда просто «расход» с той же привязкой филиала, что и у
        # самого правила). Переносим уже существующие привязки паттернов наверх,
        # на правило (если у правила такой привязки ещё нет), и очищаем источник —
        # иначе как в п.162 при каждом рестарте будут находиться те же старые
        # записи и путать привязку заново.
        _pattern_branch_rows = conn.execute(
            'SELECT prb.branch_id, prb.branch_group_id, p.rule_id '
            'FROM bank_parse_rule_pattern_branches prb '
            'JOIN bank_parse_rule_patterns p ON p.id = prb.pattern_id'
        ).fetchall()
        if _pattern_branch_rows:
            _rules_with_branches = set(
                r['rule_id'] for r in conn.execute('SELECT DISTINCT rule_id FROM bank_parse_rule_branches').fetchall()
            )
            _to_migrate = {}
            for _row in _pattern_branch_rows:
                _rid = _row['rule_id']
                if _rid in _rules_with_branches:
                    continue
                _entry = _to_migrate.setdefault(_rid, {'branch_ids': set(), 'group_id': None})
                if _row['branch_group_id']:
                    _entry['group_id'] = _row['branch_group_id']
                elif _row['branch_id']:
                    _entry['branch_ids'].add(_row['branch_id'])
            for _rid, _entry in _to_migrate.items():
                if _entry['group_id']:
                    conn.execute(
                        'INSERT INTO bank_parse_rule_branches (rule_id, branch_group_id) VALUES (?,?)',
                        (_rid, _entry['group_id'])
                    )
                for _bid in _entry['branch_ids']:
                    conn.execute(
                        'INSERT INTO bank_parse_rule_branches (rule_id, branch_id) VALUES (?,?)',
                        (_rid, _bid)
                    )
            conn.execute('DELETE FROM bank_parse_rule_pattern_branches')
        # Паттерн теперь всегда «расход» — нормализуем старые записи с другим направлением.
        conn.execute("UPDATE bank_parse_rule_patterns SET direction='expense' WHERE direction != 'expense'")

        try:
            conn.execute("ALTER TABLE bank_parse_rules ADD COLUMN category TEXT DEFAULT ''")
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE contractor_categories ADD COLUMN direction TEXT DEFAULT 'any'")
        except Exception:
            pass

        # One-time merge: contractor_categories -> expense_categories
        # Note: SQLite's LOWER() only handles ASCII, so case-insensitive matching
        # against Cyrillic names is done in Python, not in SQL.
        ctr_cats_rows = conn.execute("SELECT id, name, direction FROM contractor_categories").fetchall()
        existing_ec_rows = conn.execute("SELECT id, label, type FROM expense_categories").fetchall()
        old_names = []
        for ctr_cat in ctr_cats_rows:
            label = (ctr_cat['name'] or '').strip()
            if not label:
                continue
            old_names.append(label.lower())
            cat_type = 'income' if ctr_cat['direction'] == 'income' else 'expense'
            existing_ec = next(
                (r for r in existing_ec_rows if r['label'].lower() == label.lower() and r['type'] == cat_type),
                None
            )
            if existing_ec:
                continue
            code = _slugify(label) or ('cat_' + str(ctr_cat['id']))
            if conn.execute("SELECT id FROM expense_categories WHERE code=?", (code,)).fetchone():
                code = code + '_' + str(ctr_cat['id'])
            max_sort = conn.execute("SELECT COALESCE(MAX(sort_order),0) FROM expense_categories").fetchone()[0]
            conn.execute(
                "INSERT INTO expense_categories (code, label, type, parent_id, sort_order, show_contractors, show_shift) "
                "VALUES (?,?,?,NULL,?,1,0)",
                (code, label, cat_type, max_sort + 1)
            )
        if old_names:
            for _table in ('contractors', 'bank_transactions', 'bank_parse_rules'):
                _rows = conn.execute(f"SELECT DISTINCT category FROM {_table} WHERE category != ''").fetchall()
                for _r in _rows:
                    if (_r['category'] or '').lower() in old_names:
                        conn.execute(f"UPDATE {_table} SET category='' WHERE category=?", (_r['category'],))
        # Источник смержен — очищаем, иначе при каждом рестарте сервера этот блок
        # находит те же старые записи и заново создаёт удалённые пользователем
        # expense_categories (баг: категория "воскресала" после каждого деплоя).
        if ctr_cats_rows:
            conn.execute('DELETE FROM contractor_categories')

        # PnL report settings storage
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS pnl_settings (
                key TEXT PRIMARY KEY,
                value TEXT
            );
        ''')

        # Настройки отчёта «Сверка безнала»
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS reconciliation_cashless_settings (
                key TEXT PRIMARY KEY,
                value TEXT
            );
        ''')

        if old_names:
            for _setting_key in ('bank_income_ctr_cats', 'bank_expense_ctr_cats'):
                _row = conn.execute("SELECT value FROM pnl_settings WHERE key=?", (_setting_key,)).fetchone()
                if _row and _row['value']:
                    try:
                        _vals = _json_lib.loads(_row['value'])
                    except Exception:
                        _vals = None
                    if isinstance(_vals, list):
                        _cleaned = [v for v in _vals if (v or '').lower() not in old_names]
                        if _cleaned != _vals:
                            conn.execute(
                                "UPDATE pnl_settings SET value=? WHERE key=?",
                                (_json_lib.dumps(_cleaned), _setting_key)
                            )

        # One-time merge: P&L "наличные расходы" + "банковские расходы контрагентов"
        # -> единый список категорий расходов (expense_cats), т.к. расходы теперь
        # в одном месте (одна категория может встречаться и в кассе, и в банке).
        if not conn.execute("SELECT 1 FROM pnl_settings WHERE key='expense_cats'").fetchone():
            _merged_exp_cats = []
            for _old_key in ('cash_expense_cats', 'bank_expense_ctr_cats'):
                _row = conn.execute("SELECT value FROM pnl_settings WHERE key=?", (_old_key,)).fetchone()
                if _row and _row['value']:
                    try:
                        _vals = _json_lib.loads(_row['value'])
                    except Exception:
                        _vals = None
                    if isinstance(_vals, list):
                        for _v in _vals:
                            if _v and _v not in _merged_exp_cats:
                                _merged_exp_cats.append(_v)
            if _merged_exp_cats:
                conn.execute(
                    "INSERT OR REPLACE INTO pnl_settings (key, value) VALUES ('expense_cats', ?)",
                    (_json_lib.dumps(_merged_exp_cats),)
                )

        # Google Sheets export settings
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS gsheet_settings (
                key TEXT PRIMARY KEY,
                value TEXT
            );
        ''')

        # Branch restrictions per expense/income category
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS expense_category_branches (
                category_id INTEGER NOT NULL REFERENCES expense_categories(id) ON DELETE CASCADE,
                branch_id INTEGER NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
                PRIMARY KEY (category_id, branch_id)
            );
        ''')

        # Настраиваемые права ролей: видимость пунктов меню + область филиалов
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS role_menu_permissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                role TEXT NOT NULL,
                item_code TEXT NOT NULL,
                visible INTEGER DEFAULT 0,
                branch_scope TEXT DEFAULT 'own_only',
                UNIQUE(role, item_code)
            );
        ''')
        for _role in ROLE_CONFIGURABLE:
            for _code in list(MENU_ITEMS_BY_CODE.keys()) + MENU_SUBITEMS_FLAT:
                if conn.execute(
                    'SELECT 1 FROM role_menu_permissions WHERE role=? AND item_code=?', (_role, _code)
                ).fetchone():
                    continue
                _visible = 1 if _code in _DEFAULT_ROLE_VISIBLE.get(_role, set()) else 0
                conn.execute(
                    'INSERT INTO role_menu_permissions (role, item_code, visible, branch_scope) VALUES (?,?,?,?)',
                    (_role, _code, _visible, 'own_only')
                )

        # Разовая коррекция: изначальный сид ошибочно ставил visible=0 для
        # dashboard/employees/purchases/history/shifts_archive у admin/employee
        # (и учитывал только director для employees) — из-за этого администраторы
        # переставали видеть лист смены, хотя раньше эти страницы были доступны
        # любой залогиненной роли (просто без ссылки в меню). Правим один раз,
        # дальнейшие осознанные изменения владельца через Настройки не трогаем.
        if not conn.execute(
            "SELECT 1 FROM api_settings WHERE key='role_perms_visibility_fix_v1'"
        ).fetchone():
            _fix_ph = ','.join('?' * len(_ALWAYS_ACCESSIBLE_BEFORE))
            conn.execute(
                f"UPDATE role_menu_permissions SET visible=1 WHERE item_code IN ({_fix_ph}) AND visible=0",
                list(_ALWAYS_ACCESSIBLE_BEFORE)
            )
            conn.execute(
                "INSERT OR REPLACE INTO api_settings (key, value) VALUES ('role_perms_visibility_fix_v1', '1')"
            )

        # Одноразовая корректировка: смены, где "Итого выручка" не заполнилась
        # (0/пусто) при импорте, хотя нал+безнал+онлайн реально были внесены —
        # пересчитываем total_revenue из частей, если их сумма больше 1.
        if not conn.execute(
            "SELECT 1 FROM api_settings WHERE key='total_revenue_from_parts_fix_v1'"
        ).fetchone():
            conn.execute('''
                UPDATE shift_revenue
                SET total_revenue = COALESCE(cash_amount,0) + COALESCE(card_amount,0) + COALESCE(online_amount,0)
                WHERE COALESCE(total_revenue,0) <= 1
                  AND (COALESCE(cash_amount,0) + COALESCE(card_amount,0) + COALESCE(online_amount,0)) > 1
            ''')
            conn.execute(
                "INSERT OR REPLACE INTO api_settings (key, value) VALUES ('total_revenue_from_parts_fix_v1', '1')"
            )

        conn.execute('''
            CREATE TABLE IF NOT EXISTS revenue_plan (
                branch_id INTEGER NOT NULL,
                date TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (branch_id, date)
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS revenue_manual (
                branch_id INTEGER NOT NULL,
                date TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                PRIMARY KEY (branch_id, date)
            )
        ''')
        rm_cols = [r[1] for r in conn.execute("PRAGMA table_info(revenue_manual)").fetchall()]
        if 'orders_count' not in rm_cols:
            conn.execute("ALTER TABLE revenue_manual ADD COLUMN orders_count INTEGER DEFAULT 0")

        # Fix Russian-coded categories → correct English codes used by import
        _code_fixes = [
            ('стафф',                   'staff'),
            ('стаф',                    'staff'),
            ('рыба_(головы,хребты)',    'fish'),
            ('рыба_(головы_хребты)',    'fish'),
            ('за_масло_отработанное',   'oil'),
            ('другие_плюсы_в_кассу',   'cash_plus'),
        ]
        for bad_code, good_code in _code_fixes:
            row = conn.execute('SELECT id FROM expense_categories WHERE code=?', (bad_code,)).fetchone()
            if row:
                exists = conn.execute('SELECT id FROM expense_categories WHERE code=?', (good_code,)).fetchone()
                if exists:
                    # Correct code already exists — remove the duplicate
                    conn.execute('DELETE FROM expense_categories WHERE code=?', (bad_code,))
                else:
                    conn.execute('UPDATE expense_categories SET code=? WHERE code=?', (good_code, bad_code))

        # Отчёт по заказам — импорт из выгрузки iiko (CSV «отчёт по заказам»)
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS orders_import_batches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT,
                imported_count INTEGER DEFAULT 0,
                duplicate_count INTEGER DEFAULT 0,
                updated_count INTEGER DEFAULT 0,
                skipped_count INTEGER DEFAULT 0,
                created_by INTEGER,
                imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS orders_report (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_number TEXT NOT NULL,
                branch_raw TEXT NOT NULL,
                branch_id INTEGER REFERENCES branches(id),
                received_at TEXT NOT NULL,
                promised_minutes INTEGER,
                order_type_raw TEXT,
                order_type TEXT,
                ready_minutes INTEGER,
                delivery_minutes INTEGER,
                promo_code TEXT,
                amount REAL DEFAULT 0,
                new_client TEXT,
                import_batch_id INTEGER REFERENCES orders_import_batches(id) ON DELETE CASCADE,
                import_hash TEXT UNIQUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE INDEX IF NOT EXISTS idx_orders_report_received ON orders_report(received_at);
            CREATE INDEX IF NOT EXISTS idx_orders_report_branch ON orders_report(branch_id);
            CREATE INDEX IF NOT EXISTS idx_orders_report_number ON orders_report(order_number);
        ''')
        _or_cols = [r[1] for r in conn.execute("PRAGMA table_info(orders_report)").fetchall()]
        if 'new_client' not in _or_cols:
            conn.execute("ALTER TABLE orders_report ADD COLUMN new_client TEXT")
        _oib_cols = [r[1] for r in conn.execute("PRAGMA table_info(orders_import_batches)").fetchall()]
        if 'updated_count' not in _oib_cols:
            conn.execute("ALTER TABLE orders_import_batches ADD COLUMN updated_count INTEGER DEFAULT 0")

        conn.commit()


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def owner_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'owner':
            flash('Доступ только для владельца', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


def _load_role_perms():
    """Права меню текущей сессии, закэшированные на flask.g на время запроса."""
    if hasattr(g, '_role_perms'):
        return g._role_perms
    role = session.get('role')
    perms = {}
    if role in ROLE_CONFIGURABLE:
        with get_db() as conn:
            rows = conn.execute(
                'SELECT item_code, visible, branch_scope FROM role_menu_permissions WHERE role=?', (role,)
            ).fetchall()
        perms = {r['item_code']: {'visible': bool(r['visible']), 'branch_scope': r['branch_scope']} for r in rows}
    g._role_perms = perms
    return perms


def item_visible(item_code):
    """Виден ли пункт меню/страница текущей роли сессии. Owner — всегда True.
    Если у пункта есть подпункты (MENU_SUBITEMS) — виден, если разрешён весь
    раздел целиком, либо доступ дан хотя бы на один подпункт (см. subitem_visible)."""
    role = session.get('role')
    if role == 'owner':
        return True
    if role not in ROLE_CONFIGURABLE:
        return False
    perms = _load_role_perms()
    if bool(perms.get(item_code, {}).get('visible')):
        return True
    return any(bool(perms.get(code, {}).get('visible')) for code, _label, _bs in MENU_SUBITEMS.get(item_code, []))


def subitem_visible(parent_code, subitem_code):
    """Доступен ли конкретный подпункт (вкладка) parent_code текущей роли —
    да, если роли открыт весь раздел целиком, либо доступ дан именно на этот подпункт."""
    role = session.get('role')
    if role == 'owner':
        return True
    if role not in ROLE_CONFIGURABLE:
        return False
    perms = _load_role_perms()
    if bool(perms.get(parent_code, {}).get('visible')):
        return True
    return bool(perms.get(subitem_code, {}).get('visible'))


def group_has_visible_item(group):
    return any(item_visible(code) for code, label, grp, scoped in MENU_ITEMS if grp == group)


def any_menu_visible():
    return any(group_has_visible_item(grp) for grp in ('dash', 'reports', 'settings'))


app.jinja_env.globals['item_visible'] = item_visible
app.jinja_env.globals['subitem_visible'] = subitem_visible
app.jinja_env.globals['group_has_visible_item'] = group_has_visible_item
app.jinja_env.globals['any_menu_visible'] = any_menu_visible


def get_role_permissions(conn, role):
    """{item_code: {'visible': bool, 'branch_scope': str}} для указанной роли (для UI настроек)."""
    rows = conn.execute(
        'SELECT item_code, visible, branch_scope FROM role_menu_permissions WHERE role=?', (role,)
    ).fetchall()
    return {r['item_code']: {'visible': bool(r['visible']), 'branch_scope': r['branch_scope']} for r in rows}


def menu_permission_required(item_code):
    """Замена @owner_required для страниц с настраиваемым доступом по ролям."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not item_visible(item_code):
                flash('Доступ запрещён', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated
    return decorator


def get_effective_branch_ids(item_code, requested_ids=None):
    """Итоговый список филиалов для страницы item_code с учётом роли и её branch_scope.
    requested_ids — то, что пользователь выбрал в селекторе (может быть пустым/None).
    owner: requested_ids как есть (None/[] означает «все», как и раньше).
    не-owner с own_only: всегда свои филиалы, requested_ids игнорируется.
    не-owner с own_default: requested_ids, если переданы, иначе свои филиалы."""
    role = session.get('role')
    if role == 'owner':
        return requested_ids
    own_ids = [str(b) for b in _session_branch_ids()]
    if role not in ROLE_CONFIGURABLE:
        return own_ids
    scope = _load_role_perms().get(item_code, {}).get('branch_scope', 'own_only')
    if scope == 'own_default' and requested_ids:
        return requested_ids
    return own_ids


def can_pick_other_branches(item_code):
    """Может ли текущая роль вообще выбирать/видеть филиалы, отличные от своих, для item_code."""
    role = session.get('role')
    if role == 'owner':
        return True
    if role not in ROLE_CONFIGURABLE:
        return False
    return _load_role_perms().get(item_code, {}).get('branch_scope') == 'own_default'


app.jinja_env.globals['can_pick_other_branches'] = can_pick_other_branches


def get_current_user():
    if 'user_id' not in session:
        return None
    with get_db() as conn:
        return conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()


def get_client_ip():
    xff = request.headers.get('X-Forwarded-For', '')
    if xff:
        return xff.split(',')[0].strip()
    return request.remote_addr


def _calc_prev_kassa_nal(conn, branch_id, before_date):
    """Итого нал в кассе последней смены филиала до указанной даты (переносится как утренняя касса следующего дня)."""
    row = conn.execute('''
        SELECT COALESCE(r.morning_cash, 0)
               + COALESCE(r.cash_amount, 0)
               + COALESCE(r.change_amount, 0)
               + COALESCE((SELECT SUM(cp.amount_cash) FROM cash_plus_entries cp WHERE cp.shift_id=s.id), 0)
               - COALESCE((SELECT SUM(e.amount_cash) FROM expenses e WHERE e.shift_id=s.id), 0)
               - COALESCE((SELECT SUM(t.amount) FROM taxi_trips t
                            WHERE t.shift_id=s.id AND t.payment_type='cash'), 0)
               - COALESCE((SELECT SUM(es.total_amount) FROM employee_shifts es
                            WHERE es.shift_id=s.id AND es.is_paid=1), 0)
               AS kassa_nal
        FROM shifts s JOIN shift_revenue r ON r.shift_id=s.id
        WHERE s.branch_id=? AND s.date<?
        ORDER BY s.date DESC LIMIT 1
    ''', (branch_id, before_date)).fetchone()
    return (row['kassa_nal'] or 0) if row else 0


def _apply_change_amount_to_shift(conn, shift_id, branch_id, shift_date):
    """Apply change_amount to shift: override table first, then schedule rules."""
    try:
        d = date.fromisoformat(shift_date) if isinstance(shift_date, str) else shift_date
        weekday = d.weekday()
        shift_date = d.isoformat()
    except Exception:
        return
    override = conn.execute(
        'SELECT amount FROM change_date_overrides WHERE branch_id=? AND date=?',
        (branch_id, shift_date)
    ).fetchone()
    if override:
        conn.execute('UPDATE shift_revenue SET change_amount=? WHERE shift_id=?', (override['amount'], shift_id))
        return
    row = conn.execute(
        '''SELECT amount FROM change_schedule
           WHERE (branch_id IS NULL OR branch_id=?)
             AND (weekday IS NULL OR weekday=?)
             AND valid_from <= ?
             AND (valid_to IS NULL OR valid_to >= ?)
           ORDER BY branch_id DESC NULLS LAST, weekday DESC NULLS LAST, id DESC
           LIMIT 1''',
        (branch_id, weekday, shift_date, shift_date)
    ).fetchone()
    if row:
        conn.execute('UPDATE shift_revenue SET change_amount=? WHERE shift_id=?', (row['amount'], shift_id))


def _apply_all_change_schedules(conn):
    """Apply all change schedules to matching existing OPEN shifts only."""
    schedules = conn.execute(
        'SELECT * FROM change_schedule ORDER BY branch_id NULLS FIRST, weekday NULLS FIRST, id'
    ).fetchall()
    for sched in schedules:
        params = [sched['amount'], sched['valid_from']]
        clauses = ["s.date >= ?", "s.status != 'closed'"]
        if sched['valid_to']:
            clauses.append('s.date <= ?')
            params.append(sched['valid_to'])
        if sched['branch_id']:
            clauses.append('s.branch_id = ?')
            params.append(sched['branch_id'])
        if sched['weekday'] is not None:
            clauses.append("CAST(strftime('%u', s.date) AS INTEGER) - 1 = ?")
            params.append(sched['weekday'])
        where = ' AND '.join(clauses)
        conn.execute(f'''
            UPDATE shift_revenue SET change_amount = ?
            WHERE shift_id IN (SELECT s.id FROM shifts s WHERE {where})
        ''', params)


# ─── AUTH ─────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']
        with get_db() as conn:
            user = conn.execute('SELECT * FROM users WHERE username=?', (username,)).fetchone()
        if user and check_password_hash(user['password_hash'], password):
            rows = conn.execute(
                'SELECT branch_id FROM user_branches WHERE user_id=? ORDER BY branch_id',
                (user['id'],)
            ).fetchall()
            branch_ids = [r['branch_id'] for r in rows]
            if not branch_ids and user['branch_id']:
                branch_ids = [user['branch_id']]
            # IP restriction: non-owners are limited to the branch matching their IP
            # Managers (управляющие) are exempt — they have access to all assigned branches
            if user['role'] not in ('owner', 'director'):
                client_ip = get_client_ip()
                ip_branch = conn.execute(
                    "SELECT id FROM branches WHERE allowed_ip=? AND is_active=1",
                    (client_ip,)
                ).fetchone()
                if ip_branch:
                    branch_ids = [ip_branch['id']]
            session['user_id'] = user['id']
            session['role'] = user['role']
            session['full_name'] = user['full_name']
            session['branch_id'] = branch_ids[0] if branch_ids else None
            session['branch_ids'] = branch_ids
            return redirect(url_for('dashboard'))
        flash('Неверный логин или пароль', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ─── DASHBOARD ────────────────────────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    role = session.get('role')
    with get_db() as conn:
        if role == 'owner':
            branches = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
            stats = conn.execute('''
                SELECT b.name, s.date, COALESCE(r.total_revenue,0) as revenue,
                       COALESCE(r.delivery_orders,0)+COALESCE(r.pickup_orders,0) as orders,
                       s.status, s.id as shift_id
                FROM branches b
                LEFT JOIN shifts s ON s.branch_id=b.id AND s.date >= date('now','-7 days')
                LEFT JOIN shift_revenue r ON r.shift_id=s.id
                WHERE b.is_active=1
                ORDER BY s.date DESC, b.name
            ''').fetchall()
            weekly = conn.execute('''
                SELECT COALESCE(SUM(r.total_revenue),0) as total,
                       COALESCE(SUM(r.delivery_revenue),0) as delivery,
                       COALESCE(SUM(r.pickup_revenue),0) as pickup,
                       COALESCE(SUM(r.delivery_orders+r.pickup_orders),0) as orders
                FROM shifts s JOIN shift_revenue r ON r.shift_id=s.id
                WHERE s.date >= date('now','weekday 0','-7 days')
            ''').fetchone()
            _week_from = conn.execute("SELECT date('now','weekday 0','-7 days')").fetchone()[0]
            _week_to   = date.today().isoformat()
            _manual_week = _manual_rev_total(conn, _week_from, _week_to)
            weekly = dict(weekly)
            weekly['total'] = float(weekly['total']) + _manual_week
            open_shifts = conn.execute('''
                SELECT s.*, b.name as branch_name
                FROM shifts s JOIN branches b ON b.id=s.branch_id
                WHERE s.date=date('now') AND s.status='open'
            ''').fetchall()
            # KPI blocks
            kpi_blocks = conn.execute(
                'SELECT * FROM kpi_blocks WHERE is_active=1 ORDER BY sort_order, id'
            ).fetchall()
            # Current month stats
            month_rev = conn.execute('''
                SELECT
                    COALESCE(SUM(r.cash_amount), 0)   AS cash_amount,
                    COALESCE(SUM(r.card_amount), 0)   AS card_amount,
                    COALESCE(SUM(r.online_amount), 0) AS online_amount,
                    COALESCE(SUM(r.total_revenue), 0) AS total_revenue
                FROM shifts s
                JOIN shift_revenue r ON r.shift_id = s.id
                WHERE s.date >= date('now', 'start of month')
            ''').fetchone()
            _month_from = conn.execute("SELECT date('now','start of month')").fetchone()[0]
            _manual_month = _manual_rev_total(conn, _month_from, date.today().isoformat())
            month_rev = dict(month_rev)
            month_rev['total_revenue'] = float(month_rev['total_revenue']) + _manual_month
            month_fot = conn.execute('''
                SELECT COALESCE(SUM(es.total_amount), 0) AS fot
                FROM employee_shifts es
                JOIN shifts s ON s.id = es.shift_id
                WHERE s.date >= date('now', 'start of month')
            ''').fetchone()
            branch_groups = get_branch_groups(conn)
            return render_template('dashboard_owner.html',
                branches=branches, stats=stats, weekly=weekly,
                open_shifts=open_shifts, kpi_blocks=kpi_blocks,
                month_rev=month_rev, month_fot=month_fot['fot'] or 0,
                branch_groups=branch_groups, today=date.today().isoformat())
        else:
            if not item_visible('dashboard'):
                # "dashboard" — единственная всегда-достижимая страница после логина,
                # поэтому при скрытом пункте не редиректим (это дало бы цикл),
                # а показываем пустую страницу без данных филиалов.
                return render_template('dashboard_admin.html', branches_shifts=[], today=date.today().isoformat(), recent=[], no_access=True)
            today = date.today().isoformat()
            _req_bids = [b for b in request.args.getlist('branch_ids') if b.isdigit()]
            bids = [int(b) for b in get_effective_branch_ids('dashboard', _req_bids) or []]
            if not bids:
                flash('У вас не назначен филиал', 'warning')
                return render_template('dashboard_admin.html', branches_shifts=[], today=today, recent=[])
            branches_shifts = []
            for bid in bids:
                branch = conn.execute('SELECT * FROM branches WHERE id=?', (bid,)).fetchone()
                if not branch:
                    continue
                shift = conn.execute(
                    'SELECT s.*, u.full_name as closed_by_name FROM shifts s '
                    'LEFT JOIN users u ON u.id=s.closed_by '
                    'WHERE s.branch_id=? AND s.date=?',
                    (bid, today)
                ).fetchone()
                suggested_morning_cash = 0 if shift else int(round(_calc_prev_kassa_nal(conn, bid, today)))
                branches_shifts.append({'branch': branch, 'shift': shift, 'suggested_morning_cash': suggested_morning_cash})
            if bids:
                ids_str = ','.join(str(int(b)) for b in bids)
                recent = conn.execute(f'''
                    SELECT s.*, b.name as branch_name, COALESCE(r.total_revenue,0) as revenue
                    FROM shifts s JOIN branches b ON b.id=s.branch_id
                    LEFT JOIN shift_revenue r ON r.shift_id=s.id
                    WHERE s.branch_id IN ({ids_str}) ORDER BY s.date DESC LIMIT 10
                ''').fetchall()
            else:
                recent = []
            all_branches = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
            branch_groups = get_branch_groups(conn)
            return render_template('dashboard_admin.html',
                branches_shifts=branches_shifts, today=today, recent=recent,
                all_branches=all_branches, branch_groups=branch_groups,
                selected_branch_ids=_req_bids)


# ─── KPI API ──────────────────────────────────────────────────────────────────

@app.route('/api/kpi-values')
@login_required
@menu_permission_required('dashboard')
def api_kpi_values():
    branch_id = request.args.get('branch_id', '')
    date_from = request.args.get('date_from', date.today().isoformat())
    date_to = request.args.get('date_to', date.today().isoformat())
    with get_db() as conn:
        vals = get_kpi_values(conn, branch_id, date_from, date_to)
        blocks = conn.execute(
            'SELECT * FROM kpi_blocks WHERE is_active=1 ORDER BY sort_order, id'
        ).fetchall()
        results = []
        for b in blocks:
            value = safe_eval(b['formula'], vals)
            results.append({
                'id': b['id'],
                'title': b['title'],
                'value': round(value, 2) if value is not None else None,
                'color': b['color'],
                'unit': b['unit'],
            })
    return jsonify({'ok': True, 'blocks': results, 'vars': {k: round(v, 2) for k, v in vals.items()}})


@app.route('/api/revenue-year')
@login_required
@menu_permission_required('dashboard')
def api_revenue_year():
    today = date.today()
    raw_bids = request.args.get('branch_ids', '')
    bids     = [int(x) for x in raw_bids.split(',') if x.strip().isdigit()]
    bf       = f"AND s.branch_id IN ({','.join('?'*len(bids))})" if bids else ''
    # скользящие 12 месяцев, заканчивающихся текущим
    start_m = today.month - 11
    start_y = today.year
    if start_m <= 0:
        start_m += 12
        start_y -= 1
    date_from = date(start_y, start_m, 1).isoformat()
    date_to   = today.isoformat()
    with get_db() as conn:
        rows = conn.execute(f'''
            SELECT CAST(strftime('%Y', s.date) AS INTEGER) AS year,
                   CAST(strftime('%m', s.date) AS INTEGER) AS month,
                   COALESCE(SUM(r.total_revenue), 0) AS revenue
            FROM shifts s JOIN shift_revenue r ON r.shift_id = s.id
            WHERE s.date BETWEEN ? AND ? {bf}
            GROUP BY year, month ORDER BY year, month
        ''', [date_from, date_to] + bids).fetchall()
        manual_map = _manual_rev_by_month(conn, date_from, date_to, bids or None)
    rev = {(r['year'], r['month']): int(r['revenue']) for r in rows}
    labels = ['', 'Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн',
              'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек']
    months_list = []
    y, m = start_y, start_m
    for _ in range(12):
        months_list.append({'year': y, 'month': m, 'label': labels[m],
                            'revenue': rev.get((y, m), 0) + int(manual_map.get((y, m), 0))})
        m += 1
        if m > 12:
            m = 1
            y += 1
    return jsonify({'ok': True, 'total': sum(x['revenue'] for x in months_list), 'months': months_list})


@app.route('/api/revenue-all')
@login_required
@menu_permission_required('dashboard')
def api_revenue_all():
    """Месячная выручка за весь период где есть данные (по выбранным филиалам)."""
    raw_bids = request.args.get('branch_ids', '')
    bids     = [int(x) for x in raw_bids.split(',') if x.strip().isdigit()]
    bid_str  = f"AND s.branch_id IN ({','.join(str(b) for b in bids)})" if bids else ''
    mbid_str = f"AND branch_id IN ({','.join(str(b) for b in bids)})" if bids else ''
    today_iso = date.today().isoformat()
    with get_db() as conn:
        min_shift = conn.execute(
            f'SELECT MIN(s.date) FROM shifts s JOIN shift_revenue r ON r.shift_id=s.id WHERE 1=1 {bid_str}'
        ).fetchone()[0]
        min_manual = conn.execute(
            f'SELECT MIN(date) FROM revenue_manual WHERE 1=1 {mbid_str}'
        ).fetchone()[0]
    candidates = [x for x in [min_shift, min_manual] if x]
    if not candidates:
        return jsonify({'ok': True, 'total': 0, 'months': []})
    date_from = min(candidates)
    bf = f"AND s.branch_id IN ({','.join('?'*len(bids))})" if bids else ''
    with get_db() as conn:
        rows = conn.execute(f'''
            SELECT CAST(strftime('%Y', s.date) AS INTEGER) AS year,
                   CAST(strftime('%m', s.date) AS INTEGER) AS month,
                   COALESCE(SUM(r.total_revenue), 0) AS revenue
            FROM shifts s JOIN shift_revenue r ON r.shift_id = s.id
            WHERE s.date BETWEEN ? AND ? {bf}
            GROUP BY year, month ORDER BY year, month
        ''', [date_from, today_iso] + bids).fetchall()
        manual_map = _manual_rev_by_month(conn, date_from, today_iso, bids or None)
    rev = {(r['year'], r['month']): int(r['revenue']) for r in rows}
    labels = ['', 'Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн',
              'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек']
    start = date.fromisoformat(date_from)
    end   = date.today()
    months_list = []
    y, m = start.year, start.month
    while (y < end.year) or (y == end.year and m <= end.month):
        label = labels[m] + " '" + str(y)[-2:]
        months_list.append({'year': y, 'month': m, 'label': label,
                            'revenue': rev.get((y, m), 0) + int(manual_map.get((y, m), 0))})
        m += 1
        if m > 12:
            m = 1
            y += 1
    return jsonify({'ok': True, 'total': sum(x['revenue'] for x in months_list), 'months': months_list})


@app.route('/api/revenue-months')
@login_required
@menu_permission_required('dashboard')
def api_revenue_months():
    date_from = request.args.get('date_from')
    date_to   = request.args.get('date_to', date.today().isoformat())
    raw_bids  = request.args.get('branch_ids', '')
    bids      = [int(x) for x in raw_bids.split(',') if x.strip().isdigit()]
    bf        = f"AND s.branch_id IN ({','.join('?'*len(bids))})" if bids else ''
    if not date_from:
        return jsonify({'ok': False, 'error': 'date_from required'}), 400
    with get_db() as conn:
        rows = conn.execute(f'''
            SELECT CAST(strftime('%Y', s.date) AS INTEGER) AS year,
                   CAST(strftime('%m', s.date) AS INTEGER) AS month,
                   COALESCE(SUM(r.total_revenue), 0) AS revenue
            FROM shifts s JOIN shift_revenue r ON r.shift_id = s.id
            WHERE s.date BETWEEN ? AND ? {bf}
            GROUP BY year, month ORDER BY year, month
        ''', [date_from, date_to] + bids).fetchall()
        manual_map = _manual_rev_by_month(conn, date_from, date_to, bids or None)
    rev = {(r['year'], r['month']): int(r['revenue']) for r in rows}
    labels = ['', 'Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн',
              'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек']
    start = date.fromisoformat(date_from)
    end   = date.fromisoformat(date_to)
    months_list = []
    y, m = start.year, start.month
    while (y < end.year) or (y == end.year and m <= end.month):
        label = labels[m] + " '" + str(y)[-2:]
        months_list.append({'year': y, 'month': m, 'label': label,
                            'revenue': rev.get((y, m), 0) + int(manual_map.get((y, m), 0))})
        m += 1
        if m > 12:
            m = 1
            y += 1
    return jsonify({'ok': True, 'total': sum(x['revenue'] for x in months_list), 'months': months_list})


@app.route('/api/lfl')
@login_required
@menu_permission_required('lfl_dashboard')
def api_lfl():
    from calendar import monthrange
    today  = date.today()
    metric = request.args.get('metric', 'revenue')  # 'revenue' or 'orders'
    raw_bids = request.args.get('branch_ids', '')
    bids = [int(x) for x in raw_bids.split(',') if x.strip().isdigit()]
    bids = [int(b) for b in get_effective_branch_ids('lfl_dashboard', [str(b) for b in bids]) or []]
    bf   = f"AND s.branch_id IN ({','.join('?'*len(bids))})" if bids else ''
    if metric == 'orders':
        agg = 'COALESCE(SUM(r.delivery_orders),0) + COALESCE(SUM(r.pickup_orders),0)'
    else:
        agg = 'COALESCE(SUM(r.total_revenue),0)'
    month_labels = ['', 'Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн',
                    'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек']
    with get_db() as conn:
        shift_where = f"WHERE s.branch_id IN ({','.join('?'*len(bids))})" if bids else ''
        manual_where = f"WHERE branch_id IN ({','.join('?'*len(bids))})" if bids else ''
        min_shift = conn.execute(
            f'SELECT MIN(s.date) FROM shifts s JOIN shift_revenue r ON r.shift_id=s.id {shift_where}', bids
        ).fetchone()[0]
        min_manual = conn.execute(
            f'SELECT MIN(date) FROM revenue_manual {manual_where}', bids
        ).fetchone()[0]
        candidates = [x for x in [min_shift, min_manual] if x]
        if candidates:
            earliest = min(candidates)
            start = date.fromisoformat(earliest[:10])
            start_y, start_m = start.year, start.month
        else:
            start_y, start_m = today.year, today.month - 11
            if start_m <= 0:
                start_m += 12
                start_y -= 1
        months_seq = []
        y, m = start_y, start_m
        while (y < today.year) or (y == today.year and m <= today.month):
            months_seq.append((y, m))
            m += 1
            if m > 12:
                m = 1
                y += 1
        result = []
        for (yr, mo) in months_seq:
            is_current = (yr == today.year and mo == today.month)
            days_in_this = monthrange(yr, mo)[1]
            days_in_last = monthrange(yr - 1, mo)[1]
            if is_current:
                d_from_this = date(yr, mo, 1).isoformat()
                d_to_this   = today.isoformat()
                d_from_last = date(yr - 1, mo, 1).isoformat()
                d_to_last   = date(yr - 1, mo, min(today.day, days_in_last)).isoformat()
            else:
                d_from_this = date(yr, mo, 1).isoformat()
                d_to_this   = date(yr, mo, days_in_this).isoformat()
                d_from_last = date(yr - 1, mo, 1).isoformat()
                d_to_last   = date(yr - 1, mo, days_in_last).isoformat()
            val_this = conn.execute(
                f'SELECT {agg} FROM shifts s JOIN shift_revenue r ON r.shift_id=s.id WHERE s.date BETWEEN ? AND ? {bf}',
                [d_from_this, d_to_this] + bids
            ).fetchone()[0] or 0
            if metric == 'revenue':
                val_this += _manual_rev_total(conn, d_from_this, d_to_this, bids or None)
            val_last = conn.execute(
                f'SELECT {agg} FROM shifts s JOIN shift_revenue r ON r.shift_id=s.id WHERE s.date BETWEEN ? AND ? {bf}',
                [d_from_last, d_to_last] + bids
            ).fetchone()[0] or 0
            if metric == 'revenue':
                val_last += _manual_rev_total(conn, d_from_last, d_to_last, bids or None)
            # val_this=0 → нет данных за этот период, не показываем как -100%
            if val_this == 0:
                lfl_pct = None
            else:
                lfl_pct = round((val_this / val_last - 1) * 100, 1) if val_last > 0 else None
            result.append({
                'year': yr, 'month': mo,
                'label': month_labels[mo] + " '" + str(yr)[-2:],
                'this_year': int(val_this),
                'last_year': int(val_last),
                'lfl_pct': lfl_pct,
                'is_current': is_current,
            })
    # Обрезаем с обоих концов месяцы где вообще нет данных (оба года = 0)
    meaningful = [i for i, r in enumerate(result) if r['this_year'] > 0 or r['last_year'] > 0]
    if meaningful:
        result = result[meaningful[0]:meaningful[-1] + 1]
    else:
        result = []
    return jsonify({'ok': True, 'months': result})


@app.route('/api/lfl-branches')
@login_required
@menu_permission_required('lfl_dashboard')
def api_lfl_branches():
    """LFL за последний месяц с данными — разбивка по каждому филиалу."""
    from calendar import monthrange as _mrange
    today  = date.today()
    metric = request.args.get('metric', 'revenue')
    if metric == 'orders':
        agg = 'COALESCE(SUM(r.delivery_orders),0) + COALESCE(SUM(r.pickup_orders),0)'
    else:
        agg = 'COALESCE(SUM(r.total_revenue),0)'

    with get_db() as conn:
        branches = conn.execute(
            'SELECT id, name, abbr FROM branches WHERE is_active=1 ORDER BY id'
        ).fetchall()
        _requested = [b for b in request.args.getlist('branch_ids') if b.isdigit()]
        _allowed = get_effective_branch_ids('lfl_dashboard', _requested)
        if _allowed:
            _allowed_set = {int(x) for x in _allowed}
            branches = [b for b in branches if b['id'] in _allowed_set]
        all_bids = [b['id'] for b in branches]
        if not all_bids:
            return jsonify({'ok': True, 'branches': [], 'month_label': ''})

        # Находим последний месяц, где есть хоть какие-то данные (по всем филиалам суммарно)
        bf_all = f"AND s.branch_id IN ({','.join('?'*len(all_bids))})"
        earliest_shift = conn.execute(
            f'SELECT MIN(s.date) FROM shifts s JOIN shift_revenue r ON r.shift_id=s.id {bf_all}',
            all_bids
        ).fetchone()[0]
        earliest_manual = conn.execute(
            f"SELECT MIN(date) FROM revenue_manual WHERE branch_id IN ({','.join('?'*len(all_bids))})",
            all_bids
        ).fetchone()[0]
        candidates = [x for x in [earliest_shift, earliest_manual] if x]
        if not candidates:
            return jsonify({'ok': True, 'branches': [], 'month_label': ''})

        start = date.fromisoformat(min(candidates)[:10])
        # Перебираем месяцы и берём последний с данными
        last_month = None
        y, m = start.year, start.month
        while (y < today.year) or (y == today.year and m <= today.month):
            is_cur = (y == today.year and m == today.month)
            days_t = _mrange(y, m)[1]
            days_l = _mrange(y - 1, m)[1]
            d_from = date(y, m, 1).isoformat()
            d_to   = today.isoformat() if is_cur else date(y, m, days_t).isoformat()
            total_this = conn.execute(
                f'SELECT {agg} FROM shifts s JOIN shift_revenue r ON r.shift_id=s.id WHERE s.date BETWEEN ? AND ? {bf_all}',
                [d_from, d_to] + all_bids
            ).fetchone()[0] or 0
            if metric == 'revenue':
                total_this += _manual_rev_total(conn, d_from, d_to, all_bids)
            if total_this > 0:
                last_month = (y, m, is_cur, d_from, d_to, days_l)
            m += 1
            if m > 12: m = 1; y += 1

        if not last_month:
            return jsonify({'ok': True, 'branches': [], 'month_label': ''})

        yr, mo, is_cur, d_from_this, d_to_this, days_l = last_month
        d_from_last = date(yr - 1, mo, 1).isoformat()
        d_to_last   = date(yr - 1, mo, min(today.day, days_l) if is_cur else days_l).isoformat()

        month_names = ['', 'Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн',
                       'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек']
        month_label = month_names[mo] + " '" + str(yr)[-2:]
        if is_cur:
            month_label = 'тек. ' + month_label

        result = []
        for b in branches:
            bid   = b['id']
            bf1   = 'AND s.branch_id = ?'
            this_ = conn.execute(
                f'SELECT {agg} FROM shifts s JOIN shift_revenue r ON r.shift_id=s.id WHERE s.date BETWEEN ? AND ? {bf1}',
                [d_from_this, d_to_this, bid]
            ).fetchone()[0] or 0
            if metric == 'revenue':
                this_ += _manual_rev_total(conn, d_from_this, d_to_this, [bid])
            last_ = conn.execute(
                f'SELECT {agg} FROM shifts s JOIN shift_revenue r ON r.shift_id=s.id WHERE s.date BETWEEN ? AND ? {bf1}',
                [d_from_last, d_to_last, bid]
            ).fetchone()[0] or 0
            if metric == 'revenue':
                last_ += _manual_rev_total(conn, d_from_last, d_to_last, [bid])

            if this_ == 0:
                lfl_pct = None
            else:
                lfl_pct = round((this_ / last_ - 1) * 100, 1) if last_ > 0 else None

            result.append({
                'id':       bid,
                'name':     b['name'],
                'abbr':     b['abbr'] or b['name'][:3].upper(),
                'this_year': int(this_),
                'last_year': int(last_),
                'lfl_pct':  lfl_pct,
            })

    return jsonify({'ok': True, 'branches': result, 'month_label': month_label})


@app.route('/api/revenue-summary')
@login_required
def api_revenue_summary():
    if not (item_visible('dashboard') or item_visible('ratings_dashboard')):
        return jsonify({'ok': False, 'error': 'forbidden'}), 403
    date_from = request.args.get('date_from', date.today().isoformat())
    date_to   = request.args.get('date_to',   date.today().isoformat())
    raw_bids  = request.args.get('branch_ids', '')
    bids      = [int(x) for x in raw_bids.split(',') if x.strip().isdigit()]
    _code = 'dashboard' if item_visible('dashboard') else 'ratings_dashboard'
    bids      = [int(b) for b in get_effective_branch_ids(_code, [str(b) for b in bids]) or []]
    bf        = f"AND s.branch_id IN ({','.join('?'*len(bids))})" if bids else ''
    with get_db() as conn:
        total_row = conn.execute(f'''
            SELECT COALESCE(SUM(r.total_revenue), 0)    AS total,
                   COALESCE(SUM(r.cash_amount),   0)    AS cash,
                   COALESCE(SUM(r.card_amount),   0)    AS card,
                   COALESCE(SUM(r.online_amount), 0)    AS online,
                   COALESCE(SUM(r.delivery_revenue), 0) AS delivery,
                   COALESCE(SUM(r.pickup_revenue),   0) AS pickup,
                   COALESCE(SUM(r.delivery_orders),  0) AS delivery_orders
            FROM shifts s JOIN shift_revenue r ON r.shift_id = s.id
            WHERE s.date BETWEEN ? AND ? {bf}
        ''', [date_from, date_to] + bids).fetchone()
        fot_row = conn.execute(f'''
            SELECT COALESCE(SUM(es.total_amount), 0) AS fot
            FROM employee_shifts es JOIN shifts s ON s.id = es.shift_id
            WHERE s.date BETWEEN ? AND ? {bf}
        ''', [date_from, date_to] + bids).fetchone()
        courier_fot_row = conn.execute(f'''
            SELECT COALESCE(SUM(es.total_amount), 0) AS courier_fot
            FROM employee_shifts es JOIN shifts s ON s.id = es.shift_id
            WHERE s.date BETWEEN ? AND ? AND es.role_snapshot = 'courier' {bf}
        ''', [date_from, date_to] + bids).fetchone()
        branch_rev_rows = conn.execute(f'''
            SELECT b.id, b.name, b.abbr,
                   COALESCE(SUM(r.total_revenue), 0)    AS revenue,
                   COALESCE(SUM(r.pickup_revenue), 0)   AS pickup,
                   COALESCE(SUM(r.delivery_revenue), 0) AS delivery_revenue,
                   COALESCE(SUM(r.delivery_orders),  0) AS delivery_orders
            FROM shifts s
            JOIN branches b ON b.id = s.branch_id
            JOIN shift_revenue r ON r.shift_id = s.id
            WHERE s.date BETWEEN ? AND ? {bf}
            GROUP BY b.id, b.name ORDER BY revenue DESC
        ''', [date_from, date_to] + bids).fetchall()
        branch_fot_rows = conn.execute(f'''
            SELECT b.name,
                   COALESCE(SUM(es.total_amount), 0) AS fot,
                   COALESCE(SUM(CASE WHEN es.role_snapshot='courier' THEN es.total_amount ELSE 0 END), 0) AS courier_fot
            FROM employee_shifts es
            JOIN shifts s ON s.id = es.shift_id
            JOIN branches b ON b.id = s.branch_id
            WHERE s.date BETWEEN ? AND ? {bf}
            GROUP BY b.id, b.name
        ''', [date_from, date_to] + bids).fetchall()
        # Manual revenue по филиалам (fallback)
        mbf = f"AND m.branch_id IN ({','.join('?'*len(bids))})" if bids else ''
        manual_branch_rows = conn.execute(f'''
            SELECT m.branch_id, b.id, b.name, b.abbr,
                   COALESCE(SUM(m.amount), 0) AS revenue
            FROM revenue_manual m
            JOIN branches b ON b.id = m.branch_id
            WHERE m.date BETWEEN ? AND ? {mbf}
            AND NOT EXISTS (
                SELECT 1 FROM shifts s JOIN shift_revenue r ON r.shift_id=s.id
                WHERE s.date=m.date AND s.branch_id=m.branch_id AND (r.total_revenue > 0 OR s.status='closed')
            )
            GROUP BY m.branch_id
        ''', [date_from, date_to] + bids).fetchall()
        manual_total = _manual_rev_total(conn, date_from, date_to, bids or None)
        # Plan
        plan_bf2 = f"AND branch_id IN ({','.join('?'*len(bids))})" if bids else ''
        plan_rows = conn.execute(f'''
            SELECT branch_id, COALESCE(SUM(amount), 0) AS plan
            FROM revenue_plan
            WHERE date BETWEEN ? AND ? {plan_bf2}
            GROUP BY branch_id
        ''', [date_from, date_to] + bids).fetchall()
        plan_by_bid = {r['branch_id']: int(r['plan'] or 0) for r in plan_rows}

    total       = int(total_row['total'] or 0) + int(manual_total)
    plan_total  = sum(plan_by_bid.values())
    fot         = int(fot_row['fot'] or 0)
    courier_fot = int(courier_fot_row['courier_fot'] or 0)
    fot_by_name = {r['name']: {'fot': int(r['fot']), 'courier_fot': int(r['courier_fot'])} for r in branch_fot_rows}
    manual_by_bid = {r['id']: {'name': r['name'], 'abbr': r['abbr'], 'revenue': int(r['revenue'])}
                     for r in manual_branch_rows}
    # Объединяем shift-филиалы и manual-филиалы
    seen_bids = set()
    branches = []
    for br in branch_rev_rows:
        name = br['name'] or ''
        abbr = (br['abbr'] or '').strip() or name[:3].upper()
        bf_   = fot_by_name.get(name, {'fot': 0, 'courier_fot': 0})
        man   = manual_by_bid.get(br['id'], {})
        bid_  = br['id']
        seen_bids.add(bid_)
        branches.append({
            'abbr': abbr, 'name': name,
            'revenue':         int(br['revenue']) + int(man.get('revenue', 0)),
            'pickup':          int(br['pickup']),
            'delivery_revenue':int(br['delivery_revenue']),
            'delivery_orders': int(br['delivery_orders']),
            'fot':             bf_['fot'],
            'courier_fot':     bf_['courier_fot'],
            'plan':            plan_by_bid.get(bid_, 0),
        })
    # Филиалы только с manual-данными (без смен)
    for bid, man in manual_by_bid.items():
        if bid not in seen_bids:
            name = man['name'] or ''
            abbr = (man['abbr'] or '').strip() or name[:3].upper()
            branches.append({
                'abbr': abbr, 'name': name,
                'revenue': man['revenue'],
                'pickup': 0, 'delivery_revenue': 0, 'delivery_orders': 0,
                'fot': 0, 'courier_fot': 0,
                'plan': plan_by_bid.get(bid, 0),
            })
    branches.sort(key=lambda x: -x['revenue'])

    return jsonify({
        'ok': True,
        'total': total,
        'plan_total': plan_total,
        'cash':  int(total_row['cash'] or 0),
        'card':  int(total_row['card'] or 0),
        'online': int(total_row['online'] or 0),
        'delivery': int(total_row['delivery'] or 0),
        'delivery_orders': int(total_row['delivery_orders'] or 0),
        'pickup': int(total_row['pickup'] or 0),
        'fot':   fot,
        'courier_fot': courier_fot,
        'branches': branches,
    })


@app.route('/api/revenue-days')
@login_required
@menu_permission_required('dashboard')
def api_revenue_days():
    date_from = request.args.get('date_from', date.today().isoformat())
    date_to   = request.args.get('date_to',   date.today().isoformat())
    raw_bids  = request.args.get('branch_ids', '')
    bids      = [int(x) for x in raw_bids.split(',') if x.strip().isdigit()]
    bfilt     = f"AND s.branch_id IN ({','.join('?'*len(bids))})" if bids else ''
    plan_bf = f"AND branch_id IN ({','.join('?'*len(bids))})" if bids else ''
    with get_db() as conn:
        rows = conn.execute(f'''
            SELECT s.date, COALESCE(SUM(r.total_revenue), 0) AS revenue
            FROM shifts s JOIN shift_revenue r ON r.shift_id = s.id
            WHERE s.date BETWEEN ? AND ? {bfilt}
            GROUP BY s.date ORDER BY s.date
        ''', [date_from, date_to] + bids).fetchall()
        manual_days = _manual_rev_by_day(conn, date_from, date_to, bids or None)
        plan_rows = conn.execute(f'''
            SELECT date, COALESCE(SUM(amount), 0) AS plan
            FROM revenue_plan
            WHERE date BETWEEN ? AND ? {plan_bf}
            GROUP BY date
        ''', [date_from, date_to] + bids).fetchall()
    rev_map = {r['date']: int(r['revenue']) for r in rows}
    for d, amt in manual_days.items():
        rev_map[d] = rev_map.get(d, 0) + int(amt)
    plan_map = {r['date']: int(r['plan'] or 0) for r in plan_rows}
    # Возвращаем все дни диапазона (включая ещё не наступившие), чтобы на
    # дашборде можно было показать план на будущие дни столбцами
    all_days = []
    _d = datetime.strptime(date_from, '%Y-%m-%d').date()
    _dt_to = datetime.strptime(date_to, '%Y-%m-%d').date()
    while _d <= _dt_to:
        ds = _d.isoformat()
        all_days.append({'date': ds, 'revenue': rev_map.get(ds, 0), 'plan': plan_map.get(ds, 0)})
        _d += timedelta(days=1)
    return jsonify({'ok': True, 'days': all_days})


@app.route('/api/fot-summary')
@login_required
@menu_permission_required('fot_dashboard')
def api_fot_summary():
    date_from = request.args.get('date_from', date.today().isoformat())
    date_to   = request.args.get('date_to',   date.today().isoformat())
    raw_bids  = request.args.get('branch_ids', '')
    bids      = [int(x) for x in raw_bids.split(',') if x.strip().isdigit()]
    bids      = [int(b) for b in get_effective_branch_ids('fot_dashboard', [str(b) for b in bids]) or []]
    bf        = f"AND s.branch_id IN ({','.join('?'*len(bids))})" if bids else ''
    with get_db() as conn:
        rev_row = conn.execute(f'''
            SELECT COALESCE(SUM(r.total_revenue), 0) AS revenue
            FROM shifts s JOIN shift_revenue r ON r.shift_id = s.id
            WHERE s.date BETWEEN ? AND ? {bf}
        ''', [date_from, date_to] + bids).fetchone()
        fot_row = conn.execute(f'''
            SELECT COALESCE(SUM(es.total_amount), 0) AS fot
            FROM employee_shifts es JOIN shifts s ON s.id = es.shift_id
            WHERE s.date BETWEEN ? AND ? {bf}
        ''', [date_from, date_to] + bids).fetchone()
        role_rows = conn.execute(f'''
            SELECT es.role_snapshot,
                   COALESCE(SUM(es.total_amount), 0) AS fot
            FROM employee_shifts es JOIN shifts s ON s.id = es.shift_id
            WHERE s.date BETWEEN ? AND ? {bf}
            GROUP BY es.role_snapshot ORDER BY fot DESC
        ''', [date_from, date_to] + bids).fetchall()
        manual_rev = _manual_rev_total(conn, date_from, date_to, bids or None)
    revenue = int(rev_row['revenue'] or 0) + int(manual_rev)
    fot     = int(fot_row['fot'] or 0)
    fot_pct = round(fot / revenue * 100, 1) if revenue > 0 else 0
    role_labels = {'admin':'Администраторы','sushi':'Сушисты','packer':'Упаковщики',
                   'courier':'Курьеры','cleaner':'Уборщицы','cook':'Повара'}
    roles = []
    for r in role_rows:
        rfot = int(r['fot'] or 0)
        role = r['role_snapshot'] or ''
        if rfot <= 0 or not role:
            continue
        roles.append({
            'role':    role,
            'label':   role_labels.get(role, role),
            'fot':     rfot,
            'pct':     round(rfot / fot * 100) if fot > 0 else 0,
            'rev_pct': round(rfot / revenue * 100, 1) if revenue > 0 else 0,
        })
    return jsonify({'ok': True, 'fot': fot, 'revenue': revenue, 'fot_pct': fot_pct, 'roles': roles})


@app.route('/api/fot-year')
@login_required
@menu_permission_required('fot_dashboard')
def api_fot_year():
    today   = date.today()
    start_m = today.month - 11
    start_y = today.year
    if start_m <= 0:
        start_m += 12
        start_y -= 1
    date_from = date(start_y, start_m, 1).isoformat()
    date_to   = today.isoformat()
    raw_bids  = request.args.get('branch_ids', '')
    bids      = [int(x) for x in raw_bids.split(',') if x.strip().isdigit()]
    bids      = [int(b) for b in get_effective_branch_ids('fot_dashboard', [str(b) for b in bids]) or []]
    bf        = f"AND s.branch_id IN ({','.join('?'*len(bids))})" if bids else ''
    with get_db() as conn:
        fot_rows = conn.execute(f'''
            SELECT CAST(strftime('%Y',s.date) AS INTEGER) AS year,
                   CAST(strftime('%m',s.date) AS INTEGER) AS month,
                   COALESCE(SUM(es.total_amount),0) AS fot
            FROM employee_shifts es JOIN shifts s ON s.id=es.shift_id
            WHERE s.date BETWEEN ? AND ? {bf} GROUP BY year,month
        ''', [date_from, date_to] + bids).fetchall()
        rev_rows = conn.execute(f'''
            SELECT CAST(strftime('%Y',s.date) AS INTEGER) AS year,
                   CAST(strftime('%m',s.date) AS INTEGER) AS month,
                   COALESCE(SUM(r.total_revenue),0) AS revenue
            FROM shifts s JOIN shift_revenue r ON r.shift_id=s.id
            WHERE s.date BETWEEN ? AND ? {bf} GROUP BY year,month
        ''', [date_from, date_to] + bids).fetchall()
        manual_map = _manual_rev_by_month(conn, date_from, date_to, bids or None)
    fot_map = {(r['year'],r['month']): int(r['fot']) for r in fot_rows}
    rev_map = {(r['year'],r['month']): int(r['revenue']) + int(manual_map.get((r['year'],r['month']), 0))
               for r in rev_rows}
    for k, v in manual_map.items():
        if k not in rev_map:
            rev_map[k] = int(v)
    labels  = ['','Янв','Фев','Мар','Апр','Май','Июн','Июл','Авг','Сен','Окт','Ноя','Дек']
    months_list = []
    y, m = start_y, start_m
    for _ in range(12):
        fv = fot_map.get((y, m), 0)
        rv = rev_map.get((y, m), 0)
        months_list.append({'year':y,'month':m,'label':labels[m],
                            'fot':fv,'revenue':rv,
                            'fot_pct': round(fv/rv*100, 1) if rv > 0 else 0})
        m += 1
        if m > 12: m = 1; y += 1
    return jsonify({'ok': True, 'months': months_list})


@app.route('/api/fot-days')
@login_required
@menu_permission_required('fot_dashboard')
def api_fot_days():
    date_from = request.args.get('date_from', date.today().isoformat())
    date_to   = request.args.get('date_to',   date.today().isoformat())
    role      = request.args.get('role', '')
    raw_bids  = request.args.get('branch_ids', '')
    bids      = [int(x) for x in raw_bids.split(',') if x.strip().isdigit()]
    bids      = [int(b) for b in get_effective_branch_ids('fot_dashboard', [str(b) for b in bids]) or []]
    bf        = f"AND s.branch_id IN ({','.join('?'*len(bids))})" if bids else ''
    with get_db() as conn:
        rev_rows = conn.execute(f'''
            SELECT s.date, COALESCE(SUM(r.total_revenue),0) AS revenue
            FROM shifts s JOIN shift_revenue r ON r.shift_id=s.id
            WHERE s.date BETWEEN ? AND ? {bf}
            GROUP BY s.date ORDER BY s.date
        ''', [date_from, date_to] + bids).fetchall()
        manual_days = _manual_rev_by_day(conn, date_from, date_to, bids or None)
        if role:
            fot_rows = conn.execute(f'''
                SELECT s.date, COALESCE(SUM(es.total_amount),0) AS fot
                FROM employee_shifts es JOIN shifts s ON s.id=es.shift_id
                WHERE s.date BETWEEN ? AND ? AND es.role_snapshot=? {bf}
                GROUP BY s.date ORDER BY s.date
            ''', [date_from, date_to, role] + bids).fetchall()
        else:
            fot_rows = conn.execute(f'''
                SELECT s.date, COALESCE(SUM(es.total_amount),0) AS fot
                FROM employee_shifts es JOIN shifts s ON s.id=es.shift_id
                WHERE s.date BETWEEN ? AND ? {bf}
                GROUP BY s.date ORDER BY s.date
            ''', [date_from, date_to] + bids).fetchall()
    rev_map = {r['date']: int(r['revenue']) for r in rev_rows}
    for d, amt in manual_days.items():
        rev_map[d] = rev_map.get(d, 0) + int(amt)
    fot_map = {r['date']: int(r['fot'])     for r in fot_rows}
    all_dates = sorted(set(list(rev_map.keys()) + list(fot_map.keys())))
    days = []
    for d in all_dates:
        rv = rev_map.get(d, 0); fv = fot_map.get(d, 0)
        days.append({'date': d, 'fot': fv, 'revenue': rv,
                     'fot_pct': round(fv / rv * 100, 1) if rv > 0 else 0})
    return jsonify({'ok': True, 'days': days})


@app.route('/fot-dashboard')
@login_required
@menu_permission_required('fot_dashboard')
def fot_dashboard():
    with get_db() as conn:
        branches = [dict(b) for b in conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()]
        if not can_pick_other_branches('fot_dashboard'):
            own = {int(b) for b in get_effective_branch_ids('fot_dashboard', []) or []}
            branches = [b for b in branches if b['id'] in own]
        branch_groups = get_branch_groups(conn)
    return render_template('fot_dashboard.html', branches=branches, branch_groups=branch_groups)


@app.route('/lfl')
@login_required
@menu_permission_required('lfl_dashboard')
def lfl_dashboard():
    with get_db() as conn:
        branches = [dict(b) for b in conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()]
        if not can_pick_other_branches('lfl_dashboard'):
            own = {int(b) for b in get_effective_branch_ids('lfl_dashboard', []) or []}
            branches = [b for b in branches if b['id'] in own]
        branch_groups = get_branch_groups(conn)
    return render_template('lfl_dashboard.html', branches=branches, branch_groups=branch_groups)


@app.route('/ratings')
@login_required
@menu_permission_required('ratings_dashboard')
def ratings_dashboard():
    with get_db() as conn:
        branches = [dict(b) for b in conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()]
        if not can_pick_other_branches('ratings_dashboard'):
            own = {int(b) for b in get_effective_branch_ids('ratings_dashboard', []) or []}
            branches = [b for b in branches if b['id'] in own]
        branch_groups = get_branch_groups(conn)
    return render_template('ratings_dashboard.html', branches=branches, branch_groups=branch_groups)


# ─── ОЖИДАНИЕ (среднее время доставки/готовности из «Отчёта по заказам») ─────

@app.route('/wait-dashboard')
@login_required
@menu_permission_required('wait_dashboard')
def wait_dashboard():
    with get_db() as conn:
        branches = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
        branch_groups = get_branch_groups(conn)
    return render_template('wait_dashboard.html', branches=branches, branch_groups=branch_groups)


def _wait_scope_where(mode):
    """Тип заказа + отсечка по времени приёма (не считаем заказы до открытия точки) + только
    заказы в статусе «Текущий» (указано обещанное время) — «Предварит.» заказы без обещанного
    времени в статистику ожидания не входят (см. колонку «Оформлен на» в Отчёте по заказам)."""
    current_only = "AND promised_minutes IS NOT NULL AND promised_minutes != 0"
    if mode == 'pickup':
        return f"order_type = 'Общий - самовывоз' AND TIME(received_at) >= '10:00:00' {current_only}"
    return f"order_type LIKE 'Доставка%' AND TIME(received_at) >= '10:30:00' {current_only}"


def _wait_metric_col(metric):
    """(колонка для AVG, доп. условие только для этой метрики). 'delivery' недоступна для mode=pickup — вызывающий код должен это отсечь."""
    if metric == 'ready':
        return 'ready_minutes', 'AND ready_minutes > 0'
    if metric == 'delivery':
        return 'delivery_minutes', 'AND delivery_minutes IS NOT NULL AND delivery_minutes <= 300'
    return 'promised_minutes', ''


def _wait_metric_for(mode, requested):
    """Самовывоз не может показывать 'доставили' — там нет доставки."""
    if mode == 'pickup' and requested == 'delivery':
        return 'ready'
    return requested


def _wait_branch_filter():
    raw = request.args.get('branch_ids', '')
    ids = [int(x) for x in raw.split(',') if x.isdigit()]
    if not ids:
        return '', []
    ph = ','.join('?' * len(ids))
    return f'AND branch_id IN ({ph})', ids


def _wait_request_params():
    mode   = request.args.get('mode', 'delivery')
    metric = _wait_metric_for(mode, request.args.get('metric', 'delivery' if mode == 'delivery' else 'promised'))
    bf, bparams = _wait_branch_filter()
    col, metric_extra = _wait_metric_col(metric)
    where = f'{_wait_scope_where(mode)} {metric_extra}'
    return mode, metric, col, where, bf, bparams


@app.route('/api/wait-summary')
@login_required
@menu_permission_required('wait_dashboard')
def api_wait_summary():
    date_from = request.args.get('date_from', date.today().isoformat())
    date_to   = request.args.get('date_to',   date.today().isoformat())
    mode, metric, col, where, bf, bparams = _wait_request_params()

    with get_db() as conn:
        total = conn.execute(f'''
            SELECT COUNT(*) AS cnt, AVG({col}) AS avg_val
            FROM orders_report
            WHERE {where} AND DATE(received_at) BETWEEN ? AND ? {bf}
        ''', [date_from, date_to] + bparams).fetchone()
        branch_rows = conn.execute(f'''
            SELECT branch_raw AS name, COUNT(*) AS cnt, AVG({col}) AS avg_val
            FROM orders_report
            WHERE {where} AND DATE(received_at) BETWEEN ? AND ? {bf}
            GROUP BY branch_raw ORDER BY name
        ''', [date_from, date_to] + bparams).fetchall()

    return jsonify({
        'ok': True, 'mode': mode, 'metric': metric,
        'count': total['cnt'] or 0,
        'avg': round(total['avg_val'] or 0, 1),
        'branches': [{
            'name': r['name'], 'abbr': (r['name'] or '')[:3].upper(),
            'count': r['cnt'], 'avg': round(r['avg_val'] or 0, 1),
        } for r in branch_rows],
    })


@app.route('/api/wait-days')
@login_required
@menu_permission_required('wait_dashboard')
def api_wait_days():
    date_from = request.args.get('date_from', date.today().isoformat())
    date_to   = request.args.get('date_to',   date.today().isoformat())
    mode, metric, col, where, bf, bparams = _wait_request_params()

    with get_db() as conn:
        rows = conn.execute(f'''
            SELECT DATE(received_at) AS d, COUNT(*) AS cnt, AVG({col}) AS avg_val
            FROM orders_report
            WHERE {where} AND DATE(received_at) BETWEEN ? AND ? {bf}
            GROUP BY d
        ''', [date_from, date_to] + bparams).fetchall()
    by_day = {r['d']: r for r in rows}

    days = []
    _d    = datetime.strptime(date_from, '%Y-%m-%d').date()
    _dend = datetime.strptime(date_to, '%Y-%m-%d').date()
    while _d <= _dend:
        ds = _d.isoformat()
        r  = by_day.get(ds)
        days.append({'date': ds, 'count': (r['cnt'] if r else 0), 'value': round(r['avg_val'] or 0, 1) if r else 0})
        _d += timedelta(days=1)
    return jsonify({'ok': True, 'mode': mode, 'metric': metric, 'days': days})


def _wait_months_query(conn, col, where, date_from, date_to, bf, bparams):
    rows = conn.execute(f'''
        SELECT CAST(strftime('%Y', received_at) AS INTEGER) AS year,
               CAST(strftime('%m', received_at) AS INTEGER) AS month,
               COUNT(*) AS cnt, AVG({col}) AS avg_val
        FROM orders_report
        WHERE {where} AND DATE(received_at) BETWEEN ? AND ? {bf}
        GROUP BY year, month ORDER BY year, month
    ''', [date_from, date_to] + bparams).fetchall()
    return {(r['year'], r['month']): r for r in rows}


def _wait_months_list(agg, date_from, date_to):
    labels = ['', 'Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн',
              'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек']
    start = date.fromisoformat(date_from)
    end   = date.fromisoformat(date_to)
    months_list = []
    y, m = start.year, start.month
    while (y < end.year) or (y == end.year and m <= end.month):
        label = labels[m] + " '" + str(y)[-2:]
        r = agg.get((y, m))
        months_list.append({'year': y, 'month': m, 'label': label,
                            'count': (r['cnt'] if r else 0),
                            'value': round(r['avg_val'] or 0, 1) if r else 0})
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months_list


@app.route('/api/wait-months')
@login_required
@menu_permission_required('wait_dashboard')
def api_wait_months():
    date_from = request.args.get('date_from')
    date_to   = request.args.get('date_to', date.today().isoformat())
    mode, metric, col, where, bf, bparams = _wait_request_params()
    if not date_from:
        return jsonify({'ok': False, 'error': 'date_from required'}), 400
    with get_db() as conn:
        agg = _wait_months_query(conn, col, where, date_from, date_to, bf, bparams)
    months = _wait_months_list(agg, date_from, date_to)
    return jsonify({'ok': True, 'mode': mode, 'metric': metric,
                    'total_count': sum(x['count'] for x in months), 'months': months})


# ─── ПРОМОКОДЫ (выручка/новые клиенты/средний чек по промокоду из «Отчёта по заказам») ─

@app.route('/promo-dashboard')
@login_required
@menu_permission_required('promo_dashboard')
def promo_dashboard():
    with get_db() as conn:
        branches = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
        branch_groups = get_branch_groups(conn)
    return render_template('promo_dashboard.html', branches=branches, branch_groups=branch_groups)


def _promo_branch_filter():
    raw = request.args.get('branch_ids', '')
    ids = [int(x) for x in raw.split(',') if x.isdigit()]
    if not ids:
        return '', []
    ph = ','.join('?' * len(ids))
    return f'AND branch_id IN ({ph})', ids


def _promo_scope(promo, bparams):
    """(доп. WHERE, параметры) для конкретного промокода либо для всех промокодов сразу."""
    if promo:
        return 'AND promo_code = ?', bparams + [promo]
    return "AND promo_code IS NOT NULL AND promo_code != ''", bparams


@app.route('/api/promo-summary')
@login_required
@menu_permission_required('promo_dashboard')
def api_promo_summary():
    date_from = request.args.get('date_from', date.today().isoformat())
    date_to   = request.args.get('date_to',   date.today().isoformat())
    promo     = request.args.get('promo', '').strip()
    bf, bparams = _promo_branch_filter()
    extra, pparams = _promo_scope(promo, bparams)

    with get_db() as conn:
        total_row = conn.execute(f'''
            SELECT COALESCE(SUM(amount),0) AS total
            FROM orders_report
            WHERE DATE(received_at) BETWEEN ? AND ? {bf}
        ''', [date_from, date_to] + bparams).fetchone()
        promo_row = conn.execute(f'''
            SELECT COUNT(*) AS cnt, COALESCE(SUM(amount),0) AS revenue, COALESCE(AVG(amount),0) AS avg_check,
                   SUM(CASE WHEN new_client='Да' THEN 1 ELSE 0 END) AS new_clients
            FROM orders_report
            WHERE DATE(received_at) BETWEEN ? AND ? {bf} {extra}
        ''', [date_from, date_to] + pparams).fetchone()

    total_revenue = total_row['total'] or 0
    revenue = promo_row['revenue'] or 0
    return jsonify({
        'ok': True, 'promo': promo or None,
        'count': promo_row['cnt'] or 0,
        'revenue': round(revenue, 2),
        'total_revenue': round(total_revenue, 2),
        'pct': round(revenue / total_revenue * 100, 1) if total_revenue > 0 else 0,
        'avg_check': round(promo_row['avg_check'] or 0, 2),
        'new_clients': promo_row['new_clients'] or 0,
    })


@app.route('/api/promo-list')
@login_required
@menu_permission_required('promo_dashboard')
def api_promo_list():
    date_from = request.args.get('date_from', date.today().isoformat())
    date_to   = request.args.get('date_to',   date.today().isoformat())
    bf, bparams = _promo_branch_filter()

    with get_db() as conn:
        total_row = conn.execute(f'''
            SELECT COALESCE(SUM(amount),0) AS total
            FROM orders_report WHERE DATE(received_at) BETWEEN ? AND ? {bf}
        ''', [date_from, date_to] + bparams).fetchone()
        total_revenue = total_row['total'] or 0

        rows = conn.execute(f'''
            SELECT promo_code AS name, COUNT(*) AS cnt, COALESCE(SUM(amount),0) AS revenue,
                   COALESCE(AVG(amount),0) AS avg_check,
                   SUM(CASE WHEN new_client='Да' THEN 1 ELSE 0 END) AS new_clients,
                   MAX(received_at) AS last_used
            FROM orders_report
            WHERE DATE(received_at) BETWEEN ? AND ? {bf}
              AND promo_code IS NOT NULL AND promo_code != ''
            GROUP BY promo_code
            ORDER BY revenue DESC
        ''', [date_from, date_to] + bparams).fetchall()

    promos = [{
        'name': r['name'], 'count': r['cnt'],
        'revenue': round(r['revenue'], 2),
        'pct': round(r['revenue'] / total_revenue * 100, 1) if total_revenue > 0 else 0,
        'avg_check': round(r['avg_check'], 2),
        'new_clients': r['new_clients'] or 0,
        'last_used': r['last_used'],
    } for r in rows]

    return jsonify({'ok': True, 'total_revenue': round(total_revenue, 2), 'promos': promos})


@app.route('/api/promo-days')
@login_required
@menu_permission_required('promo_dashboard')
def api_promo_days():
    date_from = request.args.get('date_from', date.today().isoformat())
    date_to   = request.args.get('date_to',   date.today().isoformat())
    promo     = request.args.get('promo', '').strip()
    bf, bparams = _promo_branch_filter()
    extra, pparams = _promo_scope(promo, bparams)

    with get_db() as conn:
        total_rows = conn.execute(f'''
            SELECT DATE(received_at) AS d, COALESCE(SUM(amount),0) AS total
            FROM orders_report
            WHERE DATE(received_at) BETWEEN ? AND ? {bf}
            GROUP BY d
        ''', [date_from, date_to] + bparams).fetchall()
        promo_rows = conn.execute(f'''
            SELECT DATE(received_at) AS d, COALESCE(SUM(amount),0) AS revenue
            FROM orders_report
            WHERE DATE(received_at) BETWEEN ? AND ? {bf} {extra}
            GROUP BY d
        ''', [date_from, date_to] + pparams).fetchall()
    total_by_day = {r['d']: r['total'] for r in total_rows}
    promo_by_day = {r['d']: r['revenue'] for r in promo_rows}

    days = []
    _d    = datetime.strptime(date_from, '%Y-%m-%d').date()
    _dend = datetime.strptime(date_to, '%Y-%m-%d').date()
    while _d <= _dend:
        ds  = _d.isoformat()
        rev = promo_by_day.get(ds, 0)
        tot = total_by_day.get(ds, 0)
        days.append({'date': ds, 'revenue': round(rev, 2), 'total': round(tot, 2),
                     'pct': round(rev / tot * 100, 1) if tot > 0 else 0})
        _d += timedelta(days=1)
    return jsonify({'ok': True, 'promo': promo or None, 'days': days})


def _promo_months_query(conn, bf, bparams, extra, pparams, date_from, date_to):
    total_rows = conn.execute(f'''
        SELECT CAST(strftime('%Y', received_at) AS INTEGER) AS year,
               CAST(strftime('%m', received_at) AS INTEGER) AS month,
               COALESCE(SUM(amount),0) AS total
        FROM orders_report
        WHERE DATE(received_at) BETWEEN ? AND ? {bf}
        GROUP BY year, month
    ''', [date_from, date_to] + bparams).fetchall()
    promo_rows = conn.execute(f'''
        SELECT CAST(strftime('%Y', received_at) AS INTEGER) AS year,
               CAST(strftime('%m', received_at) AS INTEGER) AS month,
               COALESCE(SUM(amount),0) AS revenue
        FROM orders_report
        WHERE DATE(received_at) BETWEEN ? AND ? {bf} {extra}
        GROUP BY year, month
    ''', [date_from, date_to] + pparams).fetchall()
    total_by_ym = {(r['year'], r['month']): r['total'] for r in total_rows}
    promo_by_ym = {(r['year'], r['month']): r['revenue'] for r in promo_rows}
    return total_by_ym, promo_by_ym


def _promo_months_list(total_by_ym, promo_by_ym, date_from, date_to):
    labels = ['', 'Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн',
              'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек']
    start = date.fromisoformat(date_from)
    end   = date.fromisoformat(date_to)
    months_list = []
    y, m = start.year, start.month
    while (y < end.year) or (y == end.year and m <= end.month):
        label = labels[m] + " '" + str(y)[-2:]
        tot = total_by_ym.get((y, m), 0)
        rev = promo_by_ym.get((y, m), 0)
        months_list.append({'year': y, 'month': m, 'label': label,
                            'revenue': round(rev, 2), 'total': round(tot, 2),
                            'pct': round(rev / tot * 100, 1) if tot > 0 else 0})
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months_list


@app.route('/api/promo-months')
@login_required
@menu_permission_required('promo_dashboard')
def api_promo_months():
    date_from = request.args.get('date_from')
    date_to   = request.args.get('date_to', date.today().isoformat())
    promo     = request.args.get('promo', '').strip()
    bf, bparams = _promo_branch_filter()
    extra, pparams = _promo_scope(promo, bparams)
    if not date_from:
        return jsonify({'ok': False, 'error': 'date_from required'}), 400
    with get_db() as conn:
        total_by_ym, promo_by_ym = _promo_months_query(conn, bf, bparams, extra, pparams, date_from, date_to)
    months = _promo_months_list(total_by_ym, promo_by_ym, date_from, date_to)
    return jsonify({'ok': True, 'promo': promo or None, 'months': months})


@app.route('/api/promo-all')
@login_required
@menu_permission_required('promo_dashboard')
def api_promo_all():
    promo = request.args.get('promo', '').strip()
    bf, bparams = _promo_branch_filter()
    today_iso = date.today().isoformat()
    with get_db() as conn:
        min_row = conn.execute(f'''
            SELECT MIN(DATE(received_at)) FROM orders_report WHERE 1=1 {bf}
        ''', bparams).fetchone()
        date_from = min_row[0]
        if not date_from:
            return jsonify({'ok': True, 'promo': promo or None, 'months': []})
        extra, pparams = _promo_scope(promo, bparams)
        total_by_ym, promo_by_ym = _promo_months_query(conn, bf, bparams, extra, pparams, date_from, today_iso)
    months = _promo_months_list(total_by_ym, promo_by_ym, date_from, today_iso)
    return jsonify({'ok': True, 'promo': promo or None, 'months': months})


# ─── SHIFTS ───────────────────────────────────────────────────────────────────

@app.route('/shift/open', methods=['POST'])
@login_required
def open_shift():
    role = session.get('role')
    if role == 'owner':
        branch_id = request.form.get('branch_id', session.get('branch_id'))
    else:
        branch_id = request.form.get('branch_id') or session.get('branch_id')
        bids = _session_branch_ids()
        if branch_id and int(branch_id) not in bids:
            branch_id = bids[0] if bids else None
    if not branch_id:
        flash('Филиал не определён', 'danger')
        return redirect(url_for('dashboard'))
    today = date.today().isoformat()
    # Владелец может открыть смену задним числом (заполнить пропущенную смену),
    # остальные роли — только на сегодня.
    target_date = request.form.get('date', '').strip()
    if role == 'owner' and target_date:
        try:
            date.fromisoformat(target_date)
        except ValueError:
            flash('Некорректная дата', 'danger')
            return redirect(url_for('dashboard'))
        if target_date > today:
            flash('Нельзя открыть смену на будущую дату', 'danger')
            return redirect(url_for('dashboard'))
    else:
        target_date = today
    with get_db() as conn:
        existing = conn.execute(
            'SELECT id FROM shifts WHERE branch_id=? AND date=?', (branch_id, target_date)
        ).fetchone()
        if existing:
            flash('Смена на этот день уже существует', 'warning')
            return redirect(url_for('shift_view', shift_id=existing['id']))
        try:
            conn.execute(
                'INSERT INTO shifts (branch_id, date, opened_by) VALUES (?,?,?)',
                (branch_id, target_date, session['user_id'])
            )
        except Exception:
            # Race condition: another request created the shift simultaneously
            existing2 = conn.execute(
                'SELECT id FROM shifts WHERE branch_id=? AND date=?', (branch_id, target_date)
            ).fetchone()
            if existing2:
                flash('Смена на этот день уже существует', 'warning')
                return redirect(url_for('shift_view', shift_id=existing2['id']))
            raise
        shift_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        log_action(conn, 'shift_open', 'Смена открыта', shift_id=shift_id)
        # Если пользователь ввёл факт вручную — используем его
        _mc_raw = request.form.get('morning_cash', '').strip().replace('\xa0', '').replace(' ', '')
        if _mc_raw:
            try:
                prev_morning = float(_mc_raw.replace(',', '.'))
            except ValueError:
                prev_morning = 0.0
        else:
            prev_morning = _calc_prev_kassa_nal(conn, int(branch_id), target_date)
        conn.execute(
            'INSERT INTO shift_revenue (shift_id, morning_cash) VALUES (?, ?)',
            (shift_id, prev_morning)
        )
        _apply_change_amount_to_shift(conn, shift_id, int(branch_id), target_date)
        conn.commit()
    return redirect(url_for('shift_view', shift_id=shift_id))


@app.route('/shift/<int:shift_id>')
@login_required
def shift_view(shift_id):
    with get_db() as conn:
        shift = conn.execute('''
            SELECT s.*, b.name as branch_name
            FROM shifts s JOIN branches b ON b.id=s.branch_id
            WHERE s.id=?
        ''', (shift_id,)).fetchone()
        if not shift:
            flash('Смена не найдена', 'danger')
            return redirect(url_for('dashboard'))
        role = session.get('role')
        if role != 'owner' and shift['branch_id'] not in _session_branch_ids():
            flash('Нет доступа к этой смене', 'danger')
            return redirect(url_for('dashboard'))

        revenue = conn.execute('SELECT * FROM shift_revenue WHERE shift_id=?', (shift_id,)).fetchone()
        expenses = conn.execute('SELECT * FROM expenses WHERE shift_id=? ORDER BY id', (shift_id,)).fetchall()
        plus_entries = conn.execute('SELECT * FROM cash_plus_entries WHERE shift_id=? ORDER BY id', (shift_id,)).fetchall()
        staff_rows = conn.execute(
            '''SELECT es.*,
                      (CASE WHEN es.role_snapshot = e.role THEN COALESCE(e.pay_monthly, 0)
                            ELSE COALESCE(er.pay_monthly, 0) END) as pay_monthly_now,
                      CASE WHEN (
                          NOT EXISTS (SELECT 1 FROM employee_pay_monthly_branches pmb
                                      WHERE pmb.employee_id = es.employee_id AND pmb.role = es.role_snapshot)
                          OR EXISTS (SELECT 1 FROM employee_pay_monthly_branches pmb
                                     WHERE pmb.employee_id = es.employee_id AND pmb.role = es.role_snapshot
                                       AND pmb.branch_id = ?)
                        )
                      THEN 1 ELSE 0 END as pay_monthly_branch_ok
               FROM employee_shifts es
               LEFT JOIN employees e ON e.id = es.employee_id
               LEFT JOIN employee_roles er ON er.employee_id = es.employee_id AND er.role = es.role_snapshot
               WHERE es.shift_id=? ORDER BY es.role_snapshot, es.full_name_snapshot''',
            (shift['branch_id'], shift_id)
        ).fetchall()
        # Доп. маршруты курьеров: «Данные из гуляша» = km/orders минус эти маршруты
        courier_ids = [s['id'] for s in staff_rows if s['role_snapshot'] == 'courier']
        courier_route_map = {}
        if courier_ids:
            ph = ','.join('?' * len(courier_ids))
            for r in conn.execute(
                f'SELECT * FROM courier_extra_routes WHERE employee_shift_id IN ({ph}) ORDER BY sort_order, id',
                courier_ids
            ).fetchall():
                courier_route_map.setdefault(r['employee_shift_id'], []).append(dict(r))
        staff = []
        for s in staff_rows:
            d = dict(s)
            if d['employee_id']:
                # Режим оплаты — действовавший на дату этой смены, а не текущий
                # (иначе переключение сегодня задним числом меняло бы вид уже прошедших смен).
                if shift['import_batch_id']:
                    # Импортированные смены: показываем как обычно выплачено да/нет —
                    # бейдж «Ежемес.» тут не нужен, статус в файле уже окончательный,
                    # настройка сотрудника «ежемесячно» на импорт не влияет.
                    d['pay_monthly'] = 0
                else:
                    pm_on_date = _effective_pay_monthly(
                        conn, d['employee_id'], d['role_snapshot'], shift['date'], d['pay_monthly_now']
                    )
                    d['pay_monthly'] = 1 if (pm_on_date and d['pay_monthly_branch_ok']) else 0
            if d['role_snapshot'] == 'courier':
                routes = courier_route_map.get(d['id'], [])
                d['extra_routes'] = routes
                d['gulyash_km'] = float(d['km'] or 0) - sum(float(r['km'] or 0) for r in routes)
                d['gulyash_orders'] = int(d['orders'] or 0) - len(routes)
            staff.append(d)
        # Find all branches in same group(s) as this branch (for employee list)
        group_branch_ids = conn.execute('''
            SELECT DISTINCT bgm2.branch_id
            FROM branch_group_members bgm1
            JOIN branch_group_members bgm2 ON bgm2.group_id = bgm1.group_id
            WHERE bgm1.branch_id = ?
        ''', (shift['branch_id'],)).fetchall()
        group_branch_ids = [r[0] for r in group_branch_ids]
        if not group_branch_ids:
            group_branch_ids = [shift['branch_id']]
        emp_branch_placeholders = ','.join('?' * len(group_branch_ids))
        employees = conn.execute(
            f'''SELECT DISTINCT e.* FROM employees e
               JOIN employee_branches eb ON eb.employee_id=e.id
               WHERE eb.branch_id IN ({emp_branch_placeholders}) AND e.is_active=1
               ORDER BY e.role, e.full_name''',
            group_branch_ids
        ).fetchall()
        # Extra roles for employees (for multi-role shift placement)
        if employees:
            emp_ids = [e['id'] for e in employees]
            ph = ','.join('?' * len(emp_ids))
            extra_rows = conn.execute(
                f'SELECT employee_id, role, rate, rate_per_km, rate_per_order, rate_template_id FROM employee_roles WHERE employee_id IN ({ph}) AND is_active=1',
                emp_ids
            ).fetchall()
            # Resolve rates from templates
            emp_extra_roles = {}
            for row in extra_rows:
                r_rate = float(row['rate'] or 0)
                r_km   = float(row['rate_per_km'] or 0)
                r_ord  = float(row['rate_per_order'] or 0)
                if row['rate_template_id']:
                    tmpl = conn.execute('SELECT rate, rate_per_km, rate_per_order FROM rate_templates WHERE id=?', (row['rate_template_id'],)).fetchone()
                    if tmpl:
                        r_rate = float(tmpl['rate'] or 0)
                        r_km   = float(tmpl['rate_per_km'] or 0)
                        r_ord  = float(tmpl['rate_per_order'] or 0)
                emp_extra_roles.setdefault(row['employee_id'], []).append({'role': row['role'], 'rate': r_rate, 'rate_per_km': r_km, 'rate_per_order': r_ord})
        else:
            emp_extra_roles = {}
        # Current address per employee (as of shift date)
        emp_addresses = {}
        for emp in employees:
            addr = conn.execute(
                'SELECT address FROM employee_address_history WHERE employee_id=? AND valid_from<=? ORDER BY valid_from DESC LIMIT 1',
                (emp['id'], shift['date'])
            ).fetchone()
            emp_addresses[emp['id']] = addr['address'] if addr else ''
        # Ставки для модалки быстрого добавления нового курьера
        courier_rate_templates = conn.execute('''
            SELECT rt.* FROM rate_templates rt
            WHERE rt.role='courier' AND rt.is_active=1
              AND (NOT EXISTS (SELECT 1 FROM rate_template_branches WHERE template_id=rt.id)
                   OR EXISTS (SELECT 1 FROM rate_template_branches WHERE template_id=rt.id AND branch_id=?))
            ORDER BY rt.name
        ''', (shift['branch_id'],)).fetchall()
        # Taxi trips for this shift
        taxi_trips = conn.execute(
            'SELECT * FROM taxi_trips WHERE shift_id=? ORDER BY id',
            (shift_id,)
        ).fetchall()
        taxi_trip_emps = {}
        for t in taxi_trips:
            tte = conn.execute(
                'SELECT * FROM taxi_trip_employees WHERE trip_id=? ORDER BY id',
                (t['id'],)
            ).fetchall()
            taxi_trip_emps[t['id']] = tte
        shift_terminals_rows = conn.execute(
            'SELECT terminal_number, amount FROM shift_terminals WHERE shift_id=? ORDER BY sort_order, id',
            (shift_id,)
        ).fetchall()
        all_cats = get_expense_categories(conn, branch_id=shift['branch_id'])
        expense_cats = [c for c in all_cats if c['type'] != 'income']
        income_cats  = [c for c in all_cats if c['type'] == 'income']
        expense_cats_groups = build_cats_groups(filter_cats_by_flag(expense_cats, 'show_shift'))
        expense_cats_flat   = [(c['code'], c['label']) for c in expense_cats]
        income_cats_groups  = build_cats_groups(filter_cats_by_flag(income_cats, 'show_shift'))
        income_cats_flat    = [(c['code'], c['label']) for c in income_cats]
        seen_taxi_ids = set()
        taxi_staff = []
        for s in staff:
            if s['role_snapshot'] != 'courier' and s['employee_id'] and s['employee_id'] not in seen_taxi_ids:
                seen_taxi_ids.add(s['employee_id'])
                taxi_staff.append(s)
        can_edit = (role == 'owner') or (shift['status'] == 'open')
        try:
            shift_weekday = date.fromisoformat(shift['date']).weekday()  # 0=Mon, 4=Fri, 5=Sat
        except Exception:
            shift_weekday = 0
        # Утром в кассе: должно всегда совпадать с "Итого нал" предыдущего дня.
        # Для ещё открытых смен подправляем сохранённое значение, если оно разошлось
        # (например, из-за правок расходов/такси предыдущего дня уже после открытия текущей смены).
        prev_actual_cash = _calc_prev_kassa_nal(conn, shift['branch_id'], shift['date'])
        if shift['status'] == 'open' and revenue is not None:
            stored_morning = revenue['morning_cash'] or 0
            if abs(stored_morning - prev_actual_cash) > 0.01:
                conn.execute('UPDATE shift_revenue SET morning_cash=? WHERE shift_id=?', (prev_actual_cash, shift_id))
                conn.commit()
                revenue = conn.execute('SELECT * FROM shift_revenue WHERE shift_id=?', (shift_id,)).fetchone()
        promokod_row = conn.execute(
            'SELECT code FROM flyer_promocodes WHERE branch_id=? AND date=?',
            (shift['branch_id'], shift['date'])
        ).fetchone()
        promokod = promokod_row['code'] if promokod_row else ''
        bonus_rules_rows = conn.execute(
            'SELECT role, threshold_pct, bonus_pct FROM bonus_rules '
            'WHERE is_active=1 AND (branch_id IS NULL OR branch_id=?) '
            'ORDER BY role, threshold_pct DESC', (shift['branch_id'],)
        ).fetchall()
        # Deduplicate by role+threshold keeping branch-specific over global
        _seen_br = {}
        for r in bonus_rules_rows:
            key = (r['role'], r['threshold_pct'])
            if key not in _seen_br:
                _seen_br[key] = {'role': r['role'], 'threshold_pct': float(r['threshold_pct']), 'bonus_pct': float(r['bonus_pct'])}
        bonus_rules_list = list(_seen_br.values())
        return render_template('shift.html',
            shift=shift, revenue=revenue, expenses=expenses, plus_entries=plus_entries,
            staff=staff, employees=employees, taxi_staff=taxi_staff,
            emp_addresses=emp_addresses, emp_extra_roles=emp_extra_roles,
            courier_rate_templates=courier_rate_templates,
            taxi_trips=taxi_trips, taxi_trip_emps=taxi_trip_emps,
            expense_categories=expense_cats_flat,
            expense_cats_groups=expense_cats_groups,
            income_categories=income_cats_flat,
            income_cats_groups=income_cats_groups,
            role_labels=ROLE_LABELS,
            can_edit=can_edit,
            is_owner=(role == 'owner'),
            shift_weekday=shift_weekday,
            prev_actual_cash=prev_actual_cash,
            promokod=promokod,
            bonus_rules_list=bonus_rules_list,
            shift_terminals=[{'terminal_number': r['terminal_number'], 'amount': r['amount']}
                             for r in shift_terminals_rows])


@app.route('/shift/<int:shift_id>/save-revenue', methods=['POST'])
@login_required
def save_revenue(shift_id):
    if not _can_edit_shift(shift_id):
        return jsonify({'error': 'Нет доступа'}), 403
    data = request.json or {}
    with get_db() as conn:
        # Preserve change_amount if not sent (it's managed by the change settings page)
        if 'change_amount' not in data:
            existing_rev = conn.execute(
                'SELECT change_amount FROM shift_revenue WHERE shift_id=?', (shift_id,)
            ).fetchone()
            change_amount = float(existing_rev['change_amount'] or 0) if existing_rev else 0.0
        else:
            change_amount = _f(data, 'change_amount')
        conn.execute('''
            UPDATE shift_revenue SET
                total_revenue=?, delivery_revenue=?, delivery_orders=?,
                pickup_revenue=?, pickup_orders=?,
                cash_amount=?, card_amount=?, online_amount=?,
                change_amount=?, actual_cash=?, terminal_last3=?, terminal_amount=?,
                morning_cash=?, kassa_nal=?
            WHERE shift_id=?
        ''', (
            _f(data, 'total_revenue'), _f(data, 'delivery_revenue'), _i(data, 'delivery_orders'),
            _f(data, 'pickup_revenue'), _i(data, 'pickup_orders'),
            _f(data, 'cash_amount'), _f(data, 'card_amount'), _f(data, 'online_amount'),
            change_amount, _f(data, 'actual_cash'),
            data.get('terminal_last3', ''), _f(data, 'terminal_amount'),
            _f(data, 'morning_cash'), _f(data, 'kassa_nal'),
            shift_id
        ))
        desc = (f"нал {_fmt_money(data.get('cash_amount'))}, "
                f"безнал {_fmt_money(data.get('card_amount'))}, "
                f"онлайн {_fmt_money(data.get('online_amount'))}, "
                f"итого {_fmt_money(data.get('total_revenue'))}")
        log_action(conn, 'revenue_update', desc, shift_id=shift_id, upsert_by_shift=True)
        auto_bonuses = calculate_bonuses(conn, shift_id)
        conn.commit()
    return jsonify({'ok': True, 'auto_bonuses': auto_bonuses})


@app.route('/shift/<int:shift_id>/save-morning-cash', methods=['POST'])
@login_required
def save_morning_cash(shift_id):
    if session.get('role') != 'owner':
        return jsonify({'error': 'Только владелец может менять сумму утром в кассе'}), 403
    data = request.json or {}
    comment = (data.get('comment') or '').strip()
    if not comment:
        return jsonify({'error': 'Укажите комментарий'}), 400
    with get_db() as conn:
        existing = conn.execute(
            'SELECT morning_cash FROM shift_revenue WHERE shift_id=?', (shift_id,)
        ).fetchone()
        old_amount = float(existing['morning_cash'] or 0) if existing else 0.0
        new_amount = _f(data, 'amount')
        conn.execute(
            'UPDATE shift_revenue SET morning_cash=?, kassa_nal=? WHERE shift_id=?',
            (new_amount, _f(data, 'kassa_nal'), shift_id)
        )
        log_action(conn, 'morning_cash_update',
            f"Утром в кассе: {_fmt_money(old_amount)} → {_fmt_money(new_amount)}. Комментарий: {comment}",
            shift_id=shift_id)
        conn.commit()
    return jsonify({'ok': True})


@app.route('/shift/<int:shift_id>/save-terminals', methods=['POST'])
@login_required
def save_terminals(shift_id):
    if not _can_edit_shift(shift_id):
        return jsonify({'error': 'Нет доступа'}), 403
    data = request.json or {}
    terminals = data.get('terminals', [])
    with get_db() as conn:
        conn.execute('DELETE FROM shift_terminals WHERE shift_id=?', (shift_id,))
        for i, t in enumerate(terminals):
            tn = str(t.get('terminal_number', '')).strip()
            raw = str(t.get('amount', 0)).replace(' ', '').replace(' ', '').replace(',', '.')
            try:
                amt = float(raw)
            except ValueError:
                amt = 0.0
            if tn or amt:
                conn.execute(
                    'INSERT INTO shift_terminals (shift_id, terminal_number, amount, sort_order) VALUES (?,?,?,?)',
                    (shift_id, tn, amt, i)
                )
        conn.commit()
    return jsonify({'ok': True})


@app.route('/shift/<int:shift_id>/save-plus', methods=['POST'])
@login_required
def save_cash_plus(shift_id):
    if not _can_edit_shift(shift_id):
        return jsonify({'error': 'Нет доступа'}), 403
    data = request.json or {}
    entry_id    = data.get('id')
    amount_cash = float(data.get('amount_cash') or 0)
    amount_card = float(data.get('amount_card') or 0)
    amount      = amount_cash + amount_card
    category    = (data.get('category') or '').strip()
    description = (data.get('description') or '').strip()
    with get_db() as conn:
        if entry_id:
            conn.execute(
                'UPDATE cash_plus_entries SET amount=?, amount_cash=?, amount_card=?, category=?, description=? WHERE id=? AND shift_id=?',
                (amount, amount_cash, amount_card, category, description, entry_id, shift_id)
            )
        else:
            conn.execute(
                'INSERT INTO cash_plus_entries (shift_id, amount, amount_cash, amount_card, category, description) VALUES (?,?,?,?,?,?)',
                (shift_id, amount, amount_cash, amount_card, category, description)
            )
            entry_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        conn.commit()
    return jsonify({'ok': True, 'id': entry_id})


@app.route('/shift/<int:shift_id>/delete-plus/<int:entry_id>', methods=['POST'])
@login_required
def delete_cash_plus(shift_id, entry_id):
    if not _can_edit_shift(shift_id):
        return jsonify({'error': 'Нет доступа'}), 403
    with get_db() as conn:
        conn.execute('DELETE FROM cash_plus_entries WHERE id=? AND shift_id=?', (entry_id, shift_id))
        conn.commit()
    return jsonify({'ok': True})


@app.route('/shift/<int:shift_id>/save-expense', methods=['POST'])
@login_required
def save_expense(shift_id):
    if not _can_edit_shift(shift_id):
        return jsonify({'error': 'Нет доступа'}), 403
    data = request.json or {}
    expense_id = data.get('id')
    with get_db() as conn:
        cats = {r['code']: r['label'] for r in get_expense_categories(conn)}
        cat_label = cats.get(data.get('category', 'other'), data.get('category', 'другое'))
        amount = _f(data, 'amount_cash') + _f(data, 'amount_card')
        pay_type = 'нал' if _f(data, 'amount_cash') > 0 else 'безнал'
        note = (data.get('description') or '').strip()
        if expense_id:
            conn.execute('''
                UPDATE expenses SET category=?, description=?, amount_cash=?, amount_card=?, is_gulash=?
                WHERE id=? AND shift_id=?
            ''', (data.get('category', 'other'), data.get('description', ''),
                  _f(data, 'amount_cash'), _f(data, 'amount_card'), 1 if data.get('is_gulash') else 0,
                  expense_id, shift_id))
            desc = f"Изменён расход: {cat_label}"
            if note:
                desc += f" «{note}»"
            desc += f", {_fmt_money(amount)} ({pay_type})"
            log_action(conn, 'expense_update', desc, shift_id=shift_id, entity_id=expense_id)
        else:
            conn.execute('''
                INSERT INTO expenses (shift_id, category, description, amount_cash, amount_card, is_gulash)
                VALUES (?,?,?,?,?,?)
            ''', (shift_id, data.get('category', 'other'), data.get('description', ''),
                  _f(data, 'amount_cash'), _f(data, 'amount_card'), 1 if data.get('is_gulash') else 0))
            expense_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            desc = f"Добавлен расход: {cat_label}"
            if note:
                desc += f" «{note}»"
            desc += f", {_fmt_money(amount)} ({pay_type})"
            log_action(conn, 'expense_add', desc, shift_id=shift_id, entity_id=expense_id)
        conn.commit()
    return jsonify({'ok': True, 'id': expense_id})


@app.route('/shift/<int:shift_id>/delete-expense/<int:expense_id>', methods=['POST'])
@login_required
def delete_expense(shift_id, expense_id):
    if not _can_edit_shift(shift_id):
        return jsonify({'error': 'Нет доступа'}), 403
    with get_db() as conn:
        exp = conn.execute('SELECT * FROM expenses WHERE id=? AND shift_id=?', (expense_id, shift_id)).fetchone()
        if exp:
            cats = {r['code']: r['label'] for r in get_expense_categories(conn)}
            cat_label = cats.get(exp['category'], exp['category'])
            amount = (exp['amount_cash'] or 0) + (exp['amount_card'] or 0)
            pay_type = 'нал' if (exp['amount_cash'] or 0) > 0 else 'безнал'
            note = (exp['description'] or '').strip()
            desc = f"Удалён расход: {cat_label}"
            if note:
                desc += f" «{note}»"
            desc += f", {_fmt_money(amount)} ({pay_type})"
        else:
            desc = f"Удалён расход #{expense_id}"
        conn.execute('DELETE FROM expenses WHERE id=? AND shift_id=?', (expense_id, shift_id))
        log_action(conn, 'expense_delete', desc, shift_id=shift_id, entity_id=expense_id)
        conn.commit()
    return jsonify({'ok': True})


@app.route('/shift/<int:shift_id>/save-staff', methods=['POST'])
@login_required
def save_staff(shift_id):
    if not _can_edit_shift(shift_id):
        return jsonify({'error': 'Нет доступа'}), 403
    data = request.json or {}
    staff_id = data.get('id')
    with get_db() as conn:
        if staff_id:
            existing = conn.execute(
                'SELECT * FROM employee_shifts WHERE id=? AND shift_id=?',
                (staff_id, shift_id)
            ).fetchone()
            if not existing:
                return jsonify({'ok': False, 'error': 'not found'}), 404
            def _pick(key, default=None):
                return data[key] if key in data else (existing[key] if existing else default)
            def _pickf(key):
                return float(data[key]) if key in data else float(existing[key] or 0)
            def _picki(key):
                return int(data[key]) if key in data else int(existing[key] or 0)
            has_bonus = 'bonus_amount' in data
            bonus_amount   = _f(data, 'bonus_amount')   if has_bonus else float(existing['bonus_amount']  or 0)
            penalty_amount = _f(data, 'penalty_amount') if has_bonus else float(existing['penalty_amount'] or 0)
            bonus_comment  = data['bonus_comment']       if 'bonus_comment' in data else (existing['bonus_comment'] or '')
            # Recompute total if base fields were sent; otherwise preserve base_pay and recompute from bonus
            if 'base_pay' in data:
                base_pay     = _f(data, 'base_pay')
                total_amount = _f(data, 'total_amount')
            else:
                base_pay     = float(existing['base_pay'] or 0)
                total_amount = base_pay + bonus_amount - penalty_amount + float(existing['auto_bonus'] or 0)
            is_paid_flag = (1 if data['is_paid'] else 0) if 'is_paid' in data else int(existing['is_paid'] or 0)
            # Keep paid_amount in sync with total_amount whenever employee is paid
            if is_paid_flag:
                paid_amount_val = total_amount
            elif 'paid_amount' in data:
                paid_amount_val = _f(data, 'paid_amount')
            else:
                paid_amount_val = float(existing['paid_amount'] or 0)
            conn.execute('''
                UPDATE employee_shifts SET
                    full_name_snapshot=?, role_snapshot=?, rate_snapshot=?,
                    rate_per_km_snapshot=?, rate_per_order_snapshot=?,
                    shift_start=?, shift_end=?, hours_worked=?,
                    km=?, orders=?, bonus_amount=?, penalty_amount=?, bonus_comment=?,
                    base_pay=?, total_amount=?, is_paid=?, paid_amount=?
                WHERE id=? AND shift_id=?
            ''', (
                _pick('full_name_snapshot', ''), _pick('role_snapshot', ''),
                _pickf('rate_snapshot'), _pickf('rate_per_km_snapshot'), _pickf('rate_per_order_snapshot'),
                _pick('shift_start', ''), _pick('shift_end', ''), _pickf('hours_worked'),
                _pickf('km'), _picki('orders'),
                bonus_amount, penalty_amount, bonus_comment,
                base_pay, total_amount, is_paid_flag, paid_amount_val,
                staff_id, shift_id
            ))
            name = data.get('full_name_snapshot', '')
            role_lbl = ROLE_LABELS.get(data.get('role_snapshot', ''), data.get('role_snapshot', ''))
            log_action(conn, 'staff_update',
                f"Обновлены данные: {name} ({role_lbl}), итого {_fmt_money(data.get('total_amount'))}",
                shift_id=shift_id, entity_id=staff_id)
            auto_bonuses = calculate_bonuses(conn, shift_id)
        else:
            emp_id = data.get('employee_id')
            role_snap = data.get('role_snapshot', '')
            if emp_id:
                already = conn.execute(
                    'SELECT id FROM employee_shifts WHERE shift_id=? AND employee_id=? AND role_snapshot=?',
                    (shift_id, emp_id, role_snap)
                ).fetchone()
                if already:
                    return jsonify({'ok': False, 'error': 'duplicate', 'existing_role': role_snap}), 200
            rate = _f(data, 'rate_snapshot')
            rate_km = _f(data, 'rate_per_km_snapshot')
            rate_ord = _f(data, 'rate_per_order_snapshot')
            if emp_id and not rate:
                # Look up rate by role: check extra roles first, then primary, then history
                extra_role = conn.execute(
                    'SELECT er.*, rt.rate AS tmpl_rate, rt.rate_per_km AS tmpl_km, rt.rate_per_order AS tmpl_ord '
                    'FROM employee_roles er LEFT JOIN rate_templates rt ON rt.id=er.rate_template_id '
                    'WHERE er.employee_id=? AND er.role=? AND er.is_active=1',
                    (emp_id, role_snap)
                ).fetchone()
                if extra_role:
                    rate = float(extra_role['tmpl_rate'] or extra_role['rate'] or 0)
                    rate_km = float(extra_role['tmpl_km'] or extra_role['rate_per_km'] or 0)
                    rate_ord = float(extra_role['tmpl_ord'] or extra_role['rate_per_order'] or 0)
                else:
                    emp = conn.execute(
                        'SELECT e.*, rt.rate AS tmpl_rate, rt.rate_per_km AS tmpl_km, rt.rate_per_order AS tmpl_ord '
                        'FROM employees e LEFT JOIN rate_templates rt ON rt.id=e.rate_template_id '
                        'WHERE e.id=?', (emp_id,)
                    ).fetchone()
                    if emp:
                        rate = float(emp['tmpl_rate'] or emp['rate'] or 0)
                        rate_km = float(emp['tmpl_km'] or emp['rate_per_km'] or 0)
                        rate_ord = float(emp['tmpl_ord'] or emp['rate_per_order'] or 0)
            conn.execute('''
                INSERT INTO employee_shifts
                (shift_id, employee_id, full_name_snapshot, role_snapshot,
                 rate_snapshot, rate_per_km_snapshot, rate_per_order_snapshot,
                 shift_start, shift_end, hours_worked, km, orders,
                 bonus_amount, penalty_amount, bonus_comment,
                 base_pay, total_amount, is_paid, paid_amount)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                shift_id, emp_id or None,
                data.get('full_name_snapshot', ''), data.get('role_snapshot', ''),
                rate, rate_km, rate_ord,
                data.get('shift_start', ''), data.get('shift_end', ''), _f(data, 'hours_worked'),
                _f(data, 'km'), _i(data, 'orders'),
                _f(data, 'bonus_amount'), _f(data, 'penalty_amount'), data.get('bonus_comment', ''),
                _f(data, 'base_pay'), _f(data, 'total_amount'),
                1 if data.get('is_paid') else 0, _f(data, 'paid_amount')
            ))
            staff_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            name = data.get('full_name_snapshot', '')
            role_lbl = ROLE_LABELS.get(data.get('role_snapshot', ''), data.get('role_snapshot', ''))
            log_action(conn, 'staff_add',
                f"Добавлен сотрудник: {name} ({role_lbl})",
                shift_id=shift_id, entity_id=staff_id)
            auto_bonuses = calculate_bonuses(conn, shift_id)
        conn.commit()
    return jsonify({'ok': True, 'id': staff_id, 'auto_bonuses': auto_bonuses})


@app.route('/shift/<int:shift_id>/quick-add-courier', methods=['POST'])
@login_required
def quick_add_courier(shift_id):
    """Создаёт нового курьера в справочнике сотрудников и сразу добавляет его в текущую смену."""
    if not _can_edit_shift(shift_id):
        return jsonify({'ok': False, 'error': 'Нет доступа'}), 403
    data = request.json or {}
    last_name = (data.get('last_name') or '').strip()
    first_name = (data.get('first_name') or '').strip()
    if not last_name:
        return jsonify({'ok': False, 'error': 'Введите фамилию'}), 400
    if not first_name:
        return jsonify({'ok': False, 'error': 'Введите имя'}), 400
    phone_digits = _normalize_phone(data.get('phone') or '')
    if not phone_digits:
        return jsonify({'ok': False, 'error': 'Введите номер телефона'}), 400
    phone = _format_phone(data.get('phone') or '')
    full_name = (last_name + (' ' + first_name if first_name else '')).strip()
    rate_template_id = data.get('rate_template_id') or None
    if not rate_template_id:
        return jsonify({'ok': False, 'error': 'Выберите ставку'}), 400
    rate_template_id = int(rate_template_id)
    with get_db() as conn:
        shift = conn.execute('SELECT branch_id FROM shifts WHERE id=?', (shift_id,)).fetchone()
        if not shift:
            return jsonify({'ok': False, 'error': 'Смена не найдена'}), 404
        branch_id = shift['branch_id']
        for row in conn.execute("SELECT phone FROM employees WHERE role='courier'").fetchall():
            if row['phone'] and _normalize_phone(row['phone']) == phone_digits:
                return jsonify({'ok': False, 'error': 'Курьер есть в списке сотрудников!'}), 400
        tmpl = conn.execute('SELECT * FROM rate_templates WHERE id=?', (rate_template_id,)).fetchone()
        if not tmpl:
            return jsonify({'ok': False, 'error': 'Ставка не найдена'}), 400
        rate = float(tmpl['rate'] or 0)
        rate_km = float(tmpl['rate_per_km'] or 0)
        rate_ord = float(tmpl['rate_per_order'] or 0)
        conn.execute(
            'INSERT INTO employees (branch_id, full_name, last_name, first_name, role, rate, rate_per_km, rate_per_order, rate_template_id, phone) '
            'VALUES (?,?,?,?,?,?,?,?,?,?)',
            (branch_id, full_name, last_name, first_name, 'courier', rate, rate_km, rate_ord, rate_template_id, phone)
        )
        emp_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        conn.execute(
            'INSERT INTO employee_rate_history (employee_id, rate, rate_per_km, rate_per_order, effective_from) VALUES (?,?,?,?,?)',
            (emp_id, rate, rate_km, rate_ord, date.today().isoformat())
        )
        # Привязка к филиалам: если филиал смены входит в группу — привязываем ко всей группе
        group_rows = conn.execute('''
            SELECT DISTINCT bgm2.branch_id
            FROM branch_group_members bgm1
            JOIN branch_group_members bgm2 ON bgm2.group_id = bgm1.group_id
            WHERE bgm1.branch_id = ?
        ''', (branch_id,)).fetchall()
        group_branch_ids = [r[0] for r in group_rows] or [branch_id]
        for bid in group_branch_ids:
            conn.execute('INSERT OR IGNORE INTO employee_branches (employee_id, branch_id) VALUES (?,?)', (emp_id, bid))

        conn.execute('''
            INSERT INTO employee_shifts
            (shift_id, employee_id, full_name_snapshot, role_snapshot,
             rate_snapshot, rate_per_km_snapshot, rate_per_order_snapshot,
             shift_start, shift_end, hours_worked, km, orders,
             bonus_amount, penalty_amount, bonus_comment,
             base_pay, total_amount, is_paid, paid_amount)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            shift_id, emp_id, full_name, 'courier',
            rate, rate_km, rate_ord,
            '', '', 0, 0, 0,
            0, 0, '',
            0, 0, 0, 0
        ))
        staff_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        log_action(conn, 'staff_add', f"Добавлен сотрудник: {full_name} (Курьер)", shift_id=shift_id, entity_id=staff_id)
        calculate_bonuses(conn, shift_id)
        conn.commit()
    return jsonify({'ok': True, 'id': staff_id, 'employee_id': emp_id})


def _recalc_courier_base_pay(existing, km, orders):
    """base_pay курьера по снапшот-тарифам строки + агрегированные км/заказы."""
    hours    = float(existing['hours_worked'] or 0)
    rate     = float(existing['rate_snapshot'] or 0)
    rate_km  = float(existing['rate_per_km_snapshot'] or 0)
    rate_ord = float(existing['rate_per_order_snapshot'] or 0)
    return hours * rate + km * rate_km + orders * rate_ord


@app.route('/shift/<int:shift_id>/staff/<int:staff_id>/route/add', methods=['POST'])
@login_required
def add_courier_route(shift_id, staff_id):
    if not _can_edit_shift(shift_id):
        return jsonify({'error': 'Нет доступа'}), 403
    data = request.json or {}
    reason = data.get('reason', 'relocation')
    if reason not in ROUTE_REASON_LABELS:
        reason = 'relocation'
    km = _f(data, 'km')
    comment = (data.get('comment') or '').strip()
    with get_db() as conn:
        existing = conn.execute(
            'SELECT * FROM employee_shifts WHERE id=? AND shift_id=?', (staff_id, shift_id)
        ).fetchone()
        if not existing:
            return jsonify({'ok': False, 'error': 'not found'}), 404
        max_sort = conn.execute(
            'SELECT COALESCE(MAX(sort_order),-1) FROM courier_extra_routes WHERE employee_shift_id=?',
            (staff_id,)
        ).fetchone()[0]
        route_id = conn.execute(
            'INSERT INTO courier_extra_routes (employee_shift_id, reason, km, comment, sort_order) VALUES (?,?,?,?,?)',
            (staff_id, reason, km, comment, max_sort + 1)
        ).lastrowid
        new_km     = float(existing['km'] or 0) + km
        new_orders = int(existing['orders'] or 0) + 1
        base_pay   = _recalc_courier_base_pay(existing, new_km, new_orders)
        conn.execute('UPDATE employee_shifts SET km=?, orders=?, base_pay=? WHERE id=?',
                     (new_km, new_orders, base_pay, staff_id))
        auto_bonuses = calculate_bonuses(conn, shift_id)
        total_row = conn.execute('SELECT total_amount FROM employee_shifts WHERE id=?', (staff_id,)).fetchone()
        log_action(conn, 'staff_update',
            f"Добавлен доп. маршрут: {ROUTE_REASON_LABELS[reason]}, {km:g} км",
            shift_id=shift_id, entity_id=staff_id)
        conn.commit()
    return jsonify({'ok': True, 'route_id': route_id, 'reason': reason, 'km': km, 'comment': comment,
                    'km_total': new_km, 'orders_total': new_orders,
                    'total_amount': total_row['total_amount'], 'auto_bonuses': auto_bonuses})


@app.route('/shift/<int:shift_id>/staff/<int:staff_id>/route/<int:route_id>/update', methods=['POST'])
@login_required
def update_courier_route(shift_id, staff_id, route_id):
    if not _can_edit_shift(shift_id):
        return jsonify({'error': 'Нет доступа'}), 403
    data = request.json or {}
    reason = data.get('reason', 'relocation')
    if reason not in ROUTE_REASON_LABELS:
        reason = 'relocation'
    new_km = _f(data, 'km')
    comment = (data.get('comment') or '').strip()
    with get_db() as conn:
        route = conn.execute(
            'SELECT * FROM courier_extra_routes WHERE id=? AND employee_shift_id=?', (route_id, staff_id)
        ).fetchone()
        existing = conn.execute(
            'SELECT * FROM employee_shifts WHERE id=? AND shift_id=?', (staff_id, shift_id)
        ).fetchone()
        if not route or not existing:
            return jsonify({'ok': False, 'error': 'not found'}), 404
        km_delta = new_km - float(route['km'] or 0)
        conn.execute('UPDATE courier_extra_routes SET reason=?, km=?, comment=? WHERE id=?',
                     (reason, new_km, comment, route_id))
        aggregate_km = float(existing['km'] or 0) + km_delta
        aggregate_orders = int(existing['orders'] or 0)  # число маршрутов не меняется
        base_pay = _recalc_courier_base_pay(existing, aggregate_km, aggregate_orders)
        conn.execute('UPDATE employee_shifts SET km=?, base_pay=? WHERE id=?',
                     (aggregate_km, base_pay, staff_id))
        auto_bonuses = calculate_bonuses(conn, shift_id)
        total_row = conn.execute('SELECT total_amount FROM employee_shifts WHERE id=?', (staff_id,)).fetchone()
        log_action(conn, 'staff_update',
            f"Изменён доп. маршрут: {ROUTE_REASON_LABELS[reason]}, {new_km:g} км",
            shift_id=shift_id, entity_id=staff_id)
        conn.commit()
    return jsonify({'ok': True, 'km_total': aggregate_km, 'orders_total': aggregate_orders,
                    'total_amount': total_row['total_amount'], 'auto_bonuses': auto_bonuses})


@app.route('/shift/<int:shift_id>/staff/<int:staff_id>/route/<int:route_id>/delete', methods=['POST'])
@login_required
def delete_courier_route(shift_id, staff_id, route_id):
    if not _can_edit_shift(shift_id):
        return jsonify({'error': 'Нет доступа'}), 403
    with get_db() as conn:
        route = conn.execute(
            'SELECT * FROM courier_extra_routes WHERE id=? AND employee_shift_id=?', (route_id, staff_id)
        ).fetchone()
        existing = conn.execute(
            'SELECT * FROM employee_shifts WHERE id=? AND shift_id=?', (staff_id, shift_id)
        ).fetchone()
        if not route or not existing:
            return jsonify({'ok': False, 'error': 'not found'}), 404
        conn.execute('DELETE FROM courier_extra_routes WHERE id=?', (route_id,))
        aggregate_km = float(existing['km'] or 0) - float(route['km'] or 0)
        aggregate_orders = int(existing['orders'] or 0) - 1
        base_pay = _recalc_courier_base_pay(existing, aggregate_km, aggregate_orders)
        conn.execute('UPDATE employee_shifts SET km=?, orders=?, base_pay=? WHERE id=?',
                     (aggregate_km, aggregate_orders, base_pay, staff_id))
        auto_bonuses = calculate_bonuses(conn, shift_id)
        total_row = conn.execute('SELECT total_amount FROM employee_shifts WHERE id=?', (staff_id,)).fetchone()
        log_action(conn, 'staff_update', "Удалён доп. маршрут", shift_id=shift_id, entity_id=staff_id)
        conn.commit()
    return jsonify({'ok': True, 'km_total': aggregate_km, 'orders_total': aggregate_orders,
                    'total_amount': total_row['total_amount'], 'auto_bonuses': auto_bonuses})


_PAY_MONTHLY_EPOCH = '2000-01-01'  # сентинел «действовало всегда до сих пор» для бэкофилла старого значения


def _seed_pay_monthly_history(conn, employee_id, role, pay_monthly):
    """Начальная запись при создании сотрудника/доп.должности — действует «всегда» (с сентинела),
    чтобы у только что созданной записи сразу был определён режим на любую дату."""
    conn.execute(
        'INSERT INTO employee_pay_monthly_history (employee_id, role, pay_monthly, effective_from) VALUES (?,?,?,?)',
        (employee_id, role, pay_monthly, _PAY_MONTHLY_EPOCH)
    )


def _log_pay_monthly_change(conn, employee_id, role, old_pay_monthly, new_pay_monthly, effective_from=None):
    """Пишет в историю смену режима ежедневно/ежемесячно (если значение реально изменилось).
    Если для этой должности ещё нет ни одной записи — сначала «подкладывает» старое значение
    задним числом (с сентинела), иначе прошлые смены задним числом посчитались бы по новому режиму."""
    if int(old_pay_monthly) == int(new_pay_monthly):
        return
    effective_from = effective_from or date.today().isoformat()
    has_history = conn.execute(
        'SELECT 1 FROM employee_pay_monthly_history WHERE employee_id=? AND role=? LIMIT 1',
        (employee_id, role)
    ).fetchone()
    if not has_history:
        conn.execute(
            'INSERT INTO employee_pay_monthly_history (employee_id, role, pay_monthly, effective_from) VALUES (?,?,?,?)',
            (employee_id, role, old_pay_monthly, _PAY_MONTHLY_EPOCH)
        )
    conn.execute(
        'INSERT INTO employee_pay_monthly_history (employee_id, role, pay_monthly, effective_from) VALUES (?,?,?,?)',
        (employee_id, role, new_pay_monthly, effective_from)
    )


def _effective_pay_monthly(conn, employee_id, role, on_date, fallback):
    """Режим оплаты (0/1), действовавший на конкретную дату — а не текущий, чтобы смена режима
    сегодня не влияла на уже прошедшие смены задним числом."""
    row = conn.execute(
        'SELECT pay_monthly FROM employee_pay_monthly_history WHERE employee_id=? AND role=? AND effective_from<=? '
        'ORDER BY effective_from DESC, id DESC LIMIT 1',
        (employee_id, role, on_date)
    ).fetchone()
    return row['pay_monthly'] if row is not None else fallback


@app.route('/shift/<int:shift_id>/pay-staff/<int:staff_id>', methods=['POST'])
@login_required
def pay_staff(shift_id, staff_id):
    if not _can_edit_shift(shift_id):
        return jsonify({'error': 'Нет доступа'}), 403
    data = request.json or {}
    amount = _f(data, 'amount')
    with get_db() as conn:
        row = conn.execute(
            'SELECT * FROM employee_shifts WHERE id=? AND shift_id=?', (staff_id, shift_id)
        ).fetchone()
        if not row:
            return jsonify({'error': 'Не найдено'}), 404
        if row['employee_id']:
            emp = conn.execute('SELECT role, pay_monthly FROM employees WHERE id=?', (row['employee_id'],)).fetchone()
            if emp:
                shift_row_dt = conn.execute('SELECT date FROM shifts WHERE id=?', (shift_id,)).fetchone()
                shift_date = shift_row_dt['date'] if shift_row_dt else date.today().isoformat()
                if row['role_snapshot'] == emp['role']:
                    fallback = emp['pay_monthly']
                else:
                    er = conn.execute(
                        'SELECT pay_monthly FROM employee_roles WHERE employee_id=? AND role=?',
                        (row['employee_id'], row['role_snapshot'])
                    ).fetchone()
                    fallback = er['pay_monthly'] if er else 0
                # Смотрим режим, действовавший на дату самой смены — а не текущий,
                # чтобы переключение сегодня не блокировало уже прошедшие смены.
                pm = _effective_pay_monthly(conn, row['employee_id'], row['role_snapshot'], shift_date, fallback)
                if pm:
                    # Правило ежемесячной оплаты может быть ограничено конкретными филиалами
                    # (employee_pay_monthly_branches). Если для этой должности ничего не выбрано —
                    # правило действует везде, как и раньше.
                    scope = conn.execute(
                        'SELECT branch_id FROM employee_pay_monthly_branches WHERE employee_id=? AND role=?',
                        (row['employee_id'], row['role_snapshot'])
                    ).fetchall()
                    scope_ids = {s['branch_id'] for s in scope}
                    shift_row = conn.execute('SELECT branch_id FROM shifts WHERE id=?', (shift_id,)).fetchone()
                    shift_branch_id = shift_row['branch_id'] if shift_row else None
                    if not scope_ids or shift_branch_id in scope_ids:
                        return jsonify({'error': 'pay_monthly'}), 400
        conn.execute(
            'UPDATE employee_shifts SET is_paid=1, paid_amount=? WHERE id=?',
            (amount, staff_id)
        )
        conn.execute('''
            INSERT INTO salary_payments
            (employee_id, employee_shift_id, amount, payment_date, paid_by, paid_by_name)
            VALUES (?,?,?,?,?,?)
        ''', (row['employee_id'], staff_id, amount, date.today().isoformat(),
              session['user_id'], session.get('full_name', '')))
        name = row['full_name_snapshot']
        role_lbl = ROLE_LABELS.get(row['role_snapshot'], row['role_snapshot'])
        log_action(conn, 'salary_paid',
            f"Выплата ЗП: {name} ({role_lbl}), {_fmt_money(amount)}",
            shift_id=shift_id, entity_id=staff_id)
        conn.commit()
    return jsonify({'ok': True})


@app.route('/shift/<int:shift_id>/delete-staff/<int:staff_id>', methods=['POST'])
@login_required
def delete_staff(shift_id, staff_id):
    if not _can_edit_shift(shift_id):
        return jsonify({'error': 'Нет доступа'}), 403
    with get_db() as conn:
        row = conn.execute(
            'SELECT * FROM employee_shifts WHERE id=? AND shift_id=?', (staff_id, shift_id)
        ).fetchone()
        if row:
            name = row['full_name_snapshot']
            role_lbl = ROLE_LABELS.get(row['role_snapshot'], row['role_snapshot'])
            desc = f"Удалён из смены: {name} ({role_lbl})"
        else:
            desc = f"Удалён сотрудник #{staff_id}"
        conn.execute('DELETE FROM employee_shifts WHERE id=? AND shift_id=?', (staff_id, shift_id))
        log_action(conn, 'staff_delete', desc, shift_id=shift_id, entity_id=staff_id)
        auto_bonuses = calculate_bonuses(conn, shift_id)
        conn.commit()
    return jsonify({'ok': True, 'auto_bonuses': auto_bonuses})


GSHEET_COLS = [
    ('date',            'Дата'),
    ('branch',          'Филиал'),
    ('revenue_total',   'Выручка итого'),
    ('revenue_cash',    'Наличные приход'),
    ('revenue_card',    'Безнал приход'),
    ('actual_cash',     'Факт в кассе'),
    ('exp_cash_total',  'Расходы нал итого'),
    ('exp_card_total',  'Расходы карта итого'),
    ('exp_by_cat',      'Расходы по категориям'),
    ('salary_total',    'ФОТ итого'),
    ('salary_by_role',  'ФОТ по должностям'),
    ('profit',          'Прибыль'),
    ('closed_by',       'Кто закрыл'),
    ('comment',         'Комментарий'),
]

def _gsheet_load_settings(conn):
    rows = conn.execute('SELECT key, value FROM gsheet_settings').fetchall()
    cfg  = {r['key']: r['value'] for r in rows}
    # По умолчанию все включены
    result = {}
    for key, _ in GSHEET_COLS:
        result[key] = int(cfg.get(f'col_{key}', '1'))
    return result


def _export_shift_to_gsheet(shift_id):
    """Экспорт смены в Google Sheets. Ошибки не прерывают работу."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        creds_json = os.environ.get('GOOGLE_CREDENTIALS_JSON')
        sheet_id   = os.environ.get('GOOGLE_SHEET_ID')
        if not creds_json or not sheet_id:
            return

        creds_dict = _json_lib.loads(creds_json)
        scopes = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive',
        ]
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        gc    = gspread.authorize(creds)
        sh    = gc.open_by_key(sheet_id)

        with get_db() as conn:
            shift = conn.execute('''
                SELECT s.*, b.name AS branch_name
                FROM shifts s JOIN branches b ON b.id = s.branch_id
                WHERE s.id = ?
            ''', (shift_id,)).fetchone()
            if not shift:
                return

            rev = dict(conn.execute(
                'SELECT * FROM shift_revenue WHERE shift_id=?', (shift_id,)
            ).fetchone() or {})

            all_exp_cats = conn.execute(
                'SELECT code, label FROM expense_categories ORDER BY sort_order, label'
            ).fetchall()

            exp_totals = conn.execute('''
                SELECT COALESCE(SUM(amount_cash),0) AS cash_total,
                       COALESCE(SUM(amount_card),0) AS card_total
                FROM expenses WHERE shift_id=?
            ''', (shift_id,)).fetchone()

            exp_by_cat_rows = conn.execute('''
                SELECT e.category,
                       COALESCE(SUM(e.amount_cash),0) AS cash_amt,
                       COALESCE(SUM(e.amount_card),0) AS card_amt
                FROM expenses e WHERE e.shift_id=?
                GROUP BY e.category
            ''', (shift_id,)).fetchall()
            exp_by_cat = {r['category']: {'cash': r['cash_amt'] or 0, 'card': r['card_amt'] or 0}
                          for r in exp_by_cat_rows}

            exp_items = conn.execute('''
                SELECT COALESCE(ec.label, e.category, '?') AS cat_label,
                       e.description, e.amount_cash, e.amount_card
                FROM expenses e
                LEFT JOIN expense_categories ec ON ec.code = e.category
                WHERE e.shift_id=?
                ORDER BY ec.sort_order, e.category, e.id
            ''', (shift_id,)).fetchall()

            staff = conn.execute('''
                SELECT full_name_snapshot, role_snapshot,
                       shift_start, shift_end, hours_worked, km, orders,
                       rate_snapshot, base_pay, bonus_amount, penalty_amount,
                       auto_bonus, total_amount, is_paid, paid_amount, bonus_comment
                FROM employee_shifts WHERE shift_id=?
                ORDER BY role_snapshot, full_name_snapshot
            ''', (shift_id,)).fetchall()

            taxi_trips = conn.execute(
                'SELECT * FROM taxi_trips WHERE shift_id=? ORDER BY id', (shift_id,)
            ).fetchall()
            taxi_emps = {}
            for t in taxi_trips:
                taxi_emps[t['id']] = conn.execute(
                    'SELECT name_snapshot, address_snapshot FROM taxi_trip_employees WHERE trip_id=? ORDER BY id',
                    (t['id'],)
                ).fetchall()

            terminals = conn.execute(
                'SELECT terminal_number, amount FROM shift_terminals WHERE shift_id=? ORDER BY sort_order, id',
                (shift_id,)
            ).fetchall()

        def _rv(col, default=0):
            v = rev.get(col, default)
            return v if v is not None else default

        total_rev      = _rv('total_revenue')
        cash_rev       = _rv('cash_amount')
        card_rev       = _rv('card_amount')
        online_rev     = _rv('online_amount')
        change_amt     = _rv('change_amount')
        morning_cash   = _rv('morning_cash', 0)
        actual_cash    = _rv('actual_cash', '')
        actual_cash_cmt= _rv('actual_cash_comment', '')
        delivery_rev   = _rv('delivery_revenue')
        delivery_orders= int(_rv('delivery_orders'))
        pickup_rev     = _rv('pickup_revenue')
        pickup_orders  = int(_rv('pickup_orders'))
        exp_cash       = round((exp_totals['cash_total'] or 0) if exp_totals else 0, 2)
        exp_card       = round((exp_totals['card_total'] or 0) if exp_totals else 0, 2)
        salary         = round(sum((s['total_amount'] or 0) for s in staff), 2)
        taxi_total     = round(sum((t['amount'] or 0) for t in taxi_trips), 2)
        profit         = round(total_rev - exp_cash - exp_card - salary, 2)

        role_labels_map = {
            'admin': 'Администратор', 'cook': 'Повар', 'sushi': 'Сушист',
            'courier': 'Курьер', 'packer': 'Упаковщик', 'cleaner': 'Уборщица',
        }
        all_roles = ['admin', 'cook', 'sushi', 'courier', 'packer', 'cleaner']
        sal_by_role = {}
        for s in staff:
            role = s['role_snapshot'] or 'other'
            sal_by_role[role] = sal_by_role.get(role, 0) + (s['total_amount'] or 0)

        # ─── Build header ──────────────────────────────────────────────────
        header = [
            'Дата', 'Филиал',
            'Выручка итого',
            'Доставка сумма', 'Доставка заказов',
            'Самовывоз сумма', 'Самовывоз заказов',
            'Наличные приход', 'Безнал приход', 'Онлайн приход',
            'Размен', 'Утренняя касса',
            'Факт в кассе', 'Комментарий к факту',
            'Терминалы',
            'Расходы нал итого', 'Расходы безнал итого',
        ]
        for ec in all_exp_cats:
            header.append(f'{ec["label"]} (нал)')
            header.append(f'{ec["label"]} (безнал)')
        header += [
            'Детали расходов',
            'ФОТ итого',
        ]
        for role in all_roles:
            header.append(f'ФОТ {role_labels_map[role]}')
        header += [
            'Такси итого',
            'Сотрудники',
            'Детали такси',
            'Прибыль', 'Кто закрыл', 'Комментарий смены',
        ]

        # ─── Build row ─────────────────────────────────────────────────────
        terminals_text = '; '.join(
            f'{t["terminal_number"]}: {round(t["amount"] or 0, 2)}р' for t in terminals
        )

        exp_details = []
        for ei in exp_items:
            total_ei = (ei['amount_cash'] or 0) + (ei['amount_card'] or 0)
            pay = 'нал' if (ei['amount_cash'] or 0) > 0 else 'безнал'
            line = f'{ei["cat_label"]}: {round(total_ei, 2)}р ({pay})'
            if ei['description']:
                line += f' — {ei["description"]}'
            exp_details.append(line)

        staff_lines = []
        for s in staff:
            role_lbl = role_labels_map.get(s['role_snapshot'], s['role_snapshot'] or '')
            hours = s['hours_worked'] or 0
            total = s['total_amount'] or 0
            paid_mark = '✓' if s['is_paid'] else '—'
            line = f'{s["full_name_snapshot"]} ({role_lbl}): {hours}ч → {round(total, 2)}р [{paid_mark}]'
            if s['bonus_comment']:
                line += f' ({s["bonus_comment"]})'
            staff_lines.append(line)

        taxi_lines = []
        for t in taxi_trips:
            emps = taxi_emps.get(t['id'], [])
            emp_names = ', '.join(e['name_snapshot'] for e in emps)
            addrs = [e['address_snapshot'] for e in emps if e['address_snapshot']]
            pay_type = 'нал' if (t['payment_type'] or 'cash') == 'cash' else 'безнал'
            line = f'{round(t["amount"] or 0, 2)}р ({pay_type})'
            if emp_names:
                line += f': {emp_names}'
            if addrs:
                line += f' → {", ".join(addrs)}'
            if t['note']:
                line += f' [{t["note"]}]'
            taxi_lines.append(line)

        row = [
            shift['date'], shift['branch_name'],
            round(total_rev, 2),
            round(delivery_rev, 2), delivery_orders,
            round(pickup_rev, 2), pickup_orders,
            round(cash_rev, 2), round(card_rev, 2), round(online_rev, 2),
            round(change_amt, 2), round(morning_cash, 2),
            round(actual_cash, 2) if actual_cash != '' else '',
            actual_cash_cmt,
            terminals_text,
            exp_cash, exp_card,
        ]
        for ec in all_exp_cats:
            cat = exp_by_cat.get(ec['code'], {'cash': 0, 'card': 0})
            row.append(round(cat['cash'], 2))
            row.append(round(cat['card'], 2))
        row += [
            '\n'.join(exp_details),
            salary,
        ]
        for role in all_roles:
            row.append(round(sal_by_role.get(role, 0), 2))
        row += [
            taxi_total,
            '\n'.join(staff_lines),
            '\n'.join(taxi_lines),
            profit,
            shift['closed_by_name'] or '',
            shift['comment'] or '',
        ]

        year = (shift['date'] or '')[:4] or str(date.today().year)
        try:
            ws = sh.worksheet(year)
        except Exception:
            ws = sh.add_worksheet(title=year, rows=3000, cols=len(header) + 5)
            ws.append_row(header)

        ws.append_row(row, value_input_option='USER_ENTERED')
        print(f'[GSheets] shift {shift_id} exported OK ({len(row)} columns)')
    except Exception as e:
        print(f'[GSheets] shift {shift_id} export error: {e}')


def _gdrive_get_oauth_token():
    """Получить OAuth2 access token через сохранённый refresh token."""
    import urllib.request as _ur
    import urllib.parse as _up
    client_id     = os.environ.get('GOOGLE_OAUTH_CLIENT_ID')
    client_secret = os.environ.get('GOOGLE_OAUTH_CLIENT_SECRET')
    with get_db() as conn:
        row = conn.execute(
            "SELECT value FROM gsheet_settings WHERE key='gdrive_refresh_token'"
        ).fetchone()
    if not client_id or not client_secret or not row:
        return None
    data = _up.urlencode({
        'client_id':     client_id,
        'client_secret': client_secret,
        'refresh_token': row['value'],
        'grant_type':    'refresh_token',
    }).encode()
    req = _ur.Request(
        'https://oauth2.googleapis.com/token',
        data=data,
        headers={'Content-Type': 'application/x-www-form-urlencoded'},
    )
    with _ur.urlopen(req, timeout=15) as resp:
        return _json_lib.loads(resp.read()).get('access_token')


def _export_shift_to_gdrive_xlsx(shift_id):
    """Экспорт смены в Google Drive как xlsx в формате импорта."""
    print(f'[GDrive] start shift {shift_id}')
    try:
        import openpyxl
        from openpyxl import Workbook
        import io as _io
        import urllib.request
        import datetime as _dt_mod

        creds_json = os.environ.get('GOOGLE_CREDENTIALS_JSON')
        folder_id  = os.environ.get('GOOGLE_DRIVE_FOLDER_ID')
        print(f'[GDrive] creds={bool(creds_json)} folder_id={folder_id!r}')
        if not creds_json or not folder_id:
            print(f'[GDrive] missing env vars, exit')
            return

        with get_db() as conn:
            shift = conn.execute('''
                SELECT s.*, b.name AS branch_name
                FROM shifts s JOIN branches b ON b.id = s.branch_id
                WHERE s.id = ?
            ''', (shift_id,)).fetchone()
            if not shift:
                return

            rev = dict(conn.execute(
                'SELECT * FROM shift_revenue WHERE shift_id=?', (shift_id,)
            ).fetchone() or {})

            exp_rows = conn.execute(
                "SELECT category, description, amount_cash, amount_card FROM expenses WHERE shift_id=? ORDER BY category, id",
                (shift_id,)
            ).fetchall()

            taxi_trips = conn.execute(
                'SELECT amount, payment_type, note FROM taxi_trips WHERE shift_id=? ORDER BY id',
                (shift_id,)
            ).fetchall()

            couriers = conn.execute(
                "SELECT full_name_snapshot, hours_worked, km, orders, rate_snapshot, "
                "rate_per_km_snapshot, rate_per_order_snapshot, bonus_comment, "
                "bonus_amount, penalty_amount, total_amount, is_paid "
                "FROM employee_shifts WHERE shift_id=? AND role_snapshot='courier' ORDER BY full_name_snapshot",
                (shift_id,)
            ).fetchall()

            non_couriers = conn.execute(
                "SELECT full_name_snapshot, role_snapshot, rate_snapshot, shift_start, shift_end, "
                "hours_worked, bonus_amount, penalty_amount, bonus_comment, base_pay, total_amount, is_paid "
                "FROM employee_shifts WHERE shift_id=? AND role_snapshot IN ('admin','sushi','cleaner','cook','packer') "
                "ORDER BY role_snapshot, full_name_snapshot",
                (shift_id,)
            ).fetchall()

            terminals = conn.execute(
                'SELECT terminal_number, amount FROM shift_terminals WHERE shift_id=? ORDER BY sort_order, id',
                (shift_id,)
            ).fetchall()

        def _rv(col, default=0):
            v = rev.get(col, default)
            return v if v is not None else default

        # Reverse map: category code → import label
        _CAT_REVERSE = {}
        for _lbl, _code in _XL_CAT_MAP.items():
            if _code not in _CAT_REVERSE:
                _CAT_REVERSE[_code] = _lbl

        # Fixed row positions for known expense categories (match import template)
        _CAT_ROW = {
            'repair_plumbing': 10, 'repair_grease':  11,
            'repair_electric': 12, 'repair_fridge':  13,
            'repair_other':    14, 'shop':            15,
        }

        # Aggregate expenses by category
        exp_by_cat = {}
        for e in exp_rows:
            cat = e['category'] or 'other'
            if cat == 'taxi':
                continue
            if cat not in exp_by_cat:
                exp_by_cat[cat] = {'cash': 0.0, 'card': 0.0, 'descs': []}
            exp_by_cat[cat]['cash'] += e['amount_cash'] or 0
            exp_by_cat[cat]['card'] += e['amount_card'] or 0
            if e['description']:
                exp_by_cat[cat]['descs'].append(e['description'])

        # Build workbook
        wb = Workbook()
        ws = wb.active
        shift_date = _dt_mod.date.fromisoformat(shift['date'])
        wd_names = ['ПН', 'ВТ', 'СР', 'ЧТ', 'ПТ', 'СБ', 'ВС']
        ws.title = wd_names[shift_date.weekday()]

        def w(row, col, val):
            ws.cell(row=row, column=col, value=val)

        # ─── Revenue (cells match what _xl_process_sheet reads) ────────────
        w(2, 6, _rv('total_revenue'))            # F2
        w(3, 2, shift_date)                       # B3 = date
        w(3, 4, _rv('morning_cash', 0))           # D3 = утренняя касса
        w(4, 7, _rv('delivery_revenue'))          # G4
        w(5, 4, _rv('cash_amount'))               # D5
        w(5, 7, int(_rv('delivery_orders')))      # G5
        w(6, 4, _rv('card_amount'))               # D6
        w(6, 7, _rv('pickup_revenue'))            # G6
        w(7, 4, _rv('online_amount'))             # D7
        w(7, 7, int(_rv('pickup_orders')))        # G7
        w(31, 4, _rv('change_amount', 0))         # D31 = размен
        w(33, 4, _rv('plus_amount', 0))           # D33 = плюс в кассу

        # ─── Couriers (rows 3-8, cols K-V) ─────────────────────────────────
        for i, c in enumerate(couriers[:6]):
            r = 3 + i
            km      = c['km'] or 0
            orders  = c['orders'] or 0
            hours   = c['hours_worked'] or 0
            km_rate = c['rate_per_km_snapshot'] or 10
            or_rate = c['rate_per_order_snapshot'] or 100
            hr_rate = c['rate_snapshot'] or 0
            bonus_amt   = c['bonus_amount'] or 0
            penalty_amt = c['penalty_amount'] or 0
            w(r, 11, c['full_name_snapshot'])          # K
            w(r, 13, km)                               # M = km
            w(r, 14, round(km * km_rate, 2))           # N = km pay
            w(r, 15, hours)                            # O = hours
            w(r, 16, round(hours * hr_rate, 2))        # P = hrs pay
            w(r, 17, orders)                           # Q = orders
            w(r, 18, round(orders * or_rate, 2))       # R = ord pay
            if bonus_amt > 0:
                w(r, 19, 'премия')                     # S = label
                w(r, 20, bonus_amt)                    # T = amount
            elif penalty_amt > 0:
                w(r, 19, 'штраф')                      # S = label
                w(r, 20, penalty_amt)                  # T = amount
            else:
                w(r, 19, c['bonus_comment'] or '')     # S = comment
            w(r, 21, 'Да' if c['is_paid'] else '')    # U = paid
            w(r, 22, c['total_amount'] or 0)           # V = total

        # ─── Expense categories (rows 10-20, cols B,C,F,G) ─────────────────
        # Write labels for known categories
        for code, row in _CAT_ROW.items():
            lbl = _CAT_REVERSE.get(code, '')
            if lbl:
                w(row, 2, lbl)
        w(20, 2, 'Другое')

        other_cash, other_card, other_descs = 0.0, 0.0, []
        for cat_code, data in exp_by_cat.items():
            cat_row = _CAT_ROW.get(cat_code)
            if cat_row:
                w(cat_row, 6, round(data['cash'], 2))    # F = нал
                w(cat_row, 7, round(data['card'], 2))    # G = безнал
                if data['descs']:
                    w(cat_row, 3, '; '.join(data['descs'][:3]))  # C = описание
            else:
                other_cash += data['cash']
                other_card += data['card']
                other_descs.extend(data['descs'])
        if other_cash or other_card:
            w(20, 6, round(other_cash, 2))
            w(20, 7, round(other_card, 2))
            if other_descs:
                w(20, 3, '; '.join(other_descs[:5]))

        # ─── Admin/Sushi/Cleaner (rows 10-22, cols K-V) ────────────────────
        # Rows 10-14 → admin role (row_idx 9-13 in importer)
        # Rows 15-21 → sushi role (row_idx 14-20)
        # Row 22     → cleaner (name must be 'Уборщица' for importer)
        admins   = [s for s in non_couriers if s['role_snapshot'] in ('admin', 'cook', 'packer')]
        sushis   = [s for s in non_couriers if s['role_snapshot'] == 'sushi']
        cleaners = [s for s in non_couriers if s['role_snapshot'] == 'cleaner']

        def _write_staff(r, s):
            bonus_val = (s['bonus_amount'] or 0) - (s['penalty_amount'] or 0)
            base_pay  = s['base_pay'] or 0
            w(r, 11, s['full_name_snapshot'])
            w(r, 12, s['role_snapshot'])              # L = роль (для корректного реимпорта)
            w(r, 13, s['rate_snapshot'] or 0)        # M = ставка
            if s['shift_start']:
                w(r, 14, s['shift_start'])            # N = начало
            if s['shift_end']:
                w(r, 15, s['shift_end'])              # O = конец
            w(r, 16, s['hours_worked'] or 0)          # P = часы
            if bonus_val:
                w(r, 18, bonus_val)                   # R = премия/штраф (legacy)
            if s['bonus_comment']:
                w(r, 19, s['bonus_comment'])          # S = комментарий
            if base_pay:
                w(r, 20, base_pay)                    # T = база до премии/штрафа
            w(r, 21, 'Да' if s['is_paid'] else '')   # U = выплачено
            w(r, 22, s['total_amount'] or 0)          # V = итого

        for i, s in enumerate(admins[:5]):
            _write_staff(10 + i, s)

        for i, s in enumerate(sushis[:7]):
            _write_staff(15 + i, s)

        # Cleaner: write at row 22; importer identifies by name 'Уборщица'
        if cleaners:
            cl = cleaners[0]
            total_cl = sum(s['total_amount'] or 0 for s in cleaners)
            w(22, 11, 'Уборщица')
            w(22, 13, cl['rate_snapshot'] or 0)
            w(22, 16, sum(s['hours_worked'] or 0 for s in cleaners))
            w(22, 21, 'Да' if cl['is_paid'] else '')
            w(22, 22, total_cl)

        # ─── Taxi section (rows 22-24) ──────────────────────────────────────
        # Row 22 col B = 'ТАКСИ' skips it as taxi data; actual trips in 23-24
        w(22, 2, 'ТАКСИ')
        for i, t in enumerate(taxi_trips[:2]):
            r = 23 + i
            cash = round(t['amount'] or 0, 2) if (t['payment_type'] or 'cash') == 'cash' else 0
            card = round(t['amount'] or 0, 2) if (t['payment_type'] or 'cash') != 'cash' else 0
            w(r, 3, t['note'] or '')   # C = описание
            w(r, 6, card)              # F = безнал (в импорте F=card, G=cash — инверсия!)
            w(r, 7, cash)              # G = нал

        # ─── Terminals (rows 26, 29-32) ─────────────────────────────────────
        if terminals:
            w(26, 5, 'По терминалам:')
            w(26, 7, round(sum(t['amount'] or 0 for t in terminals), 2))
            for i, t in enumerate(terminals[:4]):
                w(29 + i, 5, t['terminal_number'])        # E = номер
                w(29 + i, 7, round(t['amount'] or 0, 2)) # G = сумма

        # ─── Actual cash, who closed ────────────────────────────────────────
        actual_cash = _rv('actual_cash', None)
        if actual_cash is not None:
            w(27, 2, 'Факт в кассе:')
            w(27, 4, round(actual_cash, 2))

        closed_by = shift['closed_by_name'] if 'closed_by_name' in shift.keys() else None
        if closed_by:
            w(28, 2, 'Смену закрыл(а):')
            w(28, 5, closed_by)

        # ─── Descriptive labels (human-readable, do not conflict with importer) ─
        # Importer never reads col A; reads col B only for specific string matches;
        # col J (10) is never read; row 2 is above courier loop (rows[2:8]→rows 3-8);
        # row 9 is above non-courier loop (rows[9:22]→rows 10-22).

        # Branch header row 1
        w(1, 1, 'Филиал:');      w(1, 2, shift['branch_name'])

        # Revenue area — col A labels, col F context labels (importer reads D and G)
        w(2, 1, 'Итого выручка:')
        w(3, 1, 'Дата:');        w(3, 3, 'Утр. касса:')
        w(4, 1, 'Доставка:');    w(4, 6, 'Выручка:')
        w(5, 1, 'Наличные:');    w(5, 6, 'Доставка заказов:')
        w(6, 1, 'Карта:');       w(6, 6, 'Самовывоз:')
        w(7, 1, 'Онлайн:');      w(7, 6, 'Самовывоз заказов:')

        # Размен / плюс в кассу labels (B col not read by importer for values)
        w(29, 2, 'Масло:');      w(30, 2, 'Рыба:')
        w(31, 2, 'Размен:');     w(33, 2, 'Плюс в кассу:')

        # Expense section header (row 9, cols A-G; row 9 is not in expense import loop)
        w(9, 1, 'РАСХОДЫ');      w(9, 2, 'Категория')
        w(9, 6, 'Нал');          w(9, 7, 'Безнал')

        # Courier column headers (row 2, cols J-V; row 2 is above courier loop rows 3-8)
        w(2, 10, 'КУРЬЕРЫ')
        w(2, 11, 'ФИО');         w(2, 13, 'КМ');      w(2, 14, 'За КМ')
        w(2, 15, 'Часы');        w(2, 16, 'За часы')
        w(2, 17, 'Заказов');     w(2, 18, 'Оплата')
        w(2, 19, 'Бонус/Штраф'); w(2, 20, 'Сумма')
        w(2, 21, 'Выплачено');   w(2, 22, 'ИТОГО')

        # Non-courier column headers (row 9, cols J-V; row 9 not in non-courier loop)
        w(9, 10, 'СОТРУДНИКИ')
        w(9, 11, 'ФИО');         w(9, 12, 'Роль');    w(9, 13, 'Ставка')
        w(9, 14, 'Начало');      w(9, 15, 'Конец');   w(9, 16, 'Часы')
        w(9, 18, 'Бонус');       w(9, 19, 'Комментарий')
        w(9, 20, 'База');        w(9, 21, 'Выплачено'); w(9, 22, 'ИТОГО')

        # Sub-section labels in col J (never read by importer)
        w(10, 10, 'Администраторы:')
        w(15, 10, 'Сушисты:')
        w(22, 10, 'Уборщица:')

        # ─── Save & upload to Google Drive ──────────────────────────────────
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx_bytes = buf.getvalue()

        token = _gdrive_get_oauth_token()
        if not token:
            print(f'[GDrive] shift {shift_id}: OAuth2 токен не получен — авторизуйте Drive в /settings/gsheet')
            return

        branch_nm = shift['branch_name'].replace(' ', '_')
        filename = f'{branch_nm}_{shift["date"]}_{wd_names[shift_date.weekday()]}.xlsx'
        xlsx_mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        boundary = '---GDriveXlsxBnd9347'
        meta_json = _json_lib.dumps({'name': filename, 'parents': [folder_id]})
        body = (
            f'--{boundary}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n'
            f'{meta_json}\r\n'
            f'--{boundary}\r\nContent-Type: {xlsx_mime}\r\n\r\n'
        ).encode('utf-8') + xlsx_bytes + f'\r\n--{boundary}--'.encode('utf-8')

        upload_req = urllib.request.Request(
            'https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart',
            data=body,
            headers={
                'Authorization': f'Bearer {token}',
                'Content-Type': f'multipart/related; boundary="{boundary}"',
            },
            method='POST'
        )
        with urllib.request.urlopen(upload_req, timeout=60) as resp:
            result = _json_lib.loads(resp.read())
        print(f'[GDrive] shift {shift_id} xlsx uploaded: {filename} id={result.get("id","?")}')

    except Exception as e:
        print(f'[GDrive] shift {shift_id} xlsx error: {e}')


@app.route('/shift/<int:shift_id>/close', methods=['POST'])
@login_required
def close_shift(shift_id):
    if not _can_edit_shift(shift_id):
        flash('Нет доступа', 'danger')
        return redirect(url_for('shift_view', shift_id=shift_id))
    try:
        comment = request.form.get('comment', '')
        closed_by_name = request.form.get('closed_by_name', session.get('full_name', ''))
        actual_cash_comment = request.form.get('actual_cash_comment', '').strip()
        if not closed_by_name or not closed_by_name.strip():
            flash('Укажите, кто закрыл смену', 'danger')
            return redirect(url_for('shift_view', shift_id=shift_id))
        with get_db() as conn:
            rows_updated = conn.execute('''
                UPDATE shifts SET status='closed', closed_by=?, closed_at=CURRENT_TIMESTAMP,
                comment=?, closed_by_name=?
                WHERE id=? AND status='open'
            ''', (session['user_id'], comment, closed_by_name, shift_id)).rowcount
            if rows_updated == 0:
                flash('Смена не найдена или уже закрыта', 'danger')
                conn.rollback()
                return redirect(url_for('shift_view', shift_id=shift_id))
            if actual_cash_comment:
                conn.execute(
                    'UPDATE shift_revenue SET actual_cash_comment=? WHERE shift_id=?',
                    (actual_cash_comment, shift_id)
                )
            log_action(conn, 'shift_close', 'Смена закрыта', shift_id=shift_id)
            conn.commit()
        flash('Смена закрыта', 'success')
        threading.Thread(target=_export_shift_to_gsheet, args=(shift_id,), daemon=True).start()
        threading.Thread(target=_export_shift_to_gdrive_xlsx, args=(shift_id,), daemon=True).start()
    except Exception as e:
        flash(f'Ошибка при закрытии смены: {e}', 'danger')
    return redirect(url_for('shift_view', shift_id=shift_id))


@app.route('/shift/<int:shift_id>/reopen', methods=['POST'])
@login_required
@menu_permission_required('shifts_archive')
def reopen_shift(shift_id):
    with get_db() as conn:
        conn.execute(
            "UPDATE shifts SET status='open', closed_by=NULL, closed_at=NULL WHERE id=?",
            (shift_id,)
        )
        log_action(conn, 'shift_reopen', 'Смена переоткрыта', shift_id=shift_id)
        conn.commit()
    flash('Смена переоткрыта', 'success')
    return redirect(url_for('shift_view', shift_id=shift_id))


@app.route('/shift/<int:shift_id>/delete', methods=['POST'])
@login_required
@menu_permission_required('shifts_archive')
def delete_shift(shift_id):
    with get_db() as conn:
        shift = conn.execute('SELECT * FROM shifts WHERE id=?', (shift_id,)).fetchone()
        if not shift:
            flash('Смена не найдена', 'danger')
            return redirect(url_for('dashboard'))
        branch_date = f'{shift["date"]} · {shift["branch_id"]}'
        conn.execute(
            'DELETE FROM salary_payments WHERE employee_shift_id IN (SELECT id FROM employee_shifts WHERE shift_id=?)',
            (shift_id,)
        )
        conn.execute('DELETE FROM employee_shifts WHERE shift_id=?', (shift_id,))
        conn.execute('DELETE FROM expenses WHERE shift_id=?', (shift_id,))
        conn.execute(
            'DELETE FROM taxi_trip_employees WHERE trip_id IN (SELECT id FROM taxi_trips WHERE shift_id=?)',
            (shift_id,)
        )
        conn.execute('DELETE FROM taxi_trips WHERE shift_id=?', (shift_id,))
        conn.execute('DELETE FROM shift_revenue WHERE shift_id=?', (shift_id,))
        conn.execute('DELETE FROM change_log WHERE shift_id=?', (shift_id,))
        conn.execute('DELETE FROM shifts WHERE id=?', (shift_id,))
        conn.commit()
    flash(f'Смена #{shift_id} удалена', 'success')
    return redirect(url_for('shifts_archive'))


# ─── EMPLOYEES ────────────────────────────────────────────────────────────────

@app.route('/employees')
@login_required
def employees():
    role = session.get('role')
    if not item_visible('employees'):
        flash('Доступ запрещён', 'danger')
        return redirect(url_for('dashboard'))
    _req_branches = [bid for bid in request.args.getlist('branch_ids') if bid.isdigit()]
    selected_branches = _req_branches if (role == 'owner' or can_pick_other_branches('employees')) else []
    search = request.args.get('search', '').strip()
    role_filter = request.args.get('role_filter', '').strip()
    with get_db() as conn:
        def _emp_where(extra=''):
            parts = [extra] if extra else []
            if search:
                parts.append(
                    "(LOWER(e.last_name) LIKE LOWER(?) OR LOWER(e.first_name) LIKE LOWER(?) OR LOWER(e.full_name) LIKE LOWER(?))"
                )
            if role_filter:
                parts.append(
                    "(e.role = ? OR EXISTS ("
                    "SELECT 1 FROM employee_roles er "
                    "WHERE er.employee_id=e.id AND er.role=? AND er.is_active=1))"
                )
            return (' AND ' + ' AND '.join(parts)) if parts else ''

        def _emp_params(base_params):
            p = list(base_params)
            if search:
                s = f'%{search}%'
                p += [s, s, s]
            if role_filter:
                p += [role_filter, role_filter]
            return p

        if role == 'owner':
            all_branches = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
            if selected_branches:
                ids_str = ','.join(str(int(b)) for b in selected_branches)
                branch_cond = f'e.branch_id IN ({ids_str})'
                emps = conn.execute(
                    f'SELECT e.*, b.name as branch_name FROM employees e LEFT JOIN branches b ON b.id=e.branch_id '
                    f'WHERE {branch_cond} AND COALESCE(e.is_fired,0)=0{_emp_where()} ORDER BY e.last_name, e.first_name, e.full_name',
                    _emp_params([])
                ).fetchall()
                fired_emps = conn.execute(
                    f'SELECT e.*, b.name as branch_name FROM employees e LEFT JOIN branches b ON b.id=e.branch_id '
                    f'WHERE {branch_cond} AND e.is_fired=1 ORDER BY e.fired_at DESC, e.full_name',
                    []
                ).fetchall()
                branches = [b for b in all_branches if str(b['id']) in selected_branches]
            else:
                emps = conn.execute(
                    f'SELECT e.*, b.name as branch_name FROM employees e LEFT JOIN branches b ON b.id=e.branch_id '
                    f'WHERE COALESCE(e.is_fired,0)=0{_emp_where()} ORDER BY e.last_name, e.first_name, e.full_name',
                    _emp_params([])
                ).fetchall()
                fired_emps = conn.execute(
                    'SELECT e.*, b.name as branch_name FROM employees e LEFT JOIN branches b ON b.id=e.branch_id '
                    'WHERE e.is_fired=1 ORDER BY e.fired_at DESC, e.full_name'
                ).fetchall()
                branches = all_branches
        else:
            fired_emps = []
            all_branches = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall() \
                if can_pick_other_branches('employees') else []
            bids = [int(b) for b in get_effective_branch_ids('employees', _req_branches) or []]
            if bids:
                ids_str = ','.join(str(int(b)) for b in bids)
                emps = conn.execute(f'''
                    SELECT e.*, b.name as branch_name FROM employees e
                    LEFT JOIN branches b ON b.id=e.branch_id
                    WHERE e.branch_id IN ({ids_str}) AND COALESCE(e.is_fired,0)=0
                    ORDER BY b.name, e.role, e.full_name
                ''').fetchall()
                branches = conn.execute(f'SELECT * FROM branches WHERE id IN ({ids_str}) ORDER BY name').fetchall()
            else:
                emps = []
                branches = []

        # Текущий адрес развоза на каждого сотрудника (полная история — на странице «История изменений»)
        address_history = {}
        all_emps_for_hist = list(emps) + list(fired_emps)
        for emp in all_emps_for_hist:
            addr_hist = conn.execute('''
                SELECT * FROM employee_address_history WHERE employee_id=?
                ORDER BY valid_from DESC LIMIT 1
            ''', (emp['id'],)).fetchall()
            address_history[emp['id']] = addr_hist

        emp_branches_map = {}
        for row in conn.execute('''
            SELECT eb.employee_id, eb.branch_id
            FROM employee_branches eb
        ''').fetchall():
            emp_branches_map.setdefault(row['employee_id'], []).append(row['branch_id'])

        # Extra roles per employee
        emp_roles_map = {}
        for row in conn.execute('SELECT * FROM employee_roles WHERE is_active=1 ORDER BY role').fetchall():
            emp_roles_map.setdefault(row['employee_id'], []).append(dict(row))

        # Филиалы, на которых действует ежемесячная оплата — по (сотрудник, должность)
        pm_scope_map = {}
        for row in conn.execute('SELECT employee_id, role, branch_id FROM employee_pay_monthly_branches').fetchall():
            pm_scope_map.setdefault((row['employee_id'], row['role']), []).append(row['branch_id'])
        pm_branches_map = {}       # employee_id -> [branch_id] (для основной должности)
        for emp in emps:
            key = (emp['id'], emp['role'])
            if key in pm_scope_map:
                pm_branches_map[emp['id']] = pm_scope_map[key]
        pm_role_branches_map = {}  # employee_roles.id -> [branch_id] (для доп. должностей)
        for er_list in emp_roles_map.values():
            for er in er_list:
                key = (er['employee_id'], er['role'])
                if key in pm_scope_map:
                    pm_role_branches_map[er['id']] = pm_scope_map[key]

        # Position abbreviations: code -> abbr
        pos_abbr_map = {
            r['code']: (r['abbr'] or r['name'][:4]).upper()
            for r in conn.execute('SELECT code, name, abbr FROM positions').fetchall()
        }

        shift_counts = {}
        all_emp_ids = [e['id'] for e in all_emps_for_hist]
        if all_emp_ids:
            ids_str = ','.join(str(i) for i in all_emp_ids)
            for row in conn.execute(f'''
                SELECT employee_id, COUNT(*) as cnt FROM employee_shifts
                WHERE employee_id IN ({ids_str}) GROUP BY employee_id
            ''').fetchall():
                shift_counts[row['employee_id']] = row['cnt']

        all_tmpls = conn.execute(
            'SELECT * FROM rate_templates WHERE is_active=1 ORDER BY role, name'
        ).fetchall()
        tmpl_branch_sets = {}
        for row in conn.execute('SELECT template_id, branch_id FROM rate_template_branches').fetchall():
            tmpl_branch_sets.setdefault(row['template_id'], set()).add(row['branch_id'])

    def tmpls_for_emp(emp):
        bid = emp['branch_id']
        result = []
        for t in all_tmpls:
            branches_set = tmpl_branch_sets.get(t['id'], set())
            if not branches_set or bid in branches_set:
                result.append(t)
        return result

        # Build branch -> groups map for display
    bgroups = get_branch_groups(conn)
    branch_to_groups = {}  # branch_id -> [group_name, ...]
    for g in bgroups:
        for b in g.get('branches', []):
            branch_to_groups.setdefault(b['id'], []).append(g['name'])

    # Group templates by role for JS/template use
    tmpls_by_role = {}
    for t in all_tmpls:
        tmpls_by_role.setdefault(t['role'], []).append(dict(t))

    return render_template('employees.html', employees=emps, fired_employees=fired_emps,
                           branches=branches,
                           all_branches=all_branches,
                           selected_branches=selected_branches,
                           emp_branches_map=emp_branches_map,
                           emp_roles_map=emp_roles_map,
                           pm_branches_map=pm_branches_map,
                           pm_role_branches_map=pm_role_branches_map,
                           pos_abbr_map=pos_abbr_map,
                           role_labels=ROLE_LABELS, is_owner=(role == 'owner'),
                           address_history=address_history,
                           rate_templates=all_tmpls,
                           rate_templates_by_role=tmpls_by_role,
                           tmpl_branch_sets=tmpl_branch_sets,
                           tmpls_for_emp=tmpls_for_emp,
                           shift_counts=shift_counts,
                           today=date.today().isoformat(),
                           branch_groups=bgroups,
                           branch_to_groups=branch_to_groups,
                           search=search,
                           role_filter=role_filter,
                           positions=conn.execute('SELECT * FROM positions WHERE is_active=1 ORDER BY sort_order, name').fetchall())


@app.route('/employees/add', methods=['POST'])
@login_required
def add_employee():
    sess_role = session.get('role')
    if sess_role == 'owner':
        branch_ids_form = [bid for bid in request.form.getlist('branch_ids') if bid.isdigit()]
        branch_id = int(branch_ids_form[0]) if branch_ids_form else None
    else:
        branch_ids_form = [str(b) for b in _session_branch_ids()]
        branch_id = session.get('branch_id')
    last_name  = request.form.get('last_name', '').strip()
    first_name = request.form.get('first_name', '').strip()
    full_name  = (last_name + (' ' + first_name if first_name else '')).strip()
    emp_role = request.form.get('role', 'sushi')
    rate_template_id = request.form.get('rate_template_id', '').strip() or None
    if rate_template_id:
        rate_template_id = int(rate_template_id)
    effective_from = request.form.get('effective_from') or date.today().isoformat()
    pay_monthly = 1 if request.form.get('pay_monthly') else 0
    phone = _format_phone(request.form.get('phone', ''))
    if not last_name:
        flash('Введите фамилию сотрудника', 'danger')
        return redirect(url_for('employees'))
    if not first_name:
        flash('Введите имя сотрудника', 'danger')
        return redirect(url_for('employees'))
    with get_db() as conn:
        if rate_template_id:
            tmpl = conn.execute('SELECT * FROM rate_templates WHERE id=?', (rate_template_id,)).fetchone()
            rate = float(tmpl['rate'] or 0) if tmpl else 0.0
            rate_km = float(tmpl['rate_per_km'] or 0) if tmpl else 0.0
            rate_ord = float(tmpl['rate_per_order'] or 0) if tmpl else 0.0
        else:
            rate = float(request.form.get('rate', 0) or 0)
            rate_km = float(request.form.get('rate_per_km', 0) or 0)
            rate_ord = float(request.form.get('rate_per_order', 0) or 0)
        conn.execute(
            'INSERT INTO employees (branch_id, full_name, last_name, first_name, role, rate, rate_per_km, rate_per_order, pay_monthly, rate_template_id, phone) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
            (branch_id, full_name, last_name, first_name, emp_role, rate, rate_km, rate_ord, pay_monthly, rate_template_id, phone)
        )
        emp_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        conn.execute(
            'INSERT INTO employee_rate_history (employee_id, rate, rate_per_km, rate_per_order, effective_from) VALUES (?,?,?,?,?)',
            (emp_id, rate, rate_km, rate_ord, effective_from)
        )
        _seed_pay_monthly_history(conn, emp_id, emp_role, pay_monthly)
        for bid in branch_ids_form:
            conn.execute('INSERT OR IGNORE INTO employee_branches (employee_id, branch_id) VALUES (?,?)', (emp_id, int(bid)))
        address = request.form.get('address', '').strip()
        if address:
            conn.execute(
                'INSERT INTO employee_address_history (employee_id, address, valid_from) VALUES (?,?,?)',
                (emp_id, address, effective_from)
            )
        conn.commit()
    flash(f'Сотрудник {full_name} добавлен', 'success')
    return redirect(url_for('employees'))


@app.route('/employees/<int:emp_id>/update-rate', methods=['POST'])
@login_required
def update_employee_rate(emp_id):
    rate = float(request.form.get('rate', 0) or 0)
    rate_km = float(request.form.get('rate_per_km', 10) or 10)
    rate_ord = float(request.form.get('rate_per_order', 100) or 100)
    effective_from = request.form.get('effective_from') or date.today().isoformat()
    with get_db() as conn:
        emp = conn.execute('SELECT * FROM employees WHERE id=?', (emp_id,)).fetchone()
        if not emp:
            flash('Сотрудник не найден', 'danger')
            return redirect(url_for('employees'))
        conn.execute(
            'INSERT INTO employee_rate_history (employee_id, rate, rate_per_km, rate_per_order, effective_from) VALUES (?,?,?,?,?)',
            (emp_id, rate, rate_km, rate_ord, effective_from)
        )
        # Update current rate only if effective_from <= today
        if effective_from <= date.today().isoformat():
            conn.execute(
                'UPDATE employees SET rate=?, rate_per_km=?, rate_per_order=? WHERE id=?',
                (rate, rate_km, rate_ord, emp_id)
            )
        conn.commit()
    flash(f'Ставка сохранена (с {effective_from})', 'success')
    return redirect(url_for('employees'))


@app.route('/employees/<int:emp_id>/toggle', methods=['POST'])
@login_required
def toggle_employee(emp_id):
    with get_db() as conn:
        conn.execute('UPDATE employees SET is_active = 1-is_active WHERE id=?', (emp_id,))
        conn.commit()
    return redirect(url_for('employees'))


@app.route('/employees/<int:emp_id>/fire', methods=['POST'])
@login_required
@menu_permission_required('employees')
def fire_employee(emp_id):
    comment = request.form.get('comment', '').strip()
    with get_db() as conn:
        conn.execute(
            'UPDATE employees SET is_fired=1, fired_at=CURRENT_TIMESTAMP, fired_comment=?, is_active=0 WHERE id=?',
            (comment, emp_id)
        )
        conn.commit()
    flash('Сотрудник уволен', 'warning')
    return redirect(url_for('employees'))


@app.route('/employees/<int:emp_id>/restore', methods=['POST'])
@login_required
@menu_permission_required('employees')
def restore_employee(emp_id):
    with get_db() as conn:
        conn.execute(
            'UPDATE employees SET is_fired=0, fired_at=NULL, fired_comment=NULL, is_active=1 WHERE id=?',
            (emp_id,)
        )
        conn.commit()
    flash('Сотрудник восстановлен', 'success')
    return redirect(url_for('employees'))


@app.route('/employees/<int:emp_id>/address', methods=['POST'])
@login_required
def update_employee_address(emp_id):
    address = request.form.get('address', '').strip()
    valid_from = request.form.get('valid_from') or date.today().isoformat()
    if not address:
        flash('Введите адрес', 'danger')
        return redirect(url_for('employees'))
    with get_db() as conn:
        conn.execute(
            'INSERT INTO employee_address_history (employee_id, address, valid_from) VALUES (?,?,?)',
            (emp_id, address, valid_from)
        )
        conn.commit()
    flash('Адрес сохранён', 'success')
    return redirect(url_for('employees'))


@app.route('/employees/<int:emp_id>/edit', methods=['POST'])
@login_required
def edit_employee(emp_id):
    last_name  = request.form.get('last_name', '').strip()
    first_name = request.form.get('first_name', '').strip()
    full_name  = (last_name + (' ' + first_name if first_name else '')).strip()
    address    = request.form.get('address', '').strip()
    address_from = request.form.get('address_from') or date.today().isoformat()
    rate_from = request.form.get('rate_from') or date.today().isoformat()
    pay_monthly = 1 if request.form.get('pay_monthly') else 0
    pay_monthly_from = request.form.get('pay_monthly_from') or date.today().isoformat()
    if pay_monthly_from < date.today().isoformat():
        pay_monthly_from = date.today().isoformat()  # нельзя назначить датой в прошлом
    phone = _format_phone(request.form.get('phone', ''))
    emp_role = request.form.get('role', '').strip() or None
    rate_template_id = request.form.get('rate_template_id', '').strip() or None
    if rate_template_id:
        rate_template_id = int(rate_template_id)
    if not last_name:
        flash('Введите фамилию сотрудника', 'danger')
        return redirect(url_for('employees'))
    if not first_name:
        flash('Введите имя сотрудника', 'danger')
        return redirect(url_for('employees'))
    with get_db() as conn:
        emp = conn.execute('SELECT * FROM employees WHERE id=?', (emp_id,)).fetchone()
        if not emp:
            flash('Сотрудник не найден', 'danger')
            return redirect(url_for('employees'))
        if rate_template_id:
            tmpl = conn.execute('SELECT * FROM rate_templates WHERE id=?', (rate_template_id,)).fetchone()
            rate = float(tmpl['rate'] or 0) if tmpl else 0.0
            rate_km = float(tmpl['rate_per_km'] or 0) if tmpl else 0.0
            rate_ord = float(tmpl['rate_per_order'] or 0) if tmpl else 0.0
        else:
            rate = float(request.form.get('rate', 0) or 0)
            rate_km = float(request.form.get('rate_per_km', 0) or 0)
            rate_ord = float(request.form.get('rate_per_order', 0) or 0)
        final_role = emp_role if (emp_role and emp_role in ROLE_LABELS) else emp['role']
        update_fields = ('rate=?, rate_per_km=?, rate_per_order=?, rate_template_id=?, phone=?, '
                          'full_name=?, last_name=?, first_name=?, pay_monthly=?')
        update_vals = [rate, rate_km, rate_ord, rate_template_id, phone,
                       full_name, last_name, first_name, pay_monthly]
        if emp_role and emp_role in ROLE_LABELS:
            update_fields += ', role=?'
            update_vals += [emp_role]
        update_vals.append(emp_id)
        conn.execute(f'UPDATE employees SET {update_fields} WHERE id=?', update_vals)
        # Меняем режим оплаты только с текущей даты — история хранит момент смены,
        # чтобы уже прошедшие смены не блокировались задним числом (см. pay_staff).
        _log_pay_monthly_change(conn, emp_id, final_role, emp['pay_monthly'], pay_monthly, pay_monthly_from)
        if session.get('role') == 'owner':
            branch_ids_form = [bid for bid in request.form.getlist('branch_ids') if bid.isdigit()]
            if branch_ids_form:
                conn.execute('UPDATE employees SET branch_id=? WHERE id=?', (int(branch_ids_form[0]), emp_id))
                conn.execute('DELETE FROM employee_branches WHERE employee_id=?', (emp_id,))
                for bid in branch_ids_form:
                    conn.execute('INSERT OR IGNORE INTO employee_branches (employee_id, branch_id) VALUES (?,?)', (emp_id, int(bid)))
            # Филиалы, на которых действует ежемесячная оплата основной должности
            pm_branch_ids = [bid for bid in request.form.getlist('pm_branch_ids') if bid.isdigit()]
            conn.execute('DELETE FROM employee_pay_monthly_branches WHERE employee_id=? AND role=?', (emp_id, final_role))
            for bid in pm_branch_ids:
                conn.execute(
                    'INSERT OR IGNORE INTO employee_pay_monthly_branches (employee_id, role, branch_id) VALUES (?,?,?)',
                    (emp_id, final_role, int(bid))
                )
        # В историю пишем только если значение реально изменилось — иначе любое
        # сохранение карточки (например, смена телефона) плодило бы дубли записей.
        rate_changed = (
            rate != float(emp['rate'] or 0)
            or rate_km != float(emp['rate_per_km'] or 0)
            or rate_ord != float(emp['rate_per_order'] or 0)
        )
        if rate_changed:
            conn.execute(
                'INSERT INTO employee_rate_history (employee_id, rate, rate_per_km, rate_per_order, effective_from) VALUES (?,?,?,?,?)',
                (emp_id, rate, rate_km, rate_ord, rate_from)
            )
        if address:
            last_addr = conn.execute(
                'SELECT address FROM employee_address_history WHERE employee_id=? ORDER BY valid_from DESC, id DESC LIMIT 1',
                (emp_id,)
            ).fetchone()
            if not last_addr or last_addr['address'] != address:
                conn.execute(
                    'INSERT INTO employee_address_history (employee_id, address, valid_from) VALUES (?,?,?)',
                    (emp_id, address, address_from)
                )
        conn.commit()
    flash('Данные сотрудника сохранены', 'success')
    return redirect(url_for('employees'))


@app.route('/employees/<int:emp_id>/delete', methods=['POST'])
@login_required
def delete_employee(emp_id):
    if session.get('role') != 'owner':
        flash('Только владелец может удалять сотрудников', 'danger')
        return redirect(url_for('employees'))
    with get_db() as conn:
        emp = conn.execute('SELECT * FROM employees WHERE id=?', (emp_id,)).fetchone()
        if not emp:
            flash('Сотрудник не найден', 'danger')
            return redirect(url_for('employees'))
        shift_count = conn.execute(
            'SELECT COUNT(*) FROM employee_shifts WHERE employee_id=?', (emp_id,)
        ).fetchone()[0]
        if shift_count > 0:
            flash(
                f'Нельзя удалить «{emp["full_name"]}» — есть {shift_count} записей в сменах. Деактивируйте сотрудника.',
                'danger'
            )
            return redirect(url_for('employees'))
        conn.execute('UPDATE taxi_trip_employees SET employee_id=NULL WHERE employee_id=?', (emp_id,))
        conn.execute('DELETE FROM employee_rate_history WHERE employee_id=?', (emp_id,))
        conn.execute('DELETE FROM employee_address_history WHERE employee_id=?', (emp_id,))
        conn.execute('DELETE FROM employee_branches WHERE employee_id=?', (emp_id,))
        conn.execute('DELETE FROM employees WHERE id=?', (emp_id,))
        conn.commit()
    flash(f'Сотрудник «{emp["full_name"]}» удалён', 'success')
    return redirect(url_for('employees'))


@app.route('/employees/<int:emp_id>/reassign-delete', methods=['POST'])
@login_required
def reassign_and_delete_employee(emp_id):
    if session.get('role') != 'owner':
        flash('Только владелец может выполнять это действие', 'danger')
        return redirect(url_for('employees'))
    target_id = request.form.get('target_id', type=int)
    if not target_id or target_id == emp_id:
        flash('Выберите другого сотрудника для переназначения', 'danger')
        return redirect(url_for('employees'))
    with get_db() as conn:
        emp = conn.execute('SELECT * FROM employees WHERE id=?', (emp_id,)).fetchone()
        target = conn.execute('SELECT * FROM employees WHERE id=?', (target_id,)).fetchone()
        if not emp or not target:
            flash('Сотрудник не найден', 'danger')
            return redirect(url_for('employees'))
        shift_count = conn.execute(
            'SELECT COUNT(*) FROM employee_shifts WHERE employee_id=?', (emp_id,)
        ).fetchone()[0]
        conn.execute(
            'UPDATE employee_shifts SET employee_id=?, full_name_snapshot=? WHERE employee_id=?',
            (target_id, target['full_name'], emp_id)
        )
        conn.execute('UPDATE salary_payments SET employee_id=? WHERE employee_id=?', (target_id, emp_id))
        conn.execute('UPDATE taxi_trip_employees SET employee_id=? WHERE employee_id=?', (target_id, emp_id))
        conn.execute('DELETE FROM employee_rate_history WHERE employee_id=?', (emp_id,))
        conn.execute('DELETE FROM employee_address_history WHERE employee_id=?', (emp_id,))
        conn.execute('DELETE FROM employee_branches WHERE employee_id=?', (emp_id,))
        conn.execute('DELETE FROM employees WHERE id=?', (emp_id,))
        conn.commit()
    flash(
        f'Переназначено {shift_count} смен: «{emp["full_name"]}» → «{target["full_name"]}». Дубликат удалён.',
        'success'
    )
    return redirect(url_for('employees'))


# ─── TAXI ─────────────────────────────────────────────────────────────────────

@app.route('/shift/<int:shift_id>/taxi/add', methods=['POST'])
@login_required
def add_taxi_trip(shift_id):
    if not _can_edit_shift(shift_id):
        return jsonify({'error': 'Нет доступа'}), 403
    data = request.json or {}
    amount = float(data.get('amount', 0) or 0)
    payment_type = data.get('payment_type', 'cash')
    in_gulyash = 1 if data.get('in_gulyash') else 0
    note = data.get('note', '') or ''
    emps = data.get('employees', [])
    with get_db() as conn:
        conn.execute(
            'INSERT INTO taxi_trips (shift_id, amount, payment_type, in_gulyash, note) VALUES (?,?,?,?,?)',
            (shift_id, amount, payment_type, in_gulyash, note)
        )
        trip_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        for emp in emps:
            conn.execute(
                'INSERT INTO taxi_trip_employees (trip_id, employee_id, name_snapshot, address_snapshot) VALUES (?,?,?,?)',
                (trip_id, emp.get('id') or None, emp.get('name', ''), emp.get('address', ''))
            )
        conn.commit()
    return jsonify({'ok': True, 'trip_id': trip_id})


@app.route('/taxi/<int:trip_id>/delete', methods=['POST'])
@login_required
def delete_taxi_trip(trip_id):
    with get_db() as conn:
        trip = conn.execute('SELECT * FROM taxi_trips WHERE id=?', (trip_id,)).fetchone()
        if not trip:
            return jsonify({'error': 'Not found'}), 404
        if not _can_edit_shift(trip['shift_id']):
            return jsonify({'error': 'Нет доступа'}), 403
        conn.execute('DELETE FROM taxi_trip_employees WHERE trip_id=?', (trip_id,))
        conn.execute('DELETE FROM taxi_trips WHERE id=?', (trip_id,))
        conn.commit()
    return jsonify({'ok': True})


@app.route('/taxi/<int:trip_id>/toggle-gulyash', methods=['POST'])
@login_required
def toggle_taxi_gulyash(trip_id):
    with get_db() as conn:
        trip = conn.execute('SELECT * FROM taxi_trips WHERE id=?', (trip_id,)).fetchone()
        if not trip:
            return jsonify({'error': 'Not found'}), 404
        if not _can_edit_shift(trip['shift_id']):
            return jsonify({'error': 'Нет доступа'}), 403
        new_val = 0 if trip['in_gulyash'] else 1
        conn.execute('UPDATE taxi_trips SET in_gulyash=? WHERE id=?', (new_val, trip_id))
        conn.commit()
    return jsonify({'ok': True, 'in_gulyash': new_val})


@app.route('/shift/<int:shift_id>/taxi/create', methods=['POST'])
@login_required
def create_taxi_trip(shift_id):
    if not _can_edit_shift(shift_id):
        return jsonify({'error': 'Нет доступа'}), 403
    with get_db() as conn:
        conn.execute('INSERT INTO taxi_trips (shift_id) VALUES (?)', (shift_id,))
        trip_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        conn.commit()
    return jsonify({'ok': True, 'trip_id': trip_id})


@app.route('/taxi/<int:trip_id>/employee/add', methods=['POST'])
@login_required
def add_taxi_employee(trip_id):
    with get_db() as conn:
        trip = conn.execute('SELECT * FROM taxi_trips WHERE id=?', (trip_id,)).fetchone()
        if not trip or not _can_edit_shift(trip['shift_id']):
            return jsonify({'error': 'Нет доступа'}), 403
        data = request.json or {}
        employee_id = data.get('employee_id') or None
        if employee_id:
            already = conn.execute('''
                SELECT COUNT(*) FROM taxi_trip_employees tte
                JOIN taxi_trips tt ON tt.id = tte.trip_id
                WHERE tt.shift_id = ? AND tte.employee_id = ?
            ''', (trip['shift_id'], employee_id)).fetchone()[0]
            if already:
                return jsonify({'error': 'Сотрудник уже добавлен в рейс этой смены', 'duplicate': True}), 400
        conn.execute(
            'INSERT INTO taxi_trip_employees (trip_id, employee_id, name_snapshot, address_snapshot) VALUES (?,?,?,?)',
            (trip_id, employee_id, data.get('name', ''), data.get('address', ''))
        )
        tte_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        conn.commit()
    return jsonify({'ok': True, 'tte_id': tte_id})


@app.route('/taxi/employee/<int:tte_id>/remove', methods=['POST'])
@login_required
def remove_taxi_employee(tte_id):
    with get_db() as conn:
        tte = conn.execute('SELECT * FROM taxi_trip_employees WHERE id=?', (tte_id,)).fetchone()
        if not tte:
            return jsonify({'error': 'Not found'}), 404
        trip = conn.execute('SELECT * FROM taxi_trips WHERE id=?', (tte['trip_id'],)).fetchone()
        if not _can_edit_shift(trip['shift_id']):
            return jsonify({'error': 'Нет доступа'}), 403
        conn.execute('DELETE FROM taxi_trip_employees WHERE id=?', (tte_id,))
        conn.commit()
    return jsonify({'ok': True})


@app.route('/taxi/<int:trip_id>/update', methods=['POST'])
@login_required
def update_taxi_trip(trip_id):
    with get_db() as conn:
        trip = conn.execute('SELECT * FROM taxi_trips WHERE id=?', (trip_id,)).fetchone()
        if not trip or not _can_edit_shift(trip['shift_id']):
            return jsonify({'error': 'Нет доступа'}), 403
        data = request.json or {}
        fields, vals = [], []
        if 'amount' in data:
            fields.append('amount=?'); vals.append(float(data['amount'] or 0))
        if 'payment_type' in data:
            fields.append('payment_type=?'); vals.append(data['payment_type'])
        if 'in_gulyash' in data:
            fields.append('in_gulyash=?'); vals.append(1 if data['in_gulyash'] else 0)
        if fields:
            vals.append(trip_id)
            conn.execute('UPDATE taxi_trips SET ' + ', '.join(fields) + ' WHERE id=?', vals)
            conn.commit()
    return jsonify({'ok': True})


@app.route('/taxi/employee/<int:tte_id>/update', methods=['POST'])
@login_required
def update_taxi_employee(tte_id):
    with get_db() as conn:
        tte = conn.execute('SELECT * FROM taxi_trip_employees WHERE id=?', (tte_id,)).fetchone()
        if not tte:
            return jsonify({'error': 'Not found'}), 404
        trip = conn.execute('SELECT * FROM taxi_trips WHERE id=?', (tte['trip_id'],)).fetchone()
        if not _can_edit_shift(trip['shift_id']):
            return jsonify({'error': 'Нет доступа'}), 403
        data = request.json or {}
        conn.execute('UPDATE taxi_trip_employees SET address_snapshot=? WHERE id=?',
                     (data.get('address', ''), tte_id))
        conn.commit()
    return jsonify({'ok': True})


# ─── BRANCHES ─────────────────────────────────────────────────────────────────

@app.route('/branches')
@login_required
@menu_permission_required('branches')
def branches():
    with get_db() as conn:
        blist = conn.execute('SELECT * FROM branches ORDER BY name').fetchall()
        groups = get_branch_groups(conn)
        cards_rows = conn.execute('SELECT * FROM branch_cards ORDER BY branch_id, id').fetchall()
        branches_raw = [r[0] for r in conn.execute(
            'SELECT DISTINCT branch_raw FROM orders_report ORDER BY branch_raw'
        ).fetchall()]
        branch_raw_map = get_branch_raw_map(conn)
    cards_by_branch = {}
    for c in cards_rows:
        cards_by_branch.setdefault(c['branch_id'], []).append(dict(c))
    return render_template('branches.html', branches=blist, my_ip=get_client_ip(),
                           branch_groups=groups, cards_by_branch=cards_by_branch,
                           branches_raw=branches_raw, branch_raw_map=branch_raw_map)


@app.route('/branches/groups/add', methods=['POST'])
@login_required
@menu_permission_required('branches')
def add_branch_group():
    name = request.form.get('name', '').strip()
    abbr = request.form.get('abbr', '').strip().upper()[:3]
    branch_ids = [b for b in request.form.getlist('branch_ids') if b.isdigit()]
    if not name:
        flash('Введите название группы', 'danger')
        return redirect(url_for('branches'))
    with get_db() as conn:
        sort_order = conn.execute('SELECT COUNT(*) FROM branch_groups').fetchone()[0] * 10
        conn.execute('INSERT INTO branch_groups (name, abbr, sort_order) VALUES (?,?,?)', (name, abbr, sort_order))
        group_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        for bid in branch_ids:
            conn.execute(
                'INSERT OR IGNORE INTO branch_group_members (group_id, branch_id) VALUES (?,?)',
                (group_id, int(bid))
            )
        conn.commit()
    flash(f'Группа «{name}» создана', 'success')
    return redirect(url_for('branches'))


@app.route('/branches/groups/<int:group_id>/edit', methods=['POST'])
@login_required
@menu_permission_required('branches')
def edit_branch_group(group_id):
    name = request.form.get('name', '').strip()
    abbr = request.form.get('abbr', '').strip().upper()[:3]
    branch_ids = [b for b in request.form.getlist('branch_ids') if b.isdigit()]
    if not name:
        flash('Введите название группы', 'danger')
        return redirect(url_for('branches'))
    with get_db() as conn:
        conn.execute('UPDATE branch_groups SET name=?, abbr=? WHERE id=?', (name, abbr, group_id))
        conn.execute('DELETE FROM branch_group_members WHERE group_id=?', (group_id,))
        for bid in branch_ids:
            conn.execute(
                'INSERT OR IGNORE INTO branch_group_members (group_id, branch_id) VALUES (?,?)',
                (group_id, int(bid))
            )
        conn.commit()
    flash(f'Группа «{name}» обновлена', 'success')
    return redirect(url_for('branches'))


@app.route('/branches/groups/<int:group_id>/delete', methods=['POST'])
@login_required
@menu_permission_required('branches')
def delete_branch_group(group_id):
    with get_db() as conn:
        g = conn.execute('SELECT name FROM branch_groups WHERE id=?', (group_id,)).fetchone()
        if g:
            conn.execute('DELETE FROM branch_groups WHERE id=?', (group_id,))
            conn.commit()
            flash(f'Группа «{g["name"]}» удалена', 'success')
    return redirect(url_for('branches'))


# ─── Сопоставление подразделений «Отчёта по заказам» (branch_raw) с филиалами из Настроек ─────
# Один общий справочник филиалов и групп (branches/branch_groups) — Ожидание/Промокоды/Отчёт по
# заказам используют branch_id, полученный через это сопоставление, а не текст branch_raw напрямую.

@app.route('/branches/mapping/save', methods=['POST'])
@login_required
@menu_permission_required('branches')
def save_branch_raw_mapping():
    raws = request.form.getlist('map_raw')
    branch_ids = request.form.getlist('map_branch_id')
    with get_db() as conn:
        valid_ids = {b['id'] for b in conn.execute('SELECT id FROM branches').fetchall()}
        for raw, bid in zip(raws, branch_ids):
            if not raw:
                continue
            if bid and bid.isdigit() and int(bid) in valid_ids:
                conn.execute(
                    'INSERT INTO branch_raw_map (branch_raw, branch_id) VALUES (?,?) '
                    'ON CONFLICT(branch_raw) DO UPDATE SET branch_id=excluded.branch_id',
                    (raw, int(bid))
                )
                conn.execute('UPDATE orders_report SET branch_id=? WHERE branch_raw=?', (int(bid), raw))
            else:
                conn.execute('DELETE FROM branch_raw_map WHERE branch_raw=?', (raw,))
                conn.execute('UPDATE orders_report SET branch_id=NULL WHERE branch_raw=?', (raw,))
        conn.commit()
    flash('Сопоставление подразделений сохранено', 'success')
    return redirect(url_for('branches'))


@app.route('/branches/add', methods=['POST'])
@login_required
@menu_permission_required('branches')
def add_branch():
    name = request.form.get('name', '').strip().upper()
    if not name:
        flash('Введите название филиала', 'danger')
        return redirect(url_for('branches'))
    allowed_ip = request.form.get('allowed_ip', '').strip() or None
    abbr = request.form.get('abbr', '').strip().upper()[:3]
    with get_db() as conn:
        conn.execute('INSERT INTO branches (name, allowed_ip, abbr) VALUES (?,?,?)', (name, allowed_ip, abbr))
        conn.commit()
    flash(f'Филиал {name} добавлен', 'success')
    return redirect(url_for('branches'))


@app.route('/branches/<int:branch_id>/edit', methods=['POST'])
@login_required
@menu_permission_required('branches')
def edit_branch(branch_id):
    allowed_ip = request.form.get('allowed_ip', '').strip() or None
    abbr = request.form.get('abbr', '').strip().upper()[:3]
    with get_db() as conn:
        conn.execute(
            'UPDATE branches SET allowed_ip=?, abbr=? WHERE id=?',
            (allowed_ip, abbr, branch_id)
        )
        conn.commit()
    flash('Настройки филиала сохранены', 'success')
    return redirect(url_for('branches'))


@app.route('/branches/<int:branch_id>/cards/add', methods=['POST'])
@login_required
@menu_permission_required('branches')
def add_branch_card(branch_id):
    card_number = request.form.get('card_number', '').strip().replace(' ', '')
    card_name   = request.form.get('card_name', '').strip()
    if not card_number:
        flash('Введите номер карты', 'danger')
        return redirect(url_for('branches'))
    with get_db() as conn:
        conn.execute(
            'INSERT OR IGNORE INTO branch_cards (branch_id, card_number, card_name) VALUES (?,?,?)',
            (branch_id, card_number, card_name)
        )
        conn.commit()
    flash(f'Карта {card_number} добавлена', 'success')
    return redirect(url_for('branches'))


@app.route('/branches/cards/<int:card_id>/delete', methods=['POST'])
@login_required
@menu_permission_required('branches')
def delete_branch_card(card_id):
    with get_db() as conn:
        conn.execute('DELETE FROM branch_cards WHERE id=?', (card_id,))
        conn.commit()
    flash('Карта удалена', 'success')
    return redirect(url_for('branches'))


# ─── USERS ────────────────────────────────────────────────────────────────────

@app.route('/users')
@login_required
@menu_permission_required('users')
def users():
    with get_db() as conn:
        ulist = conn.execute('''
            SELECT u.*, b.name as branch_name
            FROM users u LEFT JOIN branches b ON b.id=u.branch_id
            ORDER BY u.role, u.full_name
        ''').fetchall()
        branches = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
        branch_groups = get_branch_groups(conn)
        user_branches_map = {}
        for row in conn.execute('''
            SELECT ub.user_id, b.name, b.id
            FROM user_branches ub JOIN branches b ON b.id=ub.branch_id
            ORDER BY b.name
        ''').fetchall():
            user_branches_map.setdefault(row['user_id'], []).append({'id': row['id'], 'name': row['name']})
    user_groups_map = {}
    for uid, branch_list in user_branches_map.items():
        ub_ids = set(b['id'] for b in branch_list)
        grps = [g['abbr'] or g['name'] for g in branch_groups if set(g['branch_ids']) & ub_ids]
        user_groups_map[uid] = grps
    new_invite_url = None
    new_invite_token = request.args.get('new_invite', '')
    if new_invite_token:
        new_invite_url = url_for('accept_invite', token=new_invite_token, _external=True)
    return render_template('users.html', users=ulist, branches=branches,
                           branch_groups=branch_groups, user_branches_map=user_branches_map,
                           user_groups_map=user_groups_map, new_invite_url=new_invite_url)


@app.route('/users/<int:user_id>/branches', methods=['POST'])
@login_required
@menu_permission_required('users')
def edit_user_branches(user_id):
    branch_ids = [int(bid) for bid in request.form.getlist('branch_ids') if bid.isdigit()]
    with get_db() as conn:
        u = conn.execute('SELECT full_name FROM users WHERE id=?', (user_id,)).fetchone()
        if not u:
            flash('Пользователь не найден', 'danger')
            return redirect(url_for('users'))
        conn.execute('DELETE FROM user_branches WHERE user_id=?', (user_id,))
        primary = branch_ids[0] if branch_ids else None
        conn.execute('UPDATE users SET branch_id=? WHERE id=?', (primary, user_id))
        for bid in branch_ids:
            conn.execute('INSERT OR IGNORE INTO user_branches (user_id, branch_id) VALUES (?,?)', (user_id, bid))
        conn.commit()
    flash(f'Филиалы пользователя {u["full_name"]} обновлены', 'success')
    return redirect(url_for('users'))


@app.route('/users/add', methods=['POST'])
@login_required
@menu_permission_required('users')
def add_user():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    full_name = request.form.get('full_name', '').strip()
    role = request.form.get('role', 'admin')
    branch_ids = [bid for bid in request.form.getlist('branch_ids') if bid.isdigit()]
    primary_branch_id = int(branch_ids[0]) if branch_ids else None
    if not username or not password or not full_name:
        flash('Заполните все поля', 'danger')
        return redirect(url_for('users'))
    with get_db() as conn:
        existing = conn.execute('SELECT id FROM users WHERE username=?', (username,)).fetchone()
        if existing:
            flash('Логин уже занят', 'danger')
            return redirect(url_for('users'))
        email = request.form.get('email', '').strip().lower()
        conn.execute(
            'INSERT INTO users (username, password_hash, role, full_name, branch_id, email) VALUES (?,?,?,?,?,?)',
            (username, generate_password_hash(password, method='pbkdf2:sha256'), role, full_name, primary_branch_id, email)
        )
        user_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        for bid in branch_ids:
            conn.execute('INSERT OR IGNORE INTO user_branches (user_id, branch_id) VALUES (?,?)', (user_id, int(bid)))
        conn.commit()
    flash(f'Пользователь {full_name} создан', 'success')
    return redirect(url_for('users'))


@app.route('/users/<int:user_id>/reset-password', methods=['POST'])
@login_required
@menu_permission_required('users')
def reset_password(user_id):
    new_password = request.form.get('password', '')
    if not new_password:
        flash('Введите пароль', 'danger')
        return redirect(url_for('users'))
    with get_db() as conn:
        conn.execute(
            'UPDATE users SET password_hash=? WHERE id=?',
            (generate_password_hash(new_password, method='pbkdf2:sha256'), user_id)
        )
        conn.commit()
    flash('Пароль обновлён', 'success')
    return redirect(url_for('users'))


def _send_reset_email(to_email, reset_url):
    msg = MIMEMultipart('alternative')
    msg['Subject'] = 'Восстановление пароля — CRMPAPA'
    msg['From']    = SMTP_FROM
    msg['To']      = to_email
    text = f'Для сброса пароля перейдите по ссылке:\n{reset_url}\n\nСсылка действительна 2 часа.'
    html = f'''<div style="font-family:sans-serif;max-width:480px;margin:0 auto;">
        <h2 style="color:#c0392b;">CRMPAPA</h2>
        <p>Вы запросили сброс пароля. Нажмите кнопку ниже:</p>
        <a href="{reset_url}" style="display:inline-block;background:#c0392b;color:#fff;padding:12px 28px;
           border-radius:8px;text-decoration:none;font-weight:700;margin:12px 0;">Сбросить пароль</a>
        <p style="color:#888;font-size:12px;margin-top:16px;">Ссылка действительна 2 часа. Если вы не запрашивали сброс — просто проигнорируйте это письмо.</p>
    </div>'''
    msg.attach(MIMEText(text, 'plain', 'utf-8'))
    msg.attach(MIMEText(html, 'html', 'utf-8'))
    raw = msg.as_string()
    def _do_send():
        try:
            srv = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15)
            srv.starttls()
            srv.login(SMTP_USER, SMTP_PASSWORD)
            srv.sendmail(SMTP_USER, to_email, raw)
            srv.quit()
        except Exception as e:
            logging.error(f'Background email error: {e}')
    threading.Thread(target=_do_send, daemon=True).start()


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        with get_db() as conn:
            user = conn.execute(
                "SELECT * FROM users WHERE LOWER(COALESCE(email,''))=? AND email!=''", (email,)
            ).fetchone()
            if user:
                token = secrets.token_urlsafe(32)
                expires = (datetime.utcnow() + timedelta(hours=2)).strftime('%Y-%m-%d %H:%M:%S')
                conn.execute('DELETE FROM password_reset_tokens WHERE user_id=?', (user['id'],))
                conn.execute(
                    'INSERT INTO password_reset_tokens (user_id, token, expires_at) VALUES (?,?,?)',
                    (user['id'], token, expires)
                )
                conn.commit()
                reset_url = url_for('reset_password_token', token=token, _external=True)
                _send_reset_email(email, reset_url)
        flash('Если аккаунт с этой почтой существует — письмо отправлено. Проверьте входящие (и папку «Спам»).', 'info')
        return redirect(url_for('login'))
    return render_template('forgot_password.html')


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password_token(token):
    conn = get_db()
    try:
        row = conn.execute(
            'SELECT * FROM password_reset_tokens WHERE token=? AND used=0',
            (token,)
        ).fetchone()
        if not row:
            flash('Ссылка недействительна или уже была использована.', 'danger')
            return redirect(url_for('forgot_password'))
        expires_at = datetime.strptime(row['expires_at'], '%Y-%m-%d %H:%M:%S')
        if datetime.utcnow() > expires_at:
            flash('Ссылка устарела. Запросите новую.', 'danger')
            return redirect(url_for('forgot_password'))
        if request.method == 'POST':
            password = request.form.get('password', '').strip()
            if len(password) < 4:
                flash('Пароль должен быть не короче 4 символов', 'danger')
                return render_template('reset_password.html', token=token)
            conn.execute(
                'UPDATE users SET password_hash=? WHERE id=?',
                (generate_password_hash(password, method='pbkdf2:sha256'), row['user_id'])
            )
            conn.execute('UPDATE password_reset_tokens SET used=1 WHERE id=?', (row['id'],))
            conn.commit()
            flash('Пароль успешно изменён. Войдите с новым паролем.', 'success')
            return redirect(url_for('login'))
        return render_template('reset_password.html', token=token)
    finally:
        conn.close()


@app.route('/users/<int:user_id>/edit', methods=['POST'])
@login_required
@menu_permission_required('users')
def edit_user(user_id):
    full_name = request.form.get('full_name', '').strip()
    username = request.form.get('username', '').strip()
    role = request.form.get('role', 'admin')
    password = request.form.get('password', '').strip()
    branch_ids = [int(bid) for bid in request.form.getlist('branch_ids') if bid.isdigit()]
    if not full_name or not username:
        flash('Заполните все обязательные поля', 'danger')
        return redirect(url_for('users'))
    with get_db() as conn:
        u = conn.execute('SELECT * FROM users WHERE id=?', (user_id,)).fetchone()
        if not u:
            flash('Пользователь не найден', 'danger')
            return redirect(url_for('users'))
        existing = conn.execute('SELECT id FROM users WHERE username=? AND id!=?', (username, user_id)).fetchone()
        if existing:
            flash('Логин уже занят', 'danger')
            return redirect(url_for('users'))
        email = request.form.get('email', '').strip().lower()
        conn.execute('UPDATE users SET full_name=?, username=?, role=?, email=? WHERE id=?',
                     (full_name, username, role, email, user_id))
        if password:
            conn.execute('UPDATE users SET password_hash=? WHERE id=?',
                         (generate_password_hash(password, method='pbkdf2:sha256'), user_id))
        conn.execute('DELETE FROM user_branches WHERE user_id=?', (user_id,))
        primary = branch_ids[0] if branch_ids else None
        conn.execute('UPDATE users SET branch_id=? WHERE id=?', (primary, user_id))
        for bid in branch_ids:
            conn.execute('INSERT OR IGNORE INTO user_branches (user_id, branch_id) VALUES (?,?)', (user_id, bid))
        conn.commit()
    flash(f'Пользователь {full_name} обновлён', 'success')
    return redirect(url_for('users'))


# ─── INVITE LINKS ─────────────────────────────────────────────────────────────

@app.route('/users/invite/create', methods=['POST'])
@login_required
@menu_permission_required('users')
def create_invite():
    role = request.form.get('role', 'admin')
    branch_ids = [int(b) for b in request.form.getlist('branch_ids') if b.isdigit()]
    token = secrets.token_urlsafe(32)
    expires = (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as conn:
        conn.execute(
            'INSERT INTO invite_tokens (token, role, branch_ids, created_by, expires_at) VALUES (?,?,?,?,?)',
            (token, role, _json_lib.dumps(branch_ids), session['user_id'], expires)
        )
        conn.commit()
    return redirect(url_for('users', new_invite=token))


@app.route('/invite/<token>', methods=['GET', 'POST'])
def accept_invite(token):
    with get_db() as conn:
        inv = conn.execute(
            "SELECT * FROM invite_tokens WHERE token=? AND used=0 AND expires_at > datetime('now')",
            (token,)
        ).fetchone()
        if not inv:
            flash('Ссылка недействительна или устарела.', 'danger')
            return redirect(url_for('login'))
        if request.method == 'POST':
            full_name = request.form.get('full_name', '').strip()
            username  = request.form.get('username', '').strip()
            email     = request.form.get('email', '').strip().lower()
            password  = request.form.get('password', '').strip()
            if not full_name or not username or not password:
                flash('Заполните все обязательные поля', 'danger')
                return render_template('accept_invite.html', token=token, inv=inv)
            if len(password) < 4:
                flash('Пароль должен быть не короче 4 символов', 'danger')
                return render_template('accept_invite.html', token=token, inv=inv)
            existing = conn.execute('SELECT id FROM users WHERE username=?', (username,)).fetchone()
            if existing:
                flash('Этот логин уже занят, выберите другой', 'danger')
                return render_template('accept_invite.html', token=token, inv=inv)
            branch_ids = _json_lib.loads(inv['branch_ids'])
            primary = branch_ids[0] if branch_ids else None
            conn.execute(
                'INSERT INTO users (username, password_hash, role, full_name, branch_id, email) VALUES (?,?,?,?,?,?)',
                (username, generate_password_hash(password, method='pbkdf2:sha256'),
                 inv['role'], full_name, primary, email)
            )
            new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            for bid in branch_ids:
                conn.execute('INSERT OR IGNORE INTO user_branches (user_id, branch_id) VALUES (?,?)', (new_id, bid))
            conn.execute('UPDATE invite_tokens SET used=1, used_by=? WHERE id=?', (new_id, inv['id']))
            conn.commit()
            flash('Аккаунт создан! Войдите с вашим логином и паролем.', 'success')
            return redirect(url_for('login'))
    return render_template('accept_invite.html', token=token, inv=inv)


# ─── SETTINGS ─────────────────────────────────────────────────────────────────

_SETTINGS_TAB_PERM = {
    'categories': 'settings_categories',
    'rates': 'settings_rates',
    'api': 'settings_api',
    'bonuses': 'settings_bonuses',
}


@app.route('/settings')
@login_required
def settings():
    tab = request.args.get('tab', 'categories')
    code = _SETTINGS_TAB_PERM.get(tab)
    allowed = item_visible(code) if code else (session.get('role') == 'owner')
    if not allowed:
        flash('Доступ запрещён', 'danger')
        return redirect(url_for('dashboard'))
    with get_db() as conn:
        exp_cats = conn.execute(
            'SELECT * FROM expense_categories ORDER BY COALESCE(parent_id,id), sort_order, label'
        ).fetchall()
        exp_cats_parents = [dict(c) for c in exp_cats if not c['parent_id']]
        cat_branches = {}
        for row in conn.execute('SELECT category_id, branch_id FROM expense_category_branches').fetchall():
            cat_branches.setdefault(row['category_id'], []).append(row['branch_id'])
        kpi_blocks = conn.execute(
            'SELECT * FROM kpi_blocks ORDER BY sort_order, id'
        ).fetchall()
        bonus_rules = conn.execute(
            'SELECT br.*, b.name AS branch_name FROM bonus_rules br '
            'LEFT JOIN branches b ON b.id=br.branch_id '
            'ORDER BY COALESCE(b.name, \'ААА\'), br.role, br.threshold_pct'
        ).fetchall()
        branches = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
        rate_templates = conn.execute(
            'SELECT * FROM rate_templates ORDER BY role, name'
        ).fetchall()
        tmpl_branches = {}
        for row in conn.execute('SELECT template_id, branch_id FROM rate_template_branches').fetchall():
            tmpl_branches.setdefault(row['template_id'], []).append(row['branch_id'])
        tmpl_history = {}
        for row in conn.execute(
            'SELECT * FROM rate_template_history ORDER BY template_id, valid_from DESC'
        ).fetchall():
            tmpl_history.setdefault(row['template_id'], []).append(row)
        positions = conn.execute(
            'SELECT * FROM positions ORDER BY sort_order, name'
        ).fetchall()
        branch_groups_for_rates = get_branch_groups(conn)
        api_tokens = conn.execute('''
            SELECT t.*, b.name AS branch_name
            FROM api_1c_tokens t JOIN branches b ON b.id=t.branch_id
            ORDER BY b.name
        ''').fetchall()
        api_log = conn.execute(
            'SELECT l.*, b.name AS branch_name FROM api_1c_log l '
            'LEFT JOIN branches b ON b.id=l.branch_id '
            'ORDER BY l.created_at DESC LIMIT 50'
        ).fetchall()
        api_revenue_tokens = conn.execute('''
            SELECT t.*, b.name AS branch_name
            FROM api_revenue_tokens t JOIN branches b ON b.id=t.branch_id
            ORDER BY b.name
        ''').fetchall()
        api_revenue_log = conn.execute(
            'SELECT l.*, b.name AS branch_name FROM api_revenue_log l '
            'LEFT JOIN branches b ON b.id=l.branch_id '
            'ORDER BY l.created_at DESC LIMIT 50'
        ).fetchall()
        api_waittime_tokens = conn.execute('''
            SELECT t.*, b.name AS branch_name
            FROM api_waittime_tokens t JOIN branches b ON b.id=t.branch_id
            ORDER BY b.name
        ''').fetchall()
        api_waittime_log = conn.execute(
            'SELECT l.*, b.name AS branch_name FROM api_waittime_log l '
            'LEFT JOIN branches b ON b.id=l.branch_id '
            'ORDER BY l.created_at DESC LIMIT 50'
        ).fetchall()
        role_perms = {r: get_role_permissions(conn, r) for r in ROLE_CONFIGURABLE}
    return render_template('settings.html',
        exp_cats=exp_cats, exp_cats_parents=exp_cats_parents,
        cat_branches=cat_branches,
        kpi_blocks=kpi_blocks,
        bonus_rules=bonus_rules, branches=branches,
        rate_templates=rate_templates,
        tmpl_branches=tmpl_branches, tmpl_history=tmpl_history,
        api_tokens=api_tokens, api_log=api_log,
        api_revenue_tokens=api_revenue_tokens, api_revenue_log=api_revenue_log,
        api_waittime_tokens=api_waittime_tokens, api_waittime_log=api_waittime_log,
        base_url=request.host_url.rstrip('/'),
        today=date.today().isoformat(),
        formula_vars=FORMULA_VARS, role_labels=ROLE_LABELS,
        positions=positions, branch_groups_for_rates=branch_groups_for_rates,
        role_perms=role_perms, menu_items=MENU_ITEMS, menu_group_labels=MENU_GROUP_LABELS,
        menu_subitems=MENU_SUBITEMS,
        role_configurable=ROLE_CONFIGURABLE, login_role_labels=LOGIN_ROLE_LABELS)


@app.route('/settings/role-permissions/save', methods=['POST'])
@login_required
@owner_required
def role_permissions_save():
    all_codes = [m[0] for m in MENU_ITEMS] + MENU_SUBITEMS_FLAT
    with get_db() as conn:
        for role in ROLE_CONFIGURABLE:
            for code in all_codes:
                visible = 1 if request.form.get(f'visible_{role}_{code}') else 0
                scope = request.form.get(f'scope_{role}_{code}', 'own_only')
                if scope not in ('own_only', 'own_default'):
                    scope = 'own_only'
                conn.execute(
                    'INSERT OR REPLACE INTO role_menu_permissions (role, item_code, visible, branch_scope) '
                    'VALUES (?,?,?,?)',
                    (role, code, visible, scope)
                )
        conn.commit()
    flash('Права доступа сохранены', 'success')
    return redirect(url_for('settings') + '?tab=roles')


@app.route('/settings/expense-cat/add', methods=['POST'])
@login_required
@menu_permission_required('settings_categories')
def add_expense_cat():
    label = request.form.get('label', '').strip()
    cat_type = request.form.get('type', 'expense')
    parent_id = request.form.get('parent_id') or None
    show_contractors = 1 if request.form.get('show_contractors') else 0
    show_shift = 1 if request.form.get('show_shift') else 0
    if cat_type not in ('expense', 'income'):
        cat_type = 'expense'
    if not label:
        flash('Введите название', 'danger')
        return redirect(url_for('settings'))
    code = _slugify(label)
    with get_db() as conn:
        if parent_id:
            parent_row = conn.execute('SELECT id, type FROM expense_categories WHERE id=?', (parent_id,)).fetchone()
            if not parent_row:
                flash('Группа не найдена', 'danger')
                return redirect(url_for('settings'))
            cat_type = parent_row['type']
            parent_id = parent_row['id']
        existing = conn.execute('SELECT id FROM expense_categories WHERE code=?', (code,)).fetchone()
        if existing:
            code = code + '_' + str(int(datetime.now().timestamp()))[-4:]
        max_sort = conn.execute('SELECT COALESCE(MAX(sort_order),0) FROM expense_categories').fetchone()[0]
        conn.execute(
            'INSERT INTO expense_categories (code, label, type, parent_id, sort_order, show_contractors, show_shift) '
            'VALUES (?,?,?,?,?,?,?)',
            (code, label, cat_type, parent_id, max_sort + 1, show_contractors, show_shift)
        )
        conn.commit()
    kind = 'Подкатегория' if parent_id else 'Категория'
    flash(f'{kind} «{label}» добавлена', 'success')
    return redirect(url_for('settings'))


@app.route('/settings/expense-cat/<int:cat_id>/toggle', methods=['POST'])
@login_required
@menu_permission_required('settings_categories')
def toggle_expense_cat(cat_id):
    with get_db() as conn:
        conn.execute('UPDATE expense_categories SET is_active=1-is_active WHERE id=?', (cat_id,))
        # Also toggle all subcategories
        conn.execute(
            'UPDATE expense_categories SET is_active=(SELECT is_active FROM expense_categories WHERE id=?) WHERE parent_id=?',
            (cat_id, cat_id)
        )
        conn.commit()
    return redirect(url_for('settings'))


@app.route('/settings/expense-cat/<int:cat_id>/delete', methods=['POST'])
@login_required
@menu_permission_required('settings_categories')
def delete_expense_cat(cat_id):
    with get_db() as conn:
        # Move subcategories to top-level before deleting parent
        conn.execute('UPDATE expense_categories SET parent_id=NULL WHERE parent_id=?', (cat_id,))
        conn.execute('DELETE FROM expense_categories WHERE id=?', (cat_id,))
        conn.commit()
    flash('Категория удалена', 'success')
    return redirect(url_for('settings'))


@app.route('/settings/expense-cat/<int:cat_id>/edit', methods=['POST'])
@login_required
@menu_permission_required('settings_categories')
def edit_expense_cat(cat_id):
    label = request.form.get('label', '').strip()
    cat_type = request.form.get('type', '').strip()
    parent_id = request.form.get('parent_id') or None
    show_contractors = 1 if request.form.get('show_contractors') else 0
    show_shift = 1 if request.form.get('show_shift') else 0
    if not label:
        flash('Введите название', 'danger')
        return redirect(url_for('settings'))
    with get_db() as conn:
        cat = conn.execute('SELECT id, type FROM expense_categories WHERE id=?', (cat_id,)).fetchone()
        if not cat:
            flash('Категория не найдена', 'danger')
            return redirect(url_for('settings'))
        if cat_type not in ('expense', 'income'):
            cat_type = cat['type']
        has_children = conn.execute(
            'SELECT COUNT(*) FROM expense_categories WHERE parent_id=?', (cat_id,)
        ).fetchone()[0]
        if parent_id:
            if int(parent_id) == cat_id:
                flash('Категория не может быть группой сама для себя', 'danger')
                return redirect(url_for('settings'))
            if has_children:
                flash('У этой категории есть свои подкатегории — сначала перенесите их, группу нельзя вложить в другую группу', 'danger')
                return redirect(url_for('settings'))
            parent_row = conn.execute(
                'SELECT id FROM expense_categories WHERE id=? AND type=? AND parent_id IS NULL',
                (parent_id, cat_type)
            ).fetchone()
            if not parent_row:
                flash('Группа не найдена (проверьте, что тип группы совпадает с типом категории)', 'danger')
                return redirect(url_for('settings'))
            parent_id = parent_row['id']
        conn.execute(
            'UPDATE expense_categories SET label=?, type=?, parent_id=?, show_contractors=?, show_shift=? WHERE id=?',
            (label, cat_type, parent_id, show_contractors, show_shift, cat_id)
        )
        # Подкатегории всегда наследуют тип группы — если тип группы сменили, тянем их за собой.
        if has_children and cat_type != cat['type']:
            conn.execute('UPDATE expense_categories SET type=? WHERE parent_id=?', (cat_type, cat_id))
        conn.commit()
    flash('Категория обновлена' + (' (подкатегории тоже переключены на новый тип)' if has_children and cat_type != cat['type'] else ''), 'success')
    return redirect(url_for('settings'))


@app.route('/settings/expense-cat/<int:cat_id>/branches', methods=['POST'])
@login_required
@menu_permission_required('settings_categories')
def set_expense_cat_branches(cat_id):
    branch_ids = request.form.getlist('branch_ids')
    with get_db() as conn:
        conn.execute('DELETE FROM expense_category_branches WHERE category_id=?', (cat_id,))
        for bid in branch_ids:
            try:
                conn.execute(
                    'INSERT OR IGNORE INTO expense_category_branches (category_id, branch_id) VALUES (?,?)',
                    (cat_id, int(bid))
                )
            except (ValueError, TypeError):
                pass
        conn.commit()
    return redirect(url_for('settings'))


@app.route('/settings/kpi/add', methods=['POST'])
@login_required
@owner_required
def add_kpi_block():
    title = request.form.get('title', '').strip()
    formula = request.form.get('formula', '').strip()
    color = request.form.get('color', 'primary')
    unit = request.form.get('unit', '₽').strip()
    if not title or not formula:
        flash('Введите название и формулу', 'danger')
        return redirect(url_for('settings'))
    # Validate formula
    test_vars = {k: 1.0 for k in FORMULA_VARS}
    result = safe_eval(formula, test_vars)
    if result is None:
        flash('Ошибка в формуле — проверьте переменные и операторы', 'danger')
        return redirect(url_for('settings'))
    with get_db() as conn:
        max_sort = conn.execute('SELECT COALESCE(MAX(sort_order),0) FROM kpi_blocks').fetchone()[0]
        conn.execute(
            'INSERT INTO kpi_blocks (title, formula, color, unit, sort_order) VALUES (?,?,?,?,?)',
            (title, formula, color, unit, max_sort + 1)
        )
        conn.commit()
    flash(f'Блок «{title}» добавлен', 'success')
    return redirect(url_for('settings'))


@app.route('/settings/kpi/<int:block_id>/delete', methods=['POST'])
@login_required
@owner_required
def delete_kpi_block(block_id):
    with get_db() as conn:
        conn.execute('DELETE FROM kpi_blocks WHERE id=?', (block_id,))
        conn.commit()
    flash('Блок удалён', 'success')
    return redirect(url_for('settings'))


@app.route('/settings/kpi/<int:block_id>/toggle', methods=['POST'])
@login_required
@owner_required
def toggle_kpi_block(block_id):
    with get_db() as conn:
        conn.execute('UPDATE kpi_blocks SET is_active=1-is_active WHERE id=?', (block_id,))
        conn.commit()
    return redirect(url_for('settings'))


@app.route('/settings/bonus-rules/add', methods=['POST'])
@login_required
@menu_permission_required('settings_bonuses')
def add_bonus_rule():
    role = request.form.get('role', '')
    threshold = request.form.get('threshold_pct', '')
    bonus = request.form.get('bonus_pct', '')
    if not role or not threshold or not bonus:
        flash('Заполните все поля', 'danger')
        return redirect(url_for('settings') + '#tab-bonuses')
    try:
        threshold = float(threshold)
        bonus = float(bonus)
    except ValueError:
        flash('Неверный формат числа', 'danger')
        return redirect(url_for('settings') + '#tab-bonuses')
    branch_id = request.form.get('branch_id') or None
    if branch_id:
        try: branch_id = int(branch_id)
        except ValueError: branch_id = None
    with get_db() as conn:
        conn.execute(
            'INSERT INTO bonus_rules (role, threshold_pct, bonus_pct, branch_id) VALUES (?,?,?,?)',
            (role, threshold, bonus, branch_id)
        )
        conn.commit()
    flash(f'Правило добавлено', 'success')
    return redirect(url_for('settings') + '#tab-bonuses')


@app.route('/settings/bonus-rules/<int:rule_id>/delete', methods=['POST'])
@login_required
@menu_permission_required('settings_bonuses')
def delete_bonus_rule(rule_id):
    with get_db() as conn:
        conn.execute('DELETE FROM bonus_rules WHERE id=?', (rule_id,))
        conn.commit()
    flash('Правило удалено', 'success')
    return redirect(url_for('settings') + '#tab-bonuses')


@app.route('/settings/bonus-rules/<int:rule_id>/toggle', methods=['POST'])
@login_required
@menu_permission_required('settings_bonuses')
def toggle_bonus_rule(rule_id):
    with get_db() as conn:
        conn.execute('UPDATE bonus_rules SET is_active=1-is_active WHERE id=?', (rule_id,))
        conn.commit()
    return redirect(url_for('settings') + '?tab=bonuses')


@app.route('/bonus_rules/<int:rule_id>/edit', methods=['POST'])
@login_required
@menu_permission_required('settings_bonuses')
def edit_bonus_rule(rule_id):
    role = request.form.get('role')
    branch_id = request.form.get('branch_id') or None
    if branch_id:
        try:
            branch_id = int(branch_id)
        except ValueError:
            branch_id = None
    try:
        threshold_pct = float(request.form.get('threshold_pct', 0))
        bonus_pct = float(request.form.get('bonus_pct', 0))
    except ValueError:
        flash('Неверные значения', 'danger')
        return redirect(url_for('settings') + '?tab=bonuses')
    with get_db() as conn:
        conn.execute(
            'UPDATE bonus_rules SET role=?, threshold_pct=?, bonus_pct=?, branch_id=? WHERE id=?',
            (role, threshold_pct, bonus_pct, branch_id, rule_id)
        )
        conn.commit()
    flash('Правило обновлено', 'success')
    return redirect(url_for('settings') + '?tab=bonuses')


# ─── POSITIONS (ДОЛЖНОСТИ) ────────────────────────────────────────────────────

@app.route('/settings/positions/add', methods=['POST'])
@login_required
@menu_permission_required('settings_rates')
def add_position():
    name = request.form.get('name', '').strip()
    if not name:
        flash('Введите название должности', 'danger')
        return redirect(url_for('settings') + '?tab=rates')
    code = _slugify(name)
    with get_db() as conn:
        existing = conn.execute('SELECT id FROM positions WHERE code=?', (code,)).fetchone()
        if existing:
            code = code + '_' + str(int(datetime.now().timestamp()))[-4:]
        max_sort = conn.execute('SELECT COALESCE(MAX(sort_order),0) FROM positions').fetchone()[0]
        conn.execute(
            'INSERT INTO positions (code, name, sort_order) VALUES (?,?,?)',
            (code, name, max_sort + 10)
        )
        conn.commit()
        _reload_role_labels(conn)
    flash(f'Должность «{name}» добавлена', 'success')
    return redirect(url_for('settings') + '?tab=rates')


@app.route('/settings/positions/<int:pos_id>/edit', methods=['POST'])
@login_required
@menu_permission_required('settings_rates')
def edit_position(pos_id):
    name = request.form.get('name', '').strip()
    abbr = request.form.get('abbr', '').strip()[:4].upper()
    if not name:
        flash('Введите название', 'danger')
        return redirect(url_for('settings') + '?tab=rates')
    with get_db() as conn:
        conn.execute('UPDATE positions SET name=?, abbr=? WHERE id=?', (name, abbr, pos_id))
        conn.commit()
        _reload_role_labels(conn)
    flash('Должность обновлена', 'success')
    return redirect(url_for('settings') + '?tab=rates')


@app.route('/settings/positions/<int:pos_id>/delete', methods=['POST'])
@login_required
@menu_permission_required('settings_rates')
def delete_position(pos_id):
    with get_db() as conn:
        pos = conn.execute('SELECT code, name FROM positions WHERE id=?', (pos_id,)).fetchone()
        if not pos:
            flash('Должность не найдена', 'danger')
            return redirect(url_for('settings') + '?tab=rates')
        used = conn.execute(
            'SELECT COUNT(*) FROM employee_shifts WHERE role_snapshot=?', (pos['code'],)
        ).fetchone()[0]
        if used:
            flash(f'Нельзя удалить — должность «{pos["name"]}» используется в {used} сменах', 'danger')
            return redirect(url_for('settings') + '?tab=rates')
        conn.execute('DELETE FROM positions WHERE id=?', (pos_id,))
        conn.commit()
        _reload_role_labels(conn)
    flash('Должность удалена', 'success')
    return redirect(url_for('settings') + '?tab=rates')


# ─── RATE TEMPLATES ───────────────────────────────────────────────────────────

@app.route('/settings/rate-templates/add', methods=['POST'])
@login_required
@menu_permission_required('settings_rates')
def add_rate_template():
    role       = request.form.get('role', '').strip()
    name       = request.form.get('name', '').strip()
    rate       = float(request.form.get('rate', 0) or 0)
    rate_km    = float(request.form.get('rate_per_km', 10) or 10)
    rate_ord   = float(request.form.get('rate_per_order', 100) or 100)
    branch_ids = request.form.getlist('branch_ids')
    valid_from = request.form.get('valid_from') or date.today().isoformat()
    if not role or not name:
        flash('Заполните роль и название', 'danger')
        return redirect(url_for('settings') + '?tab=rates')
    with get_db() as conn:
        conn.execute(
            'INSERT INTO rate_templates (role, name, rate, rate_per_km, rate_per_order) VALUES (?,?,?,?,?)',
            (role, name, rate, rate_km, rate_ord)
        )
        tmpl_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        for bid in branch_ids:
            try:
                conn.execute('INSERT OR IGNORE INTO rate_template_branches (template_id, branch_id) VALUES (?,?)',
                             (tmpl_id, int(bid)))
            except (ValueError, Exception):
                pass
        conn.execute(
            'INSERT INTO rate_template_history (template_id, rate, rate_per_km, rate_per_order, valid_from) VALUES (?,?,?,?,?)',
            (tmpl_id, rate, rate_km, rate_ord, valid_from)
        )
        conn.commit()
    flash(f'Ставка «{name}» добавлена', 'success')
    return redirect(url_for('settings') + '?tab=rates')


@app.route('/settings/rate-templates/<int:tmpl_id>/delete', methods=['POST'])
@login_required
@menu_permission_required('settings_rates')
def delete_rate_template(tmpl_id):
    with get_db() as conn:
        conn.execute('DELETE FROM rate_templates WHERE id=?', (tmpl_id,))
        conn.commit()
    flash('Ставка удалена', 'success')
    return redirect(url_for('settings') + '?tab=rates')


@app.route('/settings/rate-templates/<int:tmpl_id>/edit', methods=['POST'])
@login_required
@menu_permission_required('settings_rates')
def edit_rate_template(tmpl_id):
    name       = request.form.get('name', '').strip()
    rate       = float(request.form.get('rate', 0) or 0)
    rate_km    = float(request.form.get('rate_per_km', 10) or 10)
    rate_ord   = float(request.form.get('rate_per_order', 100) or 100)
    branch_ids = request.form.getlist('branch_ids')
    valid_from = request.form.get('valid_from') or date.today().isoformat()
    with get_db() as conn:
        # Update current rates if valid_from <= today
        if valid_from <= date.today().isoformat():
            conn.execute(
                'UPDATE rate_templates SET name=?, rate=?, rate_per_km=?, rate_per_order=? WHERE id=?',
                (name, rate, rate_km, rate_ord, tmpl_id)
            )
        else:
            conn.execute('UPDATE rate_templates SET name=? WHERE id=?', (name, tmpl_id))
        # Save rate history entry
        conn.execute(
            'INSERT INTO rate_template_history (template_id, rate, rate_per_km, rate_per_order, valid_from) VALUES (?,?,?,?,?)',
            (tmpl_id, rate, rate_km, rate_ord, valid_from)
        )
        # Replace branch associations
        conn.execute('DELETE FROM rate_template_branches WHERE template_id=?', (tmpl_id,))
        for bid in branch_ids:
            try:
                conn.execute('INSERT OR IGNORE INTO rate_template_branches (template_id, branch_id) VALUES (?,?)',
                             (tmpl_id, int(bid)))
            except (ValueError, Exception):
                pass
        conn.commit()
    flash('Ставка обновлена', 'success')
    return redirect(url_for('settings') + '?tab=rates')


# ─── REPORTS ──────────────────────────────────────────────────────────────────

@app.route('/reports')
@login_required
def reports():
    active_tab = request.args.get('tab', 'shifts')
    if not item_visible('reports_salary' if active_tab == 'salary' else 'reports_shifts'):
        flash('Доступ запрещён', 'danger')
        return redirect(url_for('dashboard'))
    branch_ids = [bid for bid in request.args.getlist('branch_ids') if bid.isdigit()]
    branch_ids = get_effective_branch_ids('reports_shifts', branch_ids) or []

    _today = date.today()
    _month_start = _today.replace(day=1).isoformat()
    _month_end   = _today.replace(day=calendar.monthrange(_today.year, _today.month)[1]).isoformat()

    r_date_from = request.args.get('r_date_from', _month_start)
    r_date_to   = request.args.get('r_date_to',   _month_end)
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', r_date_from): r_date_from = _month_start
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', r_date_to):   r_date_to   = _month_end

    today = _today.isoformat()
    month_start = _month_start
    s_date_from = request.args.get('s_date_from', month_start)
    s_date_to   = request.args.get('s_date_to',   today)
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', s_date_from): s_date_from = month_start
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', s_date_to):   s_date_to   = today
    s_branch_ids = [bid for bid in request.args.getlist('s_branch_ids') if bid.isdigit()]
    s_role      = request.args.get('s_role', '')
    s_unpaid    = request.args.get('s_unpaid', '')

    with get_db() as conn:
        branches = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()

        date_filter = f"AND s.date BETWEEN '{r_date_from}' AND '{r_date_to}'"
        if branch_ids:
            ids_str = ','.join(str(int(bid)) for bid in branch_ids)
            branch_filter = f"AND s.branch_id IN ({ids_str})"
        else:
            branch_filter = ""

        shifts_data = conn.execute(f'''
            SELECT s.date, s.branch_id, b.name as branch_name, s.status,
                   COALESCE(r.total_revenue,0)    as revenue,
                   COALESCE(r.delivery_revenue,0) as delivery,
                   COALESCE(r.pickup_revenue,0)   as pickup,
                   COALESCE(r.delivery_orders,0)+COALESCE(r.pickup_orders,0) as orders,
                   COALESCE(r.cash_amount,0)       as cash,
                   COALESCE(r.card_amount,0)       as card,
                   COALESCE(r.online_amount,0)     as online,
                   COALESCE((SELECT SUM(es.total_amount) FROM employee_shifts es
                              WHERE es.shift_id = s.id), 0) as fot,
                   s.id as shift_id
            FROM shifts s
            JOIN branches b ON b.id=s.branch_id
            LEFT JOIN shift_revenue r ON r.shift_id=s.id
            WHERE 1=1 {date_filter} {branch_filter}
            ORDER BY s.date DESC, b.name
        ''').fetchall()
        totals = conn.execute(f'''
            SELECT COALESCE(SUM(r.total_revenue),0) as revenue,
                   COALESCE(SUM(r.delivery_revenue),0) as delivery,
                   COALESCE(SUM(r.pickup_revenue),0) as pickup,
                   COALESCE(SUM(r.delivery_orders+r.pickup_orders),0) as orders,
                   COALESCE(SUM(r.cash_amount),0) as cash,
                   COALESCE(SUM(r.card_amount),0) as card,
                   COALESCE((SELECT SUM(es.total_amount) FROM employee_shifts es
                              JOIN shifts s2 ON s2.id=es.shift_id
                              WHERE 1=1 {date_filter} {branch_filter}
                              ), 0) as fot
            FROM shifts s LEFT JOIN shift_revenue r ON r.shift_id=s.id
            WHERE 1=1 {date_filter} {branch_filter}
        ''').fetchone()

        # Group shifts by date for the grouped view
        from collections import OrderedDict as _OD
        _dg = _OD()
        for _row in shifts_data:
            _d = _row['date']
            if _d not in _dg:
                _dg[_d] = {'date': _d, 'revenue': 0.0, 'pickup': 0.0,
                            'fot': 0.0, 'branches': []}
            _dg[_d]['revenue'] += _row['revenue']
            _dg[_d]['pickup']  += _row['pickup']
            _dg[_d]['fot']     += _row['fot']
            _dg[_d]['branches'].append(dict(_row))
        day_groups = list(_dg.values())

        # ── Сводная таблица выручки: даты × филиалы ──────────────────────────
        rev_branch_list = [b for b in branches if not branch_ids or str(b['id']) in branch_ids]

        # Все даты в диапазоне (включая будущие, где будет только план)
        _rf = datetime.strptime(r_date_from, '%Y-%m-%d').date()
        _rt = datetime.strptime(r_date_to,   '%Y-%m-%d').date()
        rev_dates = []
        _d = _rf
        while _d <= _rt:
            rev_dates.append(_d.isoformat())
            _d += timedelta(days=1)
        _RU_DOW = ['пн','вт','ср','чт','пт','сб','вс']
        rev_date_dow = {d: _RU_DOW[datetime.strptime(d, '%Y-%m-%d').weekday()] for d in rev_dates}

        # Факт выручки по дням и филиалам (смены имеют приоритет, fallback на revenue_manual)
        _act_bf  = f"AND s.branch_id IN ({','.join(str(b['id']) for b in rev_branch_list)})" if rev_branch_list else ""
        _man_bf  = f"AND m.branch_id IN ({','.join(str(b['id']) for b in rev_branch_list)})" if rev_branch_list else ""
        act_rows = conn.execute(f'''
            SELECT date, branch_id, SUM(actual) as actual FROM (
                SELECT s.date, s.branch_id, COALESCE(SUM(r.total_revenue), 0) as actual
                FROM shifts s LEFT JOIN shift_revenue r ON r.shift_id = s.id
                WHERE s.date BETWEEN ? AND ? {_act_bf}
                GROUP BY s.date, s.branch_id
                UNION ALL
                SELECT m.date, m.branch_id, m.amount as actual
                FROM revenue_manual m
                WHERE m.date BETWEEN ? AND ? {_man_bf}
                AND NOT EXISTS (
                    SELECT 1 FROM shifts s2 WHERE s2.date = m.date AND s2.branch_id = m.branch_id
                )
            ) combined
            GROUP BY date, branch_id
        ''', [r_date_from, r_date_to, r_date_from, r_date_to]).fetchall()

        # План по дням и филиалам
        _plan_bf = f"AND branch_id IN ({','.join(str(b['id']) for b in rev_branch_list)})" if rev_branch_list else ""
        plan_rows = conn.execute(f'''
            SELECT branch_id, date, amount FROM revenue_plan
            WHERE date BETWEEN ? AND ? {_plan_bf}
        ''', [r_date_from, r_date_to]).fetchall()

        # Сборка сводного словаря
        rev_pivot = {}
        for d in rev_dates:
            rev_pivot[d] = {b['id']: {'actual': 0.0, 'plan': 0.0} for b in rev_branch_list}
        for row in act_rows:
            if row['date'] in rev_pivot and row['branch_id'] in rev_pivot[row['date']]:
                rev_pivot[row['date']][row['branch_id']]['actual'] = float(row['actual'])
        for row in plan_rows:
            if row['date'] in rev_pivot and row['branch_id'] in rev_pivot[row['date']]:
                rev_pivot[row['date']][row['branch_id']]['plan'] = float(row['amount'])

        rev_total_actual = sum(
            rev_pivot[d][b['id']]['actual'] for d in rev_dates for b in rev_branch_list
        )
        rev_total_plan = sum(
            rev_pivot[d][b['id']]['plan'] for d in rev_dates for b in rev_branch_list
        )

        salary_data = conn.execute(f'''
            SELECT es.full_name_snapshot, es.role_snapshot,
                   SUM(es.total_amount) as earned,
                   SUM(CASE WHEN es.is_paid=1 THEN es.total_amount ELSE 0 END) as paid,
                   SUM(CASE WHEN es.is_paid=0 THEN es.total_amount ELSE 0 END) as debt,
                   COUNT(*) as shifts_count
            FROM employee_shifts es
            JOIN shifts s ON s.id=es.shift_id
            WHERE 1=1 {date_filter} {branch_filter}
            GROUP BY es.full_name_snapshot, es.role_snapshot
            ORDER BY debt DESC
        ''').fetchall()

        # ── Зарплатный отчёт ──────────────────────────────────────────────
        s_group = request.args.get('s_group', 'summary')
        s_emps  = request.args.getlist('s_emps')

        sal_conds  = ['s.date BETWEEN ? AND ?']
        sal_params = [s_date_from, s_date_to]
        sal_branch_ids = get_effective_branch_ids('reports_salary', s_branch_ids)
        if sal_branch_ids:
            sal_ph = ','.join('?' * len(sal_branch_ids))
            sal_conds.append(f's.branch_id IN ({sal_ph})')
            sal_params.extend(int(x) for x in sal_branch_ids)
        if s_role:
            sal_conds.append('es.role_snapshot = ?')
            sal_params.append(s_role)
        sal_where  = ' AND '.join(sal_conds)
        sal_having = 'HAVING SUM(CASE WHEN es.is_paid=0 THEN es.total_amount ELSE 0 END) > 0' if s_unpaid == '1' else ''

        sal_rows_raw = conn.execute(f'''
            SELECT es.employee_id,
                   COALESCE(e.full_name, es.full_name_snapshot) AS name,
                   es.role_snapshot                              AS role,
                   b.id                                          AS branch_id,
                   b.name                                        AS branch_name,
                   COUNT(*)                                      AS shifts_count,
                   COALESCE(SUM(es.total_amount), 0)             AS earned,
                   COALESCE(SUM(CASE WHEN es.is_paid=1 THEN es.total_amount ELSE 0 END), 0) AS paid,
                   COALESCE(SUM(CASE WHEN es.is_paid=0 THEN es.total_amount ELSE 0 END), 0) AS debt
            FROM employee_shifts es
            JOIN shifts    s ON s.id    = es.shift_id
            JOIN branches  b ON b.id    = s.branch_id
            LEFT JOIN employees e ON e.id = es.employee_id
            WHERE {sal_where}
            GROUP BY COALESCE(CAST(es.employee_id AS TEXT), es.full_name_snapshot),
                     es.role_snapshot, b.id
            ORDER BY COALESCE(e.full_name, es.full_name_snapshot), es.role_snapshot, b.name
        ''', sal_params).fetchall()

        # Merge by employee: collect all roles + branches per person
        from collections import OrderedDict as _OD2
        _sal_map = _OD2()
        for row in sal_rows_raw:
            key = (row['employee_id'], row['name'])
            if key not in _sal_map:
                _sal_map[key] = {
                    'employee_id': row['employee_id'],
                    'name': row['name'],
                    'roles': [],
                    'branch_names': [],
                    'shifts_count': 0,
                    'earned': 0.0,
                    'paid': 0.0,
                    'debt': 0.0,
                }
            entry = _sal_map[key]
            if row['role'] not in entry['roles']:
                entry['roles'].append(row['role'])
            if row['branch_name'] not in entry['branch_names']:
                entry['branch_names'].append(row['branch_name'])
            entry['shifts_count'] += row['shifts_count']
            entry['earned'] += row['earned']
            entry['paid']   += row['paid']
            entry['debt']   += row['debt']
        sal_report = list(_sal_map.values())
        if s_unpaid == '1':
            sal_report = [r for r in sal_report if r['debt'] > 0]

        # Список всех сотрудников в периоде (для дропдауна выбора)
        all_sal_emps = conn.execute(f'''
            SELECT DISTINCT COALESCE(e3.full_name, es.full_name_snapshot) AS name
            FROM employee_shifts es
            JOIN shifts s ON s.id = es.shift_id
            LEFT JOIN employees e3 ON e3.id = es.employee_id
            WHERE {sal_where}
            ORDER BY 1
        ''', sal_params).fetchall()
        all_sal_emps = [r['name'] for r in all_sal_emps]

        # ── Сводная таблица по периодам ───────────────────────────────────
        pivot_rows = []
        pivot_emps = []

        if s_group != 'summary':
            raw_params = sal_params[:]
            emp_filter_sql = ''
            if s_emps:
                placeholders = ','.join('?' * len(s_emps))
                emp_filter_sql = f'AND COALESCE(e2.full_name, es.full_name_snapshot) IN ({placeholders})'
                raw_params = raw_params + s_emps

            raw_rows = conn.execute(f'''
                SELECT s.date,
                       COALESCE(e2.full_name, es.full_name_snapshot) AS name,
                       COALESCE(es.total_amount, 0) AS earned,
                       CASE WHEN es.is_paid=1 THEN COALESCE(es.total_amount, 0) ELSE 0 END AS paid
                FROM employee_shifts es
                JOIN shifts s ON s.id = es.shift_id
                LEFT JOIN employees e2 ON e2.id = es.employee_id
                WHERE {sal_where} {emp_filter_sql}
                ORDER BY s.date, COALESCE(e2.full_name, es.full_name_snapshot)
            ''', raw_params).fetchall()

            def _period_key(date_str):
                d = datetime.strptime(date_str, '%Y-%m-%d')
                if s_group == 'day':
                    return date_str
                elif s_group == 'week':
                    return (d - timedelta(days=d.weekday())).strftime('%Y-%m-%d')
                else:  # month
                    return date_str[:7]

            def _period_label(key):
                if s_group == 'day':
                    d = datetime.strptime(key, '%Y-%m-%d')
                    days_ru = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
                    return f"{days_ru[d.weekday()]} {d.strftime('%d.%m.%Y')}"
                elif s_group == 'week':
                    mon = datetime.strptime(key, '%Y-%m-%d')
                    sun = mon + timedelta(days=6)
                    return f"{mon.strftime('%d.%m')}–{sun.strftime('%d.%m.%Y')}"
                else:
                    months_ru = ['', 'Январь', 'Февраль', 'Март', 'Апрель', 'Май',
                                 'Июнь', 'Июль', 'Август', 'Сентябрь', 'Октябрь',
                                 'Ноябрь', 'Декабрь']
                    y, m = key.split('-')
                    return f"{months_ru[int(m)]} {y}"

            from collections import OrderedDict
            periods = OrderedDict()
            emp_order = []

            for row in raw_rows:
                pk  = _period_key(row['date'])
                nm  = row['name']
                if pk not in periods:
                    periods[pk] = {'label': _period_label(pk), 'emps': {}, 'earned': 0.0, 'paid': 0.0}
                if nm not in periods[pk]['emps']:
                    periods[pk]['emps'][nm] = {'earned': 0.0, 'paid': 0.0}
                periods[pk]['emps'][nm]['earned'] += row['earned']
                periods[pk]['emps'][nm]['paid']   += row['paid']
                periods[pk]['earned'] += row['earned']
                periods[pk]['paid']   += row['paid']
                if nm not in emp_order:
                    emp_order.append(nm)

            for p in periods.values():
                p['debt'] = p['earned'] - p['paid']
                for cell in p['emps'].values():
                    cell['debt'] = cell['earned'] - cell['paid']

            pivot_rows = list(periods.values())
            pivot_emps = emp_order

        pos_abbr_map = {
            r['code']: (r['abbr'] or r['name'][:4]).upper()
            for r in conn.execute('SELECT code, name, abbr FROM positions').fetchall()
        }

    return render_template('reports.html',
        shifts_data=shifts_data, totals=totals, branches=branches,
        salary_data=salary_data, selected_branches=branch_ids,
        role_labels=ROLE_LABELS, active_tab=active_tab,
        sal_report=sal_report,
        s_date_from=s_date_from, s_date_to=s_date_to,
        s_branch_ids=s_branch_ids, s_role=s_role, s_unpaid=s_unpaid,
        s_group=s_group, s_emps=s_emps,
        all_sal_emps=all_sal_emps, pivot_rows=pivot_rows, pivot_emps=pivot_emps,
        day_groups=day_groups,
        r_date_from=r_date_from, r_date_to=r_date_to,
        rev_branch_list=rev_branch_list, rev_dates=rev_dates,
        rev_date_dow=rev_date_dow,
        rev_pivot=rev_pivot, rev_total_actual=rev_total_actual,
        rev_total_plan=rev_total_plan,
        branch_groups=get_branch_groups(conn),
        pos_abbr_map=pos_abbr_map)


# ─── EMPLOYEE SALARY DETAIL ───────────────────────────────────────────────────

@app.route('/reports/revenue-plan/save', methods=['POST'])
@login_required
@menu_permission_required('reports_shifts')
def save_revenue_plan():
    data = request.get_json(force=True)
    branch_id = int(data.get('branch_id', 0))
    d = str(data.get('date', ''))
    amount = float(data.get('amount', 0) or 0)
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', d) or branch_id <= 0:
        return jsonify({'ok': False, 'error': 'bad params'}), 400
    with get_db() as conn:
        conn.execute(
            'INSERT OR REPLACE INTO revenue_plan (branch_id, date, amount, updated_at) VALUES (?,?,?,CURRENT_TIMESTAMP)',
            (branch_id, d, amount)
        )
        conn.commit()
    return jsonify({'ok': True})


@app.route('/reports/revenue-plan/auto', methods=['POST'])
@login_required
@menu_permission_required('reports_shifts')
def auto_revenue_plan():
    data = request.get_json(force=True)
    total_plan  = float(data.get('total', 0) or 0)
    date_from   = str(data.get('date_from', ''))
    date_to     = str(data.get('date_to', ''))
    branch_ids  = [int(b) for b in data.get('branch_ids', []) if str(b).isdigit()]

    if total_plan <= 0 or not re.match(r'^\d{4}-\d{2}-\d{2}$', date_from) or \
       not re.match(r'^\d{4}-\d{2}-\d{2}$', date_to):
        return jsonify({'ok': False, 'error': 'bad params'}), 400

    cur_from = datetime.strptime(date_from, '%Y-%m-%d').date()
    cur_to   = datetime.strptime(date_to,   '%Y-%m-%d').date()
    cur_dates = []
    _d = cur_from
    while _d <= cur_to:
        cur_dates.append(_d)
        _d += timedelta(days=1)

    # Какой по счёту (0-based) этот день недели в своём месяце
    def _nth_occ(d):
        return sum(1 for day in range(1, d.day)
                   if date(d.year, d.month, day).weekday() == d.weekday())

    # N-е вхождение weekday в месяце (0-based); если не хватает — возвращает последнее
    def _get_nth_weekday(year, month, weekday, n):
        last = None
        count = -1
        for day in range(1, calendar.monthrange(year, month)[1] + 1):
            d = date(year, month, day)
            if d.weekday() == weekday:
                count += 1
                last = d
                if count == n:
                    return d
        return last

    with get_db() as conn:
        if not branch_ids:
            branch_ids = [r['id'] for r in conn.execute(
                'SELECT id FROM branches WHERE is_active=1').fetchall()]

        bf  = f"AND s.branch_id IN ({','.join(str(b) for b in branch_ids)})"
        mbf = f"AND branch_id IN ({','.join(str(b) for b in branch_ids)})"

        # Данные прошлого года того же месяца: смены + ручная выручка
        prev_months = {(d.year - 1, d.month) for d in cur_dates}
        prev_by_date = {}   # date_str -> {branch_id: revenue}
        for (py, pm) in prev_months:
            ps = date(py, pm, 1).isoformat()
            pe = date(py, pm, calendar.monthrange(py, pm)[1]).isoformat()

            for row in conn.execute(f'''
                SELECT s.date, s.branch_id, COALESCE(SUM(r.total_revenue), 0) as rev
                FROM shifts s LEFT JOIN shift_revenue r ON r.shift_id = s.id
                WHERE s.date BETWEEN ? AND ? {bf}
                GROUP BY s.date, s.branch_id
            ''', [ps, pe]).fetchall():
                prev_by_date.setdefault(row['date'], {})[row['branch_id']] = float(row['rev'])

            for row in conn.execute(f'''
                SELECT date, branch_id, amount FROM revenue_manual
                WHERE date BETWEEN ? AND ? {mbf}
            ''', [ps, pe]).fetchall():
                ds, bid = row['date'], row['branch_id']
                if ds not in prev_by_date:
                    prev_by_date[ds] = {}
                if prev_by_date[ds].get(bid, 0) == 0:
                    prev_by_date[ds][bid] = float(row['amount'])

        prev_total_by_date = {ds: sum(v.values()) for ds, v in prev_by_date.items()}

        # Определяем праздничные дни прошлого года:
        # не пт/сб, но выручка > 1.5× среднего для того же дня недели в том месяце
        # Это праздники — для них берём ту же календарную дату (без смещения дня недели)
        holiday_prev = set()
        for (py, pm) in prev_months:
            month_rows = [(date.fromisoformat(ds), t)
                          for ds, t in prev_total_by_date.items()
                          if ds[:7] == f'{py:04d}-{pm:02d}']
            dow_revs = {}
            for d, t in month_rows:
                dow_revs.setdefault(d.weekday(), []).append(t)
            dow_avg_m = {dow: sum(vs) / len(vs) for dow, vs in dow_revs.items() if vs}
            for d, t in month_rows:
                if d.weekday() not in (4, 5):   # не пт и не сб
                    if dow_avg_m.get(d.weekday(), 0) > 0 and t > dow_avg_m[d.weekday()] * 1.5:
                        holiday_prev.add(d)

        # Маппинг: каждому дню текущего периода — соответствующий день прошлого года
        # Пт/Сб и обычные дни: 1-я пятница → 1-я пятница прошлого года и т.д.
        # Праздники прошлого года (не пт/сб): та же календарная дата
        date_mapping = {}
        for cur_d in cur_dates:
            py = cur_d.year - 1
            m  = cur_d.month
            max_day  = calendar.monthrange(py, m)[1]
            same_cal = date(py, m, min(cur_d.day, max_day))
            if same_cal in holiday_prev:
                date_mapping[cur_d] = same_cal
            else:
                date_mapping[cur_d] = _get_nth_weekday(py, m, cur_d.weekday(), _nth_occ(cur_d))

        # Веса = выручка прошлого года за день-ориентир
        # Коэффициент роста = план / сумма весов (= факт прошлого года)
        weights = {}
        for cur_d in cur_dates:
            prev_d = date_mapping.get(cur_d)
            weights[cur_d] = prev_total_by_date.get(prev_d.isoformat(), 0) if prev_d else 0

        total_weight = sum(weights.values())
        if total_weight == 0:
            for cur_d in cur_dates:
                weights[cur_d] = 1.0
            total_weight = float(len(cur_dates))

        growth = total_plan / total_weight

        # Записываем план: каждый день × коэффициент, внутри дня — доля по филиалам
        for cur_d in cur_dates:
            d_str    = cur_d.isoformat()
            prev_d   = date_mapping.get(cur_d)
            prev_ds  = prev_d.isoformat() if prev_d else None
            day_plan = weights[cur_d] * growth

            py_branches = prev_by_date.get(prev_ds, {})
            py_total    = sum(py_branches.values())

            for bid in branch_ids:
                share  = py_branches.get(bid, 0) / py_total if py_total > 0 else 1.0 / len(branch_ids)
                amount = round(day_plan * share, 0)
                conn.execute(
                    'INSERT OR REPLACE INTO revenue_plan (branch_id, date, amount, updated_at) VALUES (?,?,?,CURRENT_TIMESTAMP)',
                    (bid, d_str, amount)
                )
        conn.commit()

        new_plan = conn.execute(f'''
            SELECT branch_id, date, amount FROM revenue_plan
            WHERE date BETWEEN ? AND ?
            AND branch_id IN ({','.join(str(b) for b in branch_ids)})
        ''', [date_from, date_to]).fetchall()

    result = {}
    for row in new_plan:
        result.setdefault(row['date'], {})[row['branch_id']] = row['amount']
    return jsonify({'ok': True, 'plan': result})


@app.route('/settings/revenue-manual')
@login_required
@menu_permission_required('revenue_manual')
def revenue_manual():
    _today = date.today()
    _prev_month = (_today.replace(day=1) - timedelta(days=1))
    _default_from = _prev_month.replace(day=1).isoformat()
    _default_to   = _prev_month.isoformat()

    df = request.args.get('df', _default_from)
    dt = request.args.get('dt', _default_to)
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', df): df = _default_from
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', dt): dt = _default_to

    _df = datetime.strptime(df, '%Y-%m-%d').date()
    _dt = datetime.strptime(dt, '%Y-%m-%d').date()
    dates = []
    _d = _df
    while _d <= _dt:
        dates.append(_d.isoformat())
        _d += timedelta(days=1)

    _RU_DOW = ['пн','вт','ср','чт','пт','сб','вс']
    date_dow = {d: _RU_DOW[datetime.strptime(d, '%Y-%m-%d').weekday()] for d in dates}

    _RU_MONTHS = ['','Январь','Февраль','Март','Апрель','Май','Июнь',
                  'Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']
    month_lbl = f"{_RU_MONTHS[_df.month]} {_df.year}"

    with get_db() as conn:
        branches = conn.execute(
            'SELECT id, name FROM branches WHERE is_active=1 ORDER BY name'
        ).fetchall()

        manual_rows = conn.execute('''
            SELECT branch_id, date, amount FROM revenue_manual
            WHERE date BETWEEN ? AND ?
        ''', [df, dt]).fetchall()

        # Смены за период (чтобы показать, где уже есть данные из смен)
        shift_rows = conn.execute('''
            SELECT s.date, s.branch_id, COALESCE(SUM(r.total_revenue), 0) as amount
            FROM shifts s LEFT JOIN shift_revenue r ON r.shift_id = s.id
            WHERE s.date BETWEEN ? AND ?
            GROUP BY s.date, s.branch_id
        ''', [df, dt]).fetchall()

    # Сборка сводного словаря
    pivot = {}
    for d in dates:
        pivot[d] = {b['id']: {'manual': 0.0, 'shift': 0.0, 'has_shift': False} for b in branches}

    for row in manual_rows:
        if row['date'] in pivot and row['branch_id'] in pivot[row['date']]:
            pivot[row['date']][row['branch_id']]['manual'] = float(row['amount'])

    for row in shift_rows:
        if row['date'] in pivot and row['branch_id'] in pivot[row['date']]:
            pivot[row['date']][row['branch_id']]['shift']     = float(row['amount'])
            pivot[row['date']][row['branch_id']]['has_shift'] = True

    return render_template('revenue_manual.html',
        branches=branches, dates=dates, date_dow=date_dow,
        pivot=pivot, df=df, dt=dt, month_lbl=month_lbl)


@app.route('/settings/revenue-manual/save', methods=['POST'])
@login_required
@menu_permission_required('revenue_manual')
def revenue_manual_save():
    items = request.get_json(force=True)
    if not isinstance(items, list):
        return jsonify({'ok': False}), 400
    with get_db() as conn:
        for item in items:
            branch_id = item.get('branch_id')
            d         = str(item.get('date', ''))
            amount    = float(item.get('amount', 0) or 0)
            if not branch_id or not re.match(r'^\d{4}-\d{2}-\d{2}$', d):
                continue
            if amount > 0:
                conn.execute(
                    'INSERT OR REPLACE INTO revenue_manual (branch_id, date, amount) VALUES (?,?,?)',
                    (branch_id, d, amount)
                )
            else:
                conn.execute(
                    'DELETE FROM revenue_manual WHERE branch_id=? AND date=?',
                    (branch_id, d)
                )
        conn.commit()
    return jsonify({'ok': True})


@app.route('/reports/employee/<int:emp_id>')
@login_required
@menu_permission_required('reports_salary')
def employee_salary_detail(emp_id):
    month_start = date.today().replace(day=1).isoformat()
    today = date.today().isoformat()
    date_from = request.args.get('date_from', month_start)
    date_to   = request.args.get('date_to',   today)

    with get_db() as conn:
        emp = conn.execute('SELECT * FROM employees WHERE id=?', (emp_id,)).fetchone()
        if not emp:
            flash('Сотрудник не найден', 'danger')
            return redirect(url_for('reports') + '?tab=salary')

        shifts_data = conn.execute('''
            SELECT es.id AS es_id,
                   s.id  AS shift_id,
                   s.date,
                   b.name AS branch_name,
                   es.role_snapshot,
                   es.hours_worked,
                   es.km,
                   es.orders,
                   es.shift_start,
                   es.shift_end,
                   es.base_pay,
                   es.bonus_amount,
                   COALESCE(es.auto_bonus, 0) AS auto_bonus,
                   es.penalty_amount,
                   es.total_amount,
                   es.paid_amount,
                   es.is_paid,
                   es.bonus_comment
            FROM employee_shifts es
            JOIN shifts s ON s.id = es.shift_id
            JOIN branches b ON b.id = s.branch_id
            WHERE es.employee_id = ? AND s.date BETWEEN ? AND ?
            ORDER BY s.date DESC, b.name
        ''', (emp_id, date_from, date_to)).fetchall()

        total_earned = sum(float(r['total_amount'] or 0) for r in shifts_data)
        total_paid   = sum(float(r['total_amount'] if r['is_paid'] else 0) for r in shifts_data)
        total_debt   = total_earned - total_paid

    return render_template('employee_salary_detail.html',
        emp=emp, shifts_data=shifts_data,
        date_from=date_from, date_to=date_to,
        total_earned=total_earned, total_paid=total_paid, total_debt=total_debt,
        role_labels=ROLE_LABELS)


@app.route('/employee/<int:emp_id>/change-history')
@login_required
def employee_change_history(emp_id):
    with get_db() as conn:
        emp = conn.execute('SELECT * FROM employees WHERE id=?', (emp_id,)).fetchone()
        if not emp:
            flash('Сотрудник не найден', 'danger')
            return redirect(url_for('employees'))
        sess_role = session.get('role')
        if sess_role != 'owner' and emp['branch_id'] not in _session_branch_ids():
            flash('Нет доступа', 'danger')
            return redirect(url_for('employees'))

        today = date.today().isoformat()

        rate_history = []
        found_current = False
        for h in conn.execute(
            'SELECT * FROM employee_rate_history WHERE employee_id=? ORDER BY effective_from DESC, id DESC',
            (emp_id,)
        ).fetchall():
            d = dict(h)
            if d['effective_from'] > today:
                d['status'] = 'scheduled'
            elif not found_current:
                d['status'] = 'current'
                found_current = True
            else:
                d['status'] = 'past'
            rate_history.append(d)

        pm_history = []
        found_current_role = set()
        for h in conn.execute(
            'SELECT * FROM employee_pay_monthly_history WHERE employee_id=? ORDER BY effective_from DESC, id DESC',
            (emp_id,)
        ).fetchall():
            d = dict(h)
            if d['effective_from'] > today:
                d['status'] = 'scheduled'
            elif d['role'] not in found_current_role:
                d['status'] = 'current'
                found_current_role.add(d['role'])
            else:
                d['status'] = 'past'
            pm_history.append(d)

        pm_branch_scope = {}
        for row in conn.execute('''
            SELECT pmb.role, b.name as branch_name
            FROM employee_pay_monthly_branches pmb JOIN branches b ON b.id=pmb.branch_id
            WHERE pmb.employee_id=? ORDER BY b.name
        ''', (emp_id,)).fetchall():
            pm_branch_scope.setdefault(row['role'], []).append(row['branch_name'])

        role_labels_map = {emp['role']: ROLE_LABELS.get(emp['role'], emp['role'])}
        for r in conn.execute('SELECT DISTINCT role FROM employee_roles WHERE employee_id=?', (emp_id,)).fetchall():
            role_labels_map[r['role']] = ROLE_LABELS.get(r['role'], r['role'])
        for r in conn.execute('SELECT DISTINCT role FROM employee_pay_monthly_history WHERE employee_id=?', (emp_id,)).fetchall():
            role_labels_map.setdefault(r['role'], ROLE_LABELS.get(r['role'], r['role']))

        cur_branches = [r['name'] for r in conn.execute('''
            SELECT b.name FROM employee_branches eb JOIN branches b ON b.id=eb.branch_id
            WHERE eb.employee_id=? ORDER BY b.name
        ''', (emp_id,)).fetchall()]

        address_history = []
        found_current_addr = False
        for h in conn.execute(
            'SELECT * FROM employee_address_history WHERE employee_id=? ORDER BY valid_from DESC, id DESC',
            (emp_id,)
        ).fetchall():
            d = dict(h)
            if d['valid_from'] > today:
                d['status'] = 'scheduled'
            elif not found_current_addr:
                d['status'] = 'current'
                found_current_addr = True
            else:
                d['status'] = 'past'
            address_history.append(d)

    return render_template('employee_change_history.html',
        emp=emp, today=today,
        rate_history=rate_history, pm_history=pm_history,
        pm_branch_scope=pm_branch_scope, role_labels_map=role_labels_map,
        cur_branches=cur_branches, address_history=address_history)


# ─── EXPENSES REPORT ──────────────────────────────────────────────────────────


# ─── CASH FLOW REPORT ────────────────────────────────────────────────────────

@app.route('/report/cash-flow')
@login_required
@menu_permission_required('cash_flow_report')
def cash_flow_report():
    from collections import defaultdict
    today      = date.today().isoformat()
    month_start = date.today().replace(day=1).isoformat()
    date_from  = request.args.get('date_from', month_start)
    date_to    = request.args.get('date_to',   today)
    branch_ids = [b for b in request.args.getlist('branch_ids') if b.isdigit()]
    branch_ids = get_effective_branch_ids('cash_flow_report', branch_ids) or []

    with get_db() as conn:
        branches      = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
        branch_groups = get_branch_groups(conn)
        if branch_ids:
            bf_ids = ','.join(branch_ids)
            bf = f'AND s.branch_id IN ({bf_ids})'
        else:
            bf = ''

        revenue_rows = conn.execute(f'''
            SELECT s.date, s.id AS shift_id, b.id AS branch_id, b.name AS branch_name,
                   COALESCE(r.cash_amount, 0)    AS cash_revenue,
                   COALESCE(r.actual_cash, 0)    AS actual_cash,
                   COALESCE(r.actual_cash_comment, '') AS actual_cash_comment,
                   COALESCE(r.change_amount, 0)  AS razmen,
                   COALESCE(r.plus_amount, 0)    AS plus_amount,
                   COALESCE(r.morning_cash, 0)   AS morning_cash
            FROM shifts s
            JOIN branches b ON b.id = s.branch_id
            LEFT JOIN shift_revenue r ON r.shift_id = s.id
            WHERE s.date BETWEEN ? AND ? {bf}
            ORDER BY s.date, b.name
        ''', (date_from, date_to)).fetchall()

        expense_rows = conn.execute(f'''
            SELECT s.date, s.branch_id,
                COALESCE(SUM(e.amount_cash), 0) AS expenses_cash
            FROM expenses e
            JOIN shifts s ON s.id = e.shift_id
            WHERE s.date BETWEEN ? AND ? {bf}
            GROUP BY s.date, s.branch_id
        ''', (date_from, date_to)).fetchall()

        salary_rows = conn.execute(f'''
            SELECT s.date, s.branch_id, COALESCE(SUM(es.total_amount), 0) AS salary_paid
            FROM employee_shifts es
            JOIN shifts s ON s.id = es.shift_id
            WHERE s.date BETWEEN ? AND ? {bf} AND es.is_paid = 1
            GROUP BY s.date, s.branch_id
        ''', (date_from, date_to)).fetchall()

        taxi_rows = conn.execute(f'''
            SELECT s.date, s.branch_id, COALESCE(SUM(tt.amount), 0) AS taxi_cash
            FROM taxi_trips tt
            JOIN shifts s ON s.id = tt.shift_id
            WHERE tt.payment_type = 'cash' AND s.date BETWEEN ? AND ? {bf}
            GROUP BY s.date, s.branch_id
        ''', (date_from, date_to)).fetchall()

    exp_map = defaultdict(float)
    for r in expense_rows:
        exp_map[(r['date'], r['branch_id'])] += r['expenses_cash']

    sal_map = defaultdict(float)
    for r in salary_rows:
        sal_map[(r['date'], r['branch_id'])] += r['salary_paid']

    taxi_map = defaultdict(float)
    for r in taxi_rows:
        taxi_map[(r['date'], r['branch_id'])] += r['taxi_cash']

    days = {}
    for r in revenue_rows:
        d = r['date']
        bid = r['branch_id']
        exp  = exp_map.get((d, bid), 0)
        sal  = sal_map.get((d, bid), 0)
        taxi = taxi_map.get((d, bid), 0)
        raz  = r['razmen']
        plus = r['plus_amount']
        mrn = r['morning_cash']
        if d not in days:
            days[d] = {'date': d, 'shifts': [], 'cash_revenue': 0.0,
                       'expenses_cash': 0.0, 'razmen': 0.0, 'plus_amount': 0.0,
                       'morning_cash': 0.0, 'salary_paid': 0.0, 'taxi_cash': 0.0,
                       'actual_cash': 0.0}
        days[d]['shifts'].append({
            'shift_id':    r['shift_id'],
            'branch_name': r['branch_name'],
            'cash_revenue': r['cash_revenue'],
            'expenses_cash': exp,
            'razmen':       raz,
            'plus_amount':  plus,
            'morning_cash': mrn,
            'salary_paid':  sal,
            'taxi_cash':    taxi,
            'actual_cash':  r['actual_cash'],
            'actual_cash_comment': r['actual_cash_comment'],
        })
        days[d]['cash_revenue']  += r['cash_revenue']
        days[d]['expenses_cash'] += exp
        days[d]['razmen']        += raz
        days[d]['plus_amount']   += plus
        days[d]['morning_cash']  += mrn
        days[d]['salary_paid']   += sal
        days[d]['taxi_cash']     += taxi
        days[d]['actual_cash']   += r['actual_cash']

    all_dates   = sorted(days.keys())
    prev_cash   = {}
    for i, d in enumerate(all_dates):
        prev_cash[d] = days[all_dates[i-1]]['actual_cash'] if i > 0 else 0.0

    sorted_days = [days[d] for d in reversed(all_dates)]
    for day in sorted_days:
        day['morning_cash'] = prev_cash.get(day['date'], 0.0)

    totals = {
        'razmen':       sum(d['razmen']       for d in sorted_days),
        'plus_amount':  sum(d['plus_amount']  for d in sorted_days),
        'cash_revenue': sum(d['cash_revenue'] for d in sorted_days),
        'expenses_cash':sum(d['expenses_cash']for d in sorted_days),
        'salary_paid':  sum(d['salary_paid']  for d in sorted_days),
        'taxi_cash':    sum(d['taxi_cash']    for d in sorted_days),
    }

    return render_template('cash_flow.html',
        days=sorted_days,
        totals=totals,
        branches=branches,
        branch_groups=branch_groups,
        branch_ids=[str(b) for b in branch_ids],
        date_from=date_from, date_to=date_to)


def _reconciliation_cashless_load_settings(conn):
    row = conn.execute(
        "SELECT value FROM reconciliation_cashless_settings WHERE key='bank_income_categories'"
    ).fetchone()
    cats = []
    if row and row['value']:
        try:
            cats = _json_lib.loads(row['value'])
        except Exception:
            cats = []
    return {'bank_income_categories': cats}


@app.route('/report/reconciliation-cashless')
@login_required
@menu_permission_required('reconciliation_cashless')
def reconciliation_cashless():
    from collections import defaultdict
    today       = date.today().isoformat()
    month_start = date.today().replace(day=1).isoformat()
    date_from   = request.args.get('date_from', month_start)
    date_to     = request.args.get('date_to', today)
    branch_ids  = [b for b in request.args.getlist('branch_ids') if b.isdigit()]
    branch_ids  = get_effective_branch_ids('reconciliation_cashless', branch_ids) or []

    with get_db() as conn:
        branches      = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
        branch_groups = get_branch_groups(conn)
        all_cats      = get_expense_categories(conn)
        cfg           = _reconciliation_cashless_load_settings(conn)

        if branch_ids:
            ph       = ','.join('?' * len(branch_ids))
            bf_shift = f'AND s.branch_id IN ({ph})'
            b_args   = [int(b) for b in branch_ids]
        else:
            bf_shift = ''
            b_args   = []

        # Выручка безналом (без онлайна) по дням/филиалам — из листов смен.
        # branch_id+date уникальны (UNIQUE(branch_id, date) у shifts), так что s.id
        # однозначно определяет смену — используем для ссылки на неё в таблице.
        card_rows = conn.execute(f'''
            SELECT s.date AS d, s.branch_id AS branch_id, s.id AS shift_id,
                   COALESCE(SUM(r.card_amount), 0) AS card_revenue
            FROM shifts s
            LEFT JOIN shift_revenue r ON r.shift_id = s.id
            WHERE s.date BETWEEN ? AND ? {bf_shift}
            GROUP BY s.date, s.branch_id
        ''', [date_from, date_to] + b_args).fetchall()

        # «Сверки» — факт безнал по терминалам, который вносится в смене (shift_terminals,
        # сумма чеков терминала) — отдельный ручной ввод, независимый от card_amount
        # из листа выручки, используется для внутренней сверки (не с банком, а с самой сменой).
        fact_rows = conn.execute(f'''
            SELECT s.date AS d, s.branch_id AS branch_id,
                   COALESCE(SUM(st.amount), 0) AS fact_beznal
            FROM shifts s
            JOIN shift_terminals st ON st.shift_id = s.id
            WHERE s.date BETWEEN ? AND ? {bf_shift}
            GROUP BY s.date, s.branch_id
        ''', [date_from, date_to] + b_args).fetchall()

        # Приход по банку выбранной(-ых) в настройках категории — категория и филиал
        # определяются ПРИОРИТЕТНО по правилу разбора операции (Банк → Правила): если
        # правило совпало, его категория (rule.category) и филиал/группа
        # (bank_parse_rule_branches) перекрывают ручную category/отсутствие привязки
        # счёта — та же логика «эффективной категории», что и в выписке и в простом
        # P&L (bt.category у операций, категоризированных только правилом, пустой —
        # см. п.157/161). Совпадение правила определяется в Python (поиск ключевого
        # слова в описании), поэтому фильтр по категории и филиалу тоже считаем в
        # Python, а не в SQL WHERE.
        bank_map = defaultdict(float)
        if cfg['bank_income_categories']:
            allowed_cats = set(cfg['bank_income_categories'])
            parse_rules = conn.execute(
                'SELECT * FROM bank_parse_rules WHERE is_active=1 ORDER BY sort_order, id'
            ).fetchall()
            rule_branch_ids_map = _rule_branch_ids_map(conn)
            patterns_by_rule = {}
            for p in conn.execute('SELECT * FROM bank_parse_rule_patterns ORDER BY sort_order, id').fetchall():
                patterns_by_rule.setdefault(p['rule_id'], []).append(p)
            account_branch_ids_map = {}
            for row in conn.execute('SELECT bank_account_id, branch_id FROM bank_account_branches').fetchall():
                account_branch_ids_map.setdefault(row['bank_account_id'], set()).add(row['branch_id'])

            bt_rows = conn.execute('''
                SELECT bt.txn_date AS d, bt.amount, bt.description, bt.bank_account_id, bt.category
                FROM bank_transactions bt
                WHERE bt.txn_date BETWEEN ? AND ? AND bt.amount > 0 AND bt.is_ignored=0
            ''', [date_from, date_to]).fetchall()

            allowed_branch_ids = set(int(b) for b in branch_ids) if branch_ids else None
            for r in bt_rows:
                desc = r['description'] or ''
                desc_l = desc.lower()
                txn_branch_ids = None
                matched_rule_category = ''
                matched_rule_id = None
                for rule in parse_rules:
                    if rule['bank_account_id'] and rule['bank_account_id'] != r['bank_account_id']:
                        continue
                    if rule['direction'] == 'expense':
                        continue  # amount > 0 уже отфильтровано выше, это доход
                    kw = (rule['keyword'] or '').lower()
                    if not kw or kw not in desc_l:
                        continue
                    matched_rule_category = rule['category'] or ''
                    matched_rule_id = rule['id']
                    if rule['id'] in rule_branch_ids_map:
                        txn_branch_ids = rule_branch_ids_map[rule['id']]
                    break
                effective_category = r['category'] or matched_rule_category
                if effective_category not in allowed_cats:
                    continue
                if txn_branch_ids is None:
                    txn_branch_ids = account_branch_ids_map.get(r['bank_account_id'], set())

                # Паттерны правила (всегда «расход» — например комиссия банка, удержанная
                # из зачисления) прибавляются к сумме прихода: банк присылает НЕТТО, а
                # выручка безнал из смены — брутто, до удержания комиссии.
                amount = r['amount']
                if matched_rule_id is not None:
                    for pat in patterns_by_rule.get(matched_rule_id, []):
                        if not pat['regex_pattern']:
                            continue
                        try:
                            m = re.search(pat['regex_pattern'], desc, re.IGNORECASE)
                        except re.error:
                            m = None
                        if m:
                            try:
                                amount += float(m.group(1).replace(',', '.'))
                            except (ValueError, IndexError):
                                pass

                for bid in txn_branch_ids:
                    if allowed_branch_ids is not None and bid not in allowed_branch_ids:
                        continue
                    bank_map[(r['d'], bid)] += amount

    branch_name_map = {b['id']: b['name'] for b in branches}

    card_map = {}
    shift_id_map = {}
    for r in card_rows:
        card_map[(r['d'], r['branch_id'])] = r['card_revenue']
        shift_id_map[(r['d'], r['branch_id'])] = r['shift_id']

    fact_map = {}
    for r in fact_rows:
        fact_map[(r['d'], r['branch_id'])] = r['fact_beznal']

    # Объединяем оба источника по (дата, филиал) — попадают и дни, где есть
    # только выручка (банк ещё не зачислил), и дни, где есть только банк
    # (например запоздавшее зачисление) — в этом и смысл сверки.
    all_keys = set(card_map.keys()) | set(bank_map.keys())

    days = {}
    for (d, bid) in all_keys:
        if bid not in branch_name_map:
            continue
        card = card_map.get((d, bid), 0.0)
        bank = bank_map.get((d, bid), 0.0)
        fact = fact_map.get((d, bid), 0.0)
        # Разница = приход банка минус выручка безнал: банк меньше выручки — минус
        # (недостача), банк больше — плюс.
        diff = bank - card
        # Сверки разн. = факт безнал минус выручка безнал (внутренняя сверка смены,
        # не связана с банком) — нейтральная, без окраски по знаку.
        recon_diff = fact - card
        if d not in days:
            days[d] = {'date': d, 'branches': [], 'card_revenue': 0.0, 'bank_income': 0.0,
                       'fact_beznal': 0.0, 'recon_diff': 0.0, 'diff': 0.0}
        days[d]['branches'].append({
            'branch_id':    bid,
            'branch_name':  branch_name_map[bid],
            'shift_id':     shift_id_map.get((d, bid)),
            'card_revenue': card,
            'fact_beznal':  fact,
            'recon_diff':   recon_diff,
            'bank_income':  bank,
            'diff':         diff,
        })
        days[d]['card_revenue'] += card
        days[d]['fact_beznal']  += fact
        days[d]['recon_diff']   += recon_diff
        days[d]['bank_income']  += bank
        days[d]['diff']         += diff

    for d in days:
        days[d]['branches'].sort(key=lambda x: x['branch_name'])

    all_dates   = sorted(days.keys())
    sorted_days = [days[d] for d in reversed(all_dates)]

    totals = {
        'card_revenue': sum(d['card_revenue'] for d in sorted_days),
        'fact_beznal':  sum(d['fact_beznal']  for d in sorted_days),
        'recon_diff':   sum(d['recon_diff']   for d in sorted_days),
        'bank_income':  sum(d['bank_income']  for d in sorted_days),
        'diff':         sum(d['diff']         for d in sorted_days),
    }

    return render_template('reconciliation_cashless.html',
        days=sorted_days,
        totals=totals,
        branches=branches,
        branch_groups=branch_groups,
        branch_ids=[str(b) for b in branch_ids],
        date_from=date_from, date_to=date_to,
        all_cats=all_cats, cfg=cfg)


@app.route('/report/reconciliation-cashless/settings', methods=['POST'])
@login_required
@menu_permission_required('reconciliation_cashless')
def reconciliation_cashless_settings_save():
    cats = request.form.getlist('bank_income_categories')
    with get_db() as conn:
        conn.execute(
            'INSERT OR REPLACE INTO reconciliation_cashless_settings (key, value) VALUES (?, ?)',
            ('bank_income_categories', _json_lib.dumps(cats))
        )
        conn.commit()
    flash('Настройки сверки безнала сохранены', 'success')
    params = {k: request.form.get(k) for k in ('date_from', 'date_to') if request.form.get(k)}
    bids = request.form.getlist('branch_ids')
    if bids:
        params['branch_ids'] = bids
    return redirect(url_for('reconciliation_cashless', **params))


@app.route('/report/wait-time')
@login_required
@menu_permission_required('wait_time_report')
def wait_time_report():
    today = date.today().isoformat()
    month_start = date.today().replace(day=1).isoformat()
    date_from = request.args.get('date_from', month_start)
    date_to   = request.args.get('date_to', today)
    branch_ids = [b for b in request.args.getlist('branch_ids') if b.isdigit()]
    branch_ids = get_effective_branch_ids('wait_time_report', branch_ids) or []

    with get_db() as conn:
        branches = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
        branch_groups = get_branch_groups(conn)
        branch_list = [b for b in branches if not branch_ids or str(b['id']) in branch_ids]

        bf = f"AND branch_id IN ({','.join(branch_ids)})" if branch_ids else ''

        rows = conn.execute(f'''
            SELECT branch_id, DATE(recorded_at) AS d,
                   AVG(promised_minutes) AS avg_promised,
                   MIN(promised_minutes) AS min_promised,
                   MAX(promised_minutes) AS max_promised,
                   COUNT(*) AS samples
            FROM wait_time_log
            WHERE DATE(recorded_at) BETWEEN ? AND ? {bf}
            GROUP BY branch_id, DATE(recorded_at)
        ''', (date_from, date_to)).fetchall()

        pivot = {}
        dates = set()
        for r in rows:
            d = r['d']
            dates.add(d)
            pivot.setdefault(d, {})[r['branch_id']] = {
                'avg': r['avg_promised'], 'min': r['min_promised'],
                'max': r['max_promised'], 'n': r['samples'],
            }
        dates = sorted(dates, reverse=True)

        overall = conn.execute(f'''
            SELECT AVG(promised_minutes) AS avg_all, COUNT(*) AS n_all
            FROM wait_time_log
            WHERE DATE(recorded_at) BETWEEN ? AND ? {bf}
        ''', (date_from, date_to)).fetchone()

        per_branch_avg = conn.execute(f'''
            SELECT branch_id, AVG(promised_minutes) AS avg_b, COUNT(*) AS n_b
            FROM wait_time_log
            WHERE DATE(recorded_at) BETWEEN ? AND ? {bf}
            GROUP BY branch_id
        ''', (date_from, date_to)).fetchall()
        per_branch_avg = {r['branch_id']: {'avg': r['avg_b'], 'n': r['n_b']} for r in per_branch_avg}

    return render_template('wait_time_report.html',
        date_from=date_from, date_to=date_to,
        branches=branches, branch_groups=branch_groups,
        selected_branches=branch_ids, branch_list=branch_list,
        dates=dates, pivot=pivot,
        overall_avg=overall['avg_all'], overall_n=overall['n_all'],
        per_branch_avg=per_branch_avg)


# ─── API 1C ИНТЕГРАЦИЯ ────────────────────────────────────────────────────────

import secrets, xml.etree.ElementTree as _ET

@app.route('/api/1c/<token>', defaults={'subpath': ''}, methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS'])
@app.route('/api/1c/<token>/<path:subpath>', methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS'])
def api_1c_endpoint(token, subpath):
    body = request.get_data(as_text=True)
    with get_db() as conn:
        rec = conn.execute(
            'SELECT * FROM api_1c_tokens WHERE token=?', (token,)
        ).fetchone()

        conn.execute(
            'INSERT INTO api_1c_log (token, branch_id, method, path, body) VALUES (?,?,?,?,?)',
            (token, rec['branch_id'] if rec else None,
             request.method, request.full_path, body[:20000])
        )
        log_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]

        if not rec:
            conn.commit()
            return 'Unauthorized', 401

        branch_id = rec['branch_id']
        parsed = 0

        if request.method == 'POST' and body.strip():
            try:
                purchases_created = _parse_1c_body(conn, branch_id, body,
                                                   request.content_type or '')
                if purchases_created:
                    parsed = 1
                    conn.execute('UPDATE api_1c_log SET parsed_ok=1, status=? WHERE id=?',
                                 (f'создано {purchases_created} накладных', log_id))
            except Exception as e:
                conn.execute('UPDATE api_1c_log SET status=? WHERE id=?',
                             (f'ошибка: {str(e)[:200]}', log_id))

        conn.commit()

    ct = request.content_type or ''
    if 'xml' in ct:
        return '<?xml version="1.0"?><result>true</result>', 200, {'Content-Type': 'application/xml'}
    return '{"result":true}', 200, {'Content-Type': 'application/json'}


def _parse_1c_body(conn, branch_id, body, content_type):
    """Парсит тело запроса от 1С и создаёт накладные. Возвращает кол-во созданных записей."""
    created = 0
    body_s = body.strip()

    # ── JSON ────────────────────────────────────────────────────────────────────
    if body_s.startswith('{') or body_s.startswith('[') or 'json' in content_type:
        import json as _json
        data = _json.loads(body_s)
        docs = data if isinstance(data, list) else data.get('Документы', data.get('documents', [data]))
        for doc in docs:
            supplier = (doc.get('Контрагент') or doc.get('supplier') or doc.get('КонтрагентНаименование') or '')
            if isinstance(supplier, dict):
                supplier = supplier.get('Наименование') or supplier.get('name') or ''
            amount = float(doc.get('СуммаДокумента') or doc.get('amount') or doc.get('Сумма') or 0)
            date_str = (doc.get('Дата') or doc.get('date') or '')[:10]
            inv_num = str(doc.get('Номер') or doc.get('number') or '')
            if supplier and amount > 0:
                conn.execute(
                    'INSERT INTO purchases (branch_id,supplier,amount,date,invoice_number,note) VALUES (?,?,?,?,?,?)',
                    (branch_id, supplier[:200], amount, date_str or date.today().isoformat(),
                     inv_num[:100] or None, '1С импорт')
                )
                conn.execute('INSERT OR IGNORE INTO purchase_suppliers (name) VALUES (?)', (supplier[:200],))
                created += 1
        return created

    # ── XML (EnterpriseData / CommerceML) ───────────────────────────────────────
    if body_s.startswith('<'):
        root = _ET.fromstring(body_s)
        ns_map = {'ed': 'http://v8.1c.ru/edi/edi_stnd/EnterpriseData/1.0',
                  'cm': 'urn:1C.ru:commerceml_3'}

        def find_text(el, *tags):
            for tag in tags:
                for ns_prefix, ns_uri in ns_map.items():
                    found = el.find(f'{{{ns_uri}}}{tag}')
                    if found is not None and found.text:
                        return found.text.strip()
                found = el.find(tag)
                if found is not None and found.text:
                    return found.text.strip()
            return ''

        # Ищем документы поступления товаров
        candidates = (
            list(root.iter('{http://v8.1c.ru/edi/edi_stnd/EnterpriseData/1.0}Документ')) +
            list(root.iter('Документ')) +
            list(root.iter('{urn:1C.ru:commerceml_3}Документ')) +
            list(root.iter('document'))
        )
        for doc in candidates:
            type_attr = doc.get('Тип') or doc.get('type') or find_text(doc, 'Тип', 'type')
            if type_attr and 'Поступ' not in type_attr and 'Receipt' not in type_attr and 'Invoice' not in type_attr:
                continue
            supplier = find_text(doc, 'КонтрагентНаименование', 'Контрагент', 'supplier', 'Поставщик')
            if not supplier:
                ct_el = doc.find('Контрагент') or doc.find('{http://v8.1c.ru/edi/edi_stnd/EnterpriseData/1.0}Контрагент')
                if ct_el is not None:
                    supplier = find_text(ct_el, 'Наименование', 'name') or (ct_el.text or '').strip()
            amount_str = find_text(doc, 'СуммаДокумента', 'Сумма', 'amount', 'Total')
            amount = float(amount_str.replace(',', '.')) if amount_str else 0
            date_str = find_text(doc, 'Дата', 'date', 'Date')[:10] if find_text(doc, 'Дата', 'date', 'Date') else ''
            inv_num  = find_text(doc, 'Номер', 'number', 'Number')
            if supplier and amount > 0:
                conn.execute(
                    'INSERT INTO purchases (branch_id,supplier,amount,date,invoice_number,note) VALUES (?,?,?,?,?,?)',
                    (branch_id, supplier[:200], amount, date_str or date.today().isoformat(),
                     inv_num[:100] or None, '1С импорт')
                )
                conn.execute('INSERT OR IGNORE INTO purchase_suppliers (name) VALUES (?)', (supplier[:200],))
                created += 1
        return created

    return 0


# ─── API ВЕБХУК ВЫРУЧКИ (внешние источники в реальном времени) ───────────────

@app.route('/api/revenue-webhook/<token>', methods=['POST'])
def api_revenue_webhook(token):
    body = request.get_data(as_text=True)
    with get_db() as conn:
        rec = conn.execute(
            'SELECT * FROM api_revenue_tokens WHERE token=?', (token,)
        ).fetchone()

        conn.execute(
            'INSERT INTO api_revenue_log (token, branch_id, method, path, body) VALUES (?,?,?,?,?)',
            (token, rec['branch_id'] if rec else None,
             request.method, request.full_path, body[:20000])
        )
        log_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]

        if not rec:
            conn.commit()
            return jsonify({'ok': False, 'error': 'invalid token'}), 401

        branch_id = rec['branch_id']
        try:
            data = _json_lib.loads(body) if body.strip() else {}
            revenue = float(data.get('revenue') or 0)
            orders_count = int(data.get('orders_count') or 0)
            ts = data.get('timestamp')
            try:
                d = datetime.fromisoformat(str(ts).replace('Z', '+00:00')).astimezone().date().isoformat() if ts else date.today().isoformat()
            except Exception:
                d = date.today().isoformat()
            conn.execute('''
                INSERT OR REPLACE INTO revenue_manual (branch_id, date, amount, orders_count)
                VALUES (?, ?, ?, ?)
            ''', (branch_id, d, revenue, orders_count))
            conn.execute('UPDATE api_revenue_log SET parsed_ok=1, status=? WHERE id=?',
                         (f'выручка {revenue} за {d}', log_id))
            conn.commit()
            return jsonify({'ok': True})
        except Exception as e:
            conn.execute('UPDATE api_revenue_log SET status=? WHERE id=?',
                         (f'ошибка: {str(e)[:200]}', log_id))
            conn.commit()
            return jsonify({'ok': False, 'error': str(e)}), 400


@app.route('/settings/api/revenue-tokens/add', methods=['POST'])
@login_required
@menu_permission_required('settings_api')
def api_revenue_token_add():
    branch_id   = request.form.get('branch_id', type=int)
    description = request.form.get('description', '').strip()
    if not branch_id:
        flash('Выберите филиал', 'danger')
        return redirect(url_for('settings') + '?tab=api')
    token = secrets.token_urlsafe(24)
    with get_db() as conn:
        conn.execute(
            'INSERT INTO api_revenue_tokens (branch_id, token, description) VALUES (?,?,?)',
            (branch_id, token, description or None)
        )
        conn.commit()
    flash('Токен создан', 'success')
    return redirect(url_for('settings') + '?tab=api')


@app.route('/settings/api/revenue-tokens/<int:tid>/delete', methods=['POST'])
@login_required
@menu_permission_required('settings_api')
def api_revenue_token_delete(tid):
    with get_db() as conn:
        conn.execute('DELETE FROM api_revenue_tokens WHERE id=?', (tid,))
        conn.commit()
    flash('Токен удалён', 'success')
    return redirect(url_for('settings') + '?tab=api')


@app.route('/settings/api/revenue-log/clear', methods=['POST'])
@login_required
@menu_permission_required('settings_api')
def api_revenue_log_clear():
    with get_db() as conn:
        conn.execute('DELETE FROM api_revenue_log')
        conn.commit()
    flash('Лог очищен', 'success')
    return redirect(url_for('settings') + '?tab=api')


# ─── Вебхук «Время ожидания» (внешний скрипт из Гуляша) ──────────────────────

@app.route('/api/waittime-webhook/<token>', methods=['POST'])
def api_waittime_webhook(token):
    body = request.get_data(as_text=True)
    with get_db() as conn:
        rec = conn.execute(
            'SELECT * FROM api_waittime_tokens WHERE token=?', (token,)
        ).fetchone()

        conn.execute(
            'INSERT INTO api_waittime_log (token, branch_id, method, path, body) VALUES (?,?,?,?,?)',
            (token, rec['branch_id'] if rec else None,
             request.method, request.full_path, body[:20000])
        )
        log_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]

        if not rec:
            conn.commit()
            return jsonify({'ok': False, 'error': 'invalid token'}), 401

        branch_id = rec['branch_id']
        try:
            data = _json_lib.loads(body) if body.strip() else {}
            promised  = data.get('promised_minutes')
            estimated = data.get('estimated_minutes')
            promised  = float(promised) if promised is not None else None
            estimated = float(estimated) if estimated is not None else None
            ts = data.get('timestamp')
            try:
                recorded_at = (
                    datetime.fromisoformat(str(ts).replace('Z', '+00:00')).astimezone().replace(tzinfo=None).isoformat(sep=' ')
                    if ts else datetime.now().isoformat(sep=' ')
                )
            except Exception:
                recorded_at = datetime.now().isoformat(sep=' ')
            conn.execute('''
                INSERT INTO wait_time_log (branch_id, promised_minutes, estimated_minutes, recorded_at)
                VALUES (?, ?, ?, ?)
            ''', (branch_id, promised, estimated, recorded_at))
            conn.execute('UPDATE api_waittime_log SET parsed_ok=1, status=? WHERE id=?',
                         (f'обещаем {promised} мин за {recorded_at}', log_id))
            conn.commit()
            return jsonify({'ok': True})
        except Exception as e:
            conn.execute('UPDATE api_waittime_log SET status=? WHERE id=?',
                         (f'ошибка: {str(e)[:200]}', log_id))
            conn.commit()
            return jsonify({'ok': False, 'error': str(e)}), 400


@app.route('/settings/api/waittime-tokens/add', methods=['POST'])
@login_required
@menu_permission_required('settings_api')
def api_waittime_token_add():
    branch_id   = request.form.get('branch_id', type=int)
    description = request.form.get('description', '').strip()
    if not branch_id:
        flash('Выберите филиал', 'danger')
        return redirect(url_for('settings') + '?tab=api')
    token = secrets.token_urlsafe(24)
    with get_db() as conn:
        conn.execute(
            'INSERT INTO api_waittime_tokens (branch_id, token, description) VALUES (?,?,?)',
            (branch_id, token, description or None)
        )
        conn.commit()
    flash('Токен создан', 'success')
    return redirect(url_for('settings') + '?tab=api')


@app.route('/settings/api/waittime-tokens/<int:tid>/delete', methods=['POST'])
@login_required
@menu_permission_required('settings_api')
def api_waittime_token_delete(tid):
    with get_db() as conn:
        conn.execute('DELETE FROM api_waittime_tokens WHERE id=?', (tid,))
        conn.commit()
    flash('Токен удалён', 'success')
    return redirect(url_for('settings') + '?tab=api')


@app.route('/settings/api/waittime-log/clear', methods=['POST'])
@login_required
@menu_permission_required('settings_api')
def api_waittime_log_clear():
    with get_db() as conn:
        conn.execute('DELETE FROM api_waittime_log')
        conn.commit()
    flash('Лог очищен', 'success')
    return redirect(url_for('settings') + '?tab=api')


@app.route('/settings/api/tokens/add', methods=['POST'])
@login_required
@menu_permission_required('settings_api')
def api_token_add():
    branch_id   = request.form.get('branch_id', type=int)
    description = request.form.get('description', '').strip()
    if not branch_id:
        flash('Выберите филиал', 'danger')
        return redirect(url_for('settings') + '?tab=api')
    token = secrets.token_urlsafe(24)
    with get_db() as conn:
        conn.execute(
            'INSERT INTO api_1c_tokens (branch_id, token, description) VALUES (?,?,?)',
            (branch_id, token, description or None)
        )
        conn.commit()
    flash('Токен создан', 'success')
    return redirect(url_for('settings') + '?tab=api')


@app.route('/settings/api/tokens/<int:tid>/delete', methods=['POST'])
@login_required
@menu_permission_required('settings_api')
def api_token_delete(tid):
    with get_db() as conn:
        conn.execute('DELETE FROM api_1c_tokens WHERE id=?', (tid,))
        conn.commit()
    flash('Токен удалён', 'success')
    return redirect(url_for('settings') + '?tab=api')


@app.route('/settings/api/log/clear', methods=['POST'])
@login_required
@menu_permission_required('settings_api')
def api_log_clear():
    with get_db() as conn:
        conn.execute('DELETE FROM api_1c_log')
        conn.commit()
    flash('Лог очищен', 'success')
    return redirect(url_for('settings') + '?tab=api')


# ─── PURCHASES (накладные) ────────────────────────────────────────────────────

@app.route('/purchases')
@login_required
def purchases():
    role = session.get('role')
    if not item_visible('purchases'):
        flash('Доступ запрещён', 'danger')
        return redirect(url_for('dashboard'))
    with get_db() as conn:
        if role == 'owner' or can_pick_other_branches('purchases'):
            all_branches  = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
            branch_groups = get_branch_groups(conn)
        else:
            all_branches  = []
            branch_groups = []

        today     = date.today().isoformat()
        first_day = date.today().replace(day=1).isoformat()
        date_from       = request.args.get('p_date_from', first_day)
        date_to         = request.args.get('p_date_to',   today)
        branch_flt_ids  = [b for b in request.args.getlist('p_branch_ids')      if b.isdigit()]
        supp_flt_names  = request.args.getlist('p_supplier_names')

        where  = ['p.date >= ?', 'p.date <= ?']
        params = [date_from, date_to]

        if role != 'owner':
            bids = get_effective_branch_ids('purchases', branch_flt_ids)
            if bids:
                where.append(f"p.branch_id IN ({','.join(str(int(b)) for b in bids)})")
        elif branch_flt_ids:
            ph = ','.join('?' * len(branch_flt_ids))
            where.append(f'p.branch_id IN ({ph})')
            params.extend(int(b) for b in branch_flt_ids)

        if supp_flt_names:
            ph = ','.join('?' * len(supp_flt_names))
            where.append(f'p.supplier IN ({ph})')
            params.extend(supp_flt_names)

        sql_where = ' AND '.join(where)

        rows = conn.execute(f'''
            SELECT p.*, b.name AS branch_name
            FROM purchases p
            JOIN branches b ON b.id = p.branch_id
            WHERE {sql_where}
            ORDER BY p.date DESC, p.id DESC
        ''', params).fetchall()

        by_supplier = conn.execute(f'''
            SELECT p.supplier,
                   COUNT(*) AS cnt,
                   SUM(p.amount) AS total
            FROM purchases p
            JOIN branches b ON b.id = p.branch_id
            WHERE {sql_where}
            GROUP BY p.supplier
            ORDER BY total DESC
        ''', params).fetchall()

        suppliers = [r['name'] for r in conn.execute(
            'SELECT name FROM purchase_suppliers ORDER BY name'
        ).fetchall()]

    return render_template('purchases.html',
                           rows=rows,
                           by_supplier=by_supplier,
                           suppliers=suppliers,
                           all_branches=all_branches,
                           branch_groups=branch_groups,
                           is_owner=(role == 'owner'),
                           date_from=date_from,
                           date_to=date_to,
                           branch_flt_ids=branch_flt_ids,
                           supp_flt_names=supp_flt_names,
                           today=today)


@app.route('/purchases/add', methods=['POST'])
@login_required
def purchases_add():
    role = session.get('role')
    if role == 'owner':
        branch_id = request.form.get('branch_id', type=int)
    else:
        branch_id = session.get('branch_id')
    supplier = request.form.get('supplier', '').strip()
    amount   = float(request.form.get('amount', 0) or 0)
    p_date   = request.form.get('date') or date.today().isoformat()
    inv_num  = request.form.get('invoice_number', '').strip()
    note     = request.form.get('note', '').strip()
    if not supplier or not branch_id or amount <= 0:
        flash('Заполните поставщика, филиал и сумму', 'danger')
        return redirect(url_for('purchases'))
    with get_db() as conn:
        conn.execute(
            'INSERT INTO purchases (branch_id, supplier, amount, date, invoice_number, note, created_by) VALUES (?,?,?,?,?,?,?)',
            (branch_id, supplier, amount, p_date, inv_num or None, note or None, session['user_id'])
        )
        conn.execute('INSERT OR IGNORE INTO purchase_suppliers (name) VALUES (?)', (supplier,))
        conn.commit()
    flash('Накладная добавлена', 'success')
    return redirect(url_for('purchases'))


@app.route('/purchases/<int:pid>/edit', methods=['POST'])
@login_required
def purchases_edit(pid):
    role = session.get('role')
    supplier = request.form.get('supplier', '').strip()
    amount   = float(request.form.get('amount', 0) or 0)
    p_date   = request.form.get('date') or date.today().isoformat()
    inv_num  = request.form.get('invoice_number', '').strip()
    note     = request.form.get('note', '').strip()
    if role == 'owner':
        branch_id = request.form.get('branch_id', type=int)
    else:
        branch_id = None
    with get_db() as conn:
        row = conn.execute('SELECT * FROM purchases WHERE id=?', (pid,)).fetchone()
        if not row:
            flash('Запись не найдена', 'danger')
            return redirect(url_for('purchases'))
        if role != 'owner' and row['branch_id'] not in _session_branch_ids():
            flash('Нет доступа', 'danger')
            return redirect(url_for('purchases'))
        bid = branch_id if (role == 'owner' and branch_id) else row['branch_id']
        conn.execute(
            'UPDATE purchases SET branch_id=?, supplier=?, amount=?, date=?, invoice_number=?, note=? WHERE id=?',
            (bid, supplier, amount, p_date, inv_num or None, note or None, pid)
        )
        conn.execute('INSERT OR IGNORE INTO purchase_suppliers (name) VALUES (?)', (supplier,))
        conn.commit()
    flash('Накладная обновлена', 'success')
    return redirect(url_for('purchases'))


@app.route('/purchases/<int:pid>/delete', methods=['POST'])
@login_required
def purchases_delete(pid):
    with get_db() as conn:
        row = conn.execute('SELECT * FROM purchases WHERE id=?', (pid,)).fetchone()
        if not row:
            flash('Запись не найдена', 'danger')
            return redirect(url_for('purchases'))
        if session.get('role') != 'owner' and row['branch_id'] not in _session_branch_ids():
            flash('Нет доступа', 'danger')
            return redirect(url_for('purchases'))
        conn.execute('DELETE FROM purchases WHERE id=?', (pid,))
        conn.commit()
    flash('Накладная удалена', 'success')
    return redirect(url_for('purchases'))


# ─── PURCHASES EXCEL IMPORT ───────────────────────────────────────────────────

def _xls_parse_invoices(file_bytes):
    """Parse .xls (BIFF8) invoices file. Returns list of row dicts."""
    import struct
    from datetime import datetime as _dt, timedelta as _td

    data = file_bytes

    def _xl_date(v):
        try:
            return (_dt(1899, 12, 30) + _td(days=float(v))).strftime('%Y-%m-%d')
        except Exception:
            return ''

    # ---- find SST and worksheet BOF offsets ----
    def _scan_bof(data):
        results = []
        for i in range(0, len(data) - 8):
            if data[i] == 0x09 and data[i+1] == 0x08:
                rlen = struct.unpack_from('<H', data, i+2)[0]
                if rlen in (8, 16):
                    ver  = struct.unpack_from('<H', data, i+4)[0]
                    typ  = struct.unpack_from('<H', data, i+6)[0]
                    results.append((i, ver, typ))
        return results

    bofs = _scan_bof(data)
    wb_offset = next((o for o, v, t in bofs if t == 0x0005), None)
    ws_offset = next((o for o, v, t in bofs if t == 0x0010), None)
    if wb_offset is None or ws_offset is None:
        raise ValueError('Не найдена структура XLS-файла')

    # ---- parse SST ----
    def _parse_sst(data, start):
        sst = []; offset = start
        while offset < len(data) - 4:
            rtype = struct.unpack_from('<H', data, offset)[0]
            rlen  = struct.unpack_from('<H', data, offset+2)[0]
            rdata = data[offset+4:offset+4+rlen]; offset += 4 + rlen
            if rtype == 0x00FC:
                unique = struct.unpack_from('<I', rdata, 4)[0]; pos = 8
                for _ in range(unique):
                    if pos + 3 > len(rdata): break
                    n = struct.unpack_from('<H', rdata, pos)[0]
                    flags = rdata[pos+2]; pos += 3
                    is_uni = flags & 1; grbit_run = (flags >> 3) & 1
                    if is_uni: s = rdata[pos:pos+n*2].decode('utf-16-le', errors='replace'); pos += n*2
                    else:      s = rdata[pos:pos+n].decode('cp1251', errors='replace'); pos += n
                    if grbit_run:
                        nr = struct.unpack_from('<H', rdata, pos)[0] if pos+2<=len(rdata) else 0
                        pos += 2 + nr*4
                    sst.append(s)
                break
            if rtype == 0x000A: break
        return sst

    # ---- parse worksheet cells ----
    def _parse_ws(data, start, sst):
        cells = {}; offset = start
        while offset < len(data) - 4:
            rtype = struct.unpack_from('<H', data, offset)[0]
            rlen  = struct.unpack_from('<H', data, offset+2)[0]
            rdata = data[offset+4:offset+4+rlen]; offset += 4 + rlen
            if rtype == 0x00FD and len(rdata) >= 10:
                r = struct.unpack_from('<H', rdata, 0)[0]; c = struct.unpack_from('<H', rdata, 2)[0]
                idx = struct.unpack_from('<I', rdata, 6)[0]
                cells[(r, c)] = sst[idx] if idx < len(sst) else ''
            elif rtype == 0x0203 and len(rdata) >= 14:
                r = struct.unpack_from('<H', rdata, 0)[0]; c = struct.unpack_from('<H', rdata, 2)[0]
                cells[(r, c)] = struct.unpack_from('<d', rdata, 6)[0]
            elif rtype == 0x027E and len(rdata) >= 10:
                r = struct.unpack_from('<H', rdata, 0)[0]; c = struct.unpack_from('<H', rdata, 2)[0]
                rk = struct.unpack_from('<I', rdata, 6)[0]; flt = rk & 2; div = rk & 1
                val = float(rk >> 2) if flt else struct.unpack_from('<d', b'\x00\x00\x00\x00' + struct.pack('<I', rk & 0xFFFFFFFC))[0]
                if div: val /= 100.0
                cells[(r, c)] = val
            elif rtype == 0x00BD and len(rdata) >= 6:
                r = struct.unpack_from('<H', rdata, 0)[0]; c0 = struct.unpack_from('<H', rdata, 2)[0]
                for k in range((len(rdata) - 6) // 6):
                    rk = struct.unpack_from('<I', rdata, 4 + k*6 + 2)[0]; flt = rk & 2; div = rk & 1
                    val = float(rk >> 2) if flt else struct.unpack_from('<d', b'\x00\x00\x00\x00' + struct.pack('<I', rk & 0xFFFFFFFC))[0]
                    if div: val /= 100.0
                    cells[(r, c0 + k)] = val
            elif rtype == 0x000A: break
        return cells

    sst   = _parse_sst(data, wb_offset)
    cells = _parse_ws(data, ws_offset, sst)
    if not cells:
        raise ValueError('Данные не найдены в файле')

    max_row = max(r for r, c in cells)

    def cv(r, c):
        v = cells.get((r, c), '')
        return str(v).strip() if v != '' else ''

    rows = []
    for r in range(1, max_row + 1):
        date_str   = _xl_date(cv(r, 1))   # col 2 → index 1
        inv_num    = cv(r, 3)              # col 4 → index 3
        branch_raw = cv(r, 8)             # col 9 → index 8
        supp_raw   = cv(r, 10)            # col 11 → index 10
        payer_raw  = cv(r, 13)            # col 14 → index 13
        amt_raw    = cv(r, 17)            # col 18 → index 17

        if not date_str or not branch_raw:
            continue
        try:
            amount = float(amt_raw) if amt_raw else 0.0
        except ValueError:
            amount = 0.0
        if amount <= 0:
            continue

        rows.append({
            'date':         date_str,
            'invoice_number': inv_num,
            'branch_raw':   branch_raw,
            'supplier_raw': supp_raw,
            'payer_raw':    payer_raw,
            'amount':       amount,
        })
    return rows


def _xlsx_parse_invoices(file_bytes):
    """Parse .xlsx invoices file."""
    import openpyxl, io as _io
    from datetime import datetime as _dt

    wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        try:
            date_v   = row[1]
            inv_num  = str(row[3] or '').strip()
            branch_r = str(row[8] or '').strip()
            supp_r   = str(row[10] or '').strip()
            payer_r  = str(row[13] or '').strip()
            amt_v    = row[17]

            if isinstance(date_v, _dt):
                date_str = date_v.strftime('%Y-%m-%d')
            elif date_v:
                date_str = _parse_date_str(str(date_v))
            else:
                continue

            amount = float(amt_v or 0)
            if amount <= 0 or not branch_r:
                continue
            rows.append({
                'date': date_str, 'invoice_number': inv_num,
                'branch_raw': branch_r, 'supplier_raw': supp_r,
                'payer_raw': payer_r, 'amount': amount,
            })
        except Exception:
            continue
    return rows


@app.route('/purchases/import-excel', methods=['GET', 'POST'])
@login_required
@menu_permission_required('purchases')
def purchases_import_excel():
    import json, uuid, os

    with get_db() as conn:
        branches    = conn.execute('SELECT id, name FROM branches WHERE is_active=1 ORDER BY name').fetchall()
        contractors = conn.execute('SELECT id, name FROM contractors WHERE is_active=1 ORDER BY name').fetchall()

    if request.method == 'GET':
        return render_template('purchases_import_excel.html',
            step=1, branches=branches, contractors=contractors)

    file = request.files.get('excel_file')
    if not file or file.filename == '':
        flash('Выберите файл', 'danger')
        return render_template('purchases_import_excel.html', step=1, branches=branches, contractors=contractors)

    file_bytes = file.read()
    fname = file.filename.lower()

    try:
        if fname.endswith('.xls'):
            rows = _xls_parse_invoices(file_bytes)
        elif fname.endswith('.xlsx'):
            rows = _xlsx_parse_invoices(file_bytes)
        else:
            flash('Поддерживаются только .xls и .xlsx файлы', 'danger')
            return render_template('purchases_import_excel.html', step=1, branches=branches, contractors=contractors)
    except Exception as e:
        flash(f'Ошибка чтения файла: {e}', 'danger')
        return render_template('purchases_import_excel.html', step=1, branches=branches, contractors=contractors)

    if not rows:
        flash('Не найдено строк с данными (сумма > 0 и указан филиал)', 'warning')
        return render_template('purchases_import_excel.html', step=1, branches=branches, contractors=contractors)

    unique_branches  = sorted(set(r['branch_raw']  for r in rows if r['branch_raw']))
    unique_suppliers = sorted(set(r['supplier_raw'] for r in rows if r['supplier_raw']))
    unique_payers    = sorted(set(r['payer_raw']    for r in rows if r['payer_raw']))

    import hashlib

    def _row_hash(row):
        key = f"{row['date']}|{row['invoice_number']}|{row['supplier_raw']}|{row['amount']:.4f}"
        return hashlib.md5(key.encode('utf-8')).hexdigest()

    hashes = [_row_hash(r) for r in rows]

    with get_db() as conn:
        existing_hashes = set()
        for h in hashes:
            if conn.execute('SELECT 1 FROM purchases WHERE import_hash=?', (h,)).fetchone():
                existing_hashes.add(h)
    already_count = len(existing_hashes)

    import_key = str(uuid.uuid4())
    temp_path  = f'/tmp/crm_inv_{import_key}.json'
    with open(temp_path, 'w', encoding='utf-8') as f:
        json.dump({
            'rows': rows, 'filename': file.filename,
            'unique_branches': unique_branches,
            'unique_suppliers': unique_suppliers,
            'unique_payers': unique_payers,
        }, f, ensure_ascii=False)
    session['inv_import_key'] = import_key

    date_min = min(r['date'] for r in rows)
    date_max = max(r['date'] for r in rows)

    return render_template('purchases_import_excel.html',
        step=2, branches=branches, contractors=contractors,
        preview=rows[:15], total_rows=len(rows),
        already_count=already_count,
        unique_branches=unique_branches,
        unique_suppliers=unique_suppliers,
        unique_payers=unique_payers,
        filename=file.filename,
        date_min=date_min, date_max=date_max)


@app.route('/purchases/import-excel/confirm', methods=['POST'])
@login_required
@menu_permission_required('purchases')
def purchases_import_excel_confirm():
    import json, os, hashlib

    import_key = session.get('inv_import_key')
    if not import_key:
        flash('Сессия истекла, загрузите файл заново', 'danger')
        return redirect(url_for('purchases_import_excel'))

    temp_path = f'/tmp/crm_inv_{import_key}.json'
    try:
        with open(temp_path, 'r', encoding='utf-8') as f:
            idata = json.load(f)
    except FileNotFoundError:
        flash('Данные не найдены, загрузите файл заново', 'danger')
        return redirect(url_for('purchases_import_excel'))

    rows             = idata['rows']
    unique_branches  = idata['unique_branches']
    unique_suppliers = idata['unique_suppliers']
    unique_payers    = idata['unique_payers']

    branch_map = {}
    for raw in unique_branches:
        val = request.form.get(f'branch_{raw}', '').strip()
        if val.isdigit():
            branch_map[raw] = int(val)

    supplier_map = {}
    for raw in unique_suppliers:
        val = request.form.get(f'supplier_{raw}', '').strip()
        supplier_map[raw] = val if val else raw

    include_payers = set()
    for raw in unique_payers:
        if request.form.get(f'payer_{raw}'):
            include_payers.add(raw)

    def _row_hash(row):
        key = f"{row['date']}|{row['invoice_number']}|{row['supplier_raw']}|{row['amount']:.4f}"
        return hashlib.md5(key.encode('utf-8')).hexdigest()

    imported = skipped = duplicates = 0
    with get_db() as conn:
        for row in rows:
            payer = row['payer_raw']
            if unique_payers and payer not in include_payers:
                skipped += 1
                continue
            branch_id = branch_map.get(row['branch_raw'])
            if not branch_id:
                skipped += 1
                continue
            supplier = supplier_map.get(row['supplier_raw'], row['supplier_raw'])
            if not supplier:
                skipped += 1
                continue

            h = _row_hash(row)
            existing = conn.execute('SELECT id FROM purchases WHERE import_hash=?', (h,)).fetchone()
            if existing:
                duplicates += 1
                continue

            conn.execute(
                'INSERT INTO purchases (branch_id, supplier, amount, date, invoice_number, payer, import_hash, created_by) VALUES (?,?,?,?,?,?,?,?)',
                (branch_id, supplier, row['amount'], row['date'],
                 row['invoice_number'] or None, payer or None, h, session.get('user_id'))
            )
            conn.execute('INSERT OR IGNORE INTO purchase_suppliers (name) VALUES (?)', (supplier,))
            if not conn.execute('SELECT id FROM contractors WHERE LOWER(name)=LOWER(?)', (supplier,)).fetchone():
                conn.execute('INSERT INTO contractors (name, keywords) VALUES (?,?)', (supplier, supplier))
            imported += 1
        conn.commit()

    try:
        os.remove(temp_path)
    except Exception:
        pass
    session.pop('inv_import_key', None)

    parts = [f'Импортировано {imported} накладных']
    if duplicates:
        parts.append(f'пропущено дублей: {duplicates}')
    if skipped:
        parts.append(f'пропущено без маппинга: {skipped}')
    flash(', '.join(parts), 'success' if imported > 0 else 'warning')
    return redirect(url_for('purchases'))


@app.route('/report/reconciliation')
@login_required
@menu_permission_required('report_reconciliation')
def report_reconciliation():
    today      = date.today().isoformat()
    month_start = date.today().replace(day=1).isoformat()
    date_from      = request.args.get('date_from', month_start)
    date_to        = request.args.get('date_to',   today)
    contractor_id  = request.args.get('contractor_id', '').strip()

    with get_db() as conn:
        contractors = conn.execute(
            'SELECT id, name FROM contractors WHERE is_active=1 ORDER BY name'
        ).fetchall()

        contractor = None
        rows = []
        opening_balance = 0.0
        total_debit = 0.0
        total_credit = 0.0

        if contractor_id.isdigit():
            contractor = conn.execute(
                'SELECT * FROM contractors WHERE id=?', (int(contractor_id),)
            ).fetchone()

        if contractor:
            # Сальдо начальное: накладные и оплаты ДО начала периода
            ob_purchases = conn.execute('''
                SELECT COALESCE(SUM(amount), 0)
                FROM purchases
                WHERE LOWER(TRIM(supplier)) = LOWER(TRIM(?))
                  AND date < ?
            ''', (contractor['name'], date_from)).fetchone()[0] or 0.0

            ob_payments = conn.execute('''
                SELECT COALESCE(SUM(amount), 0)
                FROM bank_transactions
                WHERE contractor_id = ?
                  AND txn_date < ?
                  AND is_ignored = 0
            ''', (contractor['id'], date_from)).fetchone()[0] or 0.0

            # payments хранятся как отрицательные (расход), поэтому суммируем
            opening_balance = ob_purchases + ob_payments

            # Накладные за период → Дебет (нам поставили, мы должны)
            purchases_rows = conn.execute('''
                SELECT p.date, p.invoice_number, p.supplier, p.amount,
                       b.name AS branch_name
                FROM purchases p
                JOIN branches b ON b.id = p.branch_id
                WHERE LOWER(TRIM(p.supplier)) = LOWER(TRIM(?))
                  AND p.date BETWEEN ? AND ?
                ORDER BY p.date, p.id
            ''', (contractor['name'], date_from, date_to)).fetchall()

            # Оплаты из банковских выписок → Кредит (мы заплатили)
            payment_rows = conn.execute('''
                SELECT bt.txn_date AS date, bt.description,
                       bt.counterparty, bt.amount,
                       ba.name AS account_name
                FROM bank_transactions bt
                JOIN bank_accounts ba ON ba.id = bt.bank_account_id
                WHERE bt.contractor_id = ?
                  AND bt.txn_date BETWEEN ? AND ?
                  AND bt.is_ignored = 0
                ORDER BY bt.txn_date, bt.id
            ''', (contractor['id'], date_from, date_to)).fetchall()

            for r in purchases_rows:
                rows.append({
                    'date': r['date'],
                    'doc': 'Накладная №' + (r['invoice_number'] or '—') + ' (' + r['branch_name'] + ')',
                    'debit': r['amount'],
                    'credit': None,
                })
                total_debit += r['amount']

            for r in payment_rows:
                amt = abs(r['amount'])
                desc = r['description'] or r['counterparty'] or 'Оплата'
                rows.append({
                    'date': r['date'],
                    'doc': desc,
                    'debit': None,
                    'credit': amt,
                })
                total_credit += amt

            rows.sort(key=lambda x: x['date'])

    # Сальдо конечное: сальдо начальное + обороты по дебету − обороты по кредиту
    closing_balance = opening_balance + total_debit - total_credit

    return render_template('report_reconciliation.html',
        contractors=contractors,
        contractor=contractor,
        contractor_id=contractor_id,
        rows=rows,
        opening_balance=opening_balance,
        total_debit=total_debit,
        total_credit=total_credit,
        closing_balance=closing_balance,
        date_from=date_from,
        date_to=date_to)


@app.route('/report/expenses')
@login_required
@menu_permission_required('expenses_report')
def expenses_report():
    today = date.today().isoformat()
    month_start = date.today().replace(day=1).isoformat()
    date_from   = request.args.get('date_from', month_start)
    date_to     = request.args.get('date_to', today)
    branch_ids  = [b for b in request.args.getlist('branch_ids') if b.isdigit()]
    branch_ids  = get_effective_branch_ids('expenses_report', branch_ids) or []
    cat_filter  = request.args.get('category', '')
    pay_filter  = request.args.get('pay_type', '')   # 'cash' | 'card' | ''

    with get_db() as conn:
        branches = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
        all_cats = conn.execute('SELECT * FROM expense_categories ORDER BY sort_order, label').fetchall()
        # Для фильтра показываем только категории расходов (не приход/любое) —
        # напр. «Плюсы в кассу» отмечена как приход и не должна тут выбираться.
        filter_cats = [c for c in all_cats if c['type'] == 'expense']

        conds  = ["s.date BETWEEN ? AND ?"]
        params = [date_from, date_to]
        if branch_ids:
            ph = ','.join('?' * len(branch_ids))
            conds.append(f's.branch_id IN ({ph})')
            params.extend(int(b) for b in branch_ids)
        if cat_filter:
            # Если выбрана родительская группа — подтягиваем и все её подкатегории
            # (напр. «Реклама» показывает «Реклама в лифтах», «Промоутеры» и т.д.)
            cat_row = next((c for c in all_cats if c['code'] == cat_filter), None)
            if cat_row and not cat_row['parent_id']:
                cat_codes = [cat_filter] + [c['code'] for c in all_cats if c['parent_id'] == cat_row['id']]
            else:
                cat_codes = [cat_filter]
            ph = ','.join('?' * len(cat_codes))
            conds.append(f'e.category IN ({ph})')
            params.extend(cat_codes)
        if pay_filter == 'cash':
            conds.append('e.amount_cash > 0')
        elif pay_filter == 'card':
            conds.append('e.amount_card > 0')
        where = ' AND '.join(conds)

        rows = conn.execute(f'''
            SELECT e.id, s.date, s.id as shift_id,
                   b.name as branch_name,
                   e.category, e.description,
                   COALESCE(e.amount_cash,0) as amount_cash,
                   COALESCE(e.amount_card,0) as amount_card,
                   COALESCE(e.amount_cash,0)+COALESCE(e.amount_card,0) as total,
                   e.is_gulash
            FROM expenses e
            JOIN shifts s ON s.id = e.shift_id
            JOIN branches b ON b.id = s.branch_id
            WHERE {where}
            ORDER BY s.date DESC, b.name, e.id
        ''', params).fetchall()

        tot = conn.execute(f'''
            SELECT COALESCE(SUM(e.amount_cash),0) as cash,
                   COALESCE(SUM(e.amount_card),0) as card,
                   COALESCE(SUM(e.amount_cash+e.amount_card),0) as total
            FROM expenses e
            JOIN shifts s ON s.id = e.shift_id
            WHERE {where}
        ''', params).fetchone()

        by_cat = conn.execute(f'''
            SELECT e.category,
                   COALESCE(SUM(e.amount_cash),0) as cash,
                   COALESCE(SUM(e.amount_card),0) as card,
                   COALESCE(SUM(e.amount_cash+e.amount_card),0) as total,
                   COUNT(*) as cnt
            FROM expenses e
            JOIN shifts s ON s.id = e.shift_id
            WHERE {where}
            GROUP BY e.category
            ORDER BY total DESC
        ''', params).fetchall()

    cat_map = {c['code']: c['label'] for c in all_cats}

    return render_template('expenses_report.html',
        rows=rows, tot=tot, by_cat=by_cat,
        branches=branches, all_cats=all_cats, filter_cats=filter_cats, cat_map=cat_map,
        date_from=date_from, date_to=date_to,
        branch_ids=branch_ids, cat_filter=cat_filter, pay_filter=pay_filter,
        branch_groups=get_branch_groups(conn))


# ─── PnL REPORT ───────────────────────────────────────────────────────────────

import json as _json

_MONTH_NAMES_RU = {
    1: 'Янв', 2: 'Фев', 3: 'Мар', 4: 'Апр',
    5: 'Май', 6: 'Июн', 7: 'Июл', 8: 'Авг',
    9: 'Сен', 10: 'Окт', 11: 'Ноя', 12: 'Дек',
}


def _pnl_period_label(p):
    if p == 'total':
        return 'Итого'
    try:
        dt = datetime.strptime(p, '%Y-%m')
        return f"{_MONTH_NAMES_RU[dt.month]} {dt.year}"
    except Exception:
        return p



def _pnl_load_settings(conn):
    rows = conn.execute('SELECT key, value FROM pnl_settings').fetchall()
    cfg  = {r['key']: r['value'] for r in rows}

    def _lst(key, default):
        raw = cfg.get(key)
        if raw is None:
            return default
        try:
            return _json.loads(raw)
        except Exception:
            return default

    default_expense_cats = [c['code'] for c in conn.execute(
        "SELECT code FROM expense_categories WHERE is_active=1 AND type='expense' ORDER BY sort_order, label"
    ).fetchall()]

    bi = _lst('bank_income_ctr_cats',  None)
    return {
        # Расход теперь один список категорий: сумма и наличных (из смен),
        # и банковских расходов по этой же категории — расходы в одном месте.
        'expense_cats':          _lst('expense_cats', default_expense_cats),
        'bank_income_ctr_cats':  bi or None,   # пустой список [] → None (все)
        'include_salary':           int(cfg.get('include_salary', '1')),
        'include_salary_breakdown': int(cfg.get('include_salary_breakdown', '1')),
        # Простой P&L: категории выписки, которые не считаются ни в приходе, ни в расходе
        # (например «Перераспределение средств» — внутренние переводы между счетами).
        'simple_pnl_excluded_cats': _lst('simple_pnl_excluded_cats', []),
    }


@app.route('/report/pnl')
@login_required
@menu_permission_required('pnl_report')
def pnl_report():
    from collections import defaultdict

    today       = date.today().isoformat()
    month_start = date.today().replace(day=1).isoformat()

    date_from  = request.args.get('date_from', month_start)
    date_to    = request.args.get('date_to', today)
    branch_ids = [b for b in request.args.getlist('branch_ids') if b.isdigit()]
    branch_ids = get_effective_branch_ids('pnl_report', branch_ids) or []
    group_by   = request.args.get('group_by', 'month')

    with get_db() as conn:
        branches      = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
        branch_groups = get_branch_groups(conn)
        all_cats      = get_expense_categories(conn)
        ctr_all_cats  = [c for c in all_cats if c['show_contractors']]
        ctr_cats_income  = [c for c in ctr_all_cats if c['type'] == 'income']
        expense_cats_all = [c for c in all_cats if c['type'] == 'expense']
        cfg           = _pnl_load_settings(conn)

        if branch_ids:
            ph    = ','.join('?' * len(branch_ids))
            bf    = f'AND s.branch_id IN ({ph})'
            bf_p  = f'AND p.branch_id IN ({ph})'
            bf_bt = (f'AND bt.bank_account_id IN '
                     f'(SELECT bank_account_id FROM bank_account_branches WHERE branch_id IN ({ph}))')
            b_args = [int(b) for b in branch_ids]
        else:
            bf = bf_p = bf_bt = ''
            b_args = []

        pe    = "strftime('%Y-%m', s.date)"      if group_by == 'month' else "'total'"
        pe_bt = "strftime('%Y-%m', bt.txn_date)" if group_by == 'month' else "'total'"

        # Общая выручка (база для %)
        total_rev_by_p = {}
        for r in conn.execute(f"""
            SELECT {pe} AS period, COALESCE(SUM(r.total_revenue), 0) AS amount
            FROM shifts s LEFT JOIN shift_revenue r ON r.shift_id=s.id
            WHERE s.date BETWEEN ? AND ? {bf}
            GROUP BY period ORDER BY period
        """, [date_from, date_to] + b_args).fetchall():
            total_rev_by_p[r['period']] = r['amount']

        # Наличный приход
        cash_rev_by_p = {}
        for r in conn.execute(f"""
            SELECT {pe} AS period, COALESCE(SUM(r.cash_amount), 0) AS amount
            FROM shifts s LEFT JOIN shift_revenue r ON r.shift_id=s.id
            WHERE s.date BETWEEN ? AND ? {bf}
            GROUP BY period ORDER BY period
        """, [date_from, date_to] + b_args).fetchall():
            cash_rev_by_p[r['period']] = r['amount']

        # Приход из банка по категориям контрагентов
        bank_inc_by = defaultdict(lambda: defaultdict(float))
        _bi_filter = ''
        _bi_args   = []
        if cfg['bank_income_ctr_cats'] is not None:
            if cfg['bank_income_ctr_cats']:
                ph_c = ','.join('?' * len(cfg['bank_income_ctr_cats']))
                _bi_filter = f"AND bt.category IN ({ph_c})"
                _bi_args   = cfg['bank_income_ctr_cats']
            else:
                _bi_filter = 'AND 0=1'  # пустой список = ничего
        for r in conn.execute(f"""
            SELECT {pe_bt} AS period,
                   bt.category AS cat_name,
                   SUM(bt.amount) AS amount
            FROM bank_transactions bt
            WHERE bt.txn_date BETWEEN ? AND ?
              AND bt.amount > 0 AND bt.is_ignored=0
              AND bt.category IS NOT NULL AND bt.category != ''
              {_bi_filter} {bf_bt}
            GROUP BY period, bt.category
            HAVING SUM(bt.amount) > 0
        """, [date_from, date_to] + _bi_args + b_args).fetchall():
            bank_inc_by[r['cat_name']][r['period']] += r['amount']

        # Приход из сумм, извлечённых по паттернам правил разбора (Банк → Правила) —
        # у них branch_id указан напрямую, фильтр по филиалу проще, чем у bank_transactions.
        pe_pat = "strftime('%Y-%m', pe.txn_date)" if group_by == 'month' else "'total'"
        bf_pe = f"AND pe.branch_id IN ({','.join('?' * len(branch_ids))})" if branch_ids else ''
        for r in conn.execute(f"""
            SELECT {pe_pat} AS period, pe.category AS cat_name, SUM(pe.amount) AS amount
            FROM bank_parse_extracted_entries pe
            WHERE pe.txn_date BETWEEN ? AND ? AND pe.direction='income'
              AND pe.category IS NOT NULL AND pe.category != '' {_bi_filter.replace('bt.', 'pe.')} {bf_pe}
            GROUP BY period, pe.category
            HAVING SUM(pe.amount) > 0
        """, [date_from, date_to] + _bi_args + b_args).fetchall():
            bank_inc_by[r['cat_name']][r['period']] += r['amount']

        # Расходы по категориям — расходы теперь в одном месте: для каждой
        # выбранной категории суммируем и наличные/карту из смен, и банк.
        cash_exp_by = defaultdict(lambda: defaultdict(float))
        bank_exp_by = defaultdict(lambda: defaultdict(float))
        if cfg['expense_cats']:
            ph_c = ','.join('?' * len(cfg['expense_cats']))
            for r in conn.execute(f"""
                SELECT {pe} AS period, e.category AS cat,
                       COALESCE(SUM(e.amount_cash + e.amount_card), 0) AS amount
                FROM expenses e JOIN shifts s ON s.id=e.shift_id
                WHERE s.date BETWEEN ? AND ?
                  AND e.category IN ({ph_c}) {bf}
                GROUP BY period, e.category
            """, [date_from, date_to] + cfg['expense_cats'] + b_args).fetchall():
                cash_exp_by[r['cat']][r['period']] += r['amount']
            for r in conn.execute(f"""
                SELECT {pe_bt} AS period,
                       bt.category AS cat_name,
                       SUM(-bt.amount) AS amount
                FROM bank_transactions bt
                WHERE bt.txn_date BETWEEN ? AND ?
                  AND bt.amount < 0 AND bt.is_ignored=0
                  AND bt.category IN ({ph_c}) {bf_bt}
                GROUP BY period, bt.category
                HAVING SUM(-bt.amount) > 0
            """, [date_from, date_to] + cfg['expense_cats'] + b_args).fetchall():
                bank_exp_by[r['cat_name']][r['period']] += r['amount']

            # Расход из сумм, извлечённых по паттернам правил разбора
            for r in conn.execute(f"""
                SELECT {pe_pat} AS period, pe.category AS cat_name, SUM(pe.amount) AS amount
                FROM bank_parse_extracted_entries pe
                WHERE pe.txn_date BETWEEN ? AND ? AND pe.direction='expense'
                  AND pe.category IN ({ph_c}) {bf_pe}
                GROUP BY period, pe.category
                HAVING SUM(pe.amount) > 0
            """, [date_from, date_to] + cfg['expense_cats'] + b_args).fetchall():
                bank_exp_by[r['cat_name']][r['period']] += r['amount']

        # Диагностика банковских данных
        _bt_debug = conn.execute("""
            SELECT COUNT(*) AS total,
                   SUM(CASE WHEN category IS NOT NULL AND category!='' THEN 1 ELSE 0 END) AS with_cat,
                   SUM(CASE WHEN amount>0 AND is_ignored=0 THEN 1 ELSE 0 END) AS income_rows,
                   SUM(CASE WHEN amount<0 AND is_ignored=0 THEN 1 ELSE 0 END) AS expense_rows,
                   SUM(CASE WHEN category IS NOT NULL AND category!='' AND is_ignored=0 AND amount<0 THEN 1 ELSE 0 END) AS cat_exp_active,
                   SUM(CASE WHEN category IS NOT NULL AND category!='' AND is_ignored=0 AND amount>0 THEN 1 ELSE 0 END) AS cat_inc_active,
                   SUM(CASE WHEN category IS NOT NULL AND category!='' AND is_ignored=1 THEN 1 ELSE 0 END) AS cat_ignored
            FROM bank_transactions
            WHERE txn_date BETWEEN ? AND ?
        """, [date_from, date_to]).fetchone()
        bt_debug = dict(_bt_debug) if _bt_debug else {}
        _bt_all = conn.execute("""
            SELECT COUNT(*) AS total_all,
                   MIN(txn_date) AS min_date, MAX(txn_date) AS max_date,
                   SUM(CASE WHEN category IS NOT NULL AND category!='' THEN 1 ELSE 0 END) AS with_cat_all
            FROM bank_transactions
        """).fetchone()
        bt_debug.update(dict(_bt_all) if _bt_all else {})
        _bt_cats = conn.execute("""
            SELECT DISTINCT category FROM bank_transactions
            WHERE category IS NOT NULL AND category != '' LIMIT 10
        """).fetchall()
        bt_debug['sample_cats'] = [r['category'] for r in _bt_cats]
        bt_debug['ctr_cats'] = [c['code'] for c in ctr_all_cats]

        # ФОТ по ролям
        sal_by_role = defaultdict(lambda: defaultdict(float))
        if cfg['include_salary']:
            for r in conn.execute(f"""
                SELECT {pe} AS period,
                       COALESCE(es.role_snapshot, 'other') AS role,
                       COALESCE(SUM(es.total_amount), 0) AS amount
                FROM employee_shifts es JOIN shifts s ON s.id=es.shift_id
                WHERE s.date BETWEEN ? AND ? {bf}
                GROUP BY period, role
            """, [date_from, date_to] + b_args).fetchall():
                sal_by_role[r['role']][r['period']] += r['amount']

        # ── ПРОСТОЙ P&L ──────────────────────────────────────────────────────
        # Наличные расходы в сменах — ВСЕ категории (в отличие от основного P&L,
        # где считаются только выбранные в настройках).
        simple_cash_exp_by_p = {}
        for r in conn.execute(f"""
            SELECT {pe} AS period, COALESCE(SUM(e.amount_cash),0) AS amount
            FROM expenses e JOIN shifts s ON s.id=e.shift_id
            WHERE s.date BETWEEN ? AND ? {bf}
            GROUP BY period
        """, [date_from, date_to] + b_args).fetchall():
            simple_cash_exp_by_p[r['period']] = r['amount']

        # ФОТ — всегда целиком, независимо от настройки include_salary основного P&L
        simple_fot_by_p = {}
        for r in conn.execute(f"""
            SELECT {pe} AS period, COALESCE(SUM(es.total_amount),0) AS amount
            FROM employee_shifts es JOIN shifts s ON s.id=es.shift_id
            WHERE s.date BETWEEN ? AND ? {bf}
            GROUP BY period
        """, [date_from, date_to] + b_args).fetchall():
            simple_fot_by_p[r['period']] = r['amount']

        # Такси, оплаченное наличными
        simple_taxi_cash_by_p = {}
        for r in conn.execute(f"""
            SELECT {pe} AS period, COALESCE(SUM(tt.amount),0) AS amount
            FROM taxi_trips tt JOIN shifts s ON s.id=tt.shift_id
            WHERE s.date BETWEEN ? AND ? AND tt.payment_type='cash' {bf}
            GROUP BY period
        """, [date_from, date_to] + b_args).fetchall():
            simple_taxi_cash_by_p[r['period']] = r['amount']

        # Выписки банковские — весь приход/расход как есть, кроме исключённых
        # в настройках категорий (например «Перераспределение средств»).
        # Категория берётся эффективная (ручная ИЛИ подставленная правилом разбора
        # банка — bt.category у таких транзакций пустой, сама категория есть только
        # в parse_rule_category), иначе исключение не срабатывало бы для операций,
        # категоризированных только правилом — тот же баг, что чинили в выписке (п.157).
        simple_excl_cats = set(cfg['simple_pnl_excluded_cats'])
        simple_bt_rows = conn.execute(f"""
            SELECT {pe_bt} AS period, bt.amount, bt.category, bt.description, bt.bank_account_id
            FROM bank_transactions bt
            WHERE bt.txn_date BETWEEN ? AND ? AND bt.is_ignored=0 {bf_bt}
        """, [date_from, date_to] + b_args).fetchall()
        simple_bt_txns = [dict(row) for row in simple_bt_rows]
        _apply_bank_parse_rules(conn, simple_bt_txns)

        simple_bank_income_by_p = {}
        simple_bank_exp_by_cat = defaultdict(lambda: defaultdict(float))
        for d in simple_bt_txns:
            eff_cat = d.get('category') or d.get('parse_rule_category') or ''
            if eff_cat in simple_excl_cats:
                continue
            if d['amount'] > 0:
                simple_bank_income_by_p[d['period']] = simple_bank_income_by_p.get(d['period'], 0.0) + d['amount']
            elif d['amount'] < 0:
                simple_bank_exp_by_cat[eff_cat][d['period']] += -d['amount']

    # Все периоды
    all_p = set(total_rev_by_p) | set(cash_rev_by_p)
    for d in list(bank_inc_by.values()) + list(cash_exp_by.values()) + list(bank_exp_by.values()):
        all_p.update(d)
    for d in sal_by_role.values():
        all_p.update(d)
    all_p.update(simple_cash_exp_by_p, simple_fot_by_p, simple_taxi_cash_by_p, simple_bank_income_by_p)
    for d in simple_bank_exp_by_cat.values():
        all_p.update(d)
    periods = sorted(all_p) if group_by == 'month' else ['total']

    cat_map = {c['code']: c['label'] for c in all_cats}

    def _row(label, by_p, badges=None, sub=False):
        if isinstance(badges, str):
            badges = [badges]
        r = {'label': label, 'amounts': {}, 'total': 0.0, 'badges': badges or [], 'sub': sub}
        for p in periods:
            v = by_p.get(p, 0.0)
            r['amounts'][p] = v
            r['total'] += v
        return r

    # ── ДОХОДЫ ────────────────────────────────────────────────────────────────
    inc_totals = {}
    inc_rows   = [_row('Наличные', cash_rev_by_p, 'cash')]

    for cat_name, by_p in sorted(bank_inc_by.items()):
        inc_rows.append(_row(cat_name, dict(by_p), 'bank'))

    for row in inc_rows:
        for p in periods:
            inc_totals[p] = inc_totals.get(p, 0.0) + row['amounts'].get(p, 0.0)

    inc_grand = sum(inc_totals.get(p, 0) for p in periods)

    # ── РАСХОДЫ ───────────────────────────────────────────────────────────────
    exp_totals = {}
    exp_rows   = []

    # ФОТ
    if cfg['include_salary']:
        sal_total_by_p = {}
        for role, by_p in sal_by_role.items():
            for p, v in by_p.items():
                sal_total_by_p[p] = sal_total_by_p.get(p, 0.0) + v
        sal_total_row = _row('ФОТ (итого)', sal_total_by_p, 'salary')
        exp_rows.append(sal_total_row)

        if cfg['include_salary_breakdown']:
            role_order = ['admin', 'cook', 'sushi', 'courier', 'packer', 'cleaner']
            for role in role_order:
                if role in sal_by_role:
                    exp_rows.append(_row(ROLE_LABELS.get(role, role), dict(sal_by_role[role]), 'salary', sub=True))
            # прочие роли не в списке
            for role, by_p in sal_by_role.items():
                if role not in role_order:
                    exp_rows.append(_row(ROLE_LABELS.get(role, role), dict(by_p), 'salary', sub=True))

    # Расходы по категориям — наличные/карта из смен и банк объединены в одну строку
    for cat_code in cfg['expense_cats']:
        cash_by_p = cash_exp_by.get(cat_code, {})
        bank_by_p = bank_exp_by.get(cat_code, {})
        if not cash_by_p and not bank_by_p:
            continue
        merged_by_p = defaultdict(float)
        for p, v in cash_by_p.items():
            merged_by_p[p] += v
        for p, v in bank_by_p.items():
            merged_by_p[p] += v
        if not any(v > 0 for v in merged_by_p.values()):
            continue
        badges = []
        if any(v > 0 for v in cash_by_p.values()):
            badges.append('cash')
        if any(v > 0 for v in bank_by_p.values()):
            badges.append('bank')
        exp_rows.append(_row(cat_map.get(cat_code, cat_code), dict(merged_by_p), badges))

    for row in exp_rows:
        if not row['sub']:
            for p in periods:
                exp_totals[p] = exp_totals.get(p, 0.0) + row['amounts'].get(p, 0.0)

    exp_grand = sum(exp_totals.get(p, 0) for p in periods)

    profit_by_p   = {p: inc_totals.get(p, 0) - exp_totals.get(p, 0) for p in periods}
    profit_grand  = inc_grand - exp_grand
    total_rev_grand = sum(total_rev_by_p.get(p, 0) for p in periods)
    period_labels = {p: _pnl_period_label(p) for p in periods}

    # ── ПРОСТОЙ P&L: строки для шаблона ──────────────────────────────────────
    s_income_cash_row = _row('Наличные (выручка)', cash_rev_by_p)
    s_exp_shift_row = _row('Расходы наличными в сменах', simple_cash_exp_by_p)
    s_exp_fot_row   = _row('ФОТ', simple_fot_by_p)
    s_exp_taxi_row  = _row('Такси наличными', simple_taxi_cash_by_p)
    s_cash_exp_rows = [s_exp_shift_row, s_exp_fot_row, s_exp_taxi_row]
    s_cash_exp_total_by_p = {}
    for row in s_cash_exp_rows:
        for p in periods:
            s_cash_exp_total_by_p[p] = s_cash_exp_total_by_p.get(p, 0.0) + row['amounts'].get(p, 0.0)
    s_cash_exp_total_row = _row('Итого расходы наличными', s_cash_exp_total_by_p)

    s_bank_income_row = _row('Приход по выпискам', simple_bank_income_by_p)

    s_bank_exp_rows = []
    for cat_code, by_p in sorted(simple_bank_exp_by_cat.items(),
                                  key=lambda kv: cat_map.get(kv[0], kv[0]) if kv[0] else 'яяя'):
        label = cat_map.get(cat_code, cat_code) if cat_code else 'Без категории'
        s_bank_exp_rows.append(_row(label, dict(by_p)))
    s_bank_exp_total_by_p = {}
    for row in s_bank_exp_rows:
        for p in periods:
            s_bank_exp_total_by_p[p] = s_bank_exp_total_by_p.get(p, 0.0) + row['amounts'].get(p, 0.0)
    s_bank_exp_total_row = _row('Итого расход по выпискам', s_bank_exp_total_by_p)

    s_total_income_by_p  = {p: s_income_cash_row['amounts'][p] + s_bank_income_row['amounts'][p] for p in periods}
    s_total_expense_by_p = {p: s_cash_exp_total_row['amounts'][p] + s_bank_exp_total_row['amounts'][p] for p in periods}
    s_total_income_row  = _row('ИТОГО ПРИХОД', s_total_income_by_p)
    s_total_expense_row = _row('ИТОГО РАСХОД', s_total_expense_by_p)
    s_profit_by_p = {p: s_total_income_by_p[p] - s_total_expense_by_p[p] for p in periods}
    s_profit_row  = _row('ПРИБЫЛЬ', s_profit_by_p)
    s_profit_grand = s_total_income_row['total'] - s_total_expense_row['total']

    return render_template('pnl.html',
        periods=periods, period_labels=period_labels,
        inc_rows=inc_rows, inc_totals=inc_totals, inc_grand=inc_grand,
        exp_rows=exp_rows, exp_totals=exp_totals, exp_grand=exp_grand,
        profit_by_p=profit_by_p, profit_grand=profit_grand,
        total_rev_by_p=total_rev_by_p, total_rev_grand=total_rev_grand,
        cfg=cfg,
        branches=branches, branch_groups=branch_groups,
        all_cats=all_cats, cat_map=cat_map,
        ctr_cats_income=ctr_cats_income, expense_cats_all=expense_cats_all,
        role_labels=ROLE_LABELS,
        date_from=date_from, date_to=date_to,
        branch_ids=branch_ids, group_by=group_by,
        bt_debug=bt_debug,
        s_income_cash_row=s_income_cash_row,
        s_cash_exp_rows=s_cash_exp_rows, s_cash_exp_total_row=s_cash_exp_total_row,
        s_bank_income_row=s_bank_income_row,
        s_bank_exp_rows=s_bank_exp_rows, s_bank_exp_total_row=s_bank_exp_total_row,
        s_total_income_row=s_total_income_row, s_total_expense_row=s_total_expense_row,
        s_profit_row=s_profit_row, s_profit_grand=s_profit_grand,
    )


@app.route('/report/pnl/settings/simple', methods=['POST'])
@login_required
@menu_permission_required('pnl_report')
def pnl_settings_simple_save():
    excluded = request.form.getlist('simple_excluded_cats')
    with get_db() as conn:
        conn.execute(
            'INSERT OR REPLACE INTO pnl_settings (key, value) VALUES (?, ?)',
            ('simple_pnl_excluded_cats', _json.dumps(excluded))
        )
        conn.commit()
    flash('Настройки простого P&L сохранены', 'success')
    params = {k: request.form.get(k) for k in ('date_from', 'date_to', 'group_by') if request.form.get(k)}
    bids = request.form.getlist('branch_ids')
    if bids:
        params['branch_ids'] = bids
    return redirect(url_for('pnl_report', **params))



# ─── GSHEET SETTINGS ──────────────────────────────────────────────────────────

@app.route('/settings/gsheet', methods=['GET', 'POST'])
@login_required
@menu_permission_required('gsheet_settings')
def gsheet_settings():
    with get_db() as conn:
        if request.method == 'POST':
            for key, _ in GSHEET_COLS:
                val = '1' if request.form.get(f'col_{key}') else '0'
                conn.execute(
                    'INSERT OR REPLACE INTO gsheet_settings (key, value) VALUES (?,?)',
                    (f'col_{key}', val)
                )
            conn.commit()
            flash('Настройки экспорта сохранены', 'success')
            return redirect(url_for('gsheet_settings'))
        cfg = _gsheet_load_settings(conn)
        row = conn.execute(
            "SELECT value FROM gsheet_settings WHERE key='gdrive_refresh_token'"
        ).fetchone()
        gdrive_authorized = bool(row)
        backup_row = conn.execute(
            "SELECT value FROM gsheet_settings WHERE key='last_db_backup'"
        ).fetchone()
        last_db_backup = backup_row['value'] if backup_row else None
    return render_template('gsheet_settings.html', cfg=cfg, cols=GSHEET_COLS,
                           sheet_id=os.environ.get('GOOGLE_SHEET_ID', ''),
                           drive_folder_id=os.environ.get('GOOGLE_DRIVE_FOLDER_ID', ''),
                           has_creds=bool(os.environ.get('GOOGLE_CREDENTIALS_JSON')),
                           gdrive_authorized=gdrive_authorized,
                           last_db_backup=last_db_backup)


@app.route('/settings/gdrive-auth')
@login_required
@menu_permission_required('gsheet_settings')
def gdrive_auth():
    import urllib.parse as _up
    client_id = os.environ.get('GOOGLE_OAUTH_CLIENT_ID')
    if not client_id:
        flash('Добавьте GOOGLE_OAUTH_CLIENT_ID в Railway Variables', 'danger')
        return redirect(url_for('gsheet_settings'))
    params = {
        'client_id':     client_id,
        'redirect_uri':  url_for('gdrive_callback', _external=True),
        'scope':         'https://www.googleapis.com/auth/drive',
        'response_type': 'code',
        'access_type':   'offline',
        'prompt':        'consent',
    }
    return redirect('https://accounts.google.com/o/oauth2/v2/auth?' + _up.urlencode(params))


@app.route('/settings/gdrive-callback')
@login_required
@menu_permission_required('gsheet_settings')
def gdrive_callback():
    import urllib.request as _ur
    import urllib.parse as _up
    code  = request.args.get('code')
    error = request.args.get('error')
    if error or not code:
        flash(f'Ошибка авторизации Google: {error or "нет кода"}', 'danger')
        return redirect(url_for('gsheet_settings'))
    client_id     = os.environ.get('GOOGLE_OAUTH_CLIENT_ID')
    client_secret = os.environ.get('GOOGLE_OAUTH_CLIENT_SECRET')
    data = _up.urlencode({
        'client_id':     client_id,
        'client_secret': client_secret,
        'code':          code,
        'redirect_uri':  url_for('gdrive_callback', _external=True),
        'grant_type':    'authorization_code',
    }).encode()
    try:
        req = _ur.Request(
            'https://oauth2.googleapis.com/token',
            data=data,
            headers={'Content-Type': 'application/x-www-form-urlencoded'},
        )
        with _ur.urlopen(req, timeout=15) as resp:
            tokens = _json_lib.loads(resp.read())
    except Exception as e:
        flash(f'Ошибка получения токена: {e}', 'danger')
        return redirect(url_for('gsheet_settings'))
    refresh_token = tokens.get('refresh_token')
    if not refresh_token:
        flash('Google не вернул refresh_token — попробуйте ещё раз.', 'danger')
        return redirect(url_for('gsheet_settings'))
    with get_db() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO gsheet_settings (key, value) VALUES ('gdrive_refresh_token', ?)",
            (refresh_token,)
        )
        conn.commit()
    flash('✅ Google Drive авторизован! Теперь xlsx будут загружаться автоматически.', 'success')
    return redirect(url_for('gsheet_settings'))


@app.route('/settings/gdrive-test', methods=['POST'])
@login_required
@menu_permission_required('gsheet_settings')
def gdrive_test():
    """Пошаговая диагностика подключения к Google Drive."""
    steps = []
    def ok(msg):  steps.append({'status': 'ok',    'msg': msg})
    def err(msg): steps.append({'status': 'error', 'msg': msg})

    folder_id = os.environ.get('GOOGLE_DRIVE_FOLDER_ID')

    if not folder_id:
        err('GOOGLE_DRIVE_FOLDER_ID не задана в Railway')
        return jsonify({'steps': steps})
    ok(f'GOOGLE_DRIVE_FOLDER_ID = {folder_id}')

    token = _gdrive_get_oauth_token()
    if not token:
        err('OAuth2 токен не получен — нажмите «Авторизовать Google Drive» на этой странице')
        return jsonify({'steps': steps})
    ok('OAuth2 токен получен успешно')

    try:
        import urllib.request as _ur
        import json as _j
        list_req = _ur.Request(
            f'https://www.googleapis.com/drive/v3/files?q=%27{folder_id}%27+in+parents&pageSize=1&fields=files(id,name)',
            headers={'Authorization': f'Bearer {token}'}
        )
        with _ur.urlopen(list_req, timeout=15) as resp:
            data = _j.loads(resp.read())
        count = len(data.get('files', []))
        ok(f'Доступ к папке есть (найдено файлов: {count})')
    except Exception as e:
        err(f'Нет доступа к папке Drive: {e}')
        return jsonify({'steps': steps})

    try:
        import urllib.request as _ur
        import json as _j
        test_content = b'CRM PAPA gdrive test'
        boundary = 'GDriveTestBnd1234'
        meta = _j.dumps({'name': '_crm_test_.txt', 'parents': [folder_id]})
        body = (
            f'--{boundary}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n'
            f'{meta}\r\n'
            f'--{boundary}\r\nContent-Type: text/plain\r\n\r\n'
        ).encode() + test_content + f'\r\n--{boundary}--'.encode()
        upload_req = _ur.Request(
            'https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart',
            data=body,
            headers={
                'Authorization': f'Bearer {token}',
                'Content-Type': f'multipart/related; boundary="{boundary}"',
            },
            method='POST'
        )
        with _ur.urlopen(upload_req, timeout=30) as resp:
            result = _j.loads(resp.read())
        file_id = result.get('id', '?')
        ok(f'Тестовый файл загружен в папку (id={file_id})')
        # Delete test file
        del_req = _ur.Request(
            f'https://www.googleapis.com/drive/v3/files/{file_id}',
            headers={'Authorization': f'Bearer {token}'},
            method='DELETE'
        )
        try:
            _ur.urlopen(del_req, timeout=10)
            ok('Тестовый файл удалён')
        except Exception:
            ok(f'Тестовый файл (_crm_test_.txt) остался в папке — удалите вручную')
    except Exception as e:
        import urllib.error as _ue
        if isinstance(e, _ue.HTTPError):
            try:
                body = e.read().decode('utf-8', errors='replace')
            except Exception:
                body = '(не удалось прочитать тело ответа)'
            err(f'Ошибка при загрузке файла: {e} — {body[:600]}')
        else:
            err(f'Ошибка при загрузке файла: {e}')
        return jsonify({'steps': steps})

    return jsonify({'steps': steps})


def _backup_db_to_gdrive():
    """Загрузить копию crm.db в Google Drive. Возвращает (ok: bool, msg: str)."""
    import urllib.request as _ur
    import datetime as _dt

    token = _gdrive_get_oauth_token()
    if not token:
        return False, 'Google Drive не авторизован — откройте Настройки → Google Sheets'

    folder_id = os.environ.get('GOOGLE_DRIVE_FOLDER_ID')
    if not folder_id:
        return False, 'GOOGLE_DRIVE_FOLDER_ID не задан в Railway'

    db_path = DATABASE
    if not os.path.exists(db_path):
        return False, f'Файл базы данных не найден: {db_path}'

    # Горячий снимок через SQLite backup API — гарантирует целостность файла
    import io as _io
    import tempfile as _tmp
    with _tmp.NamedTemporaryFile(suffix='.db', delete=False) as tf:
        tmp_path = tf.name
    try:
        src = sqlite3.connect(DATABASE)
        dst = sqlite3.connect(tmp_path)
        src.backup(dst)
        dst.close()
        src.close()
        with open(tmp_path, 'rb') as f:
            db_bytes = f.read()
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass

    date_str = _dt.datetime.now().strftime('%Y-%m-%d_%H-%M')
    filename = f'crm_backup_{date_str}.db'
    boundary = '---CrmDbBackupBnd8821'
    meta_json = _json_lib.dumps({'name': filename, 'parents': [folder_id]})
    body = (
        f'--{boundary}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n'
        f'{meta_json}\r\n'
        f'--{boundary}\r\nContent-Type: application/octet-stream\r\n\r\n'
    ).encode('utf-8') + db_bytes + f'\r\n--{boundary}--'.encode('utf-8')

    req = _ur.Request(
        'https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart',
        data=body,
        headers={
            'Authorization': f'Bearer {token}',
            'Content-Type': f'multipart/related; boundary="{boundary}"',
        },
        method='POST'
    )
    with _ur.urlopen(req, timeout=60) as resp:
        result = _json_lib.loads(resp.read())

    file_id = result.get('id', '?')
    size_kb = len(db_bytes) // 1024
    return True, f'Загружено: {filename} ({size_kb} КБ)'


@app.route('/settings/backup-db', methods=['POST'])
@login_required
@owner_required
def backup_db():
    """Ручной запуск резервного копирования базы в Google Drive."""
    try:
        ok, msg = _backup_db_to_gdrive()
        if ok:
            with get_db() as conn:
                import datetime as _dt
                conn.execute(
                    "INSERT OR REPLACE INTO gsheet_settings (key, value) VALUES ('last_db_backup', ?)",
                    (_dt.datetime.now().strftime('%d.%m.%Y %H:%M'),)
                )
                conn.commit()
            return jsonify({'status': 'ok', 'msg': msg})
        else:
            return jsonify({'status': 'error', 'msg': msg}), 400
    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)}), 500


# ─── HISTORY ──────────────────────────────────────────────────────────────────

@app.route('/history')
@login_required
def history():
    role = session.get('role')
    if not item_visible('history'):
        flash('Доступ запрещён', 'danger')
        return redirect(url_for('dashboard'))
    today = date.today().isoformat()
    week_ago = (date.today() - timedelta(days=7)).isoformat()
    date_from = request.args.get('date_from', week_ago)
    date_to = request.args.get('date_to', today)
    branch_id = request.args.get('branch_id', '')
    user_filter = request.args.get('user_id', '')
    action_filter = request.args.get('action', '')

    with get_db() as conn:
        # created_at хранится в UTC (SQLite CURRENT_TIMESTAMP) — сравниваем границы
        # диапазона (локальные даты НСК) с created_at, переведённым в Asia/Novosibirsk (+7ч)
        conds = ["datetime(cl.created_at, '+7 hours') >= ? AND datetime(cl.created_at, '+7 hours') <= ?"]
        params = [date_from + ' 00:00:00', date_to + ' 23:59:59']

        if role != 'owner':
            conds.append('cl.user_id = ?')
            params.append(session['user_id'])
            hist_bids = get_effective_branch_ids('history', [branch_id] if branch_id.isdigit() else [])
            if hist_bids:
                ph = ','.join('?' * len(hist_bids))
                conds.append(f'cl.branch_id IN ({ph})')
                params.extend(int(b) for b in hist_bids)
        else:
            if branch_id:
                conds.append('cl.branch_id = ?')
                params.append(int(branch_id))
            if user_filter:
                conds.append('cl.user_id = ?')
                params.append(int(user_filter))

        if action_filter:
            conds.append('cl.action = ?')
            params.append(action_filter)

        where = ' AND '.join(conds)
        logs = conn.execute(f'''
            SELECT cl.*
            FROM change_log cl
            WHERE {where}
            ORDER BY cl.created_at DESC
            LIMIT 300
        ''', params).fetchall()

        branches = []
        users = []
        if role == 'owner':
            branches = conn.execute(
                'SELECT * FROM branches WHERE is_active=1 ORDER BY name'
            ).fetchall()
            users = conn.execute(
                "SELECT id, full_name FROM users WHERE role != 'owner' ORDER BY full_name"
            ).fetchall()
        elif can_pick_other_branches('history'):
            branches = conn.execute(
                'SELECT * FROM branches WHERE is_active=1 ORDER BY name'
            ).fetchall()

    return render_template('history.html',
        logs=logs, branches=branches, users=users,
        date_from=date_from, date_to=date_to,
        selected_branch=branch_id, selected_user=user_filter,
        selected_action=action_filter,
        action_labels=ACTION_LABELS,
        is_owner=(role == 'owner'),
        branch_groups=get_branch_groups(conn))


# ─── API ──────────────────────────────────────────────────────────────────────

@app.route('/api/employee/<int:emp_id>')
@login_required
def api_employee(emp_id):
    with get_db() as conn:
        emp = conn.execute('SELECT * FROM employees WHERE id=?', (emp_id,)).fetchone()
        if not emp:
            return jsonify({}), 404
        roles = conn.execute(
            'SELECT * FROM employee_roles WHERE employee_id=? AND is_active=1 ORDER BY role',
            (emp_id,)
        ).fetchall()
        return jsonify({
            'id': emp['id'],
            'full_name': emp['full_name'],
            'role': emp['role'],
            'rate': emp['rate'],
            'rate_per_km': emp['rate_per_km'],
            'rate_per_order': emp['rate_per_order'],
            'extra_roles': [
                {'role': r['role'], 'rate': r['rate'],
                 'rate_per_km': r['rate_per_km'], 'rate_per_order': r['rate_per_order']}
                for r in roles
            ],
        })


@app.route('/employees/<int:emp_id>/roles/add', methods=['POST'])
@login_required
@menu_permission_required('employees')
def add_employee_role(emp_id):
    role = request.form.get('role', '').strip()
    rate_template_id = request.form.get('rate_template_id', '').strip() or None
    if rate_template_id:
        rate_template_id = int(rate_template_id)
    pay_monthly = 1 if request.form.get('pay_monthly') else 0
    if not role or role not in ROLE_LABELS:
        flash('Выберите должность', 'danger')
        return redirect(url_for('employees'))
    with get_db() as conn:
        if rate_template_id:
            tmpl = conn.execute('SELECT * FROM rate_templates WHERE id=?', (rate_template_id,)).fetchone()
            rate = float(tmpl['rate'] or 0) if tmpl else 0.0
            rate_km = float(tmpl['rate_per_km'] or 0) if tmpl else 0.0
            rate_ord = float(tmpl['rate_per_order'] or 0) if tmpl else 0.0
        else:
            rate = float(request.form.get('rate', 0) or 0)
            rate_km = float(request.form.get('rate_per_km', 0) or 0)
            rate_ord = float(request.form.get('rate_per_order', 0) or 0)
        try:
            conn.execute(
                'INSERT INTO employee_roles (employee_id, role, rate, rate_per_km, rate_per_order, rate_template_id, pay_monthly) VALUES (?,?,?,?,?,?,?)',
                (emp_id, role, rate, rate_km, rate_ord, rate_template_id, pay_monthly)
            )
            _seed_pay_monthly_history(conn, emp_id, role, pay_monthly)
            conn.commit()
            flash('Должность добавлена', 'success')
        except Exception:
            old_row = conn.execute(
                'SELECT pay_monthly FROM employee_roles WHERE employee_id=? AND role=?', (emp_id, role)
            ).fetchone()
            old_pm = old_row['pay_monthly'] if old_row else 0
            conn.execute(
                'UPDATE employee_roles SET rate=?, rate_per_km=?, rate_per_order=?, rate_template_id=?, pay_monthly=? WHERE employee_id=? AND role=?',
                (rate, rate_km, rate_ord, rate_template_id, pay_monthly, emp_id, role)
            )
            _log_pay_monthly_change(conn, emp_id, role, old_pm, pay_monthly, date.today().isoformat())
            conn.commit()
            flash('Ставка по должности обновлена', 'success')
    return redirect(url_for('employees'))


@app.route('/employees/roles/<int:role_id>/delete', methods=['POST'])
@login_required
@menu_permission_required('employees')
def delete_employee_role(role_id):
    with get_db() as conn:
        conn.execute('DELETE FROM employee_roles WHERE id=?', (role_id,))
        conn.commit()
    flash('Должность удалена', 'success')
    return redirect(url_for('employees'))


@app.route('/employees/roles/<int:role_id>/update-template', methods=['POST'])
@login_required
def update_employee_role_template(role_id):
    rate_template_id = request.form.get('rate_template_id', '').strip() or None
    if rate_template_id:
        rate_template_id = int(rate_template_id)
    with get_db() as conn:
        if rate_template_id:
            tmpl = conn.execute('SELECT * FROM rate_templates WHERE id=?', (rate_template_id,)).fetchone()
            if tmpl:
                conn.execute(
                    'UPDATE employee_roles SET rate=?, rate_per_km=?, rate_per_order=?, rate_template_id=? WHERE id=?',
                    (float(tmpl['rate'] or 0), float(tmpl['rate_per_km'] or 0), float(tmpl['rate_per_order'] or 0), rate_template_id, role_id)
                )
        else:
            conn.execute('UPDATE employee_roles SET rate_template_id=NULL WHERE id=?', (role_id,))
        conn.commit()
    flash('Ставка по должности обновлена', 'success')
    return redirect(url_for('employees'))


@app.route('/employees/roles/<int:role_id>/pay-monthly', methods=['POST'])
@login_required
@menu_permission_required('employees')
def update_employee_role_pay_monthly(role_id):
    pay_monthly = 1 if request.form.get('pay_monthly') else 0
    pay_monthly_from = request.form.get('pay_monthly_from') or date.today().isoformat()
    if pay_monthly_from < date.today().isoformat():
        pay_monthly_from = date.today().isoformat()  # нельзя назначить датой в прошлом
    pm_branch_ids = [bid for bid in request.form.getlist('pm_branch_ids') if bid.isdigit()]
    with get_db() as conn:
        er = conn.execute('SELECT employee_id, role, pay_monthly FROM employee_roles WHERE id=?', (role_id,)).fetchone()
        conn.execute('UPDATE employee_roles SET pay_monthly=? WHERE id=?', (pay_monthly, role_id))
        if er:
            _log_pay_monthly_change(conn, er['employee_id'], er['role'], er['pay_monthly'], pay_monthly, pay_monthly_from)
            conn.execute(
                'DELETE FROM employee_pay_monthly_branches WHERE employee_id=? AND role=?',
                (er['employee_id'], er['role'])
            )
            for bid in pm_branch_ids:
                conn.execute(
                    'INSERT OR IGNORE INTO employee_pay_monthly_branches (employee_id, role, branch_id) VALUES (?,?,?)',
                    (er['employee_id'], er['role'], int(bid))
                )
        conn.commit()
    return redirect(url_for('employees'))


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def _session_branch_ids():
    ids = session.get('branch_ids')
    if ids:
        return ids
    bid = session.get('branch_id')
    return [bid] if bid else []


def _can_edit_shift(shift_id):
    role = session.get('role')
    if role == 'owner':
        return True
    with get_db() as conn:
        shift = conn.execute('SELECT * FROM shifts WHERE id=?', (shift_id,)).fetchone()
        if not shift:
            return False
        return shift['status'] == 'open' and shift['branch_id'] in _session_branch_ids()


def _f(data, key):
    try:
        return float(data.get(key) or 0)
    except (ValueError, TypeError):
        return 0.0


def _i(data, key):
    try:
        return int(data.get(key) or 0)
    except (ValueError, TypeError):
        return 0


def _normalize_phone(raw):
    """Только цифры, приведённые к 11-значному номеру с ведущей 8 (89000000000)."""
    digits = re.sub(r'\D', '', raw or '')
    if not digits:
        return ''
    if len(digits) == 11 and digits[0] in ('7', '8'):
        digits = '8' + digits[1:]
    elif len(digits) == 10:
        digits = '8' + digits
    return digits


def _format_phone(raw):
    """8-900-000-0000; если не удалось распознать 11 цифр — возвращает как есть."""
    digits = _normalize_phone(raw)
    if len(digits) != 11:
        return (raw or '').strip()
    return f'{digits[0]}-{digits[1:4]}-{digits[4:7]}-{digits[7:11]}'


def calculate_bonuses(conn, shift_id):
    """Recalculate auto_bonus for all staff in a shift based on bonus_rules."""
    rev = conn.execute(
        'SELECT total_revenue FROM shift_revenue WHERE shift_id=?', (shift_id,)
    ).fetchone()
    total_revenue = float((rev['total_revenue'] or 0) if rev else 0)

    shift = conn.execute('SELECT branch_id FROM shifts WHERE id=?', (shift_id,)).fetchone()
    branch_id = shift['branch_id'] if shift else None

    staff = conn.execute(
        'SELECT id, role_snapshot, hours_worked, base_pay, bonus_amount, penalty_amount '
        'FROM employee_shifts WHERE shift_id=?', (shift_id,)
    ).fetchall()
    if not staff:
        return []

    rules = conn.execute(
        'SELECT role, threshold_pct, bonus_pct FROM bonus_rules '
        'WHERE is_active=1 AND (branch_id IS NULL OR branch_id=?) '
        'ORDER BY threshold_pct ASC', (branch_id,)
    ).fetchall()

    rules_by_role = {}
    for r in rules:
        rules_by_role.setdefault(r['role'], []).append((r['threshold_pct'], r['bonus_pct']))

    staff_by_role = {}
    for s in staff:
        staff_by_role.setdefault(s['role_snapshot'], []).append(s)

    auto_bonus_per_id = {}
    for role, role_rules in rules_by_role.items():
        role_staff = staff_by_role.get(role, [])
        if not role_staff or total_revenue <= 0:
            continue
        role_payroll = sum(float(s['base_pay'] or 0) for s in role_staff)
        payroll_pct = role_payroll / total_revenue * 100
        applicable_pct = 0
        for threshold, bonus_pct in role_rules:
            if payroll_pct < threshold:
                applicable_pct = bonus_pct
                break
        if applicable_pct <= 0:
            continue
        bonus_pool = total_revenue * applicable_pct / 100
        total_hours = sum(float(s['hours_worked'] or 0) for s in role_staff)
        for s in role_staff:
            h = float(s['hours_worked'] or 0)
            auto_bonus_per_id[s['id']] = round(bonus_pool * h / total_hours) if total_hours > 0 else 0

    results = []
    for s in staff:
        ab = auto_bonus_per_id.get(s['id'], 0)
        new_total = round(float(s['base_pay'] or 0) + float(s['bonus_amount'] or 0) + ab - float(s['penalty_amount'] or 0))
        conn.execute(
            'UPDATE employee_shifts SET auto_bonus=?, total_amount=?, '
            'paid_amount = CASE WHEN is_paid=1 THEN ? ELSE paid_amount END '
            'WHERE id=?',
            (ab, new_total, new_total, s['id'])
        )
        results.append({'id': s['id'], 'auto_bonus': ab, 'total_amount': new_total})
    return results


def _fmt_money(v):
    try:
        return f"{float(v or 0):,.0f}".replace(',', ' ') + ' ₽'
    except Exception:
        return '0 ₽'


def log_action(conn, action, description, shift_id=None, entity_id=None, upsert_by_shift=False):
    if not session.get('user_id'):
        return
    branch_id = branch_name = shift_date = None
    if shift_id:
        row = conn.execute(
            'SELECT s.branch_id, b.name, s.date FROM shifts s JOIN branches b ON b.id=s.branch_id WHERE s.id=?',
            (shift_id,)
        ).fetchone()
        if row:
            branch_id, branch_name, shift_date = row['branch_id'], row['name'], row['date']
    if upsert_by_shift and shift_id:
        existing = conn.execute(
            'SELECT id FROM change_log WHERE shift_id=? AND action=? ORDER BY id DESC LIMIT 1',
            (shift_id, action)
        ).fetchone()
        if existing:
            conn.execute(
                'UPDATE change_log SET description=?, user_id=?, user_name=?, created_at=CURRENT_TIMESTAMP WHERE id=?',
                (description, session['user_id'], session.get('full_name', ''), existing['id'])
            )
            return
    conn.execute('''
        INSERT INTO change_log
            (user_id, user_name, action, entity_id, shift_id, branch_id, branch_name, shift_date, description)
        VALUES (?,?,?,?,?,?,?,?,?)
    ''', (session['user_id'], session.get('full_name', ''), action, entity_id,
          shift_id, branch_id, branch_name, shift_date, description))


@app.template_filter('money')
def money_filter(value):
    try:
        return '{:,.0f}'.format(float(value or 0)).replace(',', ' ')
    except Exception:
        return str(value)


@app.template_filter('money_cents')
def money_cents_filter(value):
    try:
        v = float(value or 0)
        frac = round(abs(v) % 1, 2)
        if frac > 0:
            whole = '{:,.0f}'.format(abs(int(v))).replace(',', ' ')
            cents = '{:02d}'.format(round(frac * 100))
            sign = '-' if v < 0 else ''
            return f'{sign}{whole},{cents}'
        return '{:,.0f}'.format(v).replace(',', ' ')
    except Exception:
        return str(value)


@app.template_filter('moneyval')
def moneyval_filter(value):
    """Для значения в input: целое → '123', с копейками → '123.45'"""
    try:
        v = float(value or 0)
        if v == int(v):
            return str(int(v))
        return f'{v:.2f}'
    except Exception:
        return str(value)


@app.template_filter('datetime_fmt')
def datetime_fmt(value):
    if not value:
        return ''
    try:
        # created_at хранится через SQLite CURRENT_TIMESTAMP (всегда UTC) —
        # переводим в Asia/Novosibirsk (UTC+7, без перехода на летнее время)
        dt = datetime.strptime(str(value)[:19], '%Y-%m-%d %H:%M:%S') + timedelta(hours=7)
        return dt.strftime('%d.%m %H:%M')
    except Exception:
        return value


@app.template_filter('date_fmt')
def date_fmt(value):
    if not value:
        return ''
    try:
        dt = datetime.strptime(str(value)[:10], '%Y-%m-%d')
        return dt.strftime('%d.%m.%Y')
    except Exception:
        return value


@app.context_processor
def inject_globals():
    return {'now': datetime.now}


# ─── SHIFTS ARCHIVE ───────────────────────────────────────────────────────────

@app.route('/shifts')
@login_required
def shifts_archive():
    role = session.get('role')
    if not item_visible('shifts_archive'):
        flash('Доступ запрещён', 'danger')
        return redirect(url_for('dashboard'))
    today = date.today().isoformat()
    month_start = date.today().replace(day=1).isoformat()

    date_from  = request.args.get('date_from', month_start)
    date_to    = request.args.get('date_to',   today)
    branch_ids = [bid for bid in request.args.getlist('branch_ids') if bid.isdigit()]
    status_filter = request.args.get('status', '')

    with get_db() as conn:
        branches = conn.execute(
            'SELECT * FROM branches WHERE is_active=1 ORDER BY name'
        ).fetchall()

        query = '''
            SELECT s.id, s.date, s.status,
                   b.name as branch_name,
                   COALESCE(r.total_revenue, 0)       as revenue,
                   COALESCE(r.delivery_orders, 0) +
                   COALESCE(r.pickup_orders, 0)        as orders,
                   COALESCE(r.delivery_revenue, 0)     as delivery_revenue,
                   COALESCE(r.pickup_revenue, 0)       as pickup_revenue,
                   COALESCE(r.cash_amount, 0)          as cash_amount,
                   COALESCE(r.card_amount, 0)          as card_amount,
                   COALESCE(
                     NULLIF(r.morning_cash, 0),
                     (SELECT COALESCE(r2.morning_cash,0)+COALESCE(r2.cash_amount,0)
                             +COALESCE(r2.change_amount,0)
                             +COALESCE((SELECT SUM(cp2.amount_cash) FROM cash_plus_entries cp2 WHERE cp2.shift_id=s2.id),0)
                             -COALESCE((SELECT SUM(e2.amount_cash) FROM expenses e2 WHERE e2.shift_id=s2.id),0)
                             -COALESCE((SELECT SUM(t2.amount) FROM taxi_trips t2
                                        WHERE t2.shift_id=s2.id AND t2.payment_type='cash'),0)
                             -COALESCE((SELECT SUM(es2.total_amount) FROM employee_shifts es2
                                        WHERE es2.shift_id=s2.id AND es2.is_paid=1),0)
                      FROM shifts s2 JOIN shift_revenue r2 ON r2.shift_id=s2.id
                      WHERE s2.branch_id=s.branch_id AND s2.date<s.date
                      ORDER BY s2.date DESC LIMIT 1),
                     0
                   )                                    as morning_cash,
                   COALESCE(r.change_amount, 0)        as change_amount,
                   r.actual_cash,
                   r.kassa_nal,
                   (SELECT COALESCE(SUM(e.amount_cash),0)
                    FROM expenses e WHERE e.shift_id=s.id)  as exp_cash,
                   (SELECT COALESCE(SUM(es.total_amount),0)
                    FROM employee_shifts es
                    WHERE es.shift_id=s.id AND es.is_paid=1) as paid_salary,
                   (SELECT COALESCE(SUM(COALESCE(cp.amount_cash, cp.amount, 0)),0)
                    FROM cash_plus_entries cp WHERE cp.shift_id=s.id) as plus_cash,
                   (SELECT COALESCE(SUM(tt.amount),0)
                    FROM taxi_trips tt WHERE tt.shift_id=s.id AND tt.payment_type='cash') as taxi_cash,
                   s.opened_at, s.closed_at,
                   s.closed_by_name
            FROM shifts s
            JOIN branches b ON b.id = s.branch_id
            LEFT JOIN shift_revenue r ON r.shift_id = s.id
            WHERE s.date BETWEEN ? AND ?
        '''
        params = [date_from, date_to]

        if role != 'owner':
            bids = get_effective_branch_ids('shifts_archive', branch_ids)
            if bids:
                ids_str = ','.join(str(int(b)) for b in bids)
                query += f' AND s.branch_id IN ({ids_str})'
        elif branch_ids:
            ids_str = ','.join(str(int(bid)) for bid in branch_ids)
            query += f' AND s.branch_id IN ({ids_str})'

        if status_filter:
            query += ' AND s.status = ?'
            params.append(status_filter)

        query += ' ORDER BY s.date DESC, b.name'
        shifts = conn.execute(query, params).fetchall()

        total_revenue = sum(s['revenue'] for s in shifts)
        total_orders  = sum(s['orders']  for s in shifts)

    return render_template('shifts_archive.html',
        shifts=shifts, branches=branches,
        date_from=date_from, date_to=date_to,
        selected_branches=branch_ids, status_filter=status_filter,
        total_revenue=total_revenue, total_orders=total_orders,
        is_owner=(role == 'owner'),
        branch_groups=get_branch_groups(conn))


# ─── BANK STATEMENTS ──────────────────────────────────────────────────────────

@app.route('/bank/')
@login_required
@menu_permission_required('bank')
def bank():
    tab = request.args.get('tab', 'statements')
    date_from = request.args.get('date_from', (date.today().replace(day=1)).isoformat())
    date_to   = request.args.get('date_to', date.today().isoformat())
    with get_db() as conn:
        accounts    = conn.execute('SELECT * FROM bank_accounts WHERE is_active=1 ORDER BY name').fetchall()
        acc_branches = {}
        for row in conn.execute('''
            SELECT bab.bank_account_id, b.id, b.name FROM bank_account_branches bab
            JOIN branches b ON b.id=bab.branch_id ORDER BY b.name
        ''').fetchall():
            acc_branches.setdefault(row['bank_account_id'], []).append({'id': row['id'], 'name': row['name']})
        statements  = conn.execute('''
            SELECT bs.*, ba.name as account_name
            FROM bank_statements bs JOIN bank_accounts ba ON ba.id=bs.bank_account_id
            ORDER BY bs.uploaded_at DESC
        ''').fetchall()
        contractors = conn.execute('''
            SELECT c.*, ec.label as category_label
            FROM contractors c
            LEFT JOIN expense_categories ec ON ec.code = c.category
            WHERE c.is_active=1 ORDER BY c.name
        ''').fetchall()
        terminals   = conn.execute(
            'SELECT t.*, b.name as branch_name FROM bank_terminals t LEFT JOIN branches b ON b.id=t.branch_id ORDER BY t.terminal_number'
        ).fetchall()
        branches    = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
        exp_cats    = [c for c in get_expense_categories(conn) if c['show_contractors']]

        expense_rows = []
        expense_total = 0
        if tab == 'expenses':
            expense_rows = conn.execute('''
                SELECT
                    COALESCE(c.name, bt.counterparty, bt.description, '—') as name,
                    bt.category,
                    ec.label as cat_label,
                    SUM(-bt.amount) as total,
                    COUNT(*) as cnt
                FROM bank_transactions bt
                LEFT JOIN contractors c ON c.id=bt.contractor_id
                LEFT JOIN expense_categories ec ON ec.code=bt.category
                WHERE bt.txn_date BETWEEN ? AND ? AND bt.amount < 0 AND bt.is_ignored=0
                GROUP BY bt.contractor_id, bt.category, COALESCE(c.name, bt.counterparty, bt.description)
                UNION ALL
                SELECT
                    pr.name as name,
                    pe.category,
                    ec2.label as cat_label,
                    SUM(pe.amount) as total,
                    COUNT(*) as cnt
                FROM bank_parse_extracted_entries pe
                JOIN bank_parse_rule_patterns prp ON prp.id = pe.pattern_id
                JOIN bank_parse_rules pr ON pr.id = prp.rule_id
                LEFT JOIN expense_categories ec2 ON ec2.code = pe.category
                WHERE pe.txn_date BETWEEN ? AND ? AND pe.direction='expense'
                GROUP BY pr.name, pe.category
                ORDER BY total DESC
            ''', (date_from, date_to, date_from, date_to)).fetchall()
            expense_total = sum(r['total'] for r in expense_rows)

        compare_rows = []
        compare_bank = 0
        compare_crm  = 0
        if tab == 'compare':
            # Сверяем только операции по картам филиалов (настраиваются в Филиалы →
            # карты) — остальные банковские платежи (контрагентам, налоги и т.д.)
            # заведомо не имеют пары в расходах смены и только шумят в сверке.
            bank_txn_rows = conn.execute('''
                SELECT
                    bt.description as bank_desc, bt.counterparty as bank_counterparty,
                    bt.txn_date as dt,
                    COALESCE(c.name, bt.counterparty, bt.description, '—') as counterparty,
                    bt.category as bank_cat,
                    ec1.label as bank_cat_label,
                    -bt.amount as bank_amount
                FROM bank_transactions bt
                LEFT JOIN contractors c ON c.id=bt.contractor_id
                LEFT JOIN expense_categories ec1 ON ec1.code=bt.category
                WHERE bt.txn_date BETWEEN ? AND ? AND bt.amount < 0 AND bt.is_ignored=0
                ORDER BY bt.txn_date DESC
            ''', (date_from, date_to)).fetchall()
            bank_card_rows = [
                r for r in bank_txn_rows
                if _detect_branch_card(conn, {'description': r['bank_desc'], 'counterparty': r['bank_counterparty']})
            ]

            # CRM-сторона: обычные расходы смены + расходы на такси, отмеченные
            # оплатой картой (payment_type != 'cash') — такси хранится отдельным
            # блоком (taxi_trips), поэтому раньше сюда не попадали.
            crm_card_rows = conn.execute('''
                SELECT s.date as dt, e.amount_cash + e.amount_card as amount, e.description as description, 'expense' as src
                FROM expenses e JOIN shifts s ON s.id = e.shift_id
                WHERE s.date BETWEEN ? AND ?
                UNION ALL
                SELECT s.date as dt, t.amount as amount,
                       ('Такси' || CASE WHEN t.note IS NOT NULL AND t.note != '' THEN ': ' || t.note ELSE '' END) as description,
                       'taxi' as src
                FROM taxi_trips t JOIN shifts s ON s.id = t.shift_id
                WHERE t.payment_type != 'cash' AND s.date BETWEEN ? AND ?
            ''', (date_from, date_to, date_from, date_to)).fetchall()

            crm_pool = list(crm_card_rows)
            for b in bank_card_rows:
                match = None
                for i, c in enumerate(crm_pool):
                    if c['dt'] == b['dt'] and abs(c['amount'] - b['bank_amount']) < 1:
                        match = crm_pool.pop(i)
                        break
                compare_rows.append({
                    'dt': b['dt'], 'counterparty': b['counterparty'],
                    'bank_cat': b['bank_cat'], 'bank_cat_label': b['bank_cat_label'],
                    'bank_amount': b['bank_amount'],
                    'crm_amount': match['amount'] if match else None,
                    'crm_desc': match['description'] if match else None,
                })
            # Такси по карте без пары в банке — тоже показываем, чтобы расхождение было видно
            for c in crm_pool:
                if c['src'] == 'taxi':
                    compare_rows.append({
                        'dt': c['dt'], 'counterparty': c['description'],
                        'bank_cat': None, 'bank_cat_label': None, 'bank_amount': None,
                        'crm_amount': c['amount'], 'crm_desc': c['description'],
                    })
            compare_rows.sort(key=lambda r: r['dt'], reverse=True)
            compare_bank = sum(r['bank_amount'] or 0 for r in compare_rows)
            compare_crm  = sum(r['crm_amount'] or 0 for r in compare_rows)

        sber_auto_count = conn.execute(
            "SELECT COUNT(*) FROM bank_accounts WHERE is_active=1 AND sber_auto_sync=1 AND account_number != '' AND account_number IS NOT NULL"
        ).fetchone()[0]

        parse_rules = conn.execute('''
            SELECT pr.*, ba.name as account_name
            FROM bank_parse_rules pr
            LEFT JOIN bank_accounts ba ON ba.id=pr.bank_account_id
            ORDER BY pr.sort_order, pr.id
        ''').fetchall()

        rule_patterns_map = {}
        for p in conn.execute('SELECT * FROM bank_parse_rule_patterns ORDER BY sort_order, id').fetchall():
            rule_patterns_map.setdefault(p['rule_id'], []).append(p)

        branch_groups_list = get_branch_groups(conn)
        rule_branch_ids = {}
        rule_branch_group_id = {}
        for row in conn.execute(
            'SELECT rule_id, branch_id, branch_group_id FROM bank_parse_rule_branches'
        ).fetchall():
            if row['branch_id']:
                rule_branch_ids.setdefault(row['rule_id'], []).append(row['branch_id'])
            elif row['branch_group_id']:
                rule_branch_group_id[row['rule_id']] = row['branch_group_id']
        rule_branch_label = _rule_branch_labels(conn)

    return render_template('bank.html',
        tab=tab, date_from=date_from, date_to=date_to,
        accounts=accounts, acc_branches=acc_branches, statements=statements,
        contractors=contractors, terminals=terminals,
        branches=branches, exp_cats=exp_cats,
        expense_rows=expense_rows, expense_total=expense_total,
        compare_rows=compare_rows, compare_bank=compare_bank, compare_crm=compare_crm,
        sber_auto_count=sber_auto_count,
        parse_rules=parse_rules, branch_groups=branch_groups_list,
        rule_patterns_map=rule_patterns_map,
        rule_branch_ids=rule_branch_ids, rule_branch_group_id=rule_branch_group_id,
        rule_branch_label=rule_branch_label,
    )


@app.route('/bank/upload', methods=['POST'])
@login_required
@menu_permission_required('bank')
def bank_upload():
    account_id = request.form.get('account_id', '')
    f = request.files.get('file')
    if not account_id or not f:
        flash('Выберите счёт и файл', 'danger')
        return redirect(url_for('bank'))
    raw = f.read()
    try:
        txns = _parse_bank_csv(raw)
    except Exception as e:
        flash(f'Ошибка разбора файла: {e}', 'danger')
        return redirect(url_for('bank'))
    if not txns:
        flash('Не найдено ни одной транзакции', 'warning')
        return redirect(url_for('bank'))
    with get_db() as conn:
        # Сначала помечаем расходы по картам филиалов
        for t in txns:
            bc = _detect_branch_card(conn, t)
            if bc:
                t['_branch_card'] = bc
        _match_contractors(conn, txns)
        dates = [t['date'] for t in txns]
        stmt_id = conn.execute(
            'INSERT INTO bank_statements (bank_account_id, filename, uploaded_by, date_from, date_to, row_count) VALUES (?,?,?,?,?,?)',
            (int(account_id), f.filename, session['user_id'], min(dates), max(dates), len(txns))
        ).lastrowid
        new_txn_ids = []
        for t in txns:
            tid = _match_terminal(conn, t)
            bt_id = conn.execute(
                'INSERT INTO bank_transactions (statement_id, bank_account_id, txn_date, amount, description, counterparty, contractor_id, category, terminal_id) VALUES (?,?,?,?,?,?,?,?,?)',
                (stmt_id, int(account_id), t['date'], t['amount'], t['description'], t['counterparty'],
                 t.get('contractor_id'), t.get('category', ''), tid)
            ).lastrowid
            new_txn_ids.append(bt_id)
        _sync_parse_rule_extractions(conn, new_txn_ids)
        conn.commit()
    flash(f'Загружено {len(txns)} транзакций.', 'success')
    return redirect(url_for('bank_statement_view', stmt_id=stmt_id))


@app.route('/bank/statement/<int:stmt_id>/classify', methods=['GET', 'POST'])
@login_required
@menu_permission_required('bank')
def bank_classify(stmt_id):
    with get_db() as conn:
        stmt = conn.execute('''
            SELECT bs.*, ba.name as account_name
            FROM bank_statements bs JOIN bank_accounts ba ON ba.id=bs.bank_account_id
            WHERE bs.id=?
        ''', (stmt_id,)).fetchone()
        if not stmt:
            flash('Выписка не найдена', 'danger')
            return redirect(url_for('bank'))

        if request.method == 'POST':
            ctr_names    = request.form.getlist('ctr_name')
            ctr_cats     = request.form.getlist('ctr_category')
            ctr_keywords = request.form.getlist('ctr_keywords')
            ctr_ids      = request.form.getlist('ctr_id')
            tid_numbers  = request.form.getlist('tid_number')
            tid_branches = request.form.getlist('tid_branch')
            tid_names    = request.form.getlist('tid_name')

            for name, cat, kw, ctr_id_str in zip(ctr_names, ctr_cats, ctr_keywords, ctr_ids):
                name = (name or '').strip()
                cat  = (cat or '').strip()
                if not name or not cat:
                    continue
                kw = (kw or name).strip() or name
                pre_id = int(ctr_id_str) if ctr_id_str and ctr_id_str.isdigit() else None

                if pre_id:
                    # Contractor already matched — update its category
                    conn.execute('UPDATE contractors SET category=?, keywords=? WHERE id=?',
                                 (cat, kw, pre_id))
                    # Update all transactions for this contractor in the statement
                    conn.execute('''
                        UPDATE bank_transactions SET category=?
                        WHERE statement_id=? AND contractor_id=?
                    ''', (cat, stmt_id, pre_id))
                    # Also link any remaining unlinked rows by counterparty text
                    conn.execute('''
                        UPDATE bank_transactions SET contractor_id=?, category=?
                        WHERE statement_id=? AND LOWER(TRIM(counterparty))=LOWER(?) AND contractor_id IS NULL
                    ''', (pre_id, cat, stmt_id, name))
                else:
                    existing = conn.execute(
                        'SELECT id FROM contractors WHERE LOWER(name)=LOWER(?)', (name,)
                    ).fetchone()
                    if existing:
                        conn.execute('UPDATE contractors SET category=?, keywords=? WHERE id=?',
                                     (cat, kw, existing['id']))
                        pre_id = existing['id']
                    else:
                        pre_id = conn.execute(
                            'INSERT INTO contractors (name, category, keywords) VALUES (?,?,?)',
                            (name, cat, kw)
                        ).lastrowid
                    conn.execute('''
                        UPDATE bank_transactions SET contractor_id=?, category=?
                        WHERE statement_id=? AND LOWER(TRIM(counterparty))=LOWER(?)
                    ''', (pre_id, cat, stmt_id, name))

            for tid, branch_str, tname in zip(tid_numbers, tid_branches, tid_names):
                tid = (tid or '').strip()
                if not tid:
                    continue
                branch_id = int(branch_str) if branch_str and branch_str.isdigit() else None
                tname = (tname or f'Терминал {tid}').strip()
                existing = conn.execute(
                    'SELECT id FROM bank_terminals WHERE terminal_number=?', (tid,)
                ).fetchone()
                if existing:
                    conn.execute('UPDATE bank_terminals SET branch_id=?, name=? WHERE id=?',
                                 (branch_id, tname, existing['id']))
                    term_id = existing['id']
                else:
                    term_id = conn.execute(
                        'INSERT INTO bank_terminals (terminal_number, name, branch_id) VALUES (?,?,?)',
                        (tid, tname, branch_id)
                    ).lastrowid
                conn.execute('''
                    UPDATE bank_transactions SET terminal_id=?
                    WHERE statement_id=?
                      AND (description LIKE ? OR counterparty LIKE ?
                           OR description LIKE ? OR counterparty LIKE ?)
                ''', (term_id, stmt_id,
                      f'%TID%{tid}%', f'%TID%{tid}%',
                      f'%{tid}%', f'%{tid}%'))

            conn.commit()
            flash('Разметка сохранена', 'success')
            return redirect(url_for('bank_statement_view', stmt_id=stmt_id))

        # GET — build unique counterparty and terminal lists
        ctr_rows = conn.execute('''
            SELECT TRIM(counterparty) as name, COUNT(*) as cnt, SUM(amount) as total,
                   MIN(contractor_id) as contractor_id, MIN(category) as category
            FROM bank_transactions
            WHERE statement_id=? AND TRIM(counterparty) != ''
            GROUP BY LOWER(TRIM(counterparty))
            ORDER BY SUM(amount)
        ''', (stmt_id,)).fetchall()

        all_ctrs = conn.execute('SELECT id, name, category, keywords FROM contractors').fetchall()
        existing_ctrs_by_name = {c['name'].lower(): c for c in all_ctrs}
        existing_ctrs_by_id   = {c['id']: c for c in all_ctrs}

        counterparties = []
        for row in ctr_rows:
            name   = row['name']
            ctr_id = row['contractor_id']
            # Prefer lookup by ID (handles keyword-matched contractors with different name)
            ex = existing_ctrs_by_id.get(ctr_id) if ctr_id else None
            if ex is None:
                ex = existing_ctrs_by_name.get(name.lower())
            matched_name = ex['name'] if ex and ex['name'].lower() != name.lower() else None
            counterparties.append({
                'name':         name,
                'cnt':          row['cnt'],
                'total':        row['total'],
                'ctr_id':       ex['id'] if ex else None,
                'is_new':       ex is None,
                'current_cat':  (ex['category'] if ex else row['category']) or '',
                'keywords':     (ex['keywords'] if ex else name) or name,
                'matched_name': matched_name,
            })

        # Detect TIDs in descriptions
        all_txns = conn.execute(
            'SELECT description, counterparty FROM bank_transactions WHERE statement_id=?', (stmt_id,)
        ).fetchall()
        tid_cnts = {}
        for t in all_txns:
            text = f"{t['description'] or ''} {t['counterparty'] or ''}"
            m = re.search(r'TID\s*[:\-]?\s*(\d{6,12})', text, re.IGNORECASE)
            if m:
                tid = m.group(1)
                tid_cnts[tid] = tid_cnts.get(tid, 0) + 1

        existing_tids = {r['terminal_number']: r for r in
                         conn.execute('SELECT terminal_number, name, branch_id FROM bank_terminals').fetchall()}

        tid_list = []
        for tid, cnt in sorted(tid_cnts.items()):
            ex = existing_tids.get(tid)
            tid_list.append({
                'tid':       tid,
                'cnt':       cnt,
                'is_new':    ex is None,
                'name':      (ex['name'] if ex else '') or '',
                'branch_id': (ex['branch_id'] if ex else None),
            })

        txn_count = conn.execute(
            'SELECT COUNT(*) FROM bank_transactions WHERE statement_id=?', (stmt_id,)
        ).fetchone()[0]
        branches = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
        exp_cats = [c for c in get_expense_categories(conn) if c['show_contractors']]

    return render_template('bank_classify.html',
        stmt=stmt, counterparties=counterparties, tid_list=tid_list,
        branches=branches, exp_cats=exp_cats, txn_count=txn_count)


@app.route('/bank/statement/<int:stmt_id>/delete', methods=['POST'])
@login_required
@menu_permission_required('bank')
def bank_statement_delete(stmt_id):
    with get_db() as conn:
        conn.execute('DELETE FROM bank_statements WHERE id=?', (stmt_id,))
        conn.commit()
    flash('Выписка удалена', 'success')
    return redirect(url_for('bank'))


@app.route('/bank/statements/<int:stmt_id>/rematch', methods=['POST'])
@login_required
@menu_permission_required('bank')
def bank_statement_rematch(stmt_id):
    """Повторно применить автоматический матчинг контрагентов и терминалов."""
    with get_db() as conn:
        txns_rows = conn.execute(
            'SELECT id, description, counterparty FROM bank_transactions WHERE statement_id=?',
            (stmt_id,)
        ).fetchall()
        updated = 0
        contractors = conn.execute(
            'SELECT id, name, category, keywords FROM contractors WHERE is_active=1'
        ).fetchall()
        for row in txns_rows:
            txn = dict(row)
            matched_cid = None
            matched_cat = None
            text = ((txn.get('description') or '') + ' ' + (txn.get('counterparty') or '')).lower()
            for c in contractors:
                kws = [k.strip().lower() for k in (c['keywords'] or c['name']).split(',') if k.strip()]
                if any(kw in text for kw in kws):
                    matched_cid = c['id']
                    matched_cat = c['category']
                    break
            # Если не найден — создаём из counterparty
            if not matched_cid:
                cp = (txn.get('counterparty') or '').strip()
                if cp:
                    existing = conn.execute(
                        'SELECT id FROM contractors WHERE LOWER(name)=LOWER(?)', (cp,)
                    ).fetchone()
                    if existing:
                        matched_cid = existing['id']
                    else:
                        conn.execute('INSERT INTO contractors (name, keywords) VALUES (?,?)', (cp, cp))
                        matched_cid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
                        contractors = conn.execute(
                            'SELECT id, name, category, keywords FROM contractors WHERE is_active=1'
                        ).fetchall()
            # Матчинг терминала
            tid = _match_terminal(conn, txn)
            conn.execute(
                'UPDATE bank_transactions SET contractor_id=?, category=COALESCE(NULLIF(category,""),?), terminal_id=COALESCE(terminal_id,?) WHERE id=?',
                (matched_cid, matched_cat or '', tid, row['id'])
            )
            if matched_cid or tid:
                updated += 1
        conn.commit()
    flash(f'Переназначено: {updated} из {len(txns_rows)} транзакций', 'success')
    return redirect(url_for('bank_statement_view', stmt_id=stmt_id))


@app.route('/bank/transaction/<int:txn_id>/update', methods=['POST'])
@login_required
@menu_permission_required('bank')
def bank_txn_update(txn_id):
    contractor_id = request.form.get('contractor_id') or None
    category = request.form.get('category', '')
    is_ignored = 1 if request.form.get('is_ignored') else 0
    with get_db() as conn:
        conn.execute(
            'UPDATE bank_transactions SET contractor_id=?, category=?, is_ignored=? WHERE id=?',
            (contractor_id, category, is_ignored, txn_id)
        )
        conn.commit()
    return jsonify({'ok': True})


@app.route('/bank/accounts/add', methods=['POST'])
@login_required
@menu_permission_required('bank')
def bank_account_add():
    name = request.form.get('name', '').strip()
    bank_name = request.form.get('bank_name', '').strip()
    account_number = request.form.get('account_number', '').strip()
    branch_ids = [b for b in request.form.getlist('branch_ids') if b.isdigit()]
    if not name:
        flash('Введите название счёта', 'danger')
        return redirect(url_for('bank', tab='accounts'))
    with get_db() as conn:
        acc_id = conn.execute(
            'INSERT INTO bank_accounts (name, bank_name, account_number) VALUES (?,?,?)',
            (name, bank_name, account_number)
        ).lastrowid
        for bid in branch_ids:
            conn.execute(
                'INSERT OR IGNORE INTO bank_account_branches (bank_account_id, branch_id) VALUES (?,?)',
                (acc_id, int(bid))
            )
        conn.commit()
    flash(f'Счёт «{name}» добавлен', 'success')
    return redirect(url_for('bank', tab='accounts'))


@app.route('/bank/accounts/<int:acc_id>/delete', methods=['POST'])
@login_required
@menu_permission_required('bank')
def bank_account_delete(acc_id):
    with get_db() as conn:
        conn.execute('UPDATE bank_accounts SET is_active=0 WHERE id=?', (acc_id,))
        conn.commit()
    flash('Счёт удалён', 'success')
    return redirect(url_for('bank', tab='accounts'))


@app.route('/bank/contractors/add', methods=['POST'])
@login_required
@menu_permission_required('bank')
def bank_contractor_add():
    name = request.form.get('name', '').strip()
    category = request.form.get('category', '').strip()
    raw_kw = request.form.get('keywords', '').strip()
    keywords = ', '.join(_filter_keywords([k.strip() for k in raw_kw.split(',') if k.strip()])) if raw_kw else ''
    inn = re.sub(r'\D', '', request.form.get('inn', '').strip())
    next_url = request.form.get('next', '').strip()
    if not name:
        flash('Введите название контрагента', 'danger')
        return redirect(next_url or url_for('bank', tab='contractors'))
    with get_db() as conn:
        conn.execute(
            'INSERT INTO contractors (name, category, keywords, inn) VALUES (?,?,?,?)',
            (name, category, keywords, inn or None)
        )
        conn.commit()
    flash(f'Контрагент «{name}» добавлен', 'success')
    return redirect(next_url or url_for('bank', tab='contractors'))


@app.route('/bank/contractors/<int:ctr_id>/edit', methods=['POST'])
@login_required
@menu_permission_required('bank')
def bank_contractor_edit(ctr_id):
    name = request.form.get('name', '').strip()
    category = request.form.get('category', '').strip()
    raw_kw = request.form.get('keywords', '').strip()
    keywords = ', '.join(_filter_keywords([k.strip() for k in raw_kw.split(',') if k.strip()])) if raw_kw else ''
    inn = re.sub(r'\D', '', request.form.get('inn', '').strip())
    with get_db() as conn:
        conn.execute(
            'UPDATE contractors SET name=?, category=?, keywords=?, inn=? WHERE id=?',
            (name, category, keywords, inn or None, ctr_id)
        )
        conn.commit()
    flash('Контрагент обновлён', 'success')
    return redirect(url_for('bank', tab='contractors'))


@app.route('/bank/contractors/<int:ctr_id>/delete', methods=['POST'])
@login_required
@menu_permission_required('bank')
def bank_contractor_delete(ctr_id):
    with get_db() as conn:
        conn.execute('UPDATE contractors SET is_active=0 WHERE id=?', (ctr_id,))
        conn.commit()
    flash('Контрагент удалён', 'success')
    return redirect(url_for('bank', tab='contractors'))


@app.route('/bank/terminals/add', methods=['POST'])
@login_required
@menu_permission_required('bank')
def bank_terminal_add():
    terminal_number = request.form.get('terminal_number', '').strip()
    name = request.form.get('name', '').strip()
    branch_id = request.form.get('branch_id') or None
    if not terminal_number:
        flash('Введите номер терминала', 'danger')
        return redirect(url_for('bank', tab='terminals'))
    with get_db() as conn:
        try:
            conn.execute(
                'INSERT INTO bank_terminals (terminal_number, name, branch_id) VALUES (?,?,?)',
                (terminal_number, name, branch_id)
            )
            conn.commit()
            flash(f'Терминал {terminal_number} добавлен', 'success')
        except Exception:
            flash('Терминал с таким номером уже существует', 'danger')
    return redirect(url_for('bank', tab='terminals'))


@app.route('/bank/terminals/<int:tid>/edit', methods=['POST'])
@login_required
@menu_permission_required('bank')
def bank_terminal_edit(tid):
    name = request.form.get('name', '').strip()
    branch_id = request.form.get('branch_id') or None
    with get_db() as conn:
        conn.execute(
            'UPDATE bank_terminals SET name=?, branch_id=? WHERE id=?',
            (name, branch_id, tid)
        )
        conn.commit()
    flash('Терминал обновлён', 'success')
    return redirect(url_for('bank', tab='terminals'))


@app.route('/bank/terminals/<int:tid>/delete', methods=['POST'])
@login_required
@menu_permission_required('bank')
def bank_terminal_delete(tid):
    with get_db() as conn:
        conn.execute('UPDATE bank_terminals SET is_active=0 WHERE id=?', (tid,))
        conn.commit()
    flash('Терминал удалён', 'success')
    return redirect(url_for('bank', tab='terminals'))


def _parse_pattern_rows(form):
    """Собирает динамические строки паттернов из форм-данных: pattern[N][поле].
    Паттерн всегда «расход» (направление не выбирается) и всегда относится к тому же
    филиалу/группе, что и само правило (bank_parse_rule_branches) — своей привязки
    филиала у паттерна больше нет."""
    indices = set()
    for key in form.keys():
        m = re.match(r'pattern\[(\d+)\]\[', key)
        if m:
            indices.add(int(m.group(1)))
    rows = []
    for i in sorted(indices):
        example_text  = form.get(f'pattern[{i}][example_text]', '').strip()
        example_value = form.get(f'pattern[{i}][example_value]', '').strip()
        category      = form.get(f'pattern[{i}][category]', '').strip()
        if not example_text or not example_value:
            continue
        rows.append({
            'example_text': example_text, 'example_value': example_value,
            'direction': 'expense', 'category': category,
        })
    return rows


def _save_rule_patterns(conn, rule_id, form):
    conn.execute('DELETE FROM bank_parse_rule_patterns WHERE rule_id=?', (rule_id,))
    for i, row in enumerate(_parse_pattern_rows(form)):
        regex = _build_pattern_regex(row['example_text'], row['example_value'])
        if not regex:
            continue
        conn.execute(
            'INSERT INTO bank_parse_rule_patterns '
            '(rule_id, example_text, example_value, regex_pattern, direction, category, sort_order) '
            'VALUES (?,?,?,?,?,?,?)',
            (rule_id, row['example_text'], row['example_value'], regex,
             row['direction'], row['category'], i)
        )


def _save_rule_branches(conn, rule_id, form):
    """Филиал(ы)/группа, к которым относится вся операция при совпадении правила
    (в отличие от bank_parse_rule_pattern_branches — те привязаны к отдельным
    паттернам извлечения сумм, а не ко всему правилу)."""
    conn.execute('DELETE FROM bank_parse_rule_branches WHERE rule_id=?', (rule_id,))
    branch_ids = [b for b in form.getlist('rule_branch_ids[]') if b.isdigit()]
    branch_group_id = form.get('rule_branch_group_id', '').strip() or None
    if branch_group_id and branch_group_id.isdigit():
        conn.execute(
            'INSERT INTO bank_parse_rule_branches (rule_id, branch_group_id) VALUES (?,?)',
            (rule_id, int(branch_group_id))
        )
    for bid in branch_ids:
        conn.execute(
            'INSERT INTO bank_parse_rule_branches (rule_id, branch_id) VALUES (?,?)',
            (rule_id, int(bid))
        )


def _rule_branch_labels(conn):
    """{rule_id: 'Название группы' | 'Филиал1, Филиал2'} для правил, где выбран
    филиал/группа (к которому относится вся операция при совпадении правила).
    Правила без записи в bank_parse_rule_branches просто отсутствуют в словаре
    (в UI это читается как «Все филиалы»)."""
    branch_name_map = {b['id']: b['name'] for b in conn.execute('SELECT id, name FROM branches').fetchall()}
    group_name_map  = {g['id']: g['name'] for g in conn.execute('SELECT id, name FROM branch_groups').fetchall()}
    group_by_rule = {}
    ids_by_rule = {}
    for row in conn.execute('SELECT rule_id, branch_id, branch_group_id FROM bank_parse_rule_branches').fetchall():
        if row['branch_group_id']:
            group_by_rule[row['rule_id']] = row['branch_group_id']
        elif row['branch_id']:
            ids_by_rule.setdefault(row['rule_id'], []).append(row['branch_id'])
    labels = {rule_id: group_name_map.get(gid, '—') for rule_id, gid in group_by_rule.items()}
    for rule_id, bids in ids_by_rule.items():
        if rule_id not in labels:
            labels[rule_id] = ', '.join(branch_name_map.get(bid, '?') for bid in bids)
    return labels


def _rule_branch_ids_map(conn):
    """{rule_id: set(branch_id)} — филиалы, к которым относится правило целиком
    (группы уже развёрнуты в состав участников). Правила без записи в
    bank_parse_rule_branches отсутствуют в словаре (значит — все филиалы)."""
    result = {}
    for row in conn.execute(
        'SELECT rule_id, branch_id, branch_group_id FROM bank_parse_rule_branches'
    ).fetchall():
        ids = result.setdefault(row['rule_id'], set())
        if row['branch_id']:
            ids.add(row['branch_id'])
        elif row['branch_group_id']:
            for m in conn.execute(
                'SELECT branch_id FROM branch_group_members WHERE group_id=?', (row['branch_group_id'],)
            ).fetchall():
                ids.add(m['branch_id'])
    return result


@app.route('/bank/parse-rules/add', methods=['POST'])
@login_required
@menu_permission_required('bank')
def bank_parse_rule_add():
    name       = request.form.get('name', '').strip()
    keyword    = request.form.get('keyword', '').strip()
    direction  = request.form.get('direction', 'any')
    account_id = request.form.get('bank_account_id', '').strip() or None
    category   = request.form.get('category', '').strip()
    sort_order = int(request.form.get('sort_order', 0) or 0)
    if not name or not keyword:
        flash('Укажите название и ключевое слово', 'danger')
        return redirect(url_for('bank', tab='rules'))
    with get_db() as conn:
        conn.execute(
            'INSERT INTO bank_parse_rules (bank_account_id, name, direction, keyword, category, sort_order) '
            'VALUES (?,?,?,?,?,?)',
            (account_id, name, direction, keyword, category, sort_order)
        )
        rule_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        _save_rule_patterns(conn, rule_id, request.form)
        _save_rule_branches(conn, rule_id, request.form)
        _sync_parse_rule_extractions(conn)
        conn.commit()
    flash('Правило добавлено', 'success')
    return redirect(url_for('bank', tab='rules'))


@app.route('/bank/parse-rules/<int:rule_id>/edit', methods=['POST'])
@login_required
@menu_permission_required('bank')
def bank_parse_rule_edit(rule_id):
    name       = request.form.get('name', '').strip()
    keyword    = request.form.get('keyword', '').strip()
    direction  = request.form.get('direction', 'any')
    account_id = request.form.get('bank_account_id', '').strip() or None
    category   = request.form.get('category', '').strip()
    sort_order = int(request.form.get('sort_order', 0) or 0)
    if not name or not keyword:
        flash('Укажите название и ключевое слово', 'danger')
        return redirect(url_for('bank', tab='rules'))
    with get_db() as conn:
        conn.execute(
            'UPDATE bank_parse_rules SET bank_account_id=?, name=?, direction=?, keyword=?, category=?, sort_order=? WHERE id=?',
            (account_id, name, direction, keyword, category, sort_order, rule_id)
        )
        _save_rule_patterns(conn, rule_id, request.form)
        _save_rule_branches(conn, rule_id, request.form)
        _sync_parse_rule_extractions(conn)
        conn.commit()
    flash('Правило обновлено', 'success')
    return redirect(url_for('bank', tab='rules'))


@app.route('/bank/parse-rules/<int:rule_id>/delete', methods=['POST'])
@login_required
@menu_permission_required('bank')
def bank_parse_rule_delete(rule_id):
    with get_db() as conn:
        conn.execute('DELETE FROM bank_parse_rules WHERE id=?', (rule_id,))
        _sync_parse_rule_extractions(conn)
        conn.commit()
    flash('Правило удалено', 'success')
    return redirect(url_for('bank', tab='rules'))


def _enrich_bank_txn(d):
    """Определяет op_type/op_card4 и достраивает counterparty из описания для одной транзакции (dict)."""
    desc  = d.get('description') or ''
    desc_u = desc.upper()
    desc_l = desc.lower()
    amount = d.get('amount', 0)

    if not d.get('counterparty'):
        m = re.search(r'[Сс]бербанка\s+(.+?)\s+по\s+карте', desc)
        if not m:
            m = re.search(r'в\s+ТУ\s+(.+?)\s+по\s+(?:карте|к)', desc, re.IGNORECASE)
        if m:
            d['counterparty'] = m.group(1).strip()

    if 'PURCHASE' in desc_u:
        # Карточная покупка — извлекаем последние 4 цифры карты
        cm = re.search(r'по\s+карте\s+(\S+)', desc, re.IGNORECASE)
        if cm:
            digits = re.sub(r'\D', '', cm.group(1))
            d['op_card4'] = digits[-4:] if len(digits) >= 4 else digits
        else:
            d['op_card4'] = ''
        d['op_type'] = 'card'
    elif d.get('terminal_id') or 'эквайрин' in desc_l:
        d['op_type'] = 'acquiring'
        d['op_card4'] = ''
    elif 'перераспредел' in desc_l or 'между счет' in desc_l:
        d['op_type'] = 'transfer'
        d['op_card4'] = ''
    elif amount < 0:
        d['op_type'] = 'contractor'
        d['op_card4'] = ''
    else:
        d['op_type'] = 'transfer'
        d['op_card4'] = ''
    return d


def _build_pattern_regex(example_text, example_value):
    """Строит регулярное выражение для извлечения числа по примеру, который ввёл владелец:
    example_text='К.224.54', example_value='224.54' -> re.escape('К.') + число-группа.
    Возвращает None, если пример числа не найден внутри примера текста."""
    example_text = (example_text or '').strip()
    example_value = (example_value or '').strip()
    if not example_text or not example_value:
        return None
    candidates = [example_value]
    if '.' in example_value:
        candidates.append(example_value.replace('.', ','))
    elif ',' in example_value:
        candidates.append(example_value.replace(',', '.'))
    idx = -1
    used = None
    for cand in candidates:
        idx = example_text.find(cand)
        if idx != -1:
            used = cand
            break
    if idx == -1:
        return None
    prefix = example_text[:idx]
    has_decimals = '.' in used or ',' in used
    num_pattern = r'(\d+[.,]\d+)' if has_decimals else r'(\d+(?:[.,]\d+)?)'
    return re.escape(prefix) + num_pattern


def _apply_bank_parse_rules(conn, txns):
    """Применяет правила разбора к списку транзакций (dict-ов с полями bt.*): подставляет
    имя/категорию правила и список сумм, извлечённых по паттернам (для показа в выписке)."""
    parse_rules = conn.execute(
        'SELECT * FROM bank_parse_rules WHERE is_active=1 ORDER BY sort_order, id'
    ).fetchall()
    patterns_by_rule = {}
    for p in conn.execute('SELECT * FROM bank_parse_rule_patterns ORDER BY sort_order, id').fetchall():
        patterns_by_rule.setdefault(p['rule_id'], []).append(p)
    rule_branch_label = _rule_branch_labels(conn)
    for d in txns:
        d['parse_rule_name'] = ''
        d['parse_rule_category'] = ''
        d['parse_rule_branch_label'] = ''
        d['extracted_patterns'] = []
        for rule in parse_rules:
            if rule['bank_account_id'] and rule['bank_account_id'] != d.get('bank_account_id'):
                continue
            if rule['direction'] == 'income' and d['amount'] <= 0:
                continue
            if rule['direction'] == 'expense' and d['amount'] >= 0:
                continue
            kw = (rule['keyword'] or '').lower()
            if not kw or kw not in (d.get('description') or '').lower():
                continue
            d['parse_rule_name'] = rule['name']
            d['parse_rule_category'] = rule['category'] or ''
            d['parse_rule_branch_label'] = rule_branch_label.get(rule['id'], '')
            for pat in patterns_by_rule.get(rule['id'], []):
                if not pat['regex_pattern']:
                    continue
                try:
                    m = re.search(pat['regex_pattern'], d.get('description') or '', re.IGNORECASE)
                except re.error:
                    m = None
                if m:
                    try:
                        val = float(m.group(1).replace(',', '.'))
                        d['extracted_patterns'].append(
                            {'value': val, 'direction': pat['direction'], 'category': pat['category']}
                        )
                    except (ValueError, IndexError):
                        pass
            break
    return txns


def _sync_parse_rule_extractions(conn, txn_ids=None):
    """(Пере)строит bank_parse_extracted_entries — реальные категоризированные записи по
    филиалам, извлечённые из описаний операций по паттернам правил. Вызывается при импорте
    новых транзакций (txn_ids) и при любом изменении правил/паттернов (txn_ids=None — все)."""
    rules = conn.execute(
        'SELECT * FROM bank_parse_rules WHERE is_active=1 ORDER BY sort_order, id'
    ).fetchall()

    patterns_by_rule = {}
    for p in conn.execute('SELECT * FROM bank_parse_rule_patterns ORDER BY sort_order, id').fetchall():
        patterns_by_rule.setdefault(p['rule_id'], []).append(dict(p))

    # Паттерн больше не имеет своей привязки филиала — он всегда относится
    # к тому же филиалу/группе, что и правило целиком.
    rule_branches = _rule_branch_ids_map(conn)

    if txn_ids is not None and not txn_ids:
        return
    if txn_ids:
        ph = ','.join('?' * len(txn_ids))
        conn.execute(f'DELETE FROM bank_parse_extracted_entries WHERE bank_transaction_id IN ({ph})', txn_ids)
        txns = conn.execute(
            f'SELECT id, bank_account_id, txn_date, amount, description FROM bank_transactions WHERE id IN ({ph})',
            txn_ids
        ).fetchall()
    else:
        conn.execute('DELETE FROM bank_parse_extracted_entries')
        txns = conn.execute(
            'SELECT id, bank_account_id, txn_date, amount, description FROM bank_transactions'
        ).fetchall()

    if not rules:
        return

    for t in txns:
        desc = t['description'] or ''
        for rule in rules:
            if rule['bank_account_id'] and rule['bank_account_id'] != t['bank_account_id']:
                continue
            if rule['direction'] == 'income' and t['amount'] <= 0:
                continue
            if rule['direction'] == 'expense' and t['amount'] >= 0:
                continue
            kw = (rule['keyword'] or '').lower()
            if not kw or kw not in desc.lower():
                continue
            for pat in patterns_by_rule.get(rule['id'], []):
                if not pat['regex_pattern']:
                    continue
                try:
                    m = re.search(pat['regex_pattern'], desc, re.IGNORECASE)
                except re.error:
                    continue
                if not m:
                    continue
                try:
                    val = float(m.group(1).replace(',', '.'))
                except (ValueError, IndexError):
                    continue
                bids = rule_branches.get(rule['id']) or set()
                for bid in bids:
                    conn.execute('''
                        INSERT OR REPLACE INTO bank_parse_extracted_entries
                        (bank_transaction_id, pattern_id, branch_id, amount, direction, category, txn_date)
                        VALUES (?,?,?,?,?,?,?)
                    ''', (t['id'], pat['id'], bid, val, pat['direction'], pat['category'], t['txn_date']))
            break


@app.route('/bank/statement/<int:stmt_id>')
@login_required
@menu_permission_required('bank')
def bank_statement_view(stmt_id):
    with get_db() as conn:
        stmt = conn.execute('''
            SELECT bs.*, ba.name as account_name
            FROM bank_statements bs JOIN bank_accounts ba ON ba.id=bs.bank_account_id
            WHERE bs.id=?
        ''', (stmt_id,)).fetchone()
        if not stmt:
            flash('Выписка не найдена', 'danger')
            return redirect(url_for('bank'))
        txns_raw = conn.execute('''
            SELECT bt.*, c.name as contractor_name,
                   t.terminal_number, b.name as terminal_branch
            FROM bank_transactions bt
            LEFT JOIN contractors c ON c.id=bt.contractor_id
            LEFT JOIN bank_terminals t ON t.id=bt.terminal_id
            LEFT JOIN branches b ON b.id=t.branch_id
            WHERE bt.statement_id=?
            ORDER BY bt.txn_date DESC, bt.id DESC
        ''', (stmt_id,)).fetchall()
        # Для строк с пустым counterparty — извлекаем из описания
        # Определяем тип операции для каждой транзакции
        txns = [_enrich_bank_txn(dict(row)) for row in txns_raw]
        _apply_bank_parse_rules(conn, txns)
        for d in txns:
            d['effective_category'] = d.get('category') or d.get('parse_rule_category') or ''

        contractors = conn.execute(
            'SELECT * FROM contractors WHERE is_active=1 AND COALESCE(is_card_merchant,0)=0 ORDER BY name'
        ).fetchall()
        all_exp_cats     = [c for c in get_expense_categories(conn) if c['show_contractors']]
        exp_cats_income  = [c for c in all_exp_cats if c['type'] == 'income']
        exp_cats_expense = [c for c in all_exp_cats if c['type'] != 'income']
        cat_labels = {c['code']: c['label'] for c in get_expense_categories(conn)}
    unique_cats = sorted(
        set(d['effective_category'] for d in txns if d['effective_category']),
        key=lambda code: cat_labels.get(code, code)
    )
    return render_template('bank_statement.html',
        stmt=stmt, txns=txns, contractors=contractors,
        exp_cats_income=exp_cats_income, exp_cats_expense=exp_cats_expense,
        unique_cats=unique_cats, cat_labels=cat_labels)


@app.route('/bank/statements/all')
@login_required
@menu_permission_required('bank')
def bank_statements_all():
    today = date.today()
    date_from = request.args.get('date_from', today.replace(day=1).isoformat())
    date_to   = request.args.get('date_to', today.isoformat())
    with get_db() as conn:
        txns_raw = conn.execute('''
            SELECT bt.*, c.name as contractor_name,
                   t.terminal_number, b.name as terminal_branch,
                   ba.name as account_name
            FROM bank_transactions bt
            LEFT JOIN contractors c ON c.id=bt.contractor_id
            LEFT JOIN bank_terminals t ON t.id=bt.terminal_id
            LEFT JOIN branches b ON b.id=t.branch_id
            JOIN bank_statements bs ON bs.id=bt.statement_id
            JOIN bank_accounts ba ON ba.id=bt.bank_account_id
            WHERE bt.txn_date BETWEEN ? AND ?
            ORDER BY bt.txn_date DESC, bt.id DESC
        ''', (date_from, date_to)).fetchall()
        txns = [_enrich_bank_txn(dict(row)) for row in txns_raw]
        _apply_bank_parse_rules(conn, txns)
        for d in txns:
            d['effective_category'] = d.get('category') or d.get('parse_rule_category') or ''

        contractors = conn.execute(
            'SELECT * FROM contractors WHERE is_active=1 AND COALESCE(is_card_merchant,0)=0 ORDER BY name'
        ).fetchall()
        all_exp_cats     = [c for c in get_expense_categories(conn) if c['show_contractors']]
        exp_cats_income  = [c for c in all_exp_cats if c['type'] == 'income']
        exp_cats_expense = [c for c in all_exp_cats if c['type'] != 'income']
        cat_labels = {c['code']: c['label'] for c in get_expense_categories(conn)}
    unique_cats = sorted(
        set(d['effective_category'] for d in txns if d['effective_category']),
        key=lambda code: cat_labels.get(code, code)
    )
    stmt = {'filename': 'Выписка по всем', 'account_name': 'Все счета',
            'date_from': date_from, 'date_to': date_to, 'row_count': len(txns)}
    return render_template('bank_statement.html',
        stmt=stmt, txns=txns, contractors=contractors,
        exp_cats_income=exp_cats_income, exp_cats_expense=exp_cats_expense,
        unique_cats=unique_cats, cat_labels=cat_labels, show_bank_col=True,
        date_from=date_from, date_to=date_to)


# ─── SBERBANK API SYNC ────────────────────────────────────────────────────────

def _sber_get(conn, key, default=''):
    row = conn.execute('SELECT value FROM api_settings WHERE key=?', (key,)).fetchone()
    return row[0] if row else default

def _sber_set(conn, key, value):
    conn.execute('INSERT OR REPLACE INTO api_settings(key,value) VALUES(?,?)', (key, value))


@app.route('/bank/sber/debug')
@login_required
@menu_permission_required('bank')
def sber_debug():
    """Отладка: показать состояние токенов в базе."""
    import time
    with get_db() as conn:
        keys = ['sber_client_id','sber_account_number','sber_access_token',
                'sber_refresh_token','sber_token_expires','sber_npa_active','sber_last_sync','sber_last_result']
        info = {}
        for k in keys:
            v = _sber_get(conn, k)
            if k in ('sber_access_token','sber_refresh_token') and v:
                info[k] = v[:12] + '...' + f' (len={len(v)})'
            elif k == 'sber_token_expires' and v:
                try:
                    exp = float(v)
                    remaining = exp - time.time()
                    info[k] = f'{v} (осталось {int(remaining)}с / {int(remaining/60)}мин)'
                except:
                    info[k] = v
            else:
                info[k] = v or '(пусто)'
    return jsonify(info)


@app.route('/bank/sber/debug-tx')
@login_required
@menu_permission_required('bank')
def sber_debug_tx():
    """Показать сырую структуру первых 2 транзакций от Сбера (для отладки полей)."""
    from sber_api import get_statement, _mtls, STMT_URL
    import time, requests, urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    with get_db() as conn:
        access_token   = _sber_get(conn, 'sber_access_token')
        client_id      = _sber_get(conn, 'sber_client_id', '71154')
        account_number = _sber_get(conn, 'sber_account_number')
        if not access_token or not account_number:
            return jsonify({'error': 'Нет токена или счёта'})
        today = date.today().isoformat()
        resp = requests.get(
            STMT_URL,
            cert=_mtls(), verify=False,
            headers={'Authorization': f'Bearer {access_token}',
                     'x-ibm-client-id': str(client_id),
                     'Accept': 'application/json'},
            params={'accountNumber': account_number, 'statementDate': today, 'page': 1},
            timeout=30,
        )
        raw = resp.json() if resp.ok else {'status': resp.status_code, 'body': resp.text[:1000]}
        ops = (raw.get('transactions') or raw.get('operations') or raw.get('operationList') or
               raw.get('items') or [])
        return jsonify({'status': resp.status_code, 'keys_top': list(raw.keys()),
                        'first_2_ops': ops[:2]})


@app.route('/bank/sber/settings', methods=['GET', 'POST'])
@login_required
@menu_permission_required('bank')
def sber_settings():
    with get_db() as conn:
        if request.method == 'POST':
            _sber_set(conn, 'sber_client_id',     request.form.get('client_id', '').strip())
            _sber_set(conn, 'sber_client_secret',  request.form.get('client_secret', '').strip())
            conn.commit()
            flash('Настройки API сохранены', 'success')
            return redirect(url_for('sber_settings'))

        cfg = {
            'client_id':    _sber_get(conn, 'sber_client_id', '71154'),
            'client_secret': _sber_get(conn, 'sber_client_secret'),
            'last_sync':    _sber_get(conn, 'sber_last_sync'),
            'last_result':  _sber_get(conn, 'sber_last_result'),
            'has_token':    bool(_sber_get(conn, 'sber_refresh_token')) or _sber_get(conn, 'sber_npa_active') == '1',
        }
        accounts = conn.execute(
            'SELECT id, name, account_number, sber_auto_sync, sber_last_sync, sber_last_result '
            'FROM bank_accounts WHERE is_active=1 ORDER BY name'
        ).fetchall()
    return render_template('sber_settings.html', cfg=cfg, accounts=accounts)


@app.route('/bank/sber/auth')
@login_required
@menu_permission_required('bank')
def sber_auth():
    """Редирект на страницу авторизации СберБизнес (OAuth Authorization Code)."""
    import secrets
    from sber_api import build_auth_url
    with get_db() as conn:
        client_id = _sber_get(conn, 'sber_client_id', '71154')
        state = secrets.token_urlsafe(16)
        nonce = secrets.token_urlsafe(16)
        _sber_set(conn, 'sber_oauth_state', state)
        conn.commit()
    redirect_uri = url_for('sber_callback', _external=True)
    auth_url = build_auth_url(client_id, redirect_uri, state, nonce)
    return redirect(auth_url)


@app.route('/bank/sber/callback')
@login_required
@menu_permission_required('bank')
def sber_callback():
    """Обработчик callback после авторизации в СберБизнес."""
    from sber_api import exchange_code
    import time
    code  = request.args.get('code')
    state = request.args.get('state')
    error = request.args.get('error')

    if error:
        flash(f'Сбербанк отказал в доступе: {error}', 'danger')
        return redirect(url_for('sber_settings'))

    with get_db() as conn:
        saved_state = _sber_get(conn, 'sber_oauth_state')
        if state != saved_state:
            flash('Ошибка безопасности: state не совпадает. Попробуйте снова.', 'danger')
            return redirect(url_for('sber_settings'))

        client_id     = _sber_get(conn, 'sber_client_id', '71154')
        client_secret = _sber_get(conn, 'sber_client_secret')

        try:
            redirect_uri = url_for('sber_callback', _external=True)
            tokens = exchange_code(client_id, client_secret, code, redirect_uri)
            _sber_set(conn, 'sber_access_token',  tokens['access_token'])
            _sber_set(conn, 'sber_refresh_token', tokens.get('refresh_token', ''))
            _sber_set(conn, 'sber_token_expires',
                      str(time.time() + int(tokens.get('expires_in', 3600))))
            _sber_set(conn, 'sber_npa_active', '1')
            conn.commit()
            flash('Сбербанк успешно подключён! Можно загружать выписки.', 'success')
        except Exception as e:
            flash(f'Ошибка получения токена: {e}', 'danger')

    return redirect(url_for('sber_settings'))


def _sber_get_token(conn):
    """Вернуть актуальный access_token, обновив через refresh если нужно. None = нет токена."""
    from sber_api import refresh_access_token
    import time
    client_id     = _sber_get(conn, 'sber_client_id', '71154')
    cached_token  = _sber_get(conn, 'sber_access_token')
    token_exp_str = _sber_get(conn, 'sber_token_expires')
    refresh_token = _sber_get(conn, 'sber_refresh_token')
    now_ts = time.time()
    if cached_token and token_exp_str and float(token_exp_str) > now_ts + 30:
        return cached_token, client_id
    if not refresh_token:
        return None, client_id
    client_secret = _sber_get(conn, 'sber_client_secret')
    tokens = refresh_access_token(client_id, client_secret, refresh_token)
    access_token = tokens['access_token']
    _sber_set(conn, 'sber_access_token',  access_token)
    _sber_set(conn, 'sber_token_expires', str(now_ts + int(tokens.get('expires_in', 3600))))
    if tokens.get('refresh_token'):
        _sber_set(conn, 'sber_refresh_token', tokens['refresh_token'])
    conn.commit()
    return access_token, client_id


def _sber_do_sync(conn, access_token, client_id, bank_account_id, account_number, date_from, date_to):
    """Загрузить выписку для одного счёта. Возвращает dict с ok/added/total/stmt_id."""
    from sber_api import get_statement, parse_transactions
    stmt_json = get_statement(access_token, client_id, account_number, date_from, date_to)
    txns = parse_transactions(stmt_json)
    now_str = datetime.now().strftime('%d.%m.%Y %H:%M')

    if not txns:
        conn.execute(
            "UPDATE bank_accounts SET sber_last_sync=?, sber_last_result=? WHERE id=?",
            (now_str, '0 транзакций', bank_account_id)
        )
        _sber_set(conn, 'sber_last_sync', now_str)
        _sber_set(conn, 'sber_last_result', '0 новых транзакций')
        conn.commit()
        return {'ok': True, 'added': 0, 'total': 0, 'raw': stmt_json}

    conn.execute(
        '''INSERT INTO bank_statements
           (bank_account_id, filename, date_from, date_to, row_count, uploaded_at)
           VALUES(?,?,?,?,?,?)''',
        (bank_account_id, f'Сбербанк {date_from} — {date_to}',
         date_from, date_to, len(txns), datetime.now().isoformat())
    )
    stmt_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]

    txns = _match_contractors(conn, txns)
    for txn in txns:
        if not txn.get('terminal_id'):
            txn['terminal_id'] = _match_terminal(conn, txn)

    added = 0
    new_txn_ids = []
    for txn in txns:
        dup = conn.execute('''
            SELECT id FROM bank_transactions
            WHERE bank_account_id=? AND txn_date=? AND amount=? AND description=?
        ''', (bank_account_id, txn['date'], txn['amount'], txn.get('description', ''))).fetchone()
        if dup:
            continue
        bt_id = conn.execute('''
            INSERT INTO bank_transactions
            (statement_id, bank_account_id, txn_date, amount,
             description, counterparty, contractor_id, category, terminal_id)
            VALUES(?,?,?,?,?,?,?,?,?)
        ''', (
            stmt_id, bank_account_id, txn['date'], txn['amount'],
            txn.get('description', ''), txn.get('counterparty', ''),
            txn.get('contractor_id'), txn.get('category', ''),
            txn.get('terminal_id')
        )).lastrowid
        new_txn_ids.append(bt_id)
        added += 1
    _sync_parse_rule_extractions(conn, new_txn_ids)

    result_str = f'+{added} новых из {len(txns)}'
    conn.execute(
        "UPDATE bank_accounts SET sber_last_sync=?, sber_last_result=? WHERE id=?",
        (now_str, result_str, bank_account_id)
    )
    _sber_set(conn, 'sber_last_sync', now_str)
    _sber_set(conn, 'sber_last_result', result_str)
    conn.commit()
    return {'ok': True, 'added': added, 'total': len(txns), 'stmt_id': stmt_id}


@app.route('/bank/sber/sync', methods=['POST'])
@login_required
@menu_permission_required('bank')
def sber_sync():
    data = request.get_json(silent=True) or {}
    date_from = data.get('date_from') or (date.today() - timedelta(days=7)).isoformat()
    date_to   = data.get('date_to')   or date.today().isoformat()

    with get_db() as conn:
        bank_account_id = data.get('bank_account_id') or _sber_get(conn, 'sber_bank_account_id')

        # account_number: из записи банк. счёта (если указан) или глобальная настройка
        if bank_account_id:
            ba_row = conn.execute('SELECT account_number FROM bank_accounts WHERE id=?',
                                  (bank_account_id,)).fetchone()
            account_number = ba_row['account_number'] if ba_row else ''
        else:
            account_number = _sber_get(conn, 'sber_account_number')

        if not account_number:
            return jsonify({'ok': False, 'error': 'Не задан номер счёта. Укажите его в настройках счёта.'})

        try:
            access_token, client_id = _sber_get_token(conn)
            if not access_token:
                return jsonify({'ok': False, 'error': 'Нет токена авторизации. Нажмите «Переподключить СберБизнес» и войдите снова.'})

            # Если bank_account_id не задан — найти или создать по account_number
            if not bank_account_id:
                ba = conn.execute(
                    "SELECT id FROM bank_accounts WHERE account_number=?", (account_number,)
                ).fetchone()
                if ba:
                    bank_account_id = ba['id']
                else:
                    conn.execute(
                        "INSERT INTO bank_accounts(name, account_number, bank_name, is_active) VALUES(?,?,?,1)",
                        (f'Сбербанк {account_number[-4:]}', account_number, 'Сбербанк')
                    )
                    bank_account_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
                    conn.commit()
                _sber_set(conn, 'sber_bank_account_id', str(bank_account_id))
                conn.commit()

            result = _sber_do_sync(conn, access_token, client_id, bank_account_id,
                                   account_number, date_from, date_to)
            return jsonify(result)

        except Exception as e:
            err = str(e)
            logging.exception('sber_sync error')
            _sber_set(conn, 'sber_last_result', f'Ошибка: {err[:500]}')
            if bank_account_id:
                conn.execute(
                    "UPDATE bank_accounts SET sber_last_result=? WHERE id=?",
                    (f'Ошибка: {err[:200]}', bank_account_id)
                )
            conn.commit()
            return jsonify({'ok': False, 'error': err})


@app.route('/bank/sber/sync-all', methods=['POST'])
@login_required
@menu_permission_required('bank')
def sber_sync_all():
    """Синхронизировать все счета с включённой авто-загрузкой."""
    data = request.get_json(silent=True) or {}
    date_from = data.get('date_from') or (date.today() - timedelta(days=7)).isoformat()
    date_to   = data.get('date_to')   or date.today().isoformat()

    with get_db() as conn:
        accounts = conn.execute(
            "SELECT id, name, account_number FROM bank_accounts "
            "WHERE is_active=1 AND sber_auto_sync=1 AND account_number != '' AND account_number IS NOT NULL"
        ).fetchall()

        if not accounts:
            return jsonify({'ok': True, 'results': [], 'total_added': 0,
                            'message': 'Нет счетов с включённой автозагрузкой'})

        try:
            access_token, client_id = _sber_get_token(conn)
            if not access_token:
                return jsonify({'ok': False, 'error': 'Нет токена авторизации. Откройте «Сбербанк авто-импорт» и переподключите.'})
        except Exception as e:
            return jsonify({'ok': False, 'error': f'Ошибка получения токена: {e}'})

        results = []
        total_added = 0
        for acc in accounts:
            try:
                r = _sber_do_sync(conn, access_token, client_id, acc['id'],
                                  acc['account_number'], date_from, date_to)
                total_added += r.get('added', 0)
                results.append({'id': acc['id'], 'name': acc['name'],
                                'added': r.get('added', 0), 'ok': True})
            except Exception as e:
                err = str(e)[:200]
                conn.execute(
                    "UPDATE bank_accounts SET sber_last_result=? WHERE id=?",
                    (f'Ошибка: {err}', acc['id'])
                )
                conn.commit()
                results.append({'id': acc['id'], 'name': acc['name'], 'ok': False, 'error': err})

        return jsonify({'ok': True, 'results': results, 'total_added': total_added})


@app.route('/bank/accounts/<int:acc_id>/sber-toggle', methods=['POST'])
@login_required
@menu_permission_required('bank')
def bank_account_sber_toggle(acc_id):
    """Включить / выключить авто-синхронизацию Сбербанка для конкретного счёта."""
    with get_db() as conn:
        cur = conn.execute('SELECT sber_auto_sync FROM bank_accounts WHERE id=?', (acc_id,)).fetchone()
        if not cur:
            return jsonify({'ok': False, 'error': 'Счёт не найден'})
        new_val = 0 if cur['sber_auto_sync'] else 1
        conn.execute('UPDATE bank_accounts SET sber_auto_sync=? WHERE id=?', (new_val, acc_id))
        conn.commit()
    return jsonify({'ok': True, 'auto_sync': new_val})


@app.route('/bank/sber/test', methods=['POST'])
@login_required
@menu_permission_required('bank')
def sber_test():
    """Проверка токена: пробуем refresh, если нет — сообщаем переподключиться."""
    from sber_api import refresh_access_token
    import time
    with get_db() as conn:
        client_id     = _sber_get(conn, 'sber_client_id', '71154')
        client_secret = _sber_get(conn, 'sber_client_secret')
        refresh_token = _sber_get(conn, 'sber_refresh_token')
        cached_token  = _sber_get(conn, 'sber_access_token')
        token_exp_str = _sber_get(conn, 'sber_token_expires')
    now_ts = time.time()
    if cached_token and token_exp_str and float(token_exp_str or 0) > now_ts + 30:
        return jsonify({'ok': True, 'expires_in': int(float(token_exp_str) - now_ts)})
    if not refresh_token:
        return jsonify({'ok': False, 'error': 'Нет токена. Нажмите «Переподключить СберБизнес».'})
    try:
        tokens = refresh_access_token(client_id, client_secret, refresh_token)
        with get_db() as conn:
            _sber_set(conn, 'sber_access_token',  tokens['access_token'])
            _sber_set(conn, 'sber_token_expires', str(now_ts + int(tokens.get('expires_in', 3600))))
            if tokens.get('refresh_token'):
                _sber_set(conn, 'sber_refresh_token', tokens['refresh_token'])
            conn.commit()
        return jsonify({'ok': True, 'expires_in': tokens.get('expires_in', 3600)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


# ──────────────────────────────────────────────────────────────────────────────
# ИМПОРТ EXCEL
# ──────────────────────────────────────────────────────────────────────────────

_XL_DAY_SHEETS = {'ПН', 'ВТ', 'СР', 'ЧТ', 'ПТ', 'СБ', 'ВС'}

_XL_ROLE_MAP = {
    'Админ.': 'admin', 'Адм.': 'admin', 'Адм/Упак': 'admin', 'Адм': 'admin',
    'Администратор': 'admin', 'Упак.': 'admin', 'Упак': 'admin',
    'Упаковщик': 'admin', 'Упаковщица': 'admin',
    'Сушист': 'sushi', 'Сушист.': 'sushi',
    'Уборщица': 'cleaner', 'Уборщик': 'cleaner',
    'Повара': 'cook', 'Повар': 'cook', 'Повар.': 'cook',
}

_TRANSLIT_MAP = {
    'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'yo','ж':'zh','з':'z',
    'и':'i','й':'y','к':'k','л':'l','м':'m','н':'n','о':'o','п':'p','р':'r',
    'с':'s','т':'t','у':'u','ф':'f','х':'kh','ц':'ts','ч':'ch','ш':'sh',
    'щ':'sch','ъ':'','ы':'y','ь':'','э':'e','ю':'yu','я':'ya',
}

def _slugify(s):
    s = s.lower()
    out = ''.join(_TRANSLIT_MAP.get(c, c) for c in s)
    out = re.sub(r'[^a-z0-9]+', '_', out).strip('_')
    return out[:30]


_XL_CAT_MAP = {
    'Ремонт сантех.': 'repair_plumbing', 'Чистка жироул-ля': 'repair_grease',
    'Чистка жироуловителя': 'repair_grease', 'Ремонт электрик': 'repair_electric',
    'Ремонт холод.оборуд.': 'repair_fridge', 'Ремонт другой': 'repair_other',
    'Магазин / Апт.': 'shop', 'Магазин/Апт.': 'shop', 'Другое': 'other',
    'Стафф': 'staff', 'Стаф': 'staff', 'СТАФФ': 'staff', 'СТАФ': 'staff',
}

_XL_SKIP = {
    'Курьеры', 'Повара', 'Админ', 'Курьеры:', 'Повара:', 'ФАМИЛИЯ ИМЯ:', 'ФАМИЛИЯ:',
    'Курьер 1', 'Курьер 2', 'Курьер 3', 'Курьер 4', 'Курьер 5', 'Курьер 6',
    'Курьер 1:', 'Курьер 2:', 'Курьер 3:', 'Курьер 4:', 'Курьер 5:', 'Курьер 6:',
}


def _xf(val):
    if val is None or val is False or str(val).strip() in ('х', 'x', 'Х', ''):
        return 0.0
    try:
        return float(val)
    except Exception:
        return 0.0


def _xt(t):
    from datetime import time as _t
    if t is None:
        return 0.0
    if isinstance(t, _t):
        return t.hour + t.minute / 60.0
    try:
        return float(t)
    except Exception:
        return 0.0


def _xts(t):
    from datetime import time as _t, datetime as _dt
    if t is None:
        return None
    if isinstance(t, _t):
        return f"{t.hour:02d}:{t.minute:02d}"
    if isinstance(t, _dt):
        return f"{t.hour:02d}:{t.minute:02d}"
    if isinstance(t, float):
        # Excel stores time as fraction of day: 0.5 = 12:00
        total_min = round(t * 24 * 60)
        h, m = divmod(total_min, 60)
        h = h % 24
        return f"{h:02d}:{m:02d}"
    if isinstance(t, str):
        t = t.strip()
        if ':' in t:
            parts = t.split(':')
            try:
                return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
            except (ValueError, IndexError):
                pass
    return None


def _xl_parse_sheet(ws, branch_id):
    """Parse Excel sheet and return structured dict (no DB writes). Returns None if invalid."""
    from datetime import date as _d, datetime as _dt
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 7:
        return None

    date_val = rows[2][1] if len(rows) > 2 else None
    if isinstance(date_val, _dt):
        shift_date = date_val.date()
    elif isinstance(date_val, _d):
        shift_date = date_val
    else:
        return None

    total_revenue = _xf(rows[1][5]) if len(rows) > 1 else 0
    delivery_rev  = _xf(rows[3][6]) if len(rows) > 3 else 0
    cash_amount   = _xf(rows[4][3]) if len(rows) > 4 else 0
    delivery_ord  = int(_xf(rows[4][6])) if len(rows) > 4 else 0
    card_amount   = _xf(rows[5][3]) if len(rows) > 5 else 0
    pickup_rev    = _xf(rows[5][6]) if len(rows) > 5 else 0
    online_amount = _xf(rows[6][3]) if len(rows) > 6 else 0
    pickup_ord    = int(_xf(rows[6][6])) if len(rows) > 6 else 0

    # Ячейка "Итого выручка" (F2) иногда не заполнена/не пересчитана в файле,
    # хотя нал/безнал/онлайн реально внесены — смена должна считаться и
    # отображаться в выручке, если сумма частей больше 1
    _rev_parts_sum = cash_amount + card_amount + online_amount
    if total_revenue <= 1 and _rev_parts_sum > 1:
        total_revenue = _rev_parts_sum

    actual_cash    = 0.0
    closed_by_name = None
    for r in rows[25:]:
        if r[1] == 'Факт в кассе:':
            actual_cash = _xf(r[3])
        if r[1] == 'Смену закрыл(а):':
            closed_by_name = str(r[4]).strip() if r[4] else None

    terminal_entries = []
    terminal_codes   = []
    for r in rows[28:35]:
        if len(r) < 7:
            continue
        code1 = str(r[4]).strip() if r[4] is not None and str(r[4]).strip() else None
        amt1  = _xf(r[6])
        if code1 and amt1 > 0:
            terminal_entries.append([code1, amt1])
            terminal_codes.append(code1)
        code2 = str(r[5]).strip() if len(r) > 5 and r[5] is not None and str(r[5]).strip() else None
        amt2  = _xf(r[7]) if len(r) > 7 else 0.0
        if code2 and amt2 > 0:
            terminal_entries.append([code2, amt2])
            terminal_codes.append(code2)
    terminal_amount = sum(a for _, a in terminal_entries)

    morning_cash  = _xf(rows[2][3]) if len(rows) > 2 else 0.0
    change_amount = _xf(rows[30][3]) if len(rows) > 30 else 0.0
    plus_amount   = sum(_xf(rows[i][3]) for i in ([28, 29] + list(range(32, 36))) if len(rows) > i)

    # Cash plus entries
    cash_plus = []
    if len(rows) > 28:
        amt = _xf(rows[28][3])
        if amt > 0:
            cash_plus.append({'amount': amt, 'category': 'oil', 'description': ''})
    if len(rows) > 29:
        amt = _xf(rows[29][3])
        if amt > 0:
            cash_plus.append({'amount': amt, 'category': 'fish', 'description': ''})
    for i in range(32, 36):
        if len(rows) <= i:
            break
        amt = _xf(rows[i][3])
        desc_parts = [str(rows[i][ci]).strip() for ci in (1, 2) if len(rows[i]) > ci and rows[i][ci] is not None and str(rows[i][ci]).strip()]
        desc = ' '.join(desc_parts)
        if amt > 0:
            cash_plus.append({'amount': amt, 'category': 'cash_plus', 'description': desc})

    # Expenses
    expenses = []
    _REPAIR_ROWS = {
        9:  'repair_plumbing',
        10: 'repair_grease',
        11: 'repair_electric',
        12: 'repair_fridge',
        13: 'repair_other',
    }
    for row_idx, cat in _REPAIR_ROWS.items():
        if len(rows) <= row_idx:
            continue
        r = rows[row_idx]
        cash_e = _xf(r[5]) if len(r) > 5 else 0.0
        card_e = _xf(r[6]) if len(r) > 6 else 0.0
        if cash_e <= 0 and card_e <= 0:
            continue
        parts  = [str(r[ci]).strip() for ci in (2, 3, 4) if len(r) > ci and r[ci] is not None and str(r[ci]).strip()]
        desc   = ' '.join(parts) if parts else None
        gulash = 1 if len(r) > 7 and r[7] is True else 0
        expenses.append({'category': cat, 'description': desc, 'cash': cash_e, 'card': card_e, 'gulash': gulash})

    for r in rows[14:20]:
        cash_e = _xf(r[5]) if len(r) > 5 else 0.0
        card_e = _xf(r[6]) if len(r) > 6 else 0.0
        if cash_e <= 0 and card_e <= 0:
            continue
        parts  = [str(r[ci]).strip() for ci in (2, 3, 4) if len(r) > ci and r[ci] is not None and str(r[ci]).strip()]
        desc   = ' '.join(parts) if parts else None
        cat    = 'staff' if desc and 'стаф' in desc.lower() else 'shop'
        gulash = 1 if len(r) > 7 and r[7] is True else 0
        expenses.append({'category': cat, 'description': desc, 'cash': cash_e, 'card': card_e, 'gulash': gulash})

    # Taxi
    taxi = []
    for r in rows[21:27]:
        if not r or len(r) < 7:
            continue
        parts  = [str(r[ci]).strip() for ci in (2, 3, 4) if len(r) > ci and r[ci] is not None and str(r[ci]).strip()]
        addr   = ' '.join(parts) if parts else None
        if not addr:
            continue
        card_t   = _xf(r[5])
        cash_t   = _xf(r[6])
        gulash_t = 1 if len(r) > 7 and r[7] else 0
        if cash_t <= 0 and card_t <= 0:
            continue
        taxi.append({'addr': addr, 'cash': cash_t, 'card': card_t, 'gulash': gulash_t})

    # Courier times
    _COURIER_TIMES = [
        (25, 12, 13, 26, 12, 13),
        (25, 17, 18, 26, 17, 18),
        (25, 21, 22, 26, 21, 22),
        (75, 12, 13, 76, 12, 13),
        (75, 17, 18, 76, 17, 18),
        (75, 21, 22, 76, 21, 22),
    ]

    def _read_hm(row_idx, col_h, col_m):
        if len(rows) <= row_idx:
            return None
        r = rows[row_idx]
        h = int(_xf(r[col_h])) if len(r) > col_h and r[col_h] is not None else None
        m = int(_xf(r[col_m])) if len(r) > col_m and r[col_m] is not None else 0
        if h is None:
            return None
        return f"{h:02d}:{m:02d}"

    # Employees
    employees = []

    for ci, r in enumerate(rows[2:8]):
        name = r[10]
        if not name or not isinstance(name, str) or name.strip() in _XL_SKIP:
            continue
        name = name.strip()
        if not name:
            continue
        total = _xf(r[21])
        if total == 0:
            continue
        km      = _xf(r[12]); km_pay  = _xf(r[13])
        hours   = _xf(r[14]); hrs_pay = _xf(r[15])
        orders  = int(_xf(r[16])); ord_pay = _xf(r[17])
        s_label  = str(r[18]).strip() if len(r) > 18 and r[18] else ''
        t_amount = _xf(r[19]) if len(r) > 19 else 0.0
        if 'премия' in s_label.lower():
            bonus = t_amount; penalty = 0.0; comment = ''
        elif 'штраф' in s_label.lower():
            bonus = 0.0; penalty = t_amount; comment = ''
        else:
            bonus = 0.0; penalty = 0.0
            comment = s_label if s_label and s_label != 'Ничего' else ''
        paid     = 1 if r[20] == 'Да' else 0
        rate_km  = round(km_pay / km, 2) if km > 0 else 10.0
        rate_ord = round(ord_pay / orders, 2) if orders > 0 else 100.0
        rate_hr  = round(hrs_pay / hours, 2) if hours > 0 else 0.0
        sr, sh, sm, er, eh, em = _COURIER_TIMES[ci]
        start = _read_hm(sr, sh, sm)
        end   = _read_hm(er, eh, em)
        employees.append({
            'name': name, 'role': 'courier', 'rate': rate_hr,
            'hours': hours, 'km': km, 'orders': orders,
            'rate_km': rate_km, 'rate_ord': rate_ord,
            'start': start, 'end': end,
            'bonus': bonus, 'penalty': penalty, 'comment': comment,
            'base_pay': hrs_pay + km_pay, 'total': total, 'paid': paid,
        })

    _VALID_ROLES = {'admin', 'cook', 'sushi', 'packer', 'cleaner', 'courier'}
    for row_idx, r in enumerate(rows[9:22], start=9):
        name = r[10]
        if not name or not isinstance(name, str) or name.strip() in _XL_SKIP:
            continue
        name = name.strip()
        if not name:
            continue
        total = _xf(r[21])
        if total == 0:
            continue
        l_role = str(r[11]).strip().lower() if len(r) > 11 and r[11] else ''
        if name == 'Уборщица':
            role = 'cleaner'
        elif l_role in _VALID_ROLES:
            role = l_role
        elif row_idx <= 13:
            role = 'admin'
        else:
            role = 'sushi'
        rate    = _xf(r[12])
        start   = _xts(r[13])
        end     = _xts(r[14])
        hours   = _xt(r[15])
        comment = str(r[18]).strip() if r[18] and str(r[18]) != 'Ничего' else ''
        paid    = 1 if r[20] == 'Да' else 0
        t_base  = _xf(r[19]) if len(r) > 19 else 0.0
        if t_base > 0:
            diff    = total - t_base
            bonus   = round(max(0.0, diff), 2)
            penalty = round(max(0.0, -diff), 2)
            base    = t_base
        else:
            bval    = r[17]
            bonus   = _xf(bval) if isinstance(bval, (int, float)) and bval > 0 else 0
            penalty = abs(_xf(bval)) if isinstance(bval, (int, float)) and bval < 0 else 0
            base    = round(rate * hours, 2)
        employees.append({
            'name': name, 'role': role, 'rate': rate,
            'hours': hours, 'km': 0.0, 'orders': 0,
            'rate_km': 10.0, 'rate_ord': 100.0,
            'start': start, 'end': end,
            'bonus': bonus, 'penalty': penalty, 'comment': comment,
            'base_pay': base, 'total': total, 'paid': paid,
        })

    return {
        'date': shift_date.isoformat(),
        'branch_id': branch_id,
        'closed_by_name': closed_by_name,
        'revenue': {
            'total_revenue': total_revenue,
            'delivery_rev': delivery_rev,
            'delivery_ord': delivery_ord,
            'pickup_rev': pickup_rev,
            'pickup_ord': pickup_ord,
            'cash_amount': cash_amount,
            'card_amount': card_amount,
            'online_amount': online_amount,
            'terminal_entries': terminal_entries,
            'terminal_codes': terminal_codes,
            'terminal_amount': terminal_amount,
            'actual_cash': actual_cash,
            'change_amount': change_amount,
            'plus_amount': plus_amount,
            'morning_cash': morning_cash,
        },
        'cash_plus': cash_plus,
        'expenses': expenses,
        'taxi': taxi,
        'employees': employees,
    }


# ─── Альтернативный формат (Красноармейская, Шахтёров, 40 лет Октября, Октябрьский) ───
# Та же смена, но другая раскладка ячеек. Возвращает тот же словарь, что и _xl_parse_sheet.

_XL2_CAT_LEFT = {
    'Пополнение кассы': 'cash_plus',
    'Вывоз фритюрного масла': 'oil',
}
_XL2_CAT_RIGHT = {
    'Забрал из кассы': 'other',
    'Зарплата персоналу': 'staff',
}
_XL2_COURIER_SLOTS = [
    (36, 4, 5, 37, 4, 5),
    (36, 11, 12, 37, 11, 12),
    (36, 18, 19, 37, 18, 19),
    (84, 4, 5, 85, 4, 5),
    (84, 11, 12, 85, 11, 12),
    (84, 18, 19, 85, 18, 19),
]


def _xl_parse_sheet_alt(ws, branch_id):
    """Парсит лист альтернативного формата. Возвращает тот же словарь, что и _xl_parse_sheet, или None."""
    from datetime import date as _d, datetime as _dt
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 35:
        return None

    date_val = rows[1][1] if len(rows) > 1 and len(rows[1]) > 1 else None   # B2
    if isinstance(date_val, _dt):
        shift_date = date_val.date()
    elif isinstance(date_val, _d):
        shift_date = date_val
    else:
        return None

    total_revenue = _xf(rows[1][5]) if len(rows) > 1 and len(rows[1]) > 5 else 0   # F2
    cash_amount   = _xf(rows[3][3]) if len(rows) > 3 and len(rows[3]) > 3 else 0   # D4
    card_amount   = _xf(rows[4][3]) if len(rows) > 4 and len(rows[4]) > 3 else 0   # D5
    online_amount = _xf(rows[5][3]) if len(rows) > 5 and len(rows[5]) > 3 else 0   # D6

    _rev_parts_sum = cash_amount + card_amount + online_amount
    if total_revenue <= 1 and _rev_parts_sum > 1:
        total_revenue = _rev_parts_sum

    morning_cash  = _xf(rows[1][3]) if len(rows) > 1 and len(rows[1]) > 3 else 0.0    # D2 "РАЗМЕН" = утренняя касса
    actual_cash   = _xf(rows[33][3]) if len(rows) > 33 and len(rows[33]) > 3 else 0.0  # D34 "Итого нал. в кассе"
    change_amount = 0.0
    # Деления заказов на доставку/самовывоз в этом формате нет — общее число
    # («Кол-во заказов Папа-Суши», D3) целиком идёт в «Доставка», т.к. это сеть доставки.
    delivery_ord  = int(_xf(rows[2][3])) if len(rows) > 2 and len(rows[2]) > 3 else 0   # D3
    delivery_rev = pickup_rev = pickup_ord = 0

    # Терминалы: пары код/сумма на строках 25-27 (E/F)
    terminal_entries = []
    terminal_codes   = []
    for r in rows[24:27]:
        if len(r) < 6:
            continue
        code = str(r[4]).strip() if r[4] is not None and str(r[4]).strip() else None
        amt  = _xf(r[5])
        if code and amt > 0:
            terminal_entries.append([code, amt])
            terminal_codes.append(code)
    terminal_amount = sum(a for _, a in terminal_entries)

    # Плюсы в кассу / расходы — построчный разбор с текущей категорией (как в исходном
    # скрипте импорта): строка-заголовок категории может сразу нести сумму, а следующие
    # строки без нового заголовка — дополнительные позиции той же категории.
    cash_plus = []
    expenses  = []
    cur_left  = None
    cur_right = None
    for r in rows[14:23]:
        if len(r) < 6:
            continue
        b_text = str(r[1]).strip() if r[1] is not None and str(r[1]).strip() else None
        d_amt  = _xf(r[3])
        e_text = str(r[4]).strip() if r[4] is not None and str(r[4]).strip() else None
        f_amt  = _xf(r[5])

        desc_left = ''
        if b_text and b_text in _XL2_CAT_LEFT:
            cur_left = _XL2_CAT_LEFT[b_text]
        elif b_text and b_text != 'Комментарий:':
            desc_left = b_text
        if cur_left and d_amt > 0:
            cash_plus.append({'amount': d_amt, 'category': cur_left, 'description': desc_left})

        desc_right = ''
        if e_text and e_text in _XL2_CAT_RIGHT:
            cur_right = _XL2_CAT_RIGHT[e_text]
        elif e_text and e_text not in ('Комментарий:', 'Прочие расходы'):
            desc_right = e_text
        if cur_right and f_amt > 0:
            expenses.append({'category': cur_right, 'description': desc_right, 'cash': f_amt, 'card': 0.0, 'gulash': 0})

    plus_amount = sum(cp['amount'] for cp in cash_plus)

    # Такси: после заголовка (строка 26) — адрес в B, сумма в D (одна колонка, без деления нал/безнал)
    taxi = []
    for r in rows[26:33]:
        if len(r) < 4:
            continue
        addr = str(r[1]).strip() if r[1] is not None and str(r[1]).strip() else None
        if not addr:
            continue
        amt = _xf(r[3])
        if amt <= 0:
            continue
        taxi.append({'addr': addr, 'cash': amt, 'card': 0.0, 'gulash': 0})

    def _read_hm2(row_idx, col_h, col_m):
        if len(rows) <= row_idx:
            return None
        r = rows[row_idx]
        h = int(_xf(r[col_h])) if len(r) > col_h and r[col_h] is not None else None
        m = int(_xf(r[col_m])) if len(r) > col_m and r[col_m] is not None else 0
        if h is None:
            return None
        return f"{h:02d}:{m:02d}"

    def _is_paid(r, idx):
        # Учитывается только да/нет. Любая другая пометка (например «зп 2р/мес»)
        # для импорта — то же самое, что «нет»: за эту смену не выплачено.
        # Настройка ежемесячной оплаты в профиле сотрудника импортом не трогается.
        return 1 if len(r) > idx and r[idx] is not None and str(r[idx]).strip().lower() == 'да' else 0

    # Сотрудники
    employees = []

    # Курьеры — строки 3-8, колонки I..T
    for ci, r in enumerate(rows[2:8]):
        if len(r) < 19:
            continue
        name = r[8]   # I
        if not name or not isinstance(name, str) or name.strip() in _XL_SKIP:
            continue
        name = name.strip()
        if not name:
            continue
        total = _xf(r[18])   # S
        if total == 0:
            continue
        km      = _xf(r[9])                          # J
        hours   = _xf(r[10])                          # K
        hrs_pay = _xf(r[11])                          # L
        orders  = int(_xf(r[12]))                     # M
        ord_pay = _xf(r[14])                          # O
        s_label = str(r[15]).strip() if len(r) > 15 and r[15] else ''   # P
        t_amount = _xf(r[16]) if len(r) > 16 else 0.0                    # Q
        if 'премия' in s_label.lower():
            bonus = t_amount; penalty = 0.0; comment = ''
        elif 'штраф' in s_label.lower():
            bonus = 0.0; penalty = t_amount; comment = ''
        else:
            bonus = 0.0; penalty = 0.0
            comment = s_label if s_label and s_label != 'Ничего' else ''
        paid     = _is_paid(r, 17)                    # R
        rate_km  = 10.0
        rate_ord = round(ord_pay / orders, 2) if orders > 0 else 100.0
        rate_hr  = round(hrs_pay / hours, 2) if hours > 0 else 0.0
        sr, sh, sm, er, eh, em = _XL2_COURIER_SLOTS[ci]
        start = _read_hm2(sr, sh, sm)
        end   = _read_hm2(er, eh, em)
        employees.append({
            'name': name, 'role': 'courier', 'rate': rate_hr,
            'hours': hours, 'km': km, 'orders': orders,
            'rate_km': rate_km, 'rate_ord': rate_ord,
            'start': start, 'end': end,
            'bonus': bonus, 'penalty': penalty, 'comment': comment,
            'base_pay': hrs_pay + km * rate_km, 'total': total, 'paid': paid,
        })

    # Остальные сотрудники — строки 10-34. Роль по фиксированным диапазонам
    # (метки «Админ»/«Повара»/«Другое» стоят на строках 13/20/32 во всём шаблоне).
    for row_num, r in enumerate(rows[9:34], start=10):
        if len(r) < 19:
            continue
        name = r[8]   # I
        if not name or not isinstance(name, str) or name.strip() in _XL_SKIP:
            continue
        name = name.strip()
        if not name:
            continue
        total = _xf(r[18])   # S
        if total == 0:
            continue
        if name == 'Уборщица':
            role = 'cleaner'
        elif row_num <= 19:
            role = 'admin'
        elif row_num <= 31:
            role = 'sushi'
        else:
            role = 'admin'
        rate    = _xf(r[9])                            # J
        start   = _xts(r[10])                           # K
        end     = _xts(r[11])                           # L
        hours   = _xt(r[12])                             # M
        s_label = str(r[15]).strip() if len(r) > 15 and r[15] else ''   # P
        t_amount = _xf(r[16]) if len(r) > 16 else 0.0                    # Q
        if 'премия' in s_label.lower():
            bonus = t_amount; penalty = 0.0; comment = ''
        elif 'штраф' in s_label.lower():
            bonus = 0.0; penalty = t_amount; comment = ''
        else:
            bonus = 0.0; penalty = 0.0
            comment = s_label if s_label and s_label != 'Ничего' else ''
        paid = _is_paid(r, 17)                          # R
        base = round(rate * hours, 2)
        employees.append({
            'name': name, 'role': role, 'rate': rate,
            'hours': hours, 'km': 0.0, 'orders': 0,
            'rate_km': 10.0, 'rate_ord': 100.0,
            'start': start, 'end': end,
            'bonus': bonus, 'penalty': penalty, 'comment': comment,
            'base_pay': base, 'total': total, 'paid': paid,
        })

    return {
        'date': shift_date.isoformat(),
        'branch_id': branch_id,
        'closed_by_name': None,
        'force_closed': True,
        'revenue': {
            'total_revenue': total_revenue,
            'delivery_rev': delivery_rev,
            'delivery_ord': delivery_ord,
            'pickup_rev': pickup_rev,
            'pickup_ord': pickup_ord,
            'cash_amount': cash_amount,
            'card_amount': card_amount,
            'online_amount': online_amount,
            'terminal_entries': terminal_entries,
            'terminal_codes': terminal_codes,
            'terminal_amount': terminal_amount,
            'actual_cash': actual_cash,
            'change_amount': change_amount,
            'plus_amount': plus_amount,
            'morning_cash': morning_cash,
        },
        'cash_plus': cash_plus,
        'expenses': expenses,
        'taxi': taxi,
        'employees': employees,
    }


_XL_ALT_BRANCH_NAMES = {
    '40 лет октября', 'красноармейская', 'октябрьский', 'шахтеров', 'шахтёров',
}


def _xl_import_sheet_from_parsed(sheet, emp_map, conn, stats, batch_id=None):
    """Write a parsed sheet dict to the DB using emp_map {name -> employee_id}."""
    branch_id      = sheet['branch_id']
    shift_date     = sheet['date']
    rev            = sheet['revenue']
    closed_by_name = sheet['closed_by_name']

    existing = conn.execute(
        "SELECT id FROM shifts WHERE branch_id=? AND date=?", (branch_id, shift_date)
    ).fetchone()
    if existing:
        # Смена уже есть в системе — импорт её не трогает и не дописывает данные,
        # чтобы не задвоить расходы/премии и не затереть то, что уже внесено вручную.
        stats['skipped'] = stats.get('skipped', 0) + 1
        return

    status = 'closed' if (closed_by_name or sheet.get('force_closed')) else 'open'
    conn.execute(
        "INSERT INTO shifts (branch_id, date, status, closed_by_name, import_batch_id) VALUES (?,?,?,?,?)",
        (branch_id, shift_date, status, closed_by_name, batch_id)
    )
    shift_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    stats['shifts'] += 1

    if not conn.execute("SELECT id FROM shift_revenue WHERE shift_id=?", (shift_id,)).fetchone():
        conn.execute(
            """INSERT INTO shift_revenue
               (shift_id, total_revenue, delivery_revenue, delivery_orders,
                pickup_revenue, pickup_orders, cash_amount, card_amount,
                online_amount, terminal_last3, terminal_amount, actual_cash,
                change_amount, plus_amount, morning_cash)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (shift_id, rev['total_revenue'], rev['delivery_rev'], rev['delivery_ord'],
             rev['pickup_rev'], rev['pickup_ord'], rev['cash_amount'], rev['card_amount'],
             rev['online_amount'], ','.join(rev['terminal_codes']), rev['terminal_amount'],
             rev['actual_cash'], rev['change_amount'], rev['plus_amount'], rev['morning_cash'])
        )

    if rev['terminal_entries'] and not conn.execute(
            "SELECT id FROM shift_terminals WHERE shift_id=?", (shift_id,)).fetchone():
        for si, (tn, ta) in enumerate(rev['terminal_entries']):
            conn.execute(
                "INSERT INTO shift_terminals (shift_id, terminal_number, amount, sort_order) VALUES (?,?,?,?)",
                (shift_id, tn, ta, si)
            )

    if not conn.execute("SELECT id FROM cash_plus_entries WHERE shift_id=?", (shift_id,)).fetchone():
        for cp in sheet['cash_plus']:
            conn.execute(
                "INSERT INTO cash_plus_entries (shift_id, amount, amount_cash, category, description) VALUES (?,?,?,?,?)",
                (shift_id, cp['amount'], cp['amount'], cp['category'], cp.get('description', ''))
            )

    for exp in sheet['expenses']:
        conn.execute(
            "INSERT INTO expenses (shift_id,category,description,amount_cash,amount_card,is_gulash) VALUES (?,?,?,?,?,?)",
            (shift_id, exp['category'], exp.get('description'), exp['cash'], exp['card'], exp['gulash'])
        )
        stats['expenses'] += 1

    if not conn.execute("SELECT id FROM taxi_trips WHERE shift_id=?", (shift_id,)).fetchone():
        for t in sheet['taxi']:
            pay_type = 'cash' if t['cash'] > 0 else 'card'
            amount   = t['cash'] if t['cash'] > 0 else t['card']
            conn.execute(
                "INSERT INTO taxi_trips (shift_id, amount, payment_type, in_gulyash) VALUES (?,?,?,?)",
                (shift_id, amount, pay_type, t['gulash'])
            )
            trip_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            conn.execute(
                "INSERT INTO taxi_trip_employees (trip_id, name_snapshot, address_snapshot) VALUES (?,?,?)",
                (trip_id, '—', t['addr'])
            )
            stats['expenses'] += 1

    for emp in sheet['employees']:
        name   = emp['name']
        emp_id = emp_map.get(name)
        if emp_id is None:
            continue
        if conn.execute(
            "SELECT id FROM employee_shifts WHERE shift_id=? AND employee_id=?", (shift_id, emp_id)
        ).fetchone():
            continue
        conn.execute(
            """INSERT INTO employee_shifts
               (shift_id,employee_id,full_name_snapshot,role_snapshot,
                rate_snapshot,rate_per_km_snapshot,rate_per_order_snapshot,
                hours_worked,km,orders,shift_start,shift_end,
                bonus_amount,penalty_amount,bonus_comment,
                base_pay,total_amount,is_paid,paid_amount)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (shift_id, emp_id, name, emp['role'], emp['rate'], emp['rate_km'], emp['rate_ord'],
             emp['hours'], emp['km'], emp['orders'], emp['start'], emp['end'],
             emp['bonus'], emp['penalty'], emp['comment'],
             emp['base_pay'], emp['total'], emp['paid'], emp['total'] if emp['paid'] else 0)
        )
        stats['employee_shifts'] += 1


@app.route('/excel-import', methods=['GET', 'POST'])
@login_required
@menu_permission_required('import_shifts')
def excel_import():
    return redirect(url_for('import_shifts'))


@app.route('/excel-import/batch/<int:batch_id>/delete', methods=['POST'])
@login_required
@menu_permission_required('import_shifts')
def delete_import_batch(batch_id):
    with get_db() as conn:
        batch = conn.execute('SELECT * FROM import_batches WHERE id=?', (batch_id,)).fetchone()
        if not batch:
            flash('Импорт не найден', 'danger')
            return redirect(url_for('excel_import'))

        shift_ids = [r[0] for r in conn.execute(
            'SELECT id FROM shifts WHERE import_batch_id=?', (batch_id,)
        ).fetchall()]

        if shift_ids:
            ids_str = ','.join(str(i) for i in shift_ids)
            conn.execute(f'DELETE FROM salary_payments WHERE employee_shift_id IN (SELECT id FROM employee_shifts WHERE shift_id IN ({ids_str}))')
            conn.execute(f'DELETE FROM employee_shifts WHERE shift_id IN ({ids_str})')
            conn.execute(f'DELETE FROM expenses WHERE shift_id IN ({ids_str})')
            conn.execute(f'DELETE FROM taxi_trip_employees WHERE trip_id IN (SELECT id FROM taxi_trips WHERE shift_id IN ({ids_str}))')
            conn.execute(f'DELETE FROM taxi_trips WHERE shift_id IN ({ids_str})')
            conn.execute(f'DELETE FROM shift_revenue WHERE shift_id IN ({ids_str})')
            conn.execute(f'DELETE FROM change_log WHERE shift_id IN ({ids_str})')
            conn.execute(f'DELETE FROM shifts WHERE id IN ({ids_str})')

        conn.execute('DELETE FROM import_batches WHERE id=?', (batch_id,))
        conn.commit()

    flash(f'Импорт «{batch["filename"]}» удалён ({len(shift_ids)} смен)', 'success')
    return redirect(url_for('import_shifts'))


@app.route('/shifts/bulk-delete', methods=['GET', 'POST'])
@login_required
@menu_permission_required('import_shifts')
def shifts_bulk_delete():
    with get_db() as conn:
        branches = conn.execute(
            "SELECT id, name FROM branches WHERE is_active=1 ORDER BY name"
        ).fetchall()

    date_from = request.values.get('date_from', '')
    date_to   = request.values.get('date_to', '')
    branch_id = request.values.get('branch_id', '')
    action    = request.values.get('action', '')

    preview_shifts = []
    if (date_from or date_to) and action in ('preview', 'delete'):
        conditions = []
        params = []
        if date_from:
            conditions.append('s.date >= ?')
            params.append(date_from)
        if date_to:
            conditions.append('s.date <= ?')
            params.append(date_to)
        if branch_id:
            conditions.append('s.branch_id = ?')
            params.append(int(branch_id))
        where = 'WHERE ' + ' AND '.join(conditions) if conditions else ''

        with get_db() as conn:
            preview_shifts = conn.execute(f'''
                SELECT s.id, s.date, s.status, b.name as branch_name,
                       COUNT(DISTINCT es.id) as staff_count,
                       COUNT(DISTINCT e.id)  as expense_count
                FROM shifts s
                JOIN branches b ON b.id = s.branch_id
                LEFT JOIN employee_shifts es ON es.shift_id = s.id
                LEFT JOIN expenses e ON e.shift_id = s.id
                {where}
                GROUP BY s.id
                ORDER BY s.date DESC, b.name
            ''', params).fetchall()

            if action == 'delete' and preview_shifts:
                shift_ids = [r['id'] for r in preview_shifts]
                ids_str = ','.join(str(i) for i in shift_ids)
                conn.execute(f'DELETE FROM salary_payments WHERE employee_shift_id IN (SELECT id FROM employee_shifts WHERE shift_id IN ({ids_str}))')
                conn.execute(f'DELETE FROM employee_shifts WHERE shift_id IN ({ids_str})')
                conn.execute(f'DELETE FROM expenses WHERE shift_id IN ({ids_str})')
                conn.execute(f'DELETE FROM shift_revenue WHERE shift_id IN ({ids_str})')
                conn.execute(f'DELETE FROM shifts WHERE id IN ({ids_str})')
                conn.commit()
                flash(f'Удалено {len(shift_ids)} смен(ы)', 'success')
                return redirect(url_for('shifts_bulk_delete'))

    return render_template('bulk_delete_shifts.html',
                           branches=branches,
                           date_from=date_from,
                           date_to=date_to,
                           branch_id=branch_id,
                           preview_shifts=preview_shifts)


@app.route('/gdrive-import', methods=['GET', 'POST'])
@login_required
@menu_permission_required('import_shifts')
def import_shifts():
    import io as _io, json as _json, uuid as _uuid

    with get_db() as conn:
        branches = conn.execute(
            "SELECT id, name FROM branches WHERE is_active=1 ORDER BY name"
        ).fetchall()
        import_batches = conn.execute('''
            SELECT ib.*, b.name as branch_name
            FROM import_batches ib
            JOIN branches b ON b.id = ib.branch_id
            ORDER BY ib.imported_at DESC LIMIT 50
        ''').fetchall()

    if request.method == 'POST':
        branch_id     = request.form.get('branch_id', '')
        branch_id_int = int(branch_id) if branch_id else None
        files         = request.files.getlist('files')

        if not branch_id_int:
            flash('Выберите филиал', 'danger')
            return redirect(url_for('import_shifts'))
        if not files or all(f.filename == '' for f in files):
            flash('Выберите хотя бы один файл', 'danger')
            return redirect(url_for('import_shifts'))

        try:
            import openpyxl
        except ImportError:
            flash('Библиотека openpyxl не установлена на сервере', 'danger')
            return redirect(url_for('import_shifts'))

        with get_db() as conn:
            branch_row = conn.execute(
                "SELECT name FROM branches WHERE id=?", (branch_id_int,)
            ).fetchone()
        branch_name_norm = (branch_row['name'] if branch_row else '').strip().lower()
        parse_fn = _xl_parse_sheet_alt if branch_name_norm in _XL_ALT_BRANCH_NAMES else _xl_parse_sheet

        # Parse all sheets without writing to DB
        parsed_files = []
        for file in files:
            if not file.filename.lower().endswith('.xlsx'):
                continue
            data = file.read()
            wb   = openpyxl.load_workbook(_io.BytesIO(data), data_only=True)
            sheets = []
            for sheet_name in wb.sheetnames:
                if sheet_name not in _XL_DAY_SHEETS:
                    continue
                parsed = parse_fn(wb[sheet_name], branch_id_int)
                if parsed:
                    parsed['sheet_name'] = sheet_name
                    sheets.append(parsed)
            if sheets:
                parsed_files.append({'filename': file.filename, 'sheets': sheets})

        if not parsed_files:
            flash('Подходящих листов (ПН–ВС) с данными не найдено в файлах', 'warning')
            return redirect(url_for('import_shifts'))

        # Collect unique employees (by name)
        unique_emp_map = {}   # name -> {name, roles: set, rate, first_role}
        for pf in parsed_files:
            for sheet in pf['sheets']:
                for emp in sheet['employees']:
                    n = emp['name']
                    if n not in unique_emp_map:
                        unique_emp_map[n] = {'name': n, 'roles': [], 'rate': emp['rate']}
                    if emp['role'] not in unique_emp_map[n]['roles']:
                        unique_emp_map[n]['roles'].append(emp['role'])
        unique_employees = list(unique_emp_map.values())

        # Store parsed data to staging table
        token = _uuid.uuid4().hex
        staging_data = _json.dumps({
            'branch_id': branch_id_int,
            'files': parsed_files,
            'unique_employees': unique_employees,
            'user_id': session.get('user_id'),
        }, ensure_ascii=False)

        with get_db() as conn:
            # Clean up old staging records (> 1 hour)
            conn.execute("DELETE FROM import_staging WHERE created_at < datetime('now', '-1 hour')")
            conn.execute(
                "INSERT INTO import_staging (token, branch_id, data) VALUES (?,?,?)",
                (token, branch_id_int, staging_data)
            )
            conn.commit()

        return redirect(url_for('import_shifts_confirm', token=token))

    return render_template('import_shifts.html',
                           branches=branches, import_batches=import_batches)


@app.route('/gdrive-import/confirm/<token>', methods=['GET', 'POST'])
@login_required
@menu_permission_required('import_shifts')
def import_shifts_confirm(token):
    import json as _json

    with get_db() as conn:
        row = conn.execute(
            "SELECT data FROM import_staging WHERE token=?", (token,)
        ).fetchone()
        if not row:
            flash('Сессия истекла или ссылка недействительна. Загрузите файлы снова.', 'danger')
            return redirect(url_for('import_shifts'))
        staging = _json.loads(row['data'])

        if request.method == 'GET':
            # Сотрудники только из филиалов той же группы, что и импортируемый филиал
            # (тот же принцип, что при добавлении сотрудника в смену — см. shift_view)
            group_branch_ids = conn.execute('''
                SELECT DISTINCT bgm2.branch_id
                FROM branch_group_members bgm1
                JOIN branch_group_members bgm2 ON bgm2.group_id = bgm1.group_id
                WHERE bgm1.branch_id = ?
            ''', (staging['branch_id'],)).fetchall()
            group_branch_ids = [r[0] for r in group_branch_ids] or [staging['branch_id']]
            gph = ','.join('?' * len(group_branch_ids))
            employees_db = conn.execute(
                f"""SELECT DISTINCT e.id, e.full_name, e.role, b.name as branch_name
                   FROM employees e
                   JOIN employee_branches eb ON eb.employee_id = e.id
                   LEFT JOIN branches b ON b.id = e.branch_id
                   WHERE eb.branch_id IN ({gph}) AND e.is_active=1 AND COALESCE(e.is_fired,0)=0
                   ORDER BY e.full_name""",
                group_branch_ids
            ).fetchall()
            branch = conn.execute(
                "SELECT name FROM branches WHERE id=?", (staging['branch_id'],)
            ).fetchone()

            # Build auto-suggestion: match by full_name (case-insensitive)
            db_names = {e['full_name'].lower(): e['id'] for e in employees_db}
            suggestions = {}
            for emp in staging['unique_employees']:
                match_id = db_names.get(emp['name'].lower())
                suggestions[emp['name']] = match_id  # None if no match

            # Count sheets/dates across all files
            total_sheets = sum(len(pf['sheets']) for pf in staging['files'])

            # Плоский список (файл, лист, сотрудник) для ручной правки сумм на подтверждении
            payout_rows = []
            for fi, pf in enumerate(staging['files']):
                for si, sheet in enumerate(pf['sheets']):
                    for ei, emp in enumerate(sheet['employees']):
                        payout_rows.append({
                            'fi': fi, 'si': si, 'ei': ei,
                            'filename': pf['filename'],
                            'date': sheet['date'],
                            'name': emp['name'],
                            'role': emp['role'],
                            'total': emp['total'],
                            'bonus': emp['bonus'],
                        })
            payout_rows.sort(key=lambda x: (x['date'], x['name']))

            return render_template(
                'import_shifts_confirm.html',
                staging=staging,
                employees_db=employees_db,
                branch_name=branch['name'] if branch else '—',
                suggestions=suggestions,
                total_sheets=total_sheets,
                token=token,
                role_labels=ROLE_LABELS,
                payout_rows=payout_rows,
            )

        # POST: process mapping and do the actual import
        unique_employees = staging['unique_employees']
        emp_map = {}   # name -> employee_id

        # Пропускать сотрудников нельзя — для каждого нужно выбрать существующего
        # или создать нового. Проверяем это до того, как что-либо записано в базу.
        unresolved = [
            emp['name'] for i, emp in enumerate(unique_employees)
            if not (request.form.get(f'action_{i}', '') or '').strip()
            or (request.form.get(f'action_{i}', '') not in ('new',)
                and not request.form.get(f'action_{i}', '').startswith('existing:'))
        ]
        if unresolved:
            flash(
                'Для этих сотрудников нужно выбрать существующего или создать нового: '
                + ', '.join(unresolved),
                'danger'
            )
            return redirect(url_for('import_shifts_confirm', token=token))

        for i, emp in enumerate(unique_employees):
            name   = emp['name']
            action = request.form.get(f'action_{i}', '')

            if action.startswith('existing:'):
                emp_id = int(action.split(':', 1)[1])
                emp_map[name] = emp_id
                # Должности из файла, которых у сотрудника ещё нет, добавляются как
                # дополнительные автоматически — без отдельного вопроса пользователю.
                cur_role_row = conn.execute("SELECT role FROM employees WHERE id=?", (emp_id,)).fetchone()
                cur_role = cur_role_row['role'] if cur_role_row else None
                for role in emp['roles']:
                    if role == cur_role:
                        continue
                    conn.execute(
                        "INSERT OR IGNORE INTO employee_roles (employee_id, role, is_active) VALUES (?,?,1)",
                        (emp_id, role)
                    )
            elif action == 'new':
                new_name = request.form.get(f'new_name_{i}', name).strip() or name
                new_rate = float(request.form.get(f'new_rate_{i}', '0') or '0')
                primary_role = emp['roles'][0] if emp['roles'] else 'admin'
                conn.execute(
                    "INSERT INTO employees (branch_id, full_name, role, rate, is_active) VALUES (?,?,?,?,1)",
                    (staging['branch_id'], new_name, primary_role, new_rate)
                )
                new_emp_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                emp_map[name] = new_emp_id
                conn.execute(
                    "INSERT OR IGNORE INTO employee_branches (employee_id, branch_id) VALUES (?,?)",
                    (new_emp_id, staging['branch_id'])
                )
                for role in emp['roles'][1:]:
                    conn.execute(
                        "INSERT OR IGNORE INTO employee_roles (employee_id, role, is_active) VALUES (?,?,1)",
                        (new_emp_id, role)
                    )
            # action == 'skip': emp_map[name] stays absent → employee skipped

        # Ручные правки «Итого к выплате» / «Премия» с экрана подтверждения
        for fi, pf in enumerate(staging['files']):
            for si, sheet in enumerate(pf['sheets']):
                for ei, emp in enumerate(sheet['employees']):
                    total_raw = request.form.get(f'total_override_{fi}_{si}_{ei}', '').strip()
                    bonus_raw = request.form.get(f'bonus_override_{fi}_{si}_{ei}', '').strip()
                    if total_raw != '':
                        try:
                            emp['total'] = float(total_raw)
                        except ValueError:
                            pass
                    if bonus_raw != '':
                        try:
                            emp['bonus'] = float(bonus_raw)
                        except ValueError:
                            pass

        # Import all sheets
        total_stats = {'shifts': 0, 'expenses': 0, 'employees': 0, 'employee_shifts': 0, 'skipped': 0}
        for pf in staging['files']:
            conn.execute(
                "INSERT INTO import_batches (branch_id, filename, imported_by) VALUES (?,?,?)",
                (staging['branch_id'], pf['filename'], staging.get('user_id'))
            )
            batch_id   = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            file_stats = {'shifts': 0, 'expenses': 0, 'employees': 0, 'employee_shifts': 0, 'skipped': 0}

            for sheet in pf['sheets']:
                _xl_import_sheet_from_parsed(sheet, emp_map, conn, file_stats, batch_id=batch_id)

            conn.execute(
                '''UPDATE import_batches SET shifts_created=?, expenses_created=?,
                   employees_created=?, employee_shifts_created=? WHERE id=?''',
                (file_stats['shifts'], file_stats['expenses'],
                 file_stats['employees'], file_stats['employee_shifts'], batch_id)
            )
            for k in total_stats:
                total_stats[k] += file_stats[k]
            skipped_note = f', пропущено (уже есть в системе) {file_stats["skipped"]}' if file_stats['skipped'] else ''
            flash(
                f'«{pf["filename"]}»: смен +{file_stats["shifts"]}, '
                f'расходов +{file_stats["expenses"]}, '
                f'записей ЗП +{file_stats["employee_shifts"]}'
                f'{skipped_note}',
                'success'
            )

        conn.execute("DELETE FROM import_staging WHERE token=?", (token,))
        conn.commit()

    return redirect(url_for('import_shifts'))


# ─── ОТЧЁТ ПО ЗАКАЗАМ (импорт выгрузки «отчёт по заказам» из iiko) ───────────

def _normalize_order_type(raw):
    """Зал и С собой считаем одним типом — «Общий - самовывоз». Доставка остаётся как есть."""
    raw = (raw or '').strip()
    if raw in ('Зал', 'С собой'):
        return 'Общий - самовывоз'
    return raw or '—'


def _orders_csv_col(header, *names):
    def _norm(s):
        return re.sub(r'\s+', ' ', (s or '').strip().lower().replace('ё', 'е'))
    normed = [_norm(h) for h in header]
    for n in names:
        n2 = _norm(n)
        if n2 in normed:
            return normed.index(n2)
    return None


def _parse_orders_csv(file_bytes):
    """Разбирает CSV-выгрузку «Отчёт по заказам» (iiko, разделитель ';')."""
    text = None
    for enc in ('utf-8-sig', 'cp1251', 'utf-8'):
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = file_bytes.decode('utf-8', errors='replace')

    lines = text.splitlines()
    header_idx = next((i for i, l in enumerate(lines) if 'Номер заказа' in l), None)
    if header_idx is None:
        return []

    reader = csv.reader(lines[header_idx:], delimiter=';')
    header = next(reader)

    def idx_or(default, *names):
        v = _orders_csv_col(header, *names)
        return v if v is not None else default

    idx_number   = idx_or(1, 'Номер заказа')
    idx_branch   = idx_or(2, 'Подразделение')
    idx_received = idx_or(3, 'Время приема')
    idx_promised = _orders_csv_col(header, 'Обещанное время доставки с момента приема заказа')
    idx_type     = idx_or(5, 'Тип заказа')
    idx_ready    = _orders_csv_col(header, 'Время между приемом и готовностью заказа')
    idx_delivery = _orders_csv_col(header, 'Время между приемом и фактической доставкой заказа')
    idx_promo    = _orders_csv_col(header, 'Промокод')
    idx_amount   = idx_or(9, 'К оплате', 'Оплачено')
    idx_new_cli  = _orders_csv_col(header, 'Новый клиент')

    def cell(row, idx):
        if idx is None or idx >= len(row):
            return ''
        return (row[idx] or '').strip()

    def to_int(v):
        return int(v) if v.isdigit() else None

    rows = []
    for row in reader:
        received_raw = cell(row, idx_received)
        if not received_raw:
            continue  # пустая строка или итоговая строка с суммой в конце файла
        try:
            dt = datetime.strptime(received_raw, '%d.%m.%Y %H:%M')
        except ValueError:
            continue
        order_number = cell(row, idx_number)
        if not order_number:
            continue
        amount_raw = cell(row, idx_amount).replace(' ', '').replace(',', '.')
        try:
            amount = float(amount_raw) if amount_raw else 0.0
        except ValueError:
            amount = 0.0
        type_raw = cell(row, idx_type)

        rows.append({
            'order_number':     order_number,
            'branch_raw':       cell(row, idx_branch) or '—',
            'received_at':      dt.strftime('%Y-%m-%d %H:%M:%S'),
            'promised_minutes': to_int(cell(row, idx_promised)),
            'order_type_raw':   type_raw,
            'order_type':       _normalize_order_type(type_raw),
            'ready_minutes':    to_int(cell(row, idx_ready)),
            'delivery_minutes': to_int(cell(row, idx_delivery)),
            'promo_code':       cell(row, idx_promo) or None,
            'amount':           amount,
            'new_client':       cell(row, idx_new_cli) or None,
        })
    return rows


@app.route('/orders-report')
@login_required
@menu_permission_required('orders_report')
def orders_report():
    with get_db() as conn:
        branches = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
        branch_groups = get_branch_groups(conn)

        bounds = conn.execute(
            'SELECT MIN(received_at), MAX(received_at) FROM orders_report'
        ).fetchone()
        data_min = (bounds[0] or '')[:10]
        data_max = (bounds[1] or '')[:10]

        today       = date.today().isoformat()
        month_start = date.today().replace(day=1).isoformat()
        date_from  = request.args.get('date_from', month_start)
        date_to    = request.args.get('date_to', today)
        branch_flt = [int(b) for b in request.args.getlist('branch_ids') if b.isdigit()]
        type_flt   = request.args.get('type', '')
        q          = request.args.get('q', '').strip()

        where  = []
        params = []

        if q:
            where.append('(order_number LIKE ? OR promo_code LIKE ?)')
            like = f'%{q}%'
            params.extend([like, like])
        else:
            where.append('received_at >= ?')
            params.append(date_from + ' 00:00:00')
            where.append('received_at <= ?')
            params.append(date_to + ' 23:59:59')

        if branch_flt:
            ph = ','.join('?' * len(branch_flt))
            where.append(f'branch_id IN ({ph})')
            params.extend(branch_flt)

        if type_flt == 'delivery':
            where.append("order_type LIKE 'Доставка%'")
        elif type_flt == 'pickup':
            where.append("order_type = 'Общий - самовывоз'")

        sql_where = ' AND '.join(where) if where else '1=1'

        rows = conn.execute(f'''
            SELECT *,
                CASE
                    WHEN (order_type LIKE 'Доставка%' OR order_type = 'Общий - самовывоз')
                         AND (promised_minutes IS NULL OR promised_minutes = 0) THEN 'Предварит.'
                    WHEN (order_type LIKE 'Доставка%' OR order_type = 'Общий - самовывоз') THEN 'Текущий'
                    ELSE NULL
                END AS order_status
            FROM orders_report
            WHERE {sql_where}
            ORDER BY received_at DESC
            LIMIT 20000
        ''', params).fetchall()

        stats = conn.execute(f'''
            SELECT COUNT(*), COALESCE(SUM(amount),0),
                   SUM(CASE WHEN order_type = 'Общий - самовывоз' THEN 1 ELSE 0 END),
                   SUM(CASE WHEN order_type LIKE 'Доставка%' THEN 1 ELSE 0 END)
            FROM orders_report
            WHERE {sql_where}
        ''', params).fetchone()

    return render_template('orders_report.html',
        rows=rows,
        branches=branches,
        branch_groups=branch_groups,
        date_from=date_from, date_to=date_to,
        branch_flt=branch_flt, type_flt=type_flt, q=q,
        total_count=stats[0] or 0, total_amount=stats[1] or 0,
        pickup_count=stats[2] or 0, delivery_count=stats[3] or 0,
        data_min=data_min, data_max=data_max,
        row_limit_hit=(len(rows) >= 20000))


@app.route('/orders-report/import', methods=['GET', 'POST'])
@login_required
@menu_permission_required('orders_report')
def orders_report_import():
    import base64, hashlib

    def _row_hash(r):
        # Идентификатор заказа для повторной загрузки — дата приёма + номер заказа.
        # Совпадение по этому ключу не создаёт дубль, а обновляет уже загруженную строку.
        key = f"{r['order_number']}|{r['received_at'][:10]}"
        return hashlib.md5(key.encode('utf-8')).hexdigest()

    def _batches():
        with get_db() as conn:
            return conn.execute(
                'SELECT * FROM orders_import_batches ORDER BY imported_at DESC LIMIT 50'
            ).fetchall()

    if request.method == 'GET':
        return render_template('orders_report_import.html', batches=_batches())

    # Повторная отправка после сопоставления подразделений прямо на этой странице —
    # файлы уже разобраны один раз и закодированы в форме, заново их не запрашиваем.
    if request.form.get('resubmit') == '1':
        names = request.form.getlist('file_name')
        b64s  = request.form.getlist('file_b64')
        file_items = [(n, base64.b64decode(b)) for n, b in zip(names, b64s)]
        with get_db() as conn:
            valid_ids = {b['id'] for b in conn.execute('SELECT id FROM branches').fetchall()}
            for raw, bid in zip(request.form.getlist('map_raw'), request.form.getlist('map_branch_id')):
                if raw and bid and bid.isdigit() and int(bid) in valid_ids:
                    conn.execute(
                        'INSERT INTO branch_raw_map (branch_raw, branch_id) VALUES (?,?) '
                        'ON CONFLICT(branch_raw) DO UPDATE SET branch_id=excluded.branch_id',
                        (raw, int(bid))
                    )
                    conn.execute('UPDATE orders_report SET branch_id=? WHERE branch_raw=?', (int(bid), raw))
            conn.commit()
    else:
        files = [f for f in request.files.getlist('orders_file') if f and f.filename]
        if not files:
            flash('Выберите файл', 'danger')
            return redirect(url_for('orders_report_import'))
        file_items = [(f.filename, f.read()) for f in files]

    parsed = []
    for fname, fbytes in file_items:
        try:
            frows = _parse_orders_csv(fbytes)
        except Exception as e:
            flash(f'Ошибка чтения файла «{fname}»: {e}', 'danger')
            return redirect(url_for('orders_report_import'))
        if not frows:
            flash(f'В файле «{fname}» не найдено строк с заказами. Проверьте, что это CSV-выгрузка «Отчёт по заказам»', 'warning')
            continue
        parsed.append((fname, frows))

    if not parsed:
        return redirect(url_for('orders_report_import'))

    with get_db() as conn:
        branch_map = get_branch_raw_map(conn)
        all_raws = set()
        for _, frows in parsed:
            all_raws.update(r['branch_raw'] for r in frows)
        unmapped = sorted(all_raws - set(branch_map))

        if unmapped:
            branches = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
            pending_files = [
                {'name': fname, 'b64': base64.b64encode(fbytes).decode('ascii')}
                for fname, fbytes in file_items
            ]
            return render_template('orders_report_import.html', batches=_batches(),
                                    unmapped=unmapped, branches=branches, pending_files=pending_files)

        existing_keys = set(
            (er['order_number'], er['received_at'][:10])
            for er in conn.execute('SELECT order_number, received_at FROM orders_report').fetchall()
        )

        total_imported = total_updated = total_rows = 0
        for fname, frows in parsed:
            cur = conn.execute(
                'INSERT INTO orders_import_batches (filename, created_by) VALUES (?,?)',
                (fname, session.get('user_id'))
            )
            batch_id = cur.lastrowid

            imported = updated = 0
            for r in frows:
                key = (r['order_number'], r['received_at'][:10])
                is_new = key not in existing_keys
                existing_keys.add(key)
                h = _row_hash(r)
                conn.execute('''
                    INSERT INTO orders_report
                        (order_number, branch_raw, branch_id, received_at, promised_minutes,
                         order_type_raw, order_type, ready_minutes, delivery_minutes,
                         promo_code, amount, new_client, import_batch_id, import_hash)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    ON CONFLICT(import_hash) DO UPDATE SET
                        branch_raw=excluded.branch_raw, branch_id=excluded.branch_id,
                        received_at=excluded.received_at, promised_minutes=excluded.promised_minutes,
                        order_type_raw=excluded.order_type_raw, order_type=excluded.order_type,
                        ready_minutes=excluded.ready_minutes, delivery_minutes=excluded.delivery_minutes,
                        promo_code=excluded.promo_code, amount=excluded.amount,
                        new_client=excluded.new_client, import_batch_id=excluded.import_batch_id
                ''', (
                    r['order_number'], r['branch_raw'], branch_map.get(r['branch_raw']),
                    r['received_at'], r['promised_minutes'], r['order_type_raw'], r['order_type'],
                    r['ready_minutes'], r['delivery_minutes'], r['promo_code'], r['amount'],
                    r['new_client'], batch_id, h
                ))
                if is_new:
                    imported += 1
                else:
                    updated += 1

            conn.execute(
                'UPDATE orders_import_batches SET imported_count=?, updated_count=? WHERE id=?',
                (imported, updated, batch_id)
            )
            total_imported += imported
            total_updated  += updated
            total_rows     += len(frows)

        conn.commit()

    parts = [f'Импортировано {total_imported} новых заказов']
    if total_updated:
        parts.append(f'обновлено: {total_updated}')
    suffix = f' из {len(parsed)} файлов' if len(parsed) > 1 else ''
    flash(', '.join(parts) + f' (всего строк: {total_rows}{suffix})', 'success' if (total_imported or total_updated) else 'warning')
    return redirect(url_for('orders_report'))


@app.route('/orders-report/batch/<int:batch_id>/delete', methods=['POST'])
@login_required
@menu_permission_required('orders_report')
def orders_report_batch_delete(batch_id):
    with get_db() as conn:
        batch = conn.execute('SELECT * FROM orders_import_batches WHERE id=?', (batch_id,)).fetchone()
        if not batch:
            flash('Импорт не найден', 'danger')
            return redirect(url_for('orders_report_import'))
        cnt = conn.execute(
            'SELECT COUNT(*) FROM orders_report WHERE import_batch_id=?', (batch_id,)
        ).fetchone()[0]
        conn.execute('DELETE FROM orders_report WHERE import_batch_id=?', (batch_id,))
        conn.execute('DELETE FROM orders_import_batches WHERE id=?', (batch_id,))
        conn.commit()
    flash(f'Импорт «{batch["filename"]}» удалён ({cnt} заказов)', 'success')
    return redirect(url_for('orders_report_import'))


# ─── CHANGE (РАЗМЕН) SETTINGS ────────────────────────────────────────────────

@app.route('/settings/change')
@login_required
@menu_permission_required('change_settings')
def change_settings():
    WD_LABELS = ['Пн','Вт','Ср','Чт','Пт','Сб','Вс']
    today_str = date.today().isoformat()
    date_from = request.args.get('date_from', (date.today() - timedelta(days=6)).isoformat())
    date_to   = request.args.get('date_to',   date.today().isoformat())
    branch_ids_filter = request.args.getlist('branch_ids')

    # Build full date range list first
    dates_list = []
    try:
        d_cur = date.fromisoformat(date_from)
        d_end = date.fromisoformat(date_to)
        while d_cur <= d_end:
            wd = d_cur.weekday()
            dates_list.append({'iso': d_cur.isoformat(), 'weekday': wd,
                                'wd_label': WD_LABELS[wd], 'is_weekend': wd >= 5})
            d_cur += timedelta(days=1)
        dates_list.reverse()
    except Exception:
        pass

    with get_db() as conn:
        branches = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
        branch_groups_raw = conn.execute('SELECT * FROM branch_groups ORDER BY name').fetchall()
        branch_group_members = {}
        for m in conn.execute('SELECT * FROM branch_group_members').fetchall():
            branch_group_members.setdefault(m['group_id'], []).append(m['branch_id'])
        q_branches = branch_ids_filter if branch_ids_filter else [str(b['id']) for b in branches]
        q_branch_ids = [int(x) for x in q_branches]
        sel_branches = [b for b in branches if b['id'] in q_branch_ids]
        placeholders = ','.join('?' * len(q_branches))
        rows = conn.execute(f'''
            SELECT s.id as shift_id, s.date, b.id as branch_id,
                   s.status, COALESCE(r.change_amount, 0) as change_amount
            FROM shifts s
            JOIN branches b ON b.id = s.branch_id
            LEFT JOIN shift_revenue r ON r.shift_id = s.id
            WHERE s.date BETWEEN ? AND ?
              AND s.branch_id IN ({placeholders})
            ORDER BY s.date DESC, b.name
        ''', [date_from, date_to] + list(q_branches)).fetchall()
        schedules = conn.execute('''
            SELECT cs.*, b.name AS branch_name FROM change_schedule cs
            LEFT JOIN branches b ON b.id = cs.branch_id ORDER BY cs.id DESC
        ''').fetchall()

        # Build cross-tab grid
        grid = {}
        for r in rows:
            grid.setdefault(r['date'], {})[r['branch_id']] = {
                'shift_id': r['shift_id'], 'change_amount': r['change_amount'], 'status': r['status'],
            }

        # Add future "no-shift" cells (today and forward) with override or schedule amounts
        future_days = [d for d in dates_list if d['iso'] >= today_str]
        if future_days and sel_branches:
            b_ids = [b['id'] for b in sel_branches]
            fd_isos = [d['iso'] for d in future_days]
            ph_fd = ','.join('?' * len(fd_isos))
            ph_b  = ','.join('?' * len(b_ids))
            ov_rows = conn.execute(
                f'SELECT branch_id, date, amount FROM change_date_overrides WHERE date IN ({ph_fd}) AND branch_id IN ({ph_b})',
                fd_isos + b_ids
            ).fetchall()
            overrides = {(o['branch_id'], o['date']): o['amount'] for o in ov_rows}

            for day in future_days:
                day_iso, weekday = day['iso'], day['weekday']
                for b in sel_branches:
                    if b['id'] not in grid.get(day_iso, {}):
                        amount = overrides.get((b['id'], day_iso))
                        if amount is None:
                            srow = conn.execute(
                                '''SELECT amount FROM change_schedule
                                   WHERE (branch_id IS NULL OR branch_id=?)
                                     AND (weekday IS NULL OR weekday=?)
                                     AND valid_from <= ? AND (valid_to IS NULL OR valid_to >= ?)
                                   ORDER BY branch_id DESC NULLS LAST, weekday DESC NULLS LAST, id DESC
                                   LIMIT 1''',
                                (b['id'], weekday, day_iso, day_iso)
                            ).fetchone()
                            amount = srow['amount'] if srow else 0
                        grid.setdefault(day_iso, {})[b['id']] = {
                            'shift_id': None, 'status': 'future',
                            'change_amount': amount, 'branch_id': b['id'], 'date': day_iso,
                        }

    # Branch totals — real shifts only
    branch_totals = {}
    for cells in grid.values():
        for bid, cell in cells.items():
            if cell.get('status') != 'future':
                branch_totals[bid] = branch_totals.get(bid, 0) + (cell['change_amount'] or 0)

    return render_template('change_settings.html',
        branches=branches, branch_groups=branch_groups_raw,
        branch_group_members=branch_group_members,
        sel_branches=sel_branches, grid=grid,
        branch_totals=branch_totals, dates_list=dates_list,
        schedules=schedules,
        date_from=date_from, date_to=date_to,
        branch_ids_filter=branch_ids_filter,
        weekday_labels=WD_LABELS,
        today=today_str)


@app.route('/settings/change/manual/save', methods=['POST'])
@login_required
@menu_permission_required('change_settings')
def change_manual_save():
    data = request.json or {}
    with get_db() as conn:
        for shift_id_str, amount_str in data.items():
            try:
                sid = int(shift_id_str)
                amt = float(str(amount_str).replace(' ', '').replace(',', '.') or 0)
            except (ValueError, TypeError):
                continue
            shift = conn.execute('SELECT id FROM shifts WHERE id=?', (sid,)).fetchone()
            if not shift:
                continue
            conn.execute('UPDATE shift_revenue SET change_amount=? WHERE shift_id=?', (amt, sid))
        conn.commit()
    return jsonify({'ok': True})


@app.route('/settings/change/future/save', methods=['POST'])
@login_required
@menu_permission_required('change_settings')
def change_future_save():
    data = request.json or {}
    try:
        branch_id = int(data['branch_id'])
        date_str  = str(data['date'])
        amount    = float(str(data.get('amount', 0)).replace(' ', '').replace(',', '.') or 0)
    except (KeyError, ValueError, TypeError):
        return jsonify({'ok': False}), 400
    with get_db() as conn:
        conn.execute('''
            INSERT OR REPLACE INTO change_date_overrides (branch_id, date, amount)
            VALUES (?, ?, ?)
        ''', (branch_id, date_str, amount))
        conn.commit()
    return jsonify({'ok': True})


@app.route('/settings/change/schedule/add', methods=['POST'])
@login_required
@menu_permission_required('change_settings')
def change_schedule_add():
    branch_id  = request.form.get('branch_id') or None
    weekday    = request.form.get('weekday')
    weekday    = int(weekday) if weekday and weekday.strip() else None
    amount     = float(request.form.get('amount') or 0)
    valid_from = request.form.get('valid_from')
    valid_to   = request.form.get('valid_to') or None
    label      = request.form.get('label', '').strip()
    if valid_from and valid_from < date.today().isoformat():
        flash('Дата начала не может быть в прошлом', 'danger')
        return redirect(url_for('change_settings') + '#tab-schedule')
    with get_db() as conn:
        conn.execute(
            'INSERT INTO change_schedule (branch_id, weekday, amount, valid_from, valid_to, label) VALUES (?,?,?,?,?,?)',
            (branch_id, weekday, amount, valid_from, valid_to, label)
        )
        _apply_all_change_schedules(conn)
        conn.commit()
    flash('Расписание сохранено и применено к сменам', 'success')
    return redirect(url_for('change_settings') + '#tab-schedule')


@app.route('/settings/change/schedule/delete/<int:schedule_id>', methods=['POST'])
@login_required
@menu_permission_required('change_settings')
def change_schedule_delete(schedule_id):
    with get_db() as conn:
        conn.execute('DELETE FROM change_schedule WHERE id=?', (schedule_id,))
        conn.commit()
    flash('Расписание удалено', 'success')
    return redirect(url_for('change_settings') + '#tab-schedule')


@app.route('/settings/change/schedule/apply', methods=['POST'])
@login_required
@menu_permission_required('change_settings')
def change_schedule_apply():
    with get_db() as conn:
        _apply_all_change_schedules(conn)
        conn.commit()
    flash('Расписание применено ко всем подходящим сменам', 'success')
    return redirect(url_for('change_settings') + '#tab-schedule')


# ─── ЛИСТОВКА В ЗАКАЗ (промокод) ─────────────────────────────────────────────

@app.route('/settings/flyer-promo')
@login_required
@menu_permission_required('flyer_promo_settings')
def flyer_promo_settings():
    date_from = request.args.get('date_from', date.today().isoformat())
    date_to   = request.args.get('date_to', (date.today() + timedelta(days=13)).isoformat())
    branch_ids_filter = request.args.getlist('branch_ids')

    WD_LABELS = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
    dates_list = []
    try:
        d_cur = date.fromisoformat(date_from)
        d_end = date.fromisoformat(date_to)
        while d_cur <= d_end:
            wd = d_cur.weekday()
            dates_list.append({'iso': d_cur.isoformat(), 'weekday': wd, 'wd_label': WD_LABELS[wd]})
            d_cur += timedelta(days=1)
    except Exception:
        pass

    with get_db() as conn:
        branches = conn.execute('SELECT * FROM branches WHERE is_active=1 ORDER BY name').fetchall()
        branch_groups = get_branch_groups(conn)
        q_branch_ids = [int(x) for x in branch_ids_filter] if branch_ids_filter else [b['id'] for b in branches]
        sel_branches = [b for b in branches if b['id'] in q_branch_ids]

        grid = {}
        if dates_list and sel_branches:
            ph_b = ','.join('?' * len(q_branch_ids))
            rows = conn.execute(f'''
                SELECT branch_id, date, code FROM flyer_promocodes
                WHERE date BETWEEN ? AND ? AND branch_id IN ({ph_b})
            ''', [date_from, date_to] + q_branch_ids).fetchall()
            for r in rows:
                grid.setdefault(r['date'], {})[r['branch_id']] = r['code']

    return render_template('flyer_promo.html',
        branches=branches, branch_groups=branch_groups, branch_ids_filter=branch_ids_filter,
        sel_branches=sel_branches, dates_list=dates_list, grid=grid,
        date_from=date_from, date_to=date_to)


@app.route('/settings/flyer-promo/save', methods=['POST'])
@login_required
@menu_permission_required('flyer_promo_settings')
def flyer_promo_save():
    data = request.json or {}
    try:
        branch_id = int(data['branch_id'])
        date_str  = str(data['date'])
        code      = str(data.get('code', '')).strip()
    except (KeyError, ValueError, TypeError):
        return jsonify({'ok': False}), 400
    with get_db() as conn:
        if code:
            conn.execute('''
                INSERT INTO flyer_promocodes (branch_id, date, code) VALUES (?, ?, ?)
                ON CONFLICT(branch_id, date) DO UPDATE SET code=excluded.code
            ''', (branch_id, date_str, code))
        else:
            conn.execute('DELETE FROM flyer_promocodes WHERE branch_id=? AND date=?', (branch_id, date_str))
        conn.commit()
    return jsonify({'ok': True})


# ─── КОЛЛ-ЦЕНТР ───────────────────────────────────────────────────────────────
# Полностью отдельный набор таблиц (call_center_*) — операторы не привязаны
# к филиалу и намеренно не пересекаются с обычными сотрудниками/сменами.

def _cc_effective_rate(conn, employee_id, on_date, fallback_rate=0):
    row = conn.execute(
        'SELECT rate FROM call_center_rate_history WHERE employee_id=? AND effective_from<=? '
        'ORDER BY effective_from DESC, id DESC LIMIT 1',
        (employee_id, on_date)
    ).fetchone()
    return float(row['rate']) if row else float(fallback_rate or 0)


def _cc_duration_hours(start, end):
    """Часы между plановыми start/end (HH:MM), напр. 10:00-22:00 -> 12.0."""
    try:
        sh, sm = [int(x) for x in str(start).split(':')]
        eh, em = [int(x) for x in str(end).split(':')]
        mins = (eh * 60 + em) - (sh * 60 + sm)
        if mins < 0:
            mins += 24 * 60
        return round(mins / 60, 2)
    except Exception:
        return 0.0


def _cc_snap_quarter(hhmm, fallback='10:00'):
    """Округляет время до ближайших 15 минут (00/15/30/45)."""
    try:
        h, m = [int(x) for x in str(hhmm).split(':')]
        total = round((h * 60 + m) / 15) * 15
        total = max(0, min(23 * 60 + 45, total))
        return f'{total // 60:02d}:{total % 60:02d}'
    except Exception:
        return fallback


@app.route('/call-center')
@login_required
@menu_permission_required('call_center')
def call_center():
    WD_LABELS = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
    _cc_tabs = ('schedule', 'shifts', 'employees')
    active_tab = request.args.get('tab', 'schedule')
    if active_tab not in _cc_tabs or not subitem_visible('call_center', f'call_center_{active_tab}'):
        # запрошенная (или дефолтная) вкладка недоступна этой роли — переключаем
        # на первую, на которую доступ есть (полный доступ к разделу уже
        # разрешает все — см. subitem_visible)
        active_tab = next((t for t in _cc_tabs if subitem_visible('call_center', f'call_center_{t}')), 'schedule')
    today = date.today()
    default_month = today.replace(day=1)
    month_str = request.args.get('month', default_month.strftime('%Y-%m'))
    try:
        y, m = [int(x) for x in month_str.split('-')]
        month_start = date(y, m, 1)
    except Exception:
        month_start = default_month
    month_str = month_start.strftime('%Y-%m')
    days_in_month = calendar.monthrange(month_start.year, month_start.month)[1]
    month_end = month_start.replace(day=days_in_month)
    prev_month = (month_start - timedelta(days=1)).replace(day=1).strftime('%Y-%m')
    next_month = (month_end + timedelta(days=1)).strftime('%Y-%m')

    days_list = []
    d_cur = month_start
    while d_cur <= month_end:
        wd = d_cur.weekday()
        days_list.append({
            'iso': d_cur.isoformat(), 'day': d_cur.day,
            'wd_label': WD_LABELS[wd], 'is_weekend': wd >= 5,
            'default_end': '23:00' if wd in (4, 5) else '22:00',  # Пт и Сб — до 23:00
        })
        d_cur += timedelta(days=1)

    with get_db() as conn:
        operators = conn.execute(
            'SELECT * FROM call_center_employees WHERE is_fired=0 ORDER BY full_name'
        ).fetchall()
        fired_operators = conn.execute(
            'SELECT * FROM call_center_employees WHERE is_fired=1 ORDER BY fired_at DESC'
        ).fetchall()

        rate_history = {}
        for op in list(operators) + list(fired_operators):
            rate_history[op['id']] = conn.execute(
                'SELECT * FROM call_center_rate_history WHERE employee_id=? ORDER BY effective_from DESC LIMIT 5',
                (op['id'],)
            ).fetchall()

        schedule_grid = {}
        month_totals = {}

        if active_tab in ('schedule', 'shifts'):
            for r in conn.execute(
                'SELECT employee_id, date, planned_start, planned_end FROM call_center_schedule WHERE date BETWEEN ? AND ?',
                (month_start.isoformat(), month_end.isoformat())
            ).fetchall():
                schedule_grid.setdefault(r['employee_id'], {})[r['date']] = {
                    'start': r['planned_start'] or '10:00', 'end': r['planned_end'] or '22:00',
                }

        if active_tab == 'shifts':
            today_str = today.isoformat()
            rate_fallback = {op['id']: op['rate'] for op in operators}
            for eid, dates in schedule_grid.items():
                for diso, cell in dates.items():
                    if diso > today_str:
                        continue
                    hours = _cc_duration_hours(cell['start'], cell['end'])
                    rate = _cc_effective_rate(conn, eid, diso, rate_fallback.get(eid, 0))
                    pay = round(hours * rate, 2)
                    cell['hours'] = hours
                    cell['pay'] = pay
                    t = month_totals.setdefault(eid, {'hours': 0.0, 'pay': 0.0})
                    t['hours'] += hours
                    t['pay'] += pay

    grand_total_hours = sum(t['hours'] for t in month_totals.values())
    grand_total_pay = sum(t['pay'] for t in month_totals.values())

    return render_template(
        'call_center.html',
        active_tab=active_tab,
        operators=operators, fired_operators=fired_operators,
        rate_history=rate_history,
        month_str=month_str, prev_month=prev_month, next_month=next_month,
        days_list=days_list,
        schedule_grid=schedule_grid,
        month_totals=month_totals,
        grand_total_hours=grand_total_hours, grand_total_pay=grand_total_pay,
        today=today.isoformat(),
        is_owner=(session.get('role') == 'owner'),
    )


@app.route('/call-center/add', methods=['POST'])
@login_required
@menu_permission_required('call_center')
def call_center_add():
    full_name = request.form.get('full_name', '').strip()
    rate = float(request.form.get('rate', 0) or 0)
    effective_from = request.form.get('effective_from') or date.today().isoformat()
    if not full_name:
        flash('Введите имя оператора', 'danger')
        return redirect(url_for('call_center'))
    with get_db() as conn:
        conn.execute('INSERT INTO call_center_employees (full_name, rate) VALUES (?,?)', (full_name, rate))
        emp_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        conn.execute(
            'INSERT INTO call_center_rate_history (employee_id, rate, effective_from) VALUES (?,?,?)',
            (emp_id, rate, effective_from)
        )
        conn.commit()
    flash(f'Оператор {full_name} добавлен', 'success')
    return redirect(url_for('call_center'))


@app.route('/call-center/<int:emp_id>/update-rate', methods=['POST'])
@login_required
@menu_permission_required('call_center')
def call_center_update_rate(emp_id):
    rate = float(request.form.get('rate', 0) or 0)
    effective_from = request.form.get('effective_from') or date.today().isoformat()
    with get_db() as conn:
        emp = conn.execute('SELECT * FROM call_center_employees WHERE id=?', (emp_id,)).fetchone()
        if not emp:
            flash('Оператор не найден', 'danger')
            return redirect(url_for('call_center'))
        conn.execute(
            'INSERT INTO call_center_rate_history (employee_id, rate, effective_from) VALUES (?,?,?)',
            (emp_id, rate, effective_from)
        )
        if effective_from <= date.today().isoformat():
            conn.execute('UPDATE call_center_employees SET rate=? WHERE id=?', (rate, emp_id))
        conn.commit()
    flash(f'Ставка сохранена (с {effective_from})', 'success')
    return redirect(url_for('call_center'))


@app.route('/call-center/<int:emp_id>/fire', methods=['POST'])
@login_required
@menu_permission_required('call_center')
def call_center_fire(emp_id):
    comment = request.form.get('comment', '').strip()
    with get_db() as conn:
        conn.execute(
            'UPDATE call_center_employees SET is_fired=1, fired_at=CURRENT_TIMESTAMP, fired_comment=?, is_active=0 WHERE id=?',
            (comment, emp_id)
        )
        conn.commit()
    flash('Оператор уволен', 'warning')
    return redirect(url_for('call_center'))


@app.route('/call-center/<int:emp_id>/restore', methods=['POST'])
@login_required
@menu_permission_required('call_center')
def call_center_restore(emp_id):
    with get_db() as conn:
        conn.execute(
            'UPDATE call_center_employees SET is_fired=0, fired_at=NULL, fired_comment=NULL, is_active=1 WHERE id=?',
            (emp_id,)
        )
        conn.commit()
    flash('Оператор восстановлен', 'success')
    return redirect(url_for('call_center'))


_CC_TIME_RE = re.compile(r'^\d{2}:\d{2}$')


@app.route('/call-center/schedule/save', methods=['POST'])
@login_required
@menu_permission_required('call_center')
def call_center_schedule_save():
    """Автосохранение одной ячейки графика (день/оператор) — вызывается сразу
    при изменении галочки или времени, без отдельной кнопки «Сохранить»."""
    data = request.json or {}
    try:
        eid = int(data['employee_id'])
        d = str(data['date'])
    except (KeyError, ValueError, TypeError):
        return jsonify({'ok': False}), 400
    working = bool(data.get('working', True))
    with get_db() as conn:
        op = conn.execute('SELECT id FROM call_center_employees WHERE id=? AND is_fired=0', (eid,)).fetchone()
        if not op:
            return jsonify({'ok': False}), 400
        if not working:
            conn.execute('DELETE FROM call_center_schedule WHERE employee_id=? AND date=?', (eid, d))
        else:
            start = str(data.get('start') or '10:00')
            end = str(data.get('end') or '22:00')
            start = _cc_snap_quarter(start, '10:00') if _CC_TIME_RE.match(start) else '10:00'
            end = _cc_snap_quarter(end, '22:00') if _CC_TIME_RE.match(end) else '22:00'
            conn.execute('''
                INSERT INTO call_center_schedule (employee_id, date, planned_start, planned_end) VALUES (?,?,?,?)
                ON CONFLICT(employee_id, date) DO UPDATE SET planned_start=excluded.planned_start, planned_end=excluded.planned_end
            ''', (eid, d, start, end))
        conn.commit()
    return jsonify({'ok': True})


# ─── ОТЧЁТ «КОНТАКТ-ЦЕНТР» ─────────────────────────────────────────────────────
# ФОТ операторов колл-центра (из call_center_schedule/rate_history) + расходы по
# контрагентам, отнесённым к категории «Контакт-центр» в Банк → Контрагенты.
# Одна общая таблица по месяцам: слева список (операторы + контрагенты), в
# колонках суммы по месяцам, внизу — общий итог по контакт-центру.

CONTACT_CENTER_CAT_CODE = 'contact_center'
CONTACT_CENTER_CAT_LABEL = 'Контакт-центр'

# Цвета для строк групп филиалов в отчёте «Контакт-центр» (group_bg — строка группы,
# branch_bg — вложенные строки филиалов, accent — левая полоса-акцент).
GROUP_COLOR_PALETTE = [
    {'group_bg': '#dbe9ff', 'branch_bg': '#eef4ff', 'accent': '#0d6efd'},
    {'group_bg': '#ffe8cc', 'branch_bg': '#fff4e5', 'accent': '#fd7e14'},
    {'group_bg': '#d2f4e0', 'branch_bg': '#eafaf1', 'accent': '#198754'},
    {'group_bg': '#fbd9e4', 'branch_bg': '#fdeef2', 'accent': '#d63384'},
    {'group_bg': '#e5d9fa', 'branch_bg': '#f3eefc', 'accent': '#6f42c1'},
    {'group_bg': '#cdf1f7', 'branch_bg': '#e7f6f8', 'accent': '#0891a8'},
]


def _ensure_contact_center_category(conn):
    """Находит категорию «Контакт-центр» по названию (код мог быть создан вручную
    через Настройки → Категории расходов и не совпадать с CONTACT_CENTER_CAT_CODE).
    Если найдено несколько одноимённых дублей — оставляет тот, на который уже
    назначены контрагенты, остальные (пустые) удаляет. Возвращает актуальный code."""
    rows = conn.execute(
        "SELECT id, code FROM expense_categories WHERE LOWER(label)=LOWER(?) ORDER BY id",
        (CONTACT_CENTER_CAT_LABEL,)
    ).fetchall()
    if not rows:
        max_sort = conn.execute('SELECT COALESCE(MAX(sort_order),0) FROM expense_categories').fetchone()[0]
        conn.execute(
            'INSERT INTO expense_categories (code, label, type, sort_order, show_contractors, show_shift) '
            'VALUES (?,?,?,?,?,?)',
            (CONTACT_CENTER_CAT_CODE, CONTACT_CENTER_CAT_LABEL, 'expense', max_sort + 1, 1, 0)
        )
        conn.commit()
        return CONTACT_CENTER_CAT_CODE
    if len(rows) == 1:
        return rows[0]['code']
    keep_code = None
    for r in rows:
        cnt = conn.execute('SELECT COUNT(*) FROM contractors WHERE category=?', (r['code'],)).fetchone()[0]
        if cnt > 0:
            keep_code = r['code']
            break
    keep_code = keep_code or rows[0]['code']
    for r in rows:
        if r['code'] != keep_code:
            conn.execute('DELETE FROM expense_categories WHERE id=?', (r['id'],))
    conn.commit()
    return keep_code


def _month_range(date_from, date_to):
    y, m = int(date_from[:4]), int(date_from[5:7])
    y1, m1 = int(date_to[:4]), int(date_to[5:7])
    months = []
    while (y, m) <= (y1, m1):
        months.append(f'{y:04d}-{m:02d}')
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months


@app.route('/report/contact-center')
@login_required
@menu_permission_required('contact_center_report')
def contact_center_report():
    today = date.today()
    # По умолчанию — последние 3 месяца, включая текущий
    default_start_m = today.month - 2
    default_start_y = today.year
    if default_start_m <= 0:
        default_start_m += 12
        default_start_y -= 1
    default_from = date(default_start_y, default_start_m, 1).isoformat()
    date_from = request.args.get('date_from', default_from)
    date_to   = request.args.get('date_to', today.isoformat())
    months = _month_range(date_from, date_to)

    with get_db() as conn:
        cc_cat_code = _ensure_contact_center_category(conn)

        # ФОТ операторов колл-центра по месяцам (часы по графику × актуальная ставка,
        # только за дни не позже сегодня — как в /call-center)
        operators = conn.execute(
            "SELECT id, full_name, rate, is_fired FROM call_center_employees ORDER BY is_fired, full_name"
        ).fetchall()
        rate_fallback = {op['id']: op['rate'] for op in operators}

        op_pay = {}  # employee_id -> {month: pay}
        sched_to = min(date_to, today.isoformat())
        if date_from <= sched_to:
            for r in conn.execute(
                'SELECT employee_id, date, planned_start, planned_end FROM call_center_schedule '
                'WHERE date BETWEEN ? AND ?',
                (date_from, sched_to)
            ).fetchall():
                hours = _cc_duration_hours(r['planned_start'] or '10:00', r['planned_end'] or '22:00')
                rate = _cc_effective_rate(conn, r['employee_id'], r['date'], rate_fallback.get(r['employee_id'], 0))
                month = r['date'][:7]
                by_month = op_pay.setdefault(r['employee_id'], {})
                by_month[month] = by_month.get(month, 0.0) + round(hours * rate, 2)

        # Расходы по контрагентам категории «Контакт-центр»
        contractors = conn.execute(
            "SELECT id, name FROM contractors WHERE category=? AND is_active=1 ORDER BY name",
            (cc_cat_code,)
        ).fetchall()

        ctr_pay = {}  # contractor_id -> {month: amount}
        for r in conn.execute('''
            SELECT bt.contractor_id AS cid, strftime('%Y-%m', bt.txn_date) AS month,
                   SUM(-bt.amount) AS amt
            FROM bank_transactions bt
            JOIN contractors c ON c.id = bt.contractor_id
            WHERE c.category=? AND bt.amount<0 AND bt.is_ignored=0
              AND bt.txn_date BETWEEN ? AND ?
            GROUP BY bt.contractor_id, month
        ''', (cc_cat_code, date_from, date_to)).fetchall():
            ctr_pay.setdefault(r['cid'], {})[r['month']] = r['amt']

        # Выручка по филиалам ПО МЕСЯЦАМ за период — база для распределения расходов
        # на контакт-центр (доля филиала в выручке месяца применяется к итогу
        # контакт-центра за тот же месяц).
        branch_groups = get_branch_groups(conn)
        branch_names = {
            b['id']: b['name']
            for b in conn.execute("SELECT id, name FROM branches WHERE is_active=1 ORDER BY name").fetchall()
        }
        branch_rev_by_month = {}  # branch_id -> {month: revenue}
        for r in conn.execute('''
            SELECT s.branch_id AS bid, strftime('%Y-%m', s.date) AS month,
                   COALESCE(SUM(sr.total_revenue),0) as revenue
            FROM shifts s
            LEFT JOIN shift_revenue sr ON sr.shift_id = s.id
            WHERE s.date BETWEEN ? AND ?
            GROUP BY s.branch_id, month
        ''', (date_from, date_to)).fetchall():
            branch_rev_by_month.setdefault(r['bid'], {})[r['month']] = r['revenue']

        # Там, где нет смены с выручкой — берём из «Старой выручки» (revenue_manual),
        # тот же fallback, что и в остальных отчётах (см. _manual_rev_by_month).
        for r in conn.execute('''
            SELECT m.branch_id AS bid, strftime('%Y-%m', m.date) AS month,
                   COALESCE(SUM(m.amount),0) as revenue
            FROM revenue_manual m
            WHERE m.date BETWEEN ? AND ?
              AND NOT EXISTS (
                  SELECT 1 FROM shifts s JOIN shift_revenue sr ON sr.shift_id=s.id
                  WHERE s.date=m.date AND s.branch_id=m.branch_id AND (sr.total_revenue > 0 OR s.status='closed')
              )
            GROUP BY m.branch_id, month
        ''', (date_from, date_to)).fetchall():
            by_month = branch_rev_by_month.setdefault(r['bid'], {})
            by_month[r['month']] = by_month.get(r['month'], 0.0) + r['revenue']

    def _row(label, by_month, extra=None):
        r = {'label': label, 'amounts': {}, 'total': 0.0}
        if extra:
            r.update(extra)
        for mo in months:
            v = by_month.get(mo, 0.0)
            r['amounts'][mo] = v
            r['total'] += v
        return r

    op_rows = [
        _row(op['full_name'], op_pay.get(op['id'], {}), {'fired': bool(op['is_fired'])})
        for op in operators
        if not op['is_fired'] or op_pay.get(op['id'])
    ]
    ctr_rows = [_row(c['name'], ctr_pay.get(c['id'], {})) for c in contractors]

    def _totals_row(label, rows):
        by_month = {}
        for row in rows:
            for mo in months:
                by_month[mo] = by_month.get(mo, 0.0) + row['amounts'].get(mo, 0.0)
        return _row(label, by_month)

    fot_total_row = _totals_row('Итого ФОТ колл-центра', op_rows)
    exp_total_row = _totals_row('Итого расходы на контакт-центр', ctr_rows)
    grand_total_row = _totals_row('ИТОГО КОНТАКТ-ЦЕНТР', [fot_total_row, exp_total_row])

    month_labels = {mo: _pnl_period_label(mo) for mo in months}

    # Распределение итога контакт-центра по филиалам/группам: для каждого месяца
    # доля_филиала = выручка_филиала_за_месяц / выручка_всех_филиалов_за_месяц,
    # сумма = доля × итог контакт-центра за тот же месяц. Строки — в том же
    # формате {label, amounts{month:val}}, что и остальная таблица.
    total_rev_by_month = {}
    for mo in months:
        total_rev_by_month[mo] = sum(by_m.get(mo, 0.0) for by_m in branch_rev_by_month.values())

    def _alloc_by_month(rev_by_month):
        by_month = {}
        for mo in months:
            trm = total_rev_by_month.get(mo, 0.0)
            if trm:
                by_month[mo] = grand_total_row['amounts'].get(mo, 0.0) * (rev_by_month.get(mo, 0.0) / trm)
        return by_month

    grouped_branch_ids = set()
    group_alloc_rows = []
    for i, g in enumerate(branch_groups):
        member_ids = [bid for bid in g['branch_ids'] if bid in branch_names]
        grouped_branch_ids.update(member_ids)
        g_rev_by_month = {
            mo: sum(branch_rev_by_month.get(bid, {}).get(mo, 0.0) for bid in member_ids)
            for mo in months
        }
        branch_rows = [
            _row(branch_names[bid], _alloc_by_month(branch_rev_by_month.get(bid, {})))
            for bid in member_ids
        ]
        group_row = _row(g['name'], _alloc_by_month(g_rev_by_month))
        group_row['branches'] = branch_rows
        group_row['color'] = GROUP_COLOR_PALETTE[i % len(GROUP_COLOR_PALETTE)]
        group_alloc_rows.append(group_row)

    ungrouped_alloc_rows = sorted((
        _row(branch_names[bid], _alloc_by_month(branch_rev_by_month.get(bid, {})))
        for bid in branch_names if bid not in grouped_branch_ids
    ), key=lambda x: x['label'])

    # % от общей выручки всех групп/филиалов, которую «съедает» контакт-центр
    pct_of_revenue_row = _row('% от общей выручки', {
        mo: (grand_total_row['amounts'].get(mo, 0.0) / total_rev_by_month[mo] * 100)
            if total_rev_by_month.get(mo) else 0.0
        for mo in months
    })

    return render_template(
        'report_contact_center.html',
        months=months, month_labels=month_labels,
        op_rows=op_rows, ctr_rows=ctr_rows,
        fot_total_row=fot_total_row, exp_total_row=exp_total_row, grand_total_row=grand_total_row,
        pct_of_revenue_row=pct_of_revenue_row,
        date_from=date_from, date_to=date_to,
        group_alloc_rows=group_alloc_rows, ungrouped_alloc_rows=ungrouped_alloc_rows,
    )


# ──────────────────────────────────────────────────────────────────────────────

init_db()

# ─── АВТО-БЭКАП БАЗЫ КАЖДЫЙ ДЕНЬ В 03:00 ─────────────────────────────────────
def _scheduled_backup():
    try:
        import datetime as _dt
        ok, msg = _backup_db_to_gdrive()
        if ok:
            with get_db() as conn:
                conn.execute(
                    "INSERT OR REPLACE INTO gsheet_settings (key, value) VALUES ('last_db_backup', ?)",
                    (_dt.datetime.now().strftime('%d.%m.%Y %H:%M'),)
                )
                conn.commit()
        print(f'[Backup] {"OK" if ok else "ERR"}: {msg}')
    except Exception as e:
        print(f'[Backup] exception: {e}')

try:
    from apscheduler.schedulers.background import BackgroundScheduler
    _scheduler = BackgroundScheduler(timezone='Asia/Novosibirsk')
    _scheduler.add_job(_scheduled_backup, 'cron', hour=3, minute=0)
    _scheduler.start()
    print('[Backup] Планировщик запущен — бэкап каждый день в 03:00 НСК')
except Exception as _e:
    print(f'[Backup] Планировщик не запустился: {_e}')

if __name__ == '__main__':
    print('\n' + '=' * 50)
    print('CRM Суши запущена!')
    print('Откройте браузер: http://localhost:5050')
    print('Логин: owner | Пароль: admin123')
    print('=' * 50 + '\n')
    port = int(os.environ.get('PORT', 5050))
    app.run(host='0.0.0.0', port=port, debug=False, threaded=True)
