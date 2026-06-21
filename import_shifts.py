#!/usr/bin/env python3
import openpyxl
import sqlite3
from datetime import datetime, date, time as dt_time
import os

DB_PATH = "/Users/aleksey/Desktop/CRM sushi/crm/crm.db"
EXCEL_DIR = "/Users/aleksey/Desktop/CRM sushi/1"
MAX_DATE = date(2026, 6, 20)

DAY_SHEETS = {'ПН', 'ВТ', 'СР', 'ЧТ', 'ПТ', 'СБ', 'ВС'}

ROLE_MAP = {
    'Админ.': 'admin', 'Адм.': 'admin', 'Адм/Упак': 'admin',
    'Адм': 'admin', 'Администратор': 'admin',
    'Упак.': 'packer', 'Упак': 'packer', 'Упаковщик': 'packer',
    'Сушист': 'sushi', 'Сушист.': 'sushi',
    'Уборщица': 'cleaner', 'Уборщик': 'cleaner',
    'Повара': 'cook', 'Повар': 'cook', 'Повар.': 'cook',
}

CATEGORY_MAP = {
    'Ремонт сантех.': 'repair_plumbing',
    'Чистка жироулов-ля': 'repair_grease',
    'Чистка жироуловителя': 'repair_grease',
    'Ремонт электрик': 'repair_electric',
    'Ремонт холод.оборуд.': 'repair_fridge',
    'Ремонт другой': 'repair_other',
    'Магазин / Апт.': 'shop',
    'Магазин/Апт.': 'shop',
    'Другое': 'other',
}

SKIP_NAMES = {
    'Курьеры', 'Повара', 'Админ', 'Курьеры:', 'Повара:',
    'Курьер 1:', 'Курьер 2:', 'Курьер 3:', 'Курьер 4:', 'Курьер 5:', 'Курьер 6:',
    'ФАМИЛИЯ ИМЯ:', 'ФАМИЛИЯ:', 'Курьер 1', 'Курьер 2', 'Курьер 3',
    'Курьер 4', 'Курьер 5', 'Курьер 6',
}


def safe_float(val):
    if val is None or val is False or val == 'х' or val == 'x' or val == 'Х':
        return 0.0
    try:
        return float(val)
    except:
        return 0.0


def time_to_hours(t):
    if t is None:
        return 0.0
    if isinstance(t, dt_time):
        return t.hour + t.minute / 60.0
    try:
        return float(t)
    except:
        return 0.0


def time_to_str(t):
    if isinstance(t, dt_time):
        return f"{t.hour:02d}:{t.minute:02d}"
    return None


def get_branch_key(filename):
    name = filename.lower()
    if 'запс' in name:
        return 'запс'
    elif 'новка' in name:
        return 'новка'
    elif 'ильинка' in name:
        return 'ильинка'
    elif 'стр' in name:
        return 'стр'
    return None


def process_sheet(ws, branch_id, cur, conn, fname, sheet_name):
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 7:
        return

    # Date from row 3 (index 2), col B (index 1)
    date_val = rows[2][1] if len(rows) > 2 else None
    if isinstance(date_val, datetime):
        shift_date = date_val.date()
    elif isinstance(date_val, date):
        shift_date = date_val
    else:
        print(f"    [{sheet_name}] нет даты, пропускаем")
        return

    if shift_date > MAX_DATE:
        print(f"    [{sheet_name}] {shift_date} > 20 июня, пропускаем")
        return

    print(f"    [{sheet_name}] {shift_date}", end=" ")

    # ── REVENUE ──────────────────────────────────────────────
    total_revenue   = safe_float(rows[1][5]) if len(rows) > 1 else 0
    delivery_rev    = safe_float(rows[3][6]) if len(rows) > 3 else 0
    cash_amount     = safe_float(rows[4][3]) if len(rows) > 4 else 0
    delivery_ord    = int(safe_float(rows[4][6])) if len(rows) > 4 else 0
    card_amount     = safe_float(rows[5][3]) if len(rows) > 5 else 0
    pickup_rev      = safe_float(rows[5][6]) if len(rows) > 5 else 0
    online_amount   = safe_float(rows[6][3]) if len(rows) > 6 else 0
    pickup_ord      = int(safe_float(rows[6][6])) if len(rows) > 6 else 0

    # ── TERMINAL / ACTUAL CASH / CLOSED BY ───────────────────
    terminal_amount = 0.0
    terminal_codes  = []
    actual_cash     = 0.0
    closed_by_name  = None

    for r in rows[25:]:
        if r[4] == 'По терминалам:' and r[6] is not None:
            terminal_amount = safe_float(r[6])
        if r[1] == 'Факт в кассе:':
            actual_cash = safe_float(r[3])
        if r[1] == 'Смену закрыл(а):':
            closed_by_name = str(r[4]).strip() if r[4] else None

    # Terminal codes from rows 29–31 (indices 28–30)
    for r in rows[28:32]:
        code = r[4]
        amt  = r[6]
        if code is not None and safe_float(amt) > 0:
            c = str(code).strip()
            if c:
                terminal_codes.append(c)

    terminal_last3 = ','.join(terminal_codes)

    # ── SHIFT ─────────────────────────────────────────────────
    cur.execute("SELECT id FROM shifts WHERE branch_id=? AND date=?",
                (branch_id, shift_date.isoformat()))
    row = cur.fetchone()
    if row:
        shift_id = row[0]
        print(f"(смена уже есть id={shift_id})")
    else:
        status = 'closed' if closed_by_name else 'open'
        cur.execute("""
            INSERT INTO shifts (branch_id, date, status, closed_by_name)
            VALUES (?, ?, ?, ?)
        """, (branch_id, shift_date.isoformat(), status, closed_by_name))
        shift_id = cur.lastrowid
        print(f"→ смена id={shift_id}", end=" ")

    # ── SHIFT_REVENUE ─────────────────────────────────────────
    cur.execute("SELECT id FROM shift_revenue WHERE shift_id=?", (shift_id,))
    if not cur.fetchone():
        cur.execute("""
            INSERT INTO shift_revenue
            (shift_id, total_revenue, delivery_revenue, delivery_orders,
             pickup_revenue, pickup_orders, cash_amount, card_amount,
             online_amount, terminal_last3, terminal_amount, actual_cash)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (shift_id, total_revenue, delivery_rev, delivery_ord,
              pickup_rev, pickup_ord, cash_amount, card_amount,
              online_amount, terminal_last3, terminal_amount, actual_cash))
        print(f"выручка={total_revenue}")
    else:
        print()

    # ── EXPENSES: regular (rows 10–20, indices 9–19) ─────────
    current_category = None
    for r in rows[9:20]:
        cat_str = r[1]
        if cat_str and isinstance(cat_str, str) and cat_str in CATEGORY_MAP:
            current_category = CATEGORY_MAP[cat_str]

        if current_category is None:
            continue

        cash_e  = safe_float(r[5])
        card_e  = safe_float(r[6])
        desc    = str(r[2]).strip() if r[2] and str(r[2]).strip() else None
        gulash  = 1 if r[7] is True else 0

        if cash_e > 0 or card_e > 0:
            cur.execute("""
                INSERT INTO expenses
                (shift_id, category, description, amount_cash, amount_card, is_gulash)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (shift_id, current_category, desc, cash_e, card_e, gulash))

    # ── EXPENSES: taxi (rows 22–24, indices 21–23) ────────────
    for r in rows[21:24]:
        desc = str(r[2]).strip() if r[2] else None
        if not desc:
            continue
        if r[1] == 'ТАКСИ':
            continue
        # For TAXI header: col5='По карте:', col6='Нал:' — reversed vs regular
        cash_t = safe_float(r[6])
        card_t = safe_float(r[5])
        if cash_t > 0 or card_t > 0:
            cur.execute("""
                INSERT INTO expenses
                (shift_id, category, description, amount_cash, amount_card, is_gulash)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (shift_id, 'taxi', desc, cash_t, card_t, 0))

    # ── EMPLOYEES ─────────────────────────────────────────────
    def get_or_create_emp(name, role, rate):
        name = name.strip()
        cur.execute("SELECT id FROM employees WHERE full_name=? AND branch_id=?",
                    (name, branch_id))
        r = cur.fetchone()
        if r:
            return r[0]
        cur.execute("""
            INSERT INTO employees (branch_id, full_name, role, rate, is_active)
            VALUES (?, ?, ?, ?, 1)
        """, (branch_id, name, role, rate))
        eid = cur.lastrowid
        print(f"        + сотрудник: {name} ({role}, ставка={rate})")
        return eid

    def add_emp_shift(emp_id, name, role, rate,
                      hours=0, km=0, orders=0,
                      rate_km=10, rate_ord=100,
                      start=None, end=None,
                      bonus=0, penalty=0, comment='',
                      base_pay=0, total=0, paid=0):
        cur.execute("SELECT id FROM employee_shifts WHERE shift_id=? AND employee_id=?",
                    (shift_id, emp_id))
        if cur.fetchone():
            return
        cur.execute("""
            INSERT INTO employee_shifts
            (shift_id, employee_id, full_name_snapshot, role_snapshot,
             rate_snapshot, rate_per_km_snapshot, rate_per_order_snapshot,
             hours_worked, km, orders,
             shift_start, shift_end,
             bonus_amount, penalty_amount, bonus_comment,
             base_pay, total_amount, is_paid, paid_amount)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (shift_id, emp_id, name, role,
              rate, rate_km, rate_ord,
              hours, km, orders,
              start, end,
              bonus, penalty, comment,
              base_pay, total, paid, total if paid else 0))

    # Couriers — rows 3–8 (indices 2–7)
    for r in rows[2:8]:
        name = r[10]
        if not name or not isinstance(name, str) or name.strip() in SKIP_NAMES:
            continue
        name = name.strip()
        if not name:
            continue

        total = safe_float(r[21])
        if total == 0:
            continue

        km       = safe_float(r[12])
        km_pay   = safe_float(r[13])
        hours    = safe_float(r[14])
        hrs_pay  = safe_float(r[15])
        orders   = int(safe_float(r[16]))
        ord_pay  = safe_float(r[17])
        comment  = str(r[18]).strip() if r[18] and str(r[18]) != 'Ничего' else ''
        paid     = 1 if r[20] == 'Да' else 0

        rate_km  = round(km_pay / km, 2) if km > 0 else 10.0
        rate_ord = round(ord_pay / orders, 2) if orders > 0 else 100.0
        rate_hr  = round(hrs_pay / hours, 2) if hours > 0 else 0.0

        emp_id = get_or_create_emp(name, 'courier', rate_hr)
        add_emp_shift(emp_id, name, 'courier', rate_hr,
                      hours=hours, km=km, orders=orders,
                      rate_km=rate_km, rate_ord=rate_ord,
                      comment=comment,
                      base_pay=hrs_pay + km_pay,
                      total=total, paid=paid)

    # Non-couriers — rows 10–22 (indices 9–21)
    for r in rows[9:22]:
        name = r[10]
        if not name or not isinstance(name, str) or name.strip() in SKIP_NAMES:
            continue
        name = name.strip()
        if not name:
            continue

        total = safe_float(r[21])
        if total == 0:
            continue

        role_str = r[11]
        if name == 'Уборщица':
            role = 'cleaner'
        elif role_str and isinstance(role_str, str) and role_str.strip() in ROLE_MAP:
            role = ROLE_MAP[role_str.strip()]
        else:
            role = 'admin'

        rate   = safe_float(r[12])
        start  = time_to_str(r[13])
        end    = time_to_str(r[14])
        hours  = time_to_hours(r[15])
        bval   = r[17]
        bonus  = safe_float(bval) if isinstance(bval, (int, float)) and bval > 0 else 0
        penalty = abs(safe_float(bval)) if isinstance(bval, (int, float)) and bval < 0 else 0
        comment = str(r[18]).strip() if r[18] and str(r[18]) != 'Ничего' else ''
        paid   = 1 if r[20] == 'Да' else 0
        base   = round(rate * hours, 2)

        emp_id = get_or_create_emp(name, role, rate)
        add_emp_shift(emp_id, name, role, rate,
                      hours=hours, start=start, end=end,
                      bonus=bonus, penalty=penalty, comment=comment,
                      base_pay=base, total=total, paid=paid)

    conn.commit()


def main():
    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()

    # ── Add new branches ─────────────────────────────────────
    branch_map = {
        'стр':     1,
        'запс':    2,
        'новка':   None,
        'ильинка': None,
    }
    new_branches = {
        'новка':   'Новобайдаевская',
        'ильинка': 'Авиаторов',
    }
    for key, bname in new_branches.items():
        cur.execute("SELECT id FROM branches WHERE name=?", (bname,))
        r = cur.fetchone()
        if r:
            branch_map[key] = r[0]
        else:
            cur.execute("INSERT INTO branches (name, is_active) VALUES (?, 1)", (bname,))
            branch_map[key] = cur.lastrowid
            print(f"Создан филиал: {bname} (id={branch_map[key]})")
    conn.commit()

    # ── Process Excel files ───────────────────────────────────
    files = sorted(f for f in os.listdir(EXCEL_DIR) if f.endswith('.xlsx'))
    for fname in files:
        key = get_branch_key(fname)
        if key is None:
            print(f"Пропуск: {fname}")
            continue

        branch_id = branch_map[key]
        filepath  = os.path.join(EXCEL_DIR, fname)
        print(f"\n=== {fname} → филиал id={branch_id} ===")

        wb = openpyxl.load_workbook(filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            if sheet_name not in DAY_SHEETS:
                continue
            ws = wb[sheet_name]
            process_sheet(ws, branch_id, cur, conn, fname, sheet_name)

    conn.close()
    print("\n✓ Импорт завершён")


if __name__ == '__main__':
    main()
